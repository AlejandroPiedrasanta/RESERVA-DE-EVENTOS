# ═══════════════════════════════════════════════════════════════════════════
#  Cinema Productions — API (FastAPI)   ·   ÍNDICE / MAPA DE NAVEGACIÓN
# ═══════════════════════════════════════════════════════════════════════════
#  Consulta /app/ARCHITECTURE.md para el mapa completo del proyecto.
#  Busca los marcadores "# ───" para saltar entre secciones:
#
#   1.  Config & entorno .......... imports, env vars, Mongo client, scheduler
#   2.  Lifespan ................... arranque/paro del scheduler de recordatorios
#   3.  Pydantic Models ........... Reservation/Socio/NotificationSettings/DB
#   4.  Reminder Logic ............ HTML + envío (Gmail/Push/Telegram/ntfy)
#   5.  Routes ....................  root · deployment · data · import/export
#   6.  Backup .................... download/create/history/restore
#   7.  Stats / Reservations / Socios / Financials
#   8.  App Settings / Database Settings
#   9.  Reminders (trigger manual)
#  10.  Gmail OAuth2 / Web Push / Telegram / ntfy (endpoints de prueba)
#  11.  Desktop Package Download (plantillas en desktop_package.py)
#  12.  App Updates / Appearance sync / Saved Themes
#  13.  App Security (password lock + page protection)
#  14.  GitHub Integration & AI Context (DEFAULT_AI_CONTEXT en ai_context_default.py)
#
#  Módulos externos:
#   · desktop_package.py    → plantillas del instalador de escritorio (ZIP)
#   · ai_context_default.py → texto de contexto para la próxima IA
#   · standalone_app.py     → app de escritorio embebida (independiente)
# ═══════════════════════════════════════════════════════════════════════════

from fastapi import FastAPI, APIRouter, HTTPException, UploadFile, File, Form, Request, Body
from fastapi.responses import JSONResponse, Response, RedirectResponse, StreamingResponse
from contextlib import asynccontextmanager
from dotenv import load_dotenv
import csv
import io
import re
import asyncio
import httpx
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
import base64
import uuid
import hashlib
import secrets
from pathlib import Path
from pydantic import BaseModel, Field
from typing import List, Optional
from datetime import datetime, timezone, timedelta
from bson import ObjectId
import resend as resend_lib
import json
from apscheduler.schedulers.asyncio import AsyncIOScheduler
from apscheduler.triggers.cron import CronTrigger
from apscheduler.triggers.interval import IntervalTrigger

# ── Google / Gmail ────────────────────────────────────────────
from google_auth_oauthlib.flow import Flow
from google.oauth2.credentials import Credentials
from google.auth.transport.requests import Request as GoogleRequest
from googleapiclient.discovery import build
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart

# ── Web Push ──────────────────────────────────────────────────
from pywebpush import webpush, WebPushException


ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

CUSTOM_DB_FILE = ROOT_DIR / '.db_override'
DB_NAME = os.environ['DB_NAME']

# ── Backup directory ──────────────────────────────────────────
BACKUP_DIR = ROOT_DIR / "backups"
BACKUP_DIR.mkdir(exist_ok=True)

# ── Updates directory ─────────────────────────────────────────
UPDATES_DIR = ROOT_DIR / "uploads" / "updates"
UPDATES_DIR.mkdir(parents=True, exist_ok=True)
DESKTOP_WHEELS_DIR = ROOT_DIR / "desktop_wheels"

# ── Themes JSON backup (local file mirror of saved_themes collection) ────
THEMES_DIR = Path("/app/themes")
THEMES_DIR.mkdir(parents=True, exist_ok=True)
THEMES_JSON_PATH = THEMES_DIR / "saved_themes.json"


def _ensure_desktop_wheels():
    """Descarga (cacheado por hash de requirements) wheels win_amd64 para instalacion offline del escritorio."""
    import sys as _sys, subprocess as _sp, hashlib as _hl
    # Hash SHA-256 como clave de caché para invalidar wheels cuando cambia
    # requirements.txt (no es un uso de seguridad — reemplazado MD5 → SHA-256
    # para cumplir con la guía de código y evitar el flag del linter).
    req_hash = _hl.sha256(_REQUIREMENTS.encode("utf-8")).hexdigest()
    marker = DESKTOP_WHEELS_DIR / ".ok"
    if marker.exists() and marker.read_text().strip() == req_hash and any(DESKTOP_WHEELS_DIR.glob("*.whl")):
        return DESKTOP_WHEELS_DIR
    DESKTOP_WHEELS_DIR.mkdir(exist_ok=True)
    req = DESKTOP_WHEELS_DIR / "req.txt"
    req.write_text(_REQUIREMENTS, encoding="utf-8")
    got = False
    for ver in ("311", "312", "313"):
        try:
            _sp.run([_sys.executable, "-m", "pip", "download", "-r", str(req),
                     "--dest", str(DESKTOP_WHEELS_DIR), "--platform", "win_amd64",
                     "--python-version", ver, "--only-binary=:all:", "--implementation", "cp"],
                    check=True, timeout=300, stdout=_sp.PIPE, stderr=_sp.STDOUT)
            got = True
        except Exception as e:
            logger.warning(f"[desktop] wheel download py{ver} fallo: {e}")
    if got and any(DESKTOP_WHEELS_DIR.glob("*.whl")):
        marker.write_text(req_hash, encoding="utf-8")
    return DESKTOP_WHEELS_DIR

# ── Google / VAPID config ─────────────────────────────────────
# Public URL of this deployment (used for OAuth redirects). Configurable via env
# so it is never hardcoded to a specific preview domain.
APP_PUBLIC_URL       = os.environ.get('APP_PUBLIC_URL', '').rstrip('/')
GOOGLE_CLIENT_ID     = os.environ.get('GOOGLE_CLIENT_ID', '')
GOOGLE_CLIENT_SECRET = os.environ.get('GOOGLE_CLIENT_SECRET', '')
GOOGLE_REDIRECT_URI  = os.environ.get('GOOGLE_REDIRECT_URI') or (
    f"{APP_PUBLIC_URL}/api/oauth/gmail/callback" if APP_PUBLIC_URL else ''
)
GMAIL_SCOPES         = ['https://www.googleapis.com/auth/gmail.send', 'openid', 'https://www.googleapis.com/auth/userinfo.email']
VAPID_PUBLIC_KEY     = os.environ.get('VAPID_PUBLIC_KEY', '')
VAPID_PRIVATE_KEY    = os.environ.get('VAPID_PRIVATE_KEY', '')
VAPID_EMAIL          = os.environ.get('VAPID_EMAIL', 'mailto:admin@cinema-productions.com')

# In-memory dedup: avoid sending the same reminder twice in one day
_sent_push_today: set = set()

# Active MongoDB URL: prefer custom override file
_mongo_url = CUSTOM_DB_FILE.read_text().strip() if CUSTOM_DB_FILE.exists() else os.environ['MONGO_URL']
client = AsyncIOMotorClient(_mongo_url)
db = client[DB_NAME]

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

scheduler = AsyncIOScheduler()


# ─── Lifespan ────────────────────────────────────────────
@asynccontextmanager
async def lifespan(app_instance: FastAPI):
    # Single per-minute job handles ALL reminder channels (admin + client) and
    # fires at each configured `reminder_time`. (A previous daily 08:00 cron was
    # removed because it duplicated the admin email sent by this job.)
    scheduler.add_job(
        check_and_push_reminders,
        IntervalTrigger(minutes=1),
        id="push_reminders",
        replace_existing=True
    )
    scheduler.start()
    logger.info("Scheduler started")

    # ── Seed factory GitHub config (viene "de fábrica") ────────────────
    # Solo se escribe si no existe configuración previa; nunca sobreescribe
    # cambios del usuario. Se guarda en la BD para que las copias de seguridad
    # y la app compilada de escritorio también lo hereden automáticamente.
    try:
        current = await db.app_settings.find_one({}, {"github_config": 1}) or {}
        gh_cfg = current.get("github_config") or {}
        if not gh_cfg.get("repo_url"):
            await db.app_settings.update_one(
                {},
                {"$set": {
                    "github_config.repo_url": SUGGESTED_GITHUB_REPO,
                    "github_config.branch": DEFAULT_GITHUB_BRANCH,
                }},
                upsert=True,
            )
            logger.info(f"Factory GitHub repo seeded: {SUGGESTED_GITHUB_REPO}")
    except Exception as e:
        logger.warning(f"Could not seed factory GitHub config: {e}")

    # ── Seed default "Minimalista" theme si la BD está vacía ───────────
    # Reglas:
    #   1. Si saved_themes está vacía → intentar cargar desde
    #      /app/themes/saved_themes.json (mirror local que viene con el repo).
    #   2. Si el JSON no existe o falla → insertar un tema "Minimalista"
    #      con un snapshot mínimo por defecto.
    #   3. En cualquier caso, si algún tema tiene is_default=True, se marca
    #      como `default_theme_id` en app_settings y se aplica su snapshot
    #      como `appearance_snapshot` inicial.
    try:
        seeded_default_id = None
        seeded_default_snapshot = None

        # ── Merge de temas desde el mirror local JSON (viene con el repo/GitHub) ──
        # Garantiza que los temas subidos a GitHub queden PRE-CARGADOS tras una
        # actualización, AUNQUE la BD ya tenga otros temas. Se hace upsert por
        # nombre para no duplicar ni pisar temas locales existentes.
        if THEMES_JSON_PATH.exists():
            try:
                data = json.loads(THEMES_JSON_PATH.read_text(encoding="utf-8"))
                imported = 0
                for t in (data.get("themes") or []):
                    name = t.get("name", "Minimalista")
                    existing = await db.saved_themes.find_one({"name": name})
                    if existing:
                        tid = existing["_id"]
                    else:
                        ins = await db.saved_themes.insert_one({
                            "name": name,
                            "snapshot": t.get("snapshot", {}),
                            "created_at": t.get("created_at") or datetime.now(timezone.utc).isoformat(),
                            "updated_at": t.get("updated_at") or "",
                            "is_default": bool(t.get("is_default", False)),
                        })
                        tid = ins.inserted_id
                        imported += 1
                    if t.get("is_default") and not seeded_default_id:
                        seeded_default_id = tid
                        seeded_default_snapshot = t.get("snapshot", {})
                if imported:
                    logger.info(f"Imported {imported} theme(s) from local JSON (GitHub mirror)")
            except Exception as ex:
                logger.warning(f"No se pudo leer saved_themes.json: {ex}")

        # ── Fallback: si aún no hay NINGÚN tema, crear "Minimalista" mínimo ──
        if await db.saved_themes.count_documents({}) == 0:
            default_snapshot = {"sidebar_style": "island", "nav_config": "[]"}
            ins = await db.saved_themes.insert_one({
                "name": "Minimalista",
                "snapshot": default_snapshot,
                "created_at": datetime.now(timezone.utc).isoformat(),
                "updated_at": "",
                "is_default": True,
            })
            seeded_default_id = ins.inserted_id
            seeded_default_snapshot = default_snapshot
            logger.info("Seeded default 'Minimalista' theme (fallback snapshot)")

        # ── Asegurar que exista un default_theme_id válido ──
        cur_settings = await db.app_settings.find_one({}, {"default_theme_id": 1}) or {}
        existing_default = cur_settings.get("default_theme_id")
        # Migración: normalizar ObjectId → string (rompía backups/serialización).
        if isinstance(existing_default, ObjectId):
            await db.app_settings.update_one(
                {}, {"$set": {"default_theme_id": str(existing_default)}}
            )
            existing_default = str(existing_default)
            logger.info("Migrated default_theme_id from ObjectId → string")

        if not existing_default:
            if not seeded_default_id:
                minimal = await db.saved_themes.find_one({"name": {"$regex": "^minimal", "$options": "i"}})
                if not minimal:
                    minimal = await db.saved_themes.find_one({}, sort=[("created_at", 1)])
                if minimal:
                    seeded_default_id = minimal["_id"]
                    seeded_default_snapshot = minimal.get("snapshot", {})
            if seeded_default_id:
                doc = await db.saved_themes.find_one({"_id": seeded_default_id}) or {}
                await db.app_settings.update_one(
                    {},
                    {"$set": {
                        "default_theme_id": str(seeded_default_id),
                        "default_theme_name": doc.get("name", "Minimalista"),
                        "appearance_snapshot": seeded_default_snapshot or doc.get("snapshot", {}),
                        "appearance_updated_at": datetime.now(timezone.utc).isoformat(),
                    }},
                    upsert=True,
                )
                logger.info(f"Marcado tema por defecto: {doc.get('name', 'Minimalista')}")

        # Siempre asegurar que el JSON local refleje el estado actual
        try:
            await _write_themes_local_json()
        except Exception:
            pass
    except Exception as e:
        logger.warning(f"Could not seed/merge themes: {e}")

    yield
    scheduler.shutdown()
    client.close()
    logger.info("Scheduler stopped")


app = FastAPI(lifespan=lifespan)
api_router = APIRouter(prefix="/api")


# ─── Helper ──────────────────────────────────────────────
def doc_to_dict(doc: dict) -> dict:
    if doc is None:
        return {}
    d = {k: v for k, v in doc.items() if k != '_id'}
    if '_id' in doc:
        d['id'] = str(doc['_id'])
    return d


def _json_safe(obj):
    """Recursive fallback for json.dumps: converts ObjectId, datetime, bytes, etc."""
    from bson import ObjectId as _OID
    if isinstance(obj, _OID):
        return str(obj)
    if isinstance(obj, (datetime,)):
        try:
            return obj.isoformat()
        except Exception:
            return str(obj)
    if isinstance(obj, bytes):
        try:
            return obj.decode("utf-8", errors="replace")
        except Exception:
            return obj.hex()
    if hasattr(obj, "isoformat"):
        try:
            return obj.isoformat()
        except Exception:
            pass
    return str(obj)


# ─── Pydantic Models ─────────────────────────────────────
class ReservationCreate(BaseModel):
    client_name: str
    client_phone: Optional[str] = None
    client_email: Optional[str] = None
    event_type: str
    event_date: str
    event_time: Optional[str] = None
    venue: Optional[str] = None
    guests_count: Optional[int] = None
    total_amount: float
    advance_paid: float = 0.0
    status: str = "Reservado"
    notes: Optional[str] = None
    package_type: Optional[str] = None


class ReservationUpdate(BaseModel):
    client_name: Optional[str] = None
    client_phone: Optional[str] = None
    client_email: Optional[str] = None
    event_type: Optional[str] = None
    event_date: Optional[str] = None
    event_time: Optional[str] = None
    venue: Optional[str] = None
    guests_count: Optional[int] = None
    total_amount: Optional[float] = None
    advance_paid: Optional[float] = None
    status: Optional[str] = None
    notes: Optional[str] = None
    locations: Optional[list] = None
    assigned_partners: Optional[list] = None
    package_type: Optional[str] = None


class SocioCreate(BaseModel):
    name: str
    role: str = "Fotógrafo"
    phone: Optional[str] = None
    email: Optional[str] = None
    notes: Optional[str] = None
    rate_per_event: Optional[float] = None


class SocioUpdate(BaseModel):
    name: Optional[str] = None
    role: Optional[str] = None
    phone: Optional[str] = None
    email: Optional[str] = None
    notes: Optional[str] = None
    rate_per_event: Optional[float] = None


class NotificationSettingsModel(BaseModel):
    reminders_enabled: bool = False
    reminder_periods: List[int] = [3]        # days before: [7, 3, 1, 0]
    reminder_time: Optional[str] = "09:00"   # daily send time HH:MM
    reminder_hours_before: int = 0           # 0 = disabled; N = send N hrs before event
    admin_email: Optional[str] = None
    admin_whatsapp: Optional[str] = None
    notification_channel: str = "email"
    resend_api_key: Optional[str] = None
    # ── New email fields ───────────────────────────────────────
    sender_name: Optional[str] = "Cinema Productions"
    cc_emails: Optional[str] = None          # comma-separated CC emails
    email_subject: Optional[str] = None      # custom subject template
    notify_client: bool = False              # also email client_email field
    telegram_enabled: bool = False
    telegram_bot_token: Optional[str] = None
    telegram_chat_id: Optional[str] = None
    ntfy_enabled: bool = False
    ntfy_topic: Optional[str] = None
    # ── WhatsApp Cloud API (Meta) ──────────────────────────────
    whatsapp_enabled: bool = False
    whatsapp_access_token: Optional[str] = None
    whatsapp_phone_number_id: Optional[str] = None
    whatsapp_recipient: Optional[str] = None   # E.164 without '+'
    whatsapp_template_name: Optional[str] = None
    # ── Google OAuth (Gmail) config — self-hosted credentials ──
    google_client_id: Optional[str] = None
    google_client_secret: Optional[str] = None
    # ── SMS (placeholder, not implemented) ─────────────────────
    sms_enabled: bool = False
    # ── Business config ────────────────────────────────────────
    company_name: Optional[str] = None
    company_address: Optional[str] = None
    company_phone: Optional[str] = None
    company_website: Optional[str] = None
    company_tax_id: Optional[str] = None
    timezone: Optional[str] = "America/Guatemala"
    default_advance_pct: Optional[int] = 30
    business_hours_start: Optional[str] = "08:00"
    business_hours_end: Optional[str] = "22:00"
    backup_retention: Optional[int] = 10
    auto_cleanup_months: Optional[int] = None  # None = disabled


class DBConnectRequest(BaseModel):
    url: str
    sync_local: bool = True


# ─── Reminder Logic ──────────────────────────────────────
def _build_reminder_html(events: list, days: int, sender_name: str = "Cinema Productions") -> str:
    rows = ""
    for ev in events:
        balance = ev.get("total_amount", 0) - ev.get("advance_paid", 0)
        balance_str = f"Q{balance:,.2f}" if balance > 0 else '<span style="color:#10b981;font-weight:bold;">Pagado</span>'
        rows += f"""
        <tr>
          <td style="padding:10px 16px;border-bottom:1px solid #e5e7eb;">{ev.get('client_name','')}</td>
          <td style="padding:10px 16px;border-bottom:1px solid #e5e7eb;">{ev.get('event_type','')}</td>
          <td style="padding:10px 16px;border-bottom:1px solid #e5e7eb;">{ev.get('event_date','')} {ev.get('event_time','')}</td>
          <td style="padding:10px 16px;border-bottom:1px solid #e5e7eb;">{ev.get('venue') or '—'}</td>
          <td style="padding:10px 16px;border-bottom:1px solid #e5e7eb;">{balance_str}</td>
        </tr>"""
    return f"""
    <div style="font-family:Arial,sans-serif;max-width:620px;margin:0 auto;background:#f8fafc;padding:32px;">
      <div style="background:linear-gradient(135deg,#6366f1,#8b5cf6);border-radius:16px 16px 0 0;padding:24px 32px;">
        <h1 style="color:#fff;margin:0;font-size:22px;">{sender_name}</h1>
        <p style="color:rgba(255,255,255,0.85);margin:4px 0 0;font-size:14px;">Recordatorio de eventos próximos</p>
      </div>
      <div style="background:#fff;border-radius:0 0 16px 16px;padding:24px 32px;box-shadow:0 4px 24px rgba(0,0,0,0.06);">
        <p style="color:#374151;font-size:16px;margin-top:0;">
          Tienes <strong>{len(events)} evento(s)</strong> programado(s) en <strong>{days} día(s)</strong>:
        </p>
        <table style="width:100%;border-collapse:collapse;margin-top:16px;font-size:13px;">
          <thead>
            <tr style="background:#f1f5f9;">
              <th style="padding:10px 16px;text-align:left;color:#6366f1;">Cliente</th>
              <th style="padding:10px 16px;text-align:left;color:#6366f1;">Tipo</th>
              <th style="padding:10px 16px;text-align:left;color:#6366f1;">Fecha / Hora</th>
              <th style="padding:10px 16px;text-align:left;color:#6366f1;">Lugar</th>
              <th style="padding:10px 16px;text-align:left;color:#6366f1;">Saldo</th>
            </tr>
          </thead>
          <tbody>{rows}</tbody>
        </table>
        <p style="color:#9ca3af;font-size:11px;margin-top:24px;border-top:1px solid #e5e7eb;padding-top:16px;">
          {sender_name} — Sistema de Gestión de Reservas · Este es un recordatorio automático
        </p>
      </div>
    </div>"""


def _build_client_reminder_html(ev: dict, days: int, sender_name: str = "Cinema Productions") -> str:
    balance = ev.get("total_amount", 0) - ev.get("advance_paid", 0)
    balance_str = f"Saldo pendiente: Q{balance:,.2f}" if balance > 0 else "Pago completado ✓"
    return f"""
    <div style="font-family:Arial,sans-serif;max-width:560px;margin:0 auto;background:#f8fafc;padding:32px;">
      <div style="background:linear-gradient(135deg,#6366f1,#8b5cf6);border-radius:16px 16px 0 0;padding:24px 32px;">
        <h1 style="color:#fff;margin:0;font-size:20px;">{sender_name}</h1>
        <p style="color:rgba(255,255,255,0.85);margin:4px 0 0;font-size:14px;">Recordatorio de tu evento</p>
      </div>
      <div style="background:#fff;border-radius:0 0 16px 16px;padding:24px 32px;box-shadow:0 4px 24px rgba(0,0,0,0.06);">
        <p style="color:#374151;font-size:16px;margin-top:0;">Hola <strong>{ev.get('client_name','')}</strong>,</p>
        <p style="color:#374151;">Te recordamos que tu evento está programado en <strong>{days} día(s)</strong>:</p>
        <div style="background:#f1f5f9;border-radius:12px;padding:20px;margin:16px 0;">
          <p style="margin:0 0 8px;color:#6366f1;font-weight:bold;">{ev.get('event_type', 'Evento')}</p>
          <p style="margin:0 0 6px;color:#374151;">📅 {ev.get('event_date', '')} {ev.get('event_time', '')}</p>
          <p style="margin:0 0 6px;color:#374151;">📍 {ev.get('venue') or 'Por confirmar'}</p>
          <p style="margin:0;color:#374151;">💳 {balance_str}</p>
        </div>
        <p style="color:#9ca3af;font-size:11px;margin-top:24px;border-top:1px solid #e5e7eb;padding-top:16px;">
          {sender_name} — Este es un recordatorio automático
        </p>
      </div>
    </div>"""


async def check_and_send_reminders():
    """Daily cron job: send reminders for upcoming events."""
    try:
        settings_doc = await db.app_settings.find_one({}, {"_id": 0})
        if not settings_doc or not settings_doc.get("reminders_enabled"):
            return

        days = int(settings_doc.get("reminder_days", 3))
        target_date = (datetime.now(timezone.utc).date() + timedelta(days=days)).isoformat()

        cursor = db.reservations.find(
            {"event_date": target_date, "status": {"$nin": ["Cancelado", "Completado"]}},
            {"client_name": 1, "event_date": 1, "event_time": 1, "event_type": 1, "venue": 1, "total_amount": 1, "advance_paid": 1, "client_email": 1, "_id": 0}
        )
        upcoming = await cursor.to_list(100)

        if not upcoming:
            logger.info(f"No events on {target_date}, skipping reminders.")
            return

        channel = settings_doc.get("notification_channel", "email")
        sender_name = settings_doc.get("sender_name") or "Cinema Productions"
        custom_subject = settings_doc.get("email_subject") or f"Recordatorio: {len(upcoming)} evento(s) en {days} día(s)"
        cc_raw = settings_doc.get("cc_emails") or ""
        cc_list = [e.strip() for e in cc_raw.split(",") if e.strip()]
        notify_client = settings_doc.get("notify_client", False)
        logger.info(f"Sending reminders for {len(upcoming)} events on {target_date} via {channel}")

        if channel in ("email", "both"):
            api_key = settings_doc.get("resend_api_key")
            admin_email = settings_doc.get("admin_email")
            if api_key and admin_email:
                resend_lib.api_key = api_key
                html = _build_reminder_html(upcoming, days, sender_name)
                to_list = [admin_email] + cc_list
                params = {
                    "from": f"{sender_name} <onboarding@resend.dev>",
                    "to": to_list,
                    "subject": custom_subject,
                    "html": html,
                }
                await asyncio.to_thread(resend_lib.Emails.send, params)
                logger.info(f"Reminder email sent to {to_list}")
                # Also notify each client if they have email
                if notify_client:
                    for ev in upcoming:
                        client_email = ev.get("client_email")
                        if client_email:
                            client_params = {
                                "from": f"{sender_name} <onboarding@resend.dev>",
                                "to": [client_email],
                                "subject": f"Recordatorio de tu evento — {ev.get('event_type', '')} el {ev.get('event_date', '')}",
                                "html": _build_client_reminder_html(ev, days, sender_name),
                            }
                            await asyncio.to_thread(resend_lib.Emails.send, client_params)
            else:
                logger.warning("Email reminders enabled but api_key or admin_email missing.")

    except Exception as e:
        logger.error(f"Error in reminder job: {e}")


# ─── Gmail Helper ─────────────────────────────────────────────
async def _get_google_credentials_config():
    """Prefer DB-stored credentials (self-hosted UI config); fallback to env."""
    doc = await db.app_settings.find_one({}, {"google_client_id": 1, "google_client_secret": 1, "_id": 0})
    cid = (doc or {}).get("google_client_id") or GOOGLE_CLIENT_ID
    csec = (doc or {}).get("google_client_secret") or GOOGLE_CLIENT_SECRET
    return cid, csec


async def _get_gmail_service():
    """Return authenticated Gmail API service using stored refresh token."""
    token_doc = await db.oauth_tokens.find_one({"provider": "gmail"}, {"_id": 0})
    if not token_doc or not token_doc.get("refresh_token"):
        raise Exception("Gmail not connected. Connect via Settings → Notificaciones.")
    cid, csec = await _get_google_credentials_config()
    creds = Credentials(
        token=token_doc.get("access_token"),
        refresh_token=token_doc["refresh_token"],
        client_id=cid,
        client_secret=csec,
        token_uri="https://oauth2.googleapis.com/token",
        scopes=GMAIL_SCOPES,
    )
    if not creds.valid:
        await asyncio.to_thread(creds.refresh, GoogleRequest())
        await db.oauth_tokens.update_one(
            {"provider": "gmail"},
            {"$set": {"access_token": creds.token}},
        )
    return build("gmail", "v1", credentials=creds), token_doc.get("email", "me")


async def _send_gmail(to: str, subject: str, html: str):
    service, from_email = await _get_gmail_service()
    msg = MIMEMultipart("alternative")
    msg["Subject"] = subject
    msg["From"] = from_email
    msg["To"] = to
    msg.attach(MIMEText(html, "html"))
    raw = base64.urlsafe_b64encode(msg.as_bytes()).decode()
    await asyncio.to_thread(
        service.users().messages().send(userId="me", body={"raw": raw}).execute
    )
    logger.info(f"Gmail sent to {to}: {subject}")


# ─── Web Push Helper ──────────────────────────────────────────
async def _send_push_to_all(title: str, body: str, url: str = "/dashboard"):
    subscriptions = await db.push_subscriptions.find({}).to_list(1000)
    if not subscriptions:
        return
    expired = []
    for sub in subscriptions:
        try:
            await asyncio.to_thread(
                webpush,
                subscription_info={"endpoint": sub["endpoint"], "keys": sub["keys"]},
                data=json.dumps({"title": title, "body": body, "url": url}),
                vapid_private_key=VAPID_PRIVATE_KEY,
                vapid_claims={"sub": VAPID_EMAIL},
            )
        except WebPushException as e:
            if e.response and e.response.status_code in (404, 410):
                expired.append(sub["endpoint"])
            logger.error(f"Push send error: {e}")
    for ep in expired:
        await db.push_subscriptions.delete_one({"endpoint": ep})
    logger.info(f"Push sent to {len(subscriptions) - len(expired)} subscribers")


# ─── Telegram Helper ──────────────────────────────────────────
async def _send_telegram(bot_token: str, chat_id: str, text: str):
    async with httpx.AsyncClient(timeout=10) as c:
        r = await c.post(
            f"https://api.telegram.org/bot{bot_token}/sendMessage",
            json={"chat_id": chat_id, "text": text, "parse_mode": "HTML"},
        )
        r.raise_for_status()


def _build_telegram_text(events: list, days: int) -> str:
    lines = ["<b>Cinema Productions</b> — Recordatorio\n"]
    lines.append(f"Tienes <b>{len(events)} evento(s)</b> en <b>{days} día(s)</b>:\n")
    for ev in events:
        date = ev.get("event_date", "")
        time_str = ev.get("event_time", "")
        when = f"{date} {time_str}".strip()
        lines.append(
            f"• <b>{ev.get('client_name','?')}</b> — {ev.get('event_type','?')}\n"
            f"  {when} — {ev.get('venue') or 'Sin lugar'}"
        )
    return "\n".join(lines)


# ─── ntfy.sh Helper ───────────────────────────────────────────
async def _send_ntfy(topic: str, title: str, message: str, priority: str = "default"):
    async with httpx.AsyncClient(timeout=10) as c:
        r = await c.post(
            f"https://ntfy.sh/{topic}",
            content=message.encode("utf-8"),
            headers={
                "Title": title,
                "Priority": priority,
                "Tags": "calendar,bell",
                "Content-Type": "text/plain; charset=utf-8",
            },
        )
        r.raise_for_status()


# ─── WhatsApp Cloud API Helper (Meta) ─────────────────────────
def _normalize_e164(number: str) -> str:
    """Strip +, spaces, and dashes. Meta expects digits only."""
    if not number:
        return ""
    return "".join(ch for ch in str(number) if ch.isdigit())


async def _send_whatsapp_text(access_token: str, phone_number_id: str, to: str, body: str):
    """Send a plain WhatsApp text message via Meta Graph API v20.0."""
    url = f"https://graph.facebook.com/v20.0/{phone_number_id}/messages"
    payload = {
        "messaging_product": "whatsapp",
        "recipient_type": "individual",
        "to": _normalize_e164(to),
        "type": "text",
        "text": {"preview_url": False, "body": body[:4000]},
    }
    async with httpx.AsyncClient(timeout=15) as c:
        r = await c.post(
            url,
            headers={"Authorization": f"Bearer {access_token}", "Content-Type": "application/json"},
            json=payload,
        )
        if r.status_code != 200:
            raise Exception(f"WhatsApp API {r.status_code}: {r.text[:300]}")
        return r.json()


async def _send_whatsapp_template(access_token: str, phone_number_id: str, to: str, template_name: str, lang: str = "es"):
    """Send an approved template message (needed outside 24-hour window)."""
    url = f"https://graph.facebook.com/v20.0/{phone_number_id}/messages"
    payload = {
        "messaging_product": "whatsapp",
        "to": _normalize_e164(to),
        "type": "template",
        "template": {"name": template_name, "language": {"code": lang}},
    }
    async with httpx.AsyncClient(timeout=15) as c:
        r = await c.post(
            url,
            headers={"Authorization": f"Bearer {access_token}", "Content-Type": "application/json"},
            json=payload,
        )
        if r.status_code != 200:
            raise Exception(f"WhatsApp Template {r.status_code}: {r.text[:300]}")
        return r.json()


def _build_whatsapp_text(events: list, days_label: str) -> str:
    lines = ["*Cinema Productions* — Recordatorio", ""]
    lines.append(f"Tienes {len(events)} evento(s) en {days_label}:")
    lines.append("")
    for ev in events[:10]:
        date = ev.get("event_date", "")
        time_str = ev.get("event_time", "")
        when = f"{date} {time_str}".strip()
        lines.append(f"• *{ev.get('client_name','?')}* — {ev.get('event_type','?')}")
        lines.append(f"  {when} — {ev.get('venue') or 'Sin lugar'}")
    return "\n".join(lines)


# ─── Per-minute reminder check (all channels) ────────────────
async def _dispatch_reminders(events: list, days_label: str, settings: dict):
    """Send a reminder for `events` via every enabled channel."""
    if not events:
        return

    title = f"Recordatorio: {len(events)} evento(s) en {days_label}"
    body  = ", ".join(r.get("client_name", "?") for r in events[:3])
    if len(events) > 3:
        body += f" y {len(events) - 3} más"

    # ── Email (Resend) ──────────────────────────
    channel = settings.get("notification_channel", "email")
    if channel in ("email", "both"):
        api_key     = settings.get("resend_api_key")
        admin_email = settings.get("admin_email")
        if api_key and admin_email:
            try:
                resend_lib.api_key = api_key
                html = _build_reminder_html(events, int(days_label.split()[0]) if days_label[0].isdigit() else 0)
                params = {
                    "from": "Cinema Productions <onboarding@resend.dev>",
                    "to": [admin_email],
                    "subject": title,
                    "html": html,
                }
                await asyncio.to_thread(resend_lib.Emails.send, params)
                logger.info(f"[Reminders] Email sent to {admin_email}")
            except Exception as e:
                logger.warning(f"[Reminders] Email failed: {e}")

    # ── Telegram ────────────────────────────────
    if settings.get("telegram_enabled") and settings.get("telegram_bot_token") and settings.get("telegram_chat_id"):
        try:
            text = _build_telegram_text(events, int(days_label.split()[0]) if days_label[0].isdigit() else 0)
            await _send_telegram(settings["telegram_bot_token"], settings["telegram_chat_id"], text)
            logger.info("[Reminders] Telegram sent")
        except Exception as e:
            logger.warning(f"[Reminders] Telegram failed: {e}")

    # ── ntfy.sh ─────────────────────────────────
    if settings.get("ntfy_enabled") and settings.get("ntfy_topic"):
        try:
            await _send_ntfy(settings["ntfy_topic"], title, body, priority="high")
            logger.info("[Reminders] ntfy sent")
        except Exception as e:
            logger.warning(f"[Reminders] ntfy failed: {e}")

    # ── WhatsApp Cloud API (Meta) ───────────────
    if (
        settings.get("whatsapp_enabled")
        and settings.get("whatsapp_access_token")
        and settings.get("whatsapp_phone_number_id")
        and settings.get("whatsapp_recipient")
    ):
        try:
            wa_days_int = int(days_label.split()[0]) if days_label and days_label[0].isdigit() else 0
            wa_body = _build_whatsapp_text(events, f"{wa_days_int} día(s)" if wa_days_int else days_label)
            await _send_whatsapp_text(
                settings["whatsapp_access_token"],
                settings["whatsapp_phone_number_id"],
                settings["whatsapp_recipient"],
                wa_body,
            )
            logger.info("[Reminders] WhatsApp sent")
        except Exception as e:
            logger.warning(f"[Reminders] WhatsApp failed: {e}")

    # ── Browser Push ────────────────────────────
    await _send_push_to_all(title, body, "/dashboard")

    # ── Client emails (optional, if notify_client is enabled) ──
    if settings.get("notify_client") and channel in ("email", "both"):
        api_key = settings.get("resend_api_key")
        if api_key:
            resend_lib.api_key = api_key
            days_int = int(days_label.split()[0]) if days_label and days_label[0].isdigit() else 0
            sender_name = settings.get("sender_name") or "Cinema Productions"
            for ev in events:
                client_email = ev.get("client_email")
                if not client_email:
                    continue
                try:
                    await asyncio.to_thread(resend_lib.Emails.send, {
                        "from": f"{sender_name} <onboarding@resend.dev>",
                        "to": [client_email],
                        "subject": f"Recordatorio de tu evento — {ev.get('event_type','')} el {ev.get('event_date','')}",
                        "html": _build_client_reminder_html(ev, days_int, sender_name),
                    })
                except Exception as e:
                    logger.warning(f"[Reminders] Client email failed: {e}")


async def check_and_push_reminders():
    """Per-minute job: fires reminders for each configured period and hours-before."""
    global _sent_push_today
    try:
        settings_doc = await db.app_settings.find_one({}, {"_id": 0})
        if not settings_doc or not settings_doc.get("reminders_enabled"):
            return

        reminder_time   = settings_doc.get("reminder_time", "09:00")
        reminder_periods = settings_doc.get("reminder_periods") or [settings_doc.get("reminder_days", 3)]
        hours_before    = int(settings_doc.get("reminder_hours_before", 0))
        now             = datetime.now(timezone.utc)
        current_hhmm    = now.strftime("%H:%M")
        today_str       = now.strftime("%Y-%m-%d")

        # ── Days-based reminders ───────────────────────────────
        if current_hhmm == reminder_time:
            for days in reminder_periods:
                target_date = (now.date() + timedelta(days=days)).isoformat()
                dedup_key   = f"days_{today_str}_{days}"
                if dedup_key in _sent_push_today:
                    continue
                _sent_push_today.add(dedup_key)

                cursor = db.reservations.find(
                    {"event_date": target_date, "status": {"$nin": ["Cancelado", "Completado"]}},
                    {"client_name": 1, "event_date": 1, "event_time": 1, "event_type": 1, "venue": 1, "client_email": 1, "total_amount": 1, "advance_paid": 1, "_id": 0}
                )
                upcoming = await cursor.to_list(100)
                label = f"{days} día(s)" if days > 0 else "hoy"
                await _dispatch_reminders(upcoming, label, settings_doc)

        # ── Hours-before reminders (same day events) ───────────
        if hours_before > 0:
            cursor = db.reservations.find(
                {"event_date": today_str, "status": {"$nin": ["Cancelado", "Completado"]}},
                {"client_name": 1, "event_date": 1, "event_time": 1, "event_type": 1, "venue": 1, "client_email": 1, "total_amount": 1, "advance_paid": 1, "_id": 0}
            )
            today_events = await cursor.to_list(100)
            for ev in today_events:
                et = ev.get("event_time")
                if not et:
                    continue
                try:
                    event_dt = datetime.strptime(f"{today_str} {et}", "%Y-%m-%d %H:%M").replace(tzinfo=timezone.utc)
                    reminder_dt = event_dt - timedelta(hours=hours_before)
                    reminder_hhmm = reminder_dt.strftime("%H:%M")
                    if current_hhmm != reminder_hhmm:
                        continue
                    dedup_key = f"hours_{today_str}_{ev.get('client_name','')}_{et}"
                    if dedup_key in _sent_push_today:
                        continue
                    _sent_push_today.add(dedup_key)
                    await _dispatch_reminders([ev], f"{hours_before}h antes", settings_doc)
                except Exception:
                    continue

        # Clear dedup set daily
        if len(_sent_push_today) > 1000:
            _sent_push_today.clear()

    except Exception as e:
        logger.error(f"Error in reminder job: {e}")


# ─── Routes ──────────────────────────────────────────────

@api_router.get("/")
async def root():
    # `version` = versión local en ejecución. El frontend la compara tras una
    # auto-actualización para confirmar que habla con el binario NUEVO (versión
    # distinta) antes de recargar, evitando la pantalla en blanco.
    try:
        version = await _read_local_version()
    except Exception:
        version = ""
    return {"message": "Event Reservation API", "version": version}


@api_router.get("/deployment/env-template")
async def get_env_template():
    """Returns a .env template for deployment (with current non-sensitive values where safe)."""
    settings = await db.app_settings.find_one({}, {"_id": 0}) or {}
    template = f"""# ══════════════════════════════════════════
# Cinema Productions — Deployment Config
# ══════════════════════════════════════════
# Copy this file to .env and fill in the values

# ── Database ──────────────────────────────
MONGO_URL=mongodb+srv://<user>:<password>@cluster.mongodb.net/?retryWrites=true&w=majority
DB_NAME=cinema_productions

# ── App URLs ──────────────────────────────
REACT_APP_BACKEND_URL=https://your-domain.com

# ── Email (Resend) ─────────────────────────
RESEND_API_KEY=re_xxxxxxxxxxxxxxxxxxxxxxxx

# ── Notifications ─────────────────────────
# Telegram (optional)
TELEGRAM_BOT_TOKEN=
TELEGRAM_CHAT_ID=

# ntfy.sh (optional, free)
NTFY_TOPIC={settings.get('ntfy_topic', 'cinema-reservas')}

# ── Web Push (optional) ───────────────────
VAPID_PUBLIC_KEY=
VAPID_PRIVATE_KEY=
VAPID_SUBJECT=mailto:admin@yourdomain.com

# ══════════════════════════════════════════
# Docker Compose: docker-compose up -d
# Railway: connect GitHub repo and add these vars
# Render: add as Environment Variables in dashboard
# ══════════════════════════════════════════
"""
    return Response(
        content=template.encode(),
        media_type="text/plain",
        headers={"Content-Disposition": 'attachment; filename=".env.template"'},
    )


@api_router.get("/deployment/docker-compose")
async def get_docker_compose():
    """Returns a docker-compose.yml for self-hosting."""
    compose = """version: '3.9'
services:
  backend:
    image: python:3.11-slim
    working_dir: /app/backend
    command: uvicorn server:app --host 0.0.0.0 --port 8001 --reload
    volumes:
      - ./backend:/app/backend
    env_file: .env
    ports:
      - "8001:8001"
    restart: unless-stopped
    depends_on:
      - mongo

  frontend:
    image: node:20-alpine
    working_dir: /app/frontend
    command: sh -c "yarn install && yarn build && npx serve -s build -l 3000"
    volumes:
      - ./frontend:/app/frontend
    env_file: .env
    ports:
      - "3000:3000"
    restart: unless-stopped

  mongo:
    image: mongo:7
    volumes:
      - mongo_data:/data/db
    ports:
      - "27017:27017"
    restart: unless-stopped
    # Note: Use MongoDB Atlas instead of local mongo for production

  nginx:
    image: nginx:alpine
    ports:
      - "80:80"
      - "443:443"
    volumes:
      - ./nginx.conf:/etc/nginx/conf.d/default.conf
      - ./ssl:/etc/nginx/ssl
    depends_on:
      - frontend
      - backend
    restart: unless-stopped

volumes:
  mongo_data:

# ─── Usage ────────────────────────────────
# 1. Copy .env.template to .env and fill values
# 2. docker-compose up -d
# 3. Access: http://your-server-ip
"""
    return Response(
        content=compose.encode(),
        media_type="text/plain",
        headers={"Content-Disposition": 'attachment; filename="docker-compose.yml"'},
    )


@api_router.post("/deployment/health-check")
async def health_check_url(url: str):
    """Pings a URL and returns if it's responding."""
    if not url.startswith(("http://", "https://")):
        return {"ok": False, "error": "URL inválida"}
    try:
        async with httpx.AsyncClient(timeout=8.0, follow_redirects=True) as client:
            r = await client.get(url)
            return {
                "ok": r.status_code < 400,
                "status": r.status_code,
                "latency_ms": round(r.elapsed.total_seconds() * 1000),
                "message": f"HTTP {r.status_code} — responde en {round(r.elapsed.total_seconds()*1000)}ms",
            }
    except httpx.TimeoutException:
        return {"ok": False, "error": "Timeout — no responde (>8s)"}
    except Exception as e:
        return {"ok": False, "error": str(e)[:100]}



@api_router.delete("/data/clear-all")
async def clear_all_data_endpoint(auto_backup: bool = True):
    """Elimina todas las reservas y socios. Crea respaldo automático si auto_backup=True."""
    return await clear_all_data(auto_backup=auto_backup)


async def clear_all_data(auto_backup: bool = True):
    """Elimina todas las reservas y socios. Crea respaldo automático si auto_backup=True."""
    if auto_backup:
        try:
            await _create_backup(label="auto_pre_delete")
        except Exception as e:
            logger.warning(f"Auto-backup before clear failed: {e}")
    res_result  = await db.reservations.delete_many({})
    soc_result  = await db.socios.delete_many({})
    return {
        "ok": True,
        "deleted_reservations": res_result.deleted_count,
        "deleted_socios": soc_result.deleted_count,
        "auto_backup_created": auto_backup,
    }


@api_router.post("/data/cleanup")
async def cleanup_data(action: str = "cancelled", months_old: int = 6):
    """
    Limpieza selectiva de datos.
    action: 'cancelled' | 'old_completed' | 'preview'
    """
    await _create_backup(label="auto_pre_cleanup")

    if action == "cancelled":
        result = await db.reservations.delete_many({"status": "Cancelado"})
        return {"ok": True, "deleted": result.deleted_count, "message": f"{result.deleted_count} reservas canceladas eliminadas"}

    elif action == "old_completed":
        cutoff = datetime.now(timezone.utc) - timedelta(days=months_old * 30)
        cutoff_str = cutoff.strftime("%Y-%m-%d")
        # Find completed reservations older than cutoff
        cursor = db.reservations.find({"status": "Completado"})
        docs = await cursor.to_list(100000)
        ids_to_delete = []
        for d in docs:
            date_str = d.get("event_date", "")
            if date_str and date_str < cutoff_str:
                ids_to_delete.append(d["_id"])
        if ids_to_delete:
            result = await db.reservations.delete_many({"_id": {"$in": ids_to_delete}})
            return {"ok": True, "deleted": result.deleted_count, "message": f"{result.deleted_count} reservas completadas antiguas eliminadas"}
        return {"ok": True, "deleted": 0, "message": "No hay reservas completadas antiguas para eliminar"}

    elif action == "preview":
        # Return counts without deleting
        cancelled_count = await db.reservations.count_documents({"status": "Cancelado"})
        cutoff = datetime.now(timezone.utc) - timedelta(days=months_old * 30)
        cutoff_str = cutoff.strftime("%Y-%m-%d")
        cursor = db.reservations.find({"status": "Completado"})
        docs = await cursor.to_list(100000)
        old_completed = sum(1 for d in docs if d.get("event_date", "") < cutoff_str)
        return {
            "ok": True,
            "cancelled_count": cancelled_count,
            "old_completed_count": old_completed,
            "months_threshold": months_old,
        }

    return {"ok": False, "message": "Acción no reconocida"}


@api_router.post("/import/reservations")
async def import_reservations_csv(file: UploadFile = File(...)):
    """Import reservations from a CSV file. Returns count of imported records."""
    content = await file.read()
    try:
        text = content.decode("utf-8-sig")  # handle BOM
    except UnicodeDecodeError:
        text = content.decode("latin-1")

    reader = csv.DictReader(io.StringIO(text))

    # Column mapping: accept various header names
    FIELD_MAP = {
        "client_name":  ["client_name", "nombre", "cliente", "name"],
        "client_phone": ["client_phone", "telefono", "teléfono", "phone"],
        "client_email": ["client_email", "email", "correo"],
        "event_type":   ["event_type", "tipo_evento", "tipo", "type"],
        "event_date":   ["event_date", "fecha", "date", "fecha_evento"],
        "event_time":   ["event_time", "hora", "time"],
        "venue":        ["venue", "lugar", "location", "ubicacion"],
        "guests_count": ["guests_count", "invitados", "guests"],
        "total_amount": ["total_amount", "total", "monto_total"],
        "advance_paid": ["advance_paid", "anticipo", "advance", "pago_anticipado"],
        "status":       ["status", "estado"],
        "notes":        ["notes", "notas", "note"],
    }

    def find_col(headers, candidates):
        headers_lower = {h.lower().strip(): h for h in headers}
        for c in candidates:
            if c.lower() in headers_lower:
                return headers_lower[c.lower()]
        return None

    imported = 0
    errors   = []
    now_str  = datetime.now(timezone.utc).isoformat()

    for i, row in enumerate(reader, start=2):
        try:
            headers = list(row.keys())
            def get(field):
                col = find_col(headers, FIELD_MAP.get(field, [field]))
                return row.get(col, "").strip() if col else ""

            client_name = get("client_name")
            if not client_name:
                errors.append(f"Fila {i}: nombre vacío — omitida")
                continue

            event_date = get("event_date") or ""
            total_raw  = get("total_amount") or "0"
            advance_raw = get("advance_paid") or "0"

            # Sanitize numbers
            def to_float(s):
                s = re.sub(r"[^\d.]", "", s)
                return float(s) if s else 0.0

            doc = {
                "client_name":   client_name,
                "client_phone":  get("client_phone") or None,
                "client_email":  get("client_email") or None,
                "event_type":    get("event_type") or "Otro",
                "event_date":    event_date,
                "event_time":    get("event_time") or None,
                "venue":         get("venue") or None,
                "guests_count":  int(get("guests_count")) if get("guests_count").isdigit() else None,
                "total_amount":  to_float(total_raw),
                "advance_paid":  to_float(advance_raw),
                "status":        get("status") or "Reservado",
                "notes":         get("notes") or None,
                "receipts":      [],
                "locations":     [],
                "assigned_partners": [],
                "created_at":    now_str,
                "imported_from_csv": True,
            }
            await db.reservations.insert_one(doc)
            imported += 1
        except Exception as e:
            errors.append(f"Fila {i}: {str(e)}")

    return {
        "ok": True,
        "imported": imported,
        "errors": errors[:10],  # limit error list
        "message": f"{imported} reservas importadas correctamente" + (f" ({len(errors)} errores)" if errors else ""),
    }


@api_router.get("/export/reservations/xlsx")
async def export_reservations_xlsx():
    """Export reservations as Excel (.xlsx) file."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
        import io as _io

        cursor = db.reservations.find({}, {"_id": 0, "receipts": 0, "assigned_partners": 0, "locations": 0})
        docs   = await cursor.to_list(100000)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Reservas"

        headers = ["Nombre", "Teléfono", "Email", "Tipo Evento", "Fecha", "Hora",
                   "Lugar", "Invitados", "Total", "Anticipo", "Saldo", "Estado", "Notas", "Creado"]
        keys    = ["client_name", "client_phone", "client_email", "event_type", "event_date",
                   "event_time", "venue", "guests_count", "total_amount", "advance_paid",
                   None, "status", "notes", "created_at"]

        # Header row
        header_fill = PatternFill("solid", fgColor="4F46E5")
        header_font = Font(color="FFFFFF", bold=True, size=11)
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        # Data rows
        alt_fill = PatternFill("solid", fgColor="F8F7FF")
        for row_idx, doc in enumerate(docs, start=2):
            fill = alt_fill if row_idx % 2 == 0 else None
            for col_idx, key in enumerate(keys, start=1):
                if key is None:
                    # Saldo = total - anticipo
                    total   = doc.get("total_amount") or 0
                    advance = doc.get("advance_paid") or 0
                    value   = total - advance
                else:
                    value = doc.get(key, "")
                    if isinstance(value, float) and value == int(value):
                        value = int(value)
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                if fill:
                    cell.fill = fill

        # Auto-width
        for col in ws.columns:
            max_len = max((len(str(c.value or "")) for c in col), default=8)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 40)

        buf = _io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        ts = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M")
        return Response(
            content=buf.getvalue(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="reservaciones_{ts}.xlsx"'},
        )
    except ImportError:
        return JSONResponse({"error": "openpyxl no instalado"}, status_code=500)
    except Exception as e:
        return JSONResponse({"error": str(e)}, status_code=500)


# ─── Backup helpers ───────────────────────────────────────────
BACKUP_COLLECTIONS = ["reservations", "socios", "app_settings"]

async def _all_collections() -> list:
    """All non-system collections in the DB (full backup coverage)."""
    try:
        names = await db.list_collection_names()
    except Exception:
        names = list(BACKUP_COLLECTIONS)
    return [n for n in names if not n.startswith("system.")]


async def _create_backup(label: str = "manual") -> dict:
    """Create a COMPLETE JSON backup of every collection, store in BACKUP_DIR."""
    names = await _all_collections()
    backup_data: dict = {"_meta": {"created_at": datetime.now(timezone.utc).isoformat(), "label": label, "collections": names}}
    for cname in names:
        cursor = db[cname].find({})
        docs = await cursor.to_list(100000)
        backup_data[cname] = [doc_to_dict(d) for d in docs]

    ts = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
    filename = f"backup_{label}_{ts}.json"
    filepath = BACKUP_DIR / filename
    filepath.write_text(json.dumps(backup_data, ensure_ascii=False, indent=2, default=_json_safe), encoding="utf-8")

    # Keep only the last 15 backups (oldest removed first)
    existing = sorted(BACKUP_DIR.glob("backup_*.json"), key=lambda f: f.stat().st_mtime)
    for old in existing[:-15]:
        old.unlink(missing_ok=True)

    total_docs = sum(len(v) for k, v in backup_data.items() if k != "_meta")
    return {"filename": filename, "docs": total_docs}


# ─── Backup endpoints ─────────────────────────────────────────

@api_router.get("/backup/download")
async def download_full_backup():
    """Download a complete backup of all collections as JSON (for local PC)."""
    names = await _all_collections()
    backup_data: dict = {"_meta": {"created_at": datetime.now(timezone.utc).isoformat(), "app": "Cinema Productions", "collections": names}}
    for cname in names:
        cursor = db[cname].find({})
        docs = await cursor.to_list(100000)
        backup_data[cname] = [doc_to_dict(d) for d in docs]

    ts = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
    filename = f"cinema_backup_{ts}.json"
    content = json.dumps(backup_data, ensure_ascii=False, indent=2, default=_json_safe)
    return Response(
        content=content.encode("utf-8"),
        media_type="application/json",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@api_router.post("/backup/create")
async def create_server_backup():
    """Create and save a backup on the server (history list)."""
    try:
        result = await _create_backup(label="manual")
        return {"success": True, **result, "message": f"Respaldo creado: {result['docs']} documentos"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error al crear respaldo: {e}")


@api_router.get("/backup/history")
async def list_backups():
    """List all server-side backups (newest first)."""
    files = sorted(BACKUP_DIR.glob("backup_*.json"), key=lambda f: f.stat().st_mtime, reverse=True)
    result = []
    for f in files:
        stat = f.stat()
        size_kb = stat.st_size / 1024
        size_str = f"{size_kb:.0f} KB" if size_kb < 1024 else f"{size_kb/1024:.1f} MB"
        label = "auto" if "auto_" in f.name else "manual"
        result.append({
            "filename": f.name,
            "size": size_str,
            "label": label,
            "created_at": datetime.fromtimestamp(stat.st_mtime, tz=timezone.utc).isoformat(),
        })
    return result


@api_router.get("/backup/{filename}/download")
async def download_backup_file(filename: str):
    """Download a specific server-side backup by filename."""
    # Security: only allow .json files in BACKUP_DIR
    if not filename.endswith(".json") or "/" in filename or ".." in filename:
        raise HTTPException(status_code=400, detail="Nombre de archivo inválido")
    filepath = BACKUP_DIR / filename
    if not filepath.exists():
        raise HTTPException(status_code=404, detail="Respaldo no encontrado")
    return Response(
        content=filepath.read_bytes(),
        media_type="application/json",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@api_router.delete("/backup/{filename}")
async def delete_backup_file(filename: str):
    """Delete a specific server-side backup."""
    if not filename.endswith(".json") or "/" in filename or ".." in filename:
        raise HTTPException(status_code=400, detail="Nombre de archivo inválido")
    filepath = BACKUP_DIR / filename
    if not filepath.exists():
        raise HTTPException(status_code=404, detail="Respaldo no encontrado")
    filepath.unlink()
    return {"success": True, "message": "Respaldo eliminado"}


@api_router.post("/backup/restore")
async def restore_backup(file: UploadFile = File(...)):
    """Restore all collections from an uploaded JSON backup file."""
    if not file.filename.endswith(".json"):
        raise HTTPException(status_code=400, detail="El archivo debe ser .json")
    try:
        content = await file.read()
        backup_data = json.loads(content)
    except Exception:
        raise HTTPException(status_code=400, detail="Archivo JSON inválido o corrupto")

    # Auto-save current state before restore
    try:
        await _create_backup(label="auto_pre_restore")
    except Exception as e:
        logger.warning(f"Pre-restore backup failed: {e}")

    restored: dict = {}
    errors: list = []
    for cname, docs in backup_data.items():
        if cname == "_meta" or not isinstance(docs, list):
            continue
        try:
            # Remove serialized ids so MongoDB assigns fresh _id
            clean_docs = []
            for d in docs:
                d.pop("id", None)
                d.pop("_id", None)
                clean_docs.append(d)
            await db[cname].delete_many({})
            if clean_docs:
                await db[cname].insert_many(clean_docs)
            restored[cname] = len(clean_docs)
        except Exception as e:
            errors.append(f"{cname}: {e}")

    if errors:
        raise HTTPException(status_code=500, detail="Errores al restaurar: " + " | ".join(errors))

    total = sum(restored.values())
    return {"success": True, "restored": restored, "total": total,
            "message": f"Restaurado correctamente: {total} documentos en {len(restored)} colecciones"}


@api_router.get("/stats")
async def get_stats():
    total = await db.reservations.count_documents({})
    upcoming = await db.reservations.count_documents({
        "event_date": {"$gte": datetime.now(timezone.utc).strftime("%Y-%m-%d")},
        "status": {"$nin": ["Cancelado", "Completado"]}
    })
    active_cursor = db.reservations.find(
        {"status": {"$nin": ["Cancelado"]}},
        {"total_amount": 1, "advance_paid": 1, "assigned_partners": 1, "_id": 0}
    )
    active_docs = await active_cursor.to_list(length=None)
    total_pending = sum(
        max(0, (d.get("total_amount", 0) or 0) - (d.get("advance_paid", 0) or 0))
        for d in active_docs
    )
    total_event_amount = sum((d.get("total_amount", 0) or 0) for d in active_docs)
    total_partner_cost = sum(
        p.get("payment", 0) or 0
        for d in active_docs
        for p in (d.get("assigned_partners") or [])
    )
    return {
        "total_reservations": total,
        "upcoming_events": upcoming,
        "pending_payment": round(total_pending, 2),
        "real_income": round(total_event_amount - total_partner_cost, 2),
    }


@api_router.get("/reservations")
async def list_reservations():
    cursor = db.reservations.find({}, {"receipt_images.data": 0})
    docs = await cursor.to_list(length=None)
    return [doc_to_dict(d) for d in docs]


@api_router.post("/reservations", status_code=201)
async def create_reservation(reservation: ReservationCreate):
    doc = reservation.model_dump()
    doc["receipt_images"] = []
    doc["created_at"] = datetime.now(timezone.utc).isoformat()
    result = await db.reservations.insert_one(doc)
    doc["id"] = str(result.inserted_id)
    doc.pop("_id", None)
    return doc


@api_router.get("/reservations/{reservation_id}")
async def get_reservation(reservation_id: str):
    try:
        oid = ObjectId(reservation_id)
    except Exception:
        raise HTTPException(status_code=400, detail="ID inválido")
    doc = await db.reservations.find_one({"_id": oid})
    if not doc:
        raise HTTPException(status_code=404, detail="Reserva no encontrada")
    return doc_to_dict(doc)


@api_router.put("/reservations/{reservation_id}")
async def update_reservation(reservation_id: str, reservation: ReservationUpdate):
    try:
        oid = ObjectId(reservation_id)
    except Exception:
        raise HTTPException(status_code=400, detail="ID inválido")
    update_data = reservation.model_dump(exclude_unset=True)
    if not update_data:
        raise HTTPException(status_code=400, detail="No hay datos para actualizar")
    result = await db.reservations.update_one({"_id": oid}, {"$set": update_data})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Reserva no encontrada")
    doc = await db.reservations.find_one({"_id": oid})
    return doc_to_dict(doc)


@api_router.delete("/reservations/{reservation_id}")
async def delete_reservation(reservation_id: str):
    try:
        oid = ObjectId(reservation_id)
    except Exception:
        raise HTTPException(status_code=400, detail="ID inválido")
    result = await db.reservations.delete_one({"_id": oid})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Reserva no encontrada")
    return {"message": "Reserva eliminada"}


@api_router.post("/reservations/{reservation_id}/receipts")
async def upload_receipt(reservation_id: str, file: UploadFile = File(...)):
    try:
        oid = ObjectId(reservation_id)
    except Exception:
        raise HTTPException(status_code=400, detail="ID inválido")
    content = await file.read()
    if len(content) > 10 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="Archivo muy grande (máx 10MB)")
    b64 = base64.b64encode(content).decode("utf-8")
    receipt = {
        "id": str(uuid.uuid4()),
        "filename": file.filename,
        "content_type": file.content_type,
        "data": b64,
        "uploaded_at": datetime.now(timezone.utc).isoformat()
    }
    result = await db.reservations.update_one(
        {"_id": oid}, {"$push": {"receipt_images": receipt}}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Reserva no encontrada")
    return {k: v for k, v in receipt.items() if k != "data"}


@api_router.delete("/reservations/{reservation_id}/receipts/{receipt_id}")
async def delete_receipt(reservation_id: str, receipt_id: str):
    try:
        oid = ObjectId(reservation_id)
    except Exception:
        raise HTTPException(status_code=400, detail="ID inválido")
    result = await db.reservations.update_one(
        {"_id": oid}, {"$pull": {"receipt_images": {"id": receipt_id}}}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Reserva no encontrada")
    return {"message": "Comprobante eliminado"}


@api_router.get("/export/reservations")
async def export_reservations(format: str = "csv"):
    cursor = db.reservations.find({}, {"receipt_images": 0})
    docs = await cursor.to_list(10000)
    data = [doc_to_dict(d) for d in docs]

    if format == "json":
        return JSONResponse(
            content=data,
            headers={"Content-Disposition": "attachment; filename=reservaciones.json"}
        )

    if not data:
        return Response("", media_type="text/csv",
                        headers={"Content-Disposition": "attachment; filename=reservaciones.csv"})

    output = io.StringIO()
    fields = ["id", "client_name", "client_phone", "client_email", "event_type", "event_date",
              "event_time", "venue", "guests_count", "total_amount", "advance_paid", "status", "notes", "created_at"]
    writer = csv.DictWriter(output, fieldnames=fields, extrasaction="ignore")
    writer.writeheader()
    writer.writerows(data)
    return Response(output.getvalue(), media_type="text/csv; charset=utf-8",
                    headers={"Content-Disposition": "attachment; filename=reservaciones.csv"})


@api_router.get("/calendar")
async def get_calendar_events():
    cursor = db.reservations.find(
        {"status": {"$nin": ["Cancelado"]}},
        {"client_name": 1, "client_phone": 1, "event_date": 1, "event_time": 1, "event_type": 1,
         "venue": 1, "total_amount": 1, "advance_paid": 1, "package_type": 1, "status": 1, "_id": 1}
    )
    docs = await cursor.to_list(length=None)
    return [doc_to_dict(d) for d in docs]


# ─── Socios ──────────────────────────────────────────────

@api_router.get("/socios")
async def list_socios():
    cursor = db.socios.find({}, {"photo": 0, "photo_content_type": 0})
    docs = await cursor.to_list(length=None)
    return [doc_to_dict(d) for d in docs]


@api_router.post("/socios", status_code=201)
async def create_socio(socio: SocioCreate):
    doc = socio.model_dump()
    doc["photo"] = None
    doc["photo_content_type"] = None
    doc["created_at"] = datetime.now(timezone.utc).isoformat()
    result = await db.socios.insert_one(doc)
    doc["id"] = str(result.inserted_id)
    doc.pop("_id", None)
    return doc


@api_router.get("/socios/{socio_id}")
async def get_socio(socio_id: str):
    try:
        oid = ObjectId(socio_id)
    except Exception:
        raise HTTPException(status_code=400, detail="ID inválido")
    doc = await db.socios.find_one({"_id": oid})
    if not doc:
        raise HTTPException(status_code=404, detail="Socio no encontrado")
    return doc_to_dict(doc)


@api_router.put("/socios/{socio_id}")
async def update_socio(socio_id: str, socio: SocioUpdate):
    try:
        oid = ObjectId(socio_id)
    except Exception:
        raise HTTPException(status_code=400, detail="ID inválido")
    update_data = {k: v for k, v in socio.model_dump().items() if v is not None}
    if not update_data:
        raise HTTPException(status_code=400, detail="No hay datos para actualizar")
    result = await db.socios.update_one({"_id": oid}, {"$set": update_data})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Socio no encontrado")
    doc = await db.socios.find_one({"_id": oid})
    return doc_to_dict(doc)


@api_router.delete("/socios/{socio_id}")
async def delete_socio(socio_id: str):
    try:
        oid = ObjectId(socio_id)
    except Exception:
        raise HTTPException(status_code=400, detail="ID inválido")
    result = await db.socios.delete_one({"_id": oid})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Socio no encontrado")
    return {"message": "Socio eliminado"}


@api_router.post("/socios/{socio_id}/photo")
async def upload_socio_photo(socio_id: str, file: UploadFile = File(...)):
    try:
        oid = ObjectId(socio_id)
    except Exception:
        raise HTTPException(status_code=400, detail="ID inválido")
    content = await file.read()
    if len(content) > 5 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="Archivo muy grande (máx 5MB)")
    b64 = base64.b64encode(content).decode("utf-8")
    result = await db.socios.update_one(
        {"_id": oid}, {"$set": {"photo": b64, "photo_content_type": file.content_type}}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Socio no encontrado")
    return {"message": "Foto actualizada"}


@api_router.delete("/socios/{socio_id}/photo")
async def delete_socio_photo(socio_id: str):
    try:
        oid = ObjectId(socio_id)
    except Exception:
        raise HTTPException(status_code=400, detail="ID inválido")
    await db.socios.update_one({"_id": oid}, {"$set": {"photo": None, "photo_content_type": None}})
    return {"message": "Foto eliminada"}


@api_router.get("/financials")
async def get_financials():
    cursor = db.reservations.find(
        {"status": {"$nin": ["Cancelado"]}},
        {"total_amount": 1, "advance_paid": 1, "assigned_partners": 1}
    )
    docs = await cursor.to_list(10000)
    total_event_amount = sum((d.get("total_amount") or 0) for d in docs)
    total_advance = sum((d.get("advance_paid") or 0) for d in docs)
    total_partner_cost = 0
    total_paid_to_partners = 0
    total_pending_to_partners = 0
    for d in docs:
        for p in (d.get("assigned_partners") or []):
            amt = p.get("payment") or 0
            total_partner_cost += amt
            if p.get("payment_status") == "Pagado":
                total_paid_to_partners += amt
            else:
                total_pending_to_partners += amt
    return {
        "total_event_amount": round(total_event_amount, 2),
        "total_advance": round(total_advance, 2),
        "total_partner_cost": round(total_partner_cost, 2),
        "total_paid_to_partners": round(total_paid_to_partners, 2),
        "total_pending_to_partners": round(total_pending_to_partners, 2),
        "real_income": round(total_event_amount - total_partner_cost, 2),
    }


# ─── Metas (Goals) ────────────────────────────────────────

class MetaUpsert(BaseModel):
    year: int
    month: Optional[int] = None  # 1-12 or None for annual
    type: str  # "ventas" | "ganancias" | "gastos"
    amount: float


@api_router.get("/metas")
async def get_metas(year: int, type: str):
    if type not in ("ventas", "ganancias", "gastos"):
        raise HTTPException(status_code=400, detail="type inválido")
    cursor = db.metas.find({"year": year, "type": type}, {"_id": 0})
    goals = await cursor.to_list(200)
    return {"year": year, "type": type, "goals": goals}


@api_router.put("/metas")
async def upsert_meta(meta: MetaUpsert):
    if meta.type not in ("ventas", "ganancias", "gastos"):
        raise HTTPException(status_code=400, detail="type inválido")
    key = {"year": meta.year, "type": meta.type, "month": meta.month}
    payload = {
        **key,
        "amount": float(meta.amount or 0),
        "updated_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.metas.update_one(key, {"$set": payload, "$setOnInsert": {"id": str(uuid.uuid4())}}, upsert=True)
    return {"ok": True, **payload}


@api_router.delete("/metas")
async def delete_meta(year: int, type: str, month: Optional[int] = None):
    await db.metas.delete_one({"year": year, "type": type, "month": month})
    return {"ok": True}


@api_router.get("/metas/progress")
async def metas_progress(year: int, type: str):
    """Return per-month actual values + goals + reached flags."""
    if type not in ("ventas", "ganancias", "gastos"):
        raise HTTPException(status_code=400, detail="type inválido")

    # Fetch reservations for that year (event_date is 'YYYY-MM-DD' string)
    start = f"{year}-01-01"
    end = f"{year}-12-31"
    cursor = db.reservations.find(
        {"status": {"$nin": ["Cancelado"]}, "event_date": {"$gte": start, "$lte": end}},
        {"total_amount": 1, "event_date": 1, "assigned_partners": 1, "_id": 0}
    )
    docs = await cursor.to_list(20000)

    # Aggregate by month
    monthly = {m: {"ventas": 0.0, "gastos": 0.0} for m in range(1, 13)}
    for d in docs:
        ed = d.get("event_date") or ""
        try:
            m = int(ed.split("-")[1])
        except Exception:
            continue
        if m < 1 or m > 12:
            continue
        monthly[m]["ventas"] += float(d.get("total_amount") or 0)
        for p in (d.get("assigned_partners") or []):
            monthly[m]["gastos"] += float(p.get("payment") or 0)

    # Fetch goals for year+type
    goal_cursor = db.metas.find({"year": year, "type": type}, {"_id": 0})
    goal_docs = await goal_cursor.to_list(200)
    goal_by_month = {}
    annual_goal = 0.0
    for g in goal_docs:
        if g.get("month") is None:
            annual_goal = float(g.get("amount") or 0)
        else:
            goal_by_month[int(g["month"])] = float(g.get("amount") or 0)

    def actual_for(m):
        v = monthly[m]
        if type == "ventas":
            return round(v["ventas"], 2)
        if type == "gastos":
            return round(v["gastos"], 2)
        # ganancias
        return round(v["ventas"] - v["gastos"], 2)

    months_out = []
    annual_actual = 0.0
    # Meta mensual auto-derivada de la anual (ventas/ganancias). En gastos NO se auto-deriva.
    # Los meses con meta custom se descuentan de la anual y el remanente se reparte
    # de forma uniforme entre los meses restantes, de modo que la suma de metas
    # mensuales reconcilie con la meta anual.
    auto_monthly = 0.0
    if type != "gastos" and annual_goal > 0:
        custom_sum = sum(goal_by_month.get(m, 0.0) for m in range(1, 13))
        auto_months = [m for m in range(1, 13) if m not in goal_by_month]
        if auto_months:
            remaining = max(0.0, annual_goal - custom_sum)
            auto_monthly = round(remaining / len(auto_months), 2)

    for m in range(1, 13):
        act = actual_for(m)
        annual_actual += act
        custom = goal_by_month.get(m)
        is_custom = custom is not None
        if is_custom:
            goal = float(custom)
        else:
            goal = auto_monthly
        pct = (act / goal * 100) if goal > 0 else 0.0
        months_out.append({
            "month": m,
            "actual": round(act, 2),
            "goal": round(goal, 2),
            "percent": round(pct, 2),
            "reached": goal > 0 and act >= goal,
            "is_custom": is_custom,
            "is_auto": (not is_custom) and goal > 0,
        })

    # Meta anual efectiva: la explícita si existe, si no la suma de metas mensuales custom.
    effective_annual_goal = annual_goal if annual_goal > 0 else sum(
        goal_by_month.get(m, 0.0) for m in range(1, 13)
    )
    ann_pct = (annual_actual / effective_annual_goal * 100) if effective_annual_goal > 0 else 0.0
    return {
        "year": year,
        "type": type,
        "months": months_out,
        "annual_goal": round(effective_annual_goal, 2),
        "annual_actual": round(annual_actual, 2),
        "annual_percent": round(ann_pct, 2),
        "annual_reached": effective_annual_goal > 0 and annual_actual >= effective_annual_goal,
        "annual_goal_explicit": round(annual_goal, 2),
        "auto_monthly": auto_monthly,
    }


# ─── App Settings ─────────────────────────────────────────

@api_router.get("/settings")
async def get_app_settings():
    doc = await db.app_settings.find_one({}, {"_id": 0})
    if not doc:
        return {}
    doc.pop("app_password_hash", None)
    # Mask Resend key
    if doc.get("resend_api_key"):
        key = doc["resend_api_key"]
        doc["resend_api_key"] = "re_" + "•" * 20 + key[-4:] if len(key) > 4 else "****"
        doc["has_resend_key"] = True
    else:
        doc["has_resend_key"] = False
    # Mask Telegram token
    if doc.get("telegram_bot_token"):
        tok = doc["telegram_bot_token"]
        doc["telegram_bot_token"] = tok[:8] + "•" * 20 + tok[-4:] if len(tok) > 12 else "****"
        doc["has_telegram_token"] = True
    else:
        doc["has_telegram_token"] = False
    # Mask WhatsApp access token
    if doc.get("whatsapp_access_token"):
        wa = doc["whatsapp_access_token"]
        doc["whatsapp_access_token"] = "•" * 24 + wa[-4:] if len(wa) > 8 else "****"
        doc["has_whatsapp_token"] = True
    else:
        doc["has_whatsapp_token"] = False
    # Mask Google client secret
    if doc.get("google_client_secret"):
        gs = doc["google_client_secret"]
        doc["google_client_secret"] = "•" * 24 + gs[-4:] if len(gs) > 8 else "****"
        doc["has_google_client_secret"] = True
    else:
        doc["has_google_client_secret"] = False
    # Ensure reminder_periods is always a list
    if "reminder_periods" not in doc:
        doc["reminder_periods"] = [doc.get("reminder_days", 3)]
    return doc


@api_router.put("/settings")
async def update_app_settings(settings: NotificationSettingsModel):
    update_doc = settings.model_dump()

    # If Resend key is masked (unchanged), don't overwrite
    key = update_doc.get("resend_api_key") or ""
    if "****" in key or "•" in key:
        update_doc.pop("resend_api_key", None)

    # If Telegram token is masked (unchanged), don't overwrite
    tok = update_doc.get("telegram_bot_token") or ""
    if "****" in tok or "•" in tok:
        update_doc.pop("telegram_bot_token", None)

    # If WhatsApp token is masked, don't overwrite
    wa = update_doc.get("whatsapp_access_token") or ""
    if "****" in wa or "•" in wa:
        update_doc.pop("whatsapp_access_token", None)

    # If Google client secret is masked, don't overwrite
    gs = update_doc.get("google_client_secret") or ""
    if "****" in gs or "•" in gs:
        update_doc.pop("google_client_secret", None)

    existing = await db.app_settings.find_one({}, {"_id": 0})
    if existing:
        await db.app_settings.update_one({}, {"$set": update_doc})
    else:
        await db.app_settings.insert_one(update_doc)

    saved = await db.app_settings.find_one({}, {"_id": 0})
    if saved:
        if saved.get("resend_api_key"):
            k = saved["resend_api_key"]
            saved["resend_api_key"] = "re_" + "•" * 20 + k[-4:] if len(k) > 4 else "****"
            saved["has_resend_key"] = True
        else:
            saved["has_resend_key"] = False
        if saved.get("telegram_bot_token"):
            t = saved["telegram_bot_token"]
            saved["telegram_bot_token"] = t[:8] + "•" * 20 + t[-4:] if len(t) > 12 else "****"
            saved["has_telegram_token"] = True
        else:
            saved["has_telegram_token"] = False
        if saved.get("whatsapp_access_token"):
            wa2 = saved["whatsapp_access_token"]
            saved["whatsapp_access_token"] = "•" * 24 + wa2[-4:] if len(wa2) > 8 else "****"
            saved["has_whatsapp_token"] = True
        else:
            saved["has_whatsapp_token"] = False
        if saved.get("google_client_secret"):
            gs2 = saved["google_client_secret"]
            saved["google_client_secret"] = "•" * 24 + gs2[-4:] if len(gs2) > 8 else "****"
            saved["has_google_client_secret"] = True
        else:
            saved["has_google_client_secret"] = False
    return saved or {}


# ─── Database Settings ────────────────────────────────────

@api_router.get("/settings/database")
async def get_db_stats():
    global client, db
    try:
        raw = await db.command("dbstats")
        storage_bytes = raw.get("storageSize", 0)
        data_bytes = raw.get("dataSize", 0)
        index_bytes = raw.get("indexSize", 0)
        objects = raw.get("objects", 0)
        collections = raw.get("collections", 0)

        def fmt(b):
            if b < 1024:
                return f"{b} B"
            elif b < 1024 ** 2:
                return f"{b / 1024:.1f} KB"
            return f"{b / (1024 ** 2):.2f} MB"

        is_custom = CUSTOM_DB_FILE.exists()
        current_url = CUSTOM_DB_FILE.read_text().strip() if is_custom else os.environ['MONGO_URL']
        # Mask credentials in URL for display
        display_url = current_url
        if "@" in current_url:
            proto_end = current_url.find("://") + 3
            at_pos = current_url.rfind("@")
            display_url = current_url[:proto_end] + "***:***@" + current_url[at_pos + 1:]

        is_atlas = current_url.startswith("mongodb+srv")
        used = storage_bytes + index_bytes
        # Atlas Free (M0) tiene 512 MB; para conexiones normales estimamos 512 MB tambien.
        limit_bytes = 512 * 1024 ** 2
        free_bytes = max(0, limit_bytes - used)
        used_pct = round(min(100, used / limit_bytes * 100), 1)

        return {
            "db_name": DB_NAME,
            "collections": collections,
            "objects": objects,
            "data_size": fmt(data_bytes),
            "storage_size": fmt(storage_bytes),
            "index_size": fmt(index_bytes),
            "total_size": fmt(used),
            "used_size": fmt(used),
            "free_size": fmt(free_bytes),
            "limit_size": fmt(limit_bytes),
            "used_pct": used_pct,
            "is_atlas": is_atlas,
            "current_url": display_url,
            "is_custom": is_custom,
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error al obtener estadísticas: {e}")


@api_router.get("/settings/database/stats-detailed")
async def db_stats_detailed():
    """Contadores por colección para preview de 'subir a la nube'."""
    try:
        names = await _all_collections()
        counts = {}
        for n in names:
            try:
                counts[n] = await db[n].estimated_document_count()
            except Exception:
                counts[n] = 0
        reservations = int(counts.get("reservations", 0))
        socios = int(counts.get("socios", 0))
        settings = int(counts.get("app_settings", 0)) + int(counts.get("appearance_settings", 0))
        themes = int(counts.get("saved_themes", 0))
        total = sum(int(v) for v in counts.values())
        # Reutiliza dbStats para tamaño / is_atlas
        try:
            current_url = CUSTOM_DB_FILE.read_text().strip() if CUSTOM_DB_FILE.exists() else os.environ['MONGO_URL']
        except Exception:
            current_url = ""
        is_atlas = "mongodb+srv" in (current_url or "") or ".mongodb.net" in (current_url or "")
        size_str = "—"
        try:
            st = await db.command("dbStats")
            data_size = int(st.get("dataSize", 0))
            if data_size < 1024:
                size_str = f"{data_size} B"
            elif data_size < 1024 * 1024:
                size_str = f"{data_size/1024:.1f} KB"
            else:
                size_str = f"{data_size/(1024*1024):.2f} MB"
        except Exception:
            pass
        return {
            "reservations": reservations,
            "socios": socios,
            "settings": max(settings, 1),
            "themes": themes,
            "total": total,
            "size": size_str,
            "is_cloud": is_atlas,
            "collections": counts,
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error al obtener detalle: {e}")


@api_router.post("/settings/database/optimize")
async def optimize_database():
    result = {"indexed": [], "compacted": [], "skipped": []}
    try:
        # 1) Índices útiles (permitido en Atlas)
        try:
            await db.reservations.create_index("event_date")
            await db.reservations.create_index("status")
            await db.reservations.create_index("created_at")
            result["indexed"].append("reservations")
        except Exception as e:
            result["skipped"].append(f"idx reservations: {str(e)[:60]}")
        try:
            await db.socios.create_index("name")
            result["indexed"].append("socios")
        except Exception as e:
            result["skipped"].append(f"idx socios: {str(e)[:60]}")

        # 2) Compactar cada colección (se omite en Atlas compartido)
        for cname in await _all_collections():
            try:
                await db.command({"compact": cname})
                result["compacted"].append(cname)
            except Exception:
                result["skipped"].append(f"compact {cname} (no permitido aquí)")

        msg = f"Optimización lista. Índices: {len(result['indexed'])}"
        if result["compacted"]:
            msg += f", compactadas: {len(result['compacted'])}"
        return {"success": True, "message": msg, **result}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error al optimizar: {e}")


@api_router.post("/settings/database/test")
async def test_db_connection(req: DBConnectRequest):
    try:
        test_client = AsyncIOMotorClient(req.url, serverSelectionTimeoutMS=5000)
        test_db = test_client[DB_NAME]
        await test_db.command("ping")
        test_client.close()
        return {"success": True, "message": "Conexión exitosa"}
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"No se pudo conectar: {e}")


@api_router.post("/settings/database/compare")
async def compare_database(req: DBConnectRequest):
    """Compara los conteos de documentos de la BD actual contra una BD destino (nube).
    Sirve para indicar cuántos registros locales aún no están en la nube."""
    global db
    try:
        # Conteos actuales
        current_counts: dict[str, int] = {}
        current_total = 0
        for cname in await _all_collections():
            try:
                n = await db[cname].count_documents({})
            except Exception:
                n = 0
            current_counts[cname] = n
            current_total += n

        # Conteos remotos
        target_counts: dict[str, int] = {}
        target_total = 0
        target_client = AsyncIOMotorClient(req.url, serverSelectionTimeoutMS=5000)
        try:
            target_db = target_client[DB_NAME]
            await target_db.command("ping")
            target_names = [n for n in await target_db.list_collection_names() if not n.startswith("system.")]
            for cname in target_names:
                try:
                    n = await target_db[cname].count_documents({})
                except Exception:
                    n = 0
                target_counts[cname] = n
                target_total += n
        finally:
            target_client.close()

        # Pendientes de subir (colecciones donde local > remoto)
        pending: dict[str, int] = {}
        for k, v in current_counts.items():
            diff = v - target_counts.get(k, 0)
            if diff > 0:
                pending[k] = diff
        pending_total = sum(pending.values())

        return {
            "current": current_counts,
            "current_total": current_total,
            "target": target_counts,
            "target_total": target_total,
            "pending_upload": pending,
            "pending_total": pending_total,
            "needs_sync": pending_total > 0,
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Error al comparar: {e}")


@api_router.post("/settings/database/connect")
async def switch_database(req: DBConnectRequest):
    global client, db
    try:
        new_client = AsyncIOMotorClient(req.url, serverSelectionTimeoutMS=5000)
        new_db = new_client[DB_NAME]
        await new_db.command("ping")

        # ─── Sync local → cloud (merge, do NOT overwrite existing cloud docs) ───
        sync_stats: dict[str, int] = {}
        total_uploaded = 0
        if req.sync_local:
            try:
                local_names = await db.list_collection_names()
                for cname in local_names:
                    if cname.startswith("system."):
                        continue
                    try:
                        local_docs = await db[cname].find({}).to_list(length=None)
                    except Exception:
                        continue
                    if not local_docs:
                        continue
                    uploaded = 0
                    for doc in local_docs:
                        _id = doc.get("_id")
                        try:
                            if _id is not None:
                                exists = await new_db[cname].find_one({"_id": _id}, {"_id": 1})
                                if exists:
                                    continue
                            await new_db[cname].insert_one(doc)
                            uploaded += 1
                        except Exception:
                            # Duplicate key or validation → skip, keep cloud version
                            continue
                    if uploaded > 0:
                        sync_stats[cname] = uploaded
                        total_uploaded += uploaded
            except Exception as e:
                logger.warning(f"Sync local→cloud failed: {e}")

        old_client = client
        client = new_client
        db = new_db
        CUSTOM_DB_FILE.write_text(req.url)
        old_client.close()
        logger.info(f"Database switched to: {req.url[:30]}... (synced {total_uploaded} docs)")

        msg = "Base de datos cambiada correctamente"
        if total_uploaded > 0:
            details = ", ".join(f"{k}: {v}" for k, v in sync_stats.items())
            msg = f"Conectado a la nube · {total_uploaded} registros locales subidos ({details})"
        return {"success": True, "message": msg, "synced": sync_stats, "total_uploaded": total_uploaded}
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Error al conectar: {e}")


@api_router.post("/settings/database/reset")
async def reset_database():
    global client, db
    try:
        original_url = os.environ['MONGO_URL']
        new_client = AsyncIOMotorClient(original_url)
        new_db = new_client[DB_NAME]
        await new_db.command("ping")
        old_client = client
        client = new_client
        db = new_db
        if CUSTOM_DB_FILE.exists():
            CUSTOM_DB_FILE.unlink()
        old_client.close()
        return {"success": True, "message": "Conexión restaurada a la base de datos original"}
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Error al restaurar: {e}")


@api_router.get("/settings/database/factory-presets")
async def get_factory_presets():
    """Conexiones MongoDB que vienen DE FÁBRICA con el proyecto.
    El frontend las carga automáticamente y las mezcla con las que el usuario
    haya guardado localmente. Vienen ya listas al clonar el repositorio."""
    return {"presets": FACTORY_SAVED_CONNECTIONS}


# ─── Reminders (manual trigger) ───────────────────────────

@api_router.get("/notifications/pending")
async def get_pending_notifications():
    settings_doc = await db.app_settings.find_one({}, {"_id": 0})
    periods = (settings_doc.get("reminder_periods") if settings_doc else None) or [
        (settings_doc.get("reminder_days", 3) if settings_doc else 3)
    ]
    days = max(periods) if periods else 3
    today = datetime.now(timezone.utc).date()
    end = (today + timedelta(days=days)).isoformat()
    today_str = today.isoformat()
    cursor = db.reservations.find(
        {"event_date": {"$gte": today_str, "$lte": end}, "status": {"$nin": ["Cancelado", "Completado"]}},
        {"client_name": 1, "event_date": 1, "event_type": 1, "venue": 1, "_id": 1}
    )
    docs = await cursor.to_list(100)
    return [doc_to_dict(d) for d in docs]



@api_router.post("/reminders/test-email")
async def test_email_connection():
    """Send a test email to verify Resend API key and email config — no events needed."""
    try:
        settings_doc = await db.app_settings.find_one({}, {"_id": 0})
        if not settings_doc:
            raise HTTPException(status_code=400, detail="Configura los ajustes primero")

        api_key   = settings_doc.get("resend_api_key")
        admin_email = settings_doc.get("admin_email")
        sender_name = settings_doc.get("sender_name") or "Cinema Productions"

        if not api_key:
            raise HTTPException(status_code=400, detail="Ingresa tu API Key de Resend primero")
        if not admin_email:
            raise HTTPException(status_code=400, detail="Ingresa un Email Destino primero")

        resend_lib.api_key = api_key
        html = f"""
        <div style="font-family:Arial,sans-serif;max-width:520px;margin:0 auto;background:#f8fafc;padding:32px;">
          <div style="background:linear-gradient(135deg,#6366f1,#8b5cf6);border-radius:16px 16px 0 0;padding:24px 32px;">
            <h1 style="color:#fff;margin:0;font-size:20px;">{sender_name}</h1>
            <p style="color:rgba(255,255,255,0.85);margin:4px 0 0;font-size:14px;">Prueba de conexión de email</p>
          </div>
          <div style="background:#fff;border-radius:0 0 16px 16px;padding:24px 32px;box-shadow:0 4px 24px rgba(0,0,0,0.06);">
            <p style="color:#10b981;font-size:18px;font-weight:bold;margin-top:0;">✓ Conexión exitosa</p>
            <p style="color:#374151;">Tu configuración de email con Resend funciona correctamente.</p>
            <p style="color:#374151;">Los recordatorios automáticos se enviarán a: <strong>{admin_email}</strong></p>
            <p style="color:#9ca3af;font-size:11px;margin-top:24px;border-top:1px solid #e5e7eb;padding-top:16px;">
              {sender_name} — Prueba enviada desde el Sistema de Gestión de Reservas
            </p>
          </div>
        </div>"""
        params = {
            "from": f"{sender_name} <onboarding@resend.dev>",
            "to": [admin_email],
            "subject": f"✓ Prueba de email — {sender_name}",
            "html": html,
        }
        result = await asyncio.to_thread(resend_lib.Emails.send, params)
        return {"success": True, "message": f"Email de prueba enviado a {admin_email}", "id": getattr(result, 'id', str(result))}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"test-email error: {e}")
        raise HTTPException(status_code=500, detail=f"Error al enviar: {str(e)}")


@api_router.post("/reminders/send")
async def trigger_reminders_manual():
    """Manual trigger — send reminder email/push/telegram/ntfy for upcoming events."""
    try:
        settings_doc = await db.app_settings.find_one({}, {"_id": 0})
        if not settings_doc:
            raise HTTPException(status_code=400, detail="Primero configura los ajustes de notificaciones")

        reminder_periods = settings_doc.get("reminder_periods") or [settings_doc.get("reminder_days", 3)]
        today = datetime.now(timezone.utc).date()

        all_events = []
        for days in reminder_periods:
            target = (today + timedelta(days=days)).isoformat()
            cursor = db.reservations.find(
                {
                    "event_date": target,
                    "status": {"$nin": ["Cancelado", "Completado"]},
                },
                {"client_name": 1, "event_date": 1, "event_time": 1, "event_type": 1, "venue": 1, "total_amount": 1, "advance_paid": 1, "_id": 0}
            )
            evs = await cursor.to_list(100)
            for e in evs:
                e["_days_label"] = f"{days} día(s)" if days > 0 else "hoy"
                all_events.append(e)

        await _dispatch_reminders(all_events, f"{len(reminder_periods)} período(s)", settings_doc)

        channels_used = []
        if settings_doc.get("resend_api_key") and settings_doc.get("admin_email"):
            channels_used.append("email")
        if settings_doc.get("telegram_enabled") and settings_doc.get("telegram_bot_token"):
            channels_used.append("telegram")
        if settings_doc.get("ntfy_enabled") and settings_doc.get("ntfy_topic"):
            channels_used.append("ntfy")
        channels_used.append("push")

        msg = f"{len(all_events)} evento(s) encontrado(s) — enviado vía {', '.join(channels_used)}"
        return {
            "success": True,
            "events_found": len(all_events),
            "channels": channels_used,
            "message": msg,
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error: {e}")


# ─── Desktop Package Download ─────────────────────────────

from desktop_package import (
    _ENV_TEMPLATE, _CONFIG_PY, _CONFIG_BAT, _LAUNCHER_PYW, _START_BAT,
    _START_BAT_LEGACY, _START_SH, _REQUIREMENTS, _README, _INICIAR_VBS, _STOP_BAT,
    _CREATE_SHORTCUT_VBS, _LEEME_TXT,
)


# ─── Gmail OAuth2 Endpoints ───────────────────────────────────
@api_router.get("/oauth/gmail/start")
async def gmail_oauth_start():
    """Return Google OAuth2 authorization URL."""
    cid, csec = await _get_google_credentials_config()
    if not cid or not csec:
        raise HTTPException(status_code=400, detail="Google credentials not configured. Guarda Client ID y Secret en Ajustes.")
    if not GOOGLE_REDIRECT_URI:
        raise HTTPException(status_code=500, detail="GOOGLE_REDIRECT_URI not configured (missing APP_PUBLIC_URL)")
    flow = Flow.from_client_config(
        {"web": {
            "client_id": cid,
            "client_secret": csec,
            "auth_uri": "https://accounts.google.com/o/oauth2/auth",
            "token_uri": "https://oauth2.googleapis.com/token",
            "redirect_uris": [GOOGLE_REDIRECT_URI],
        }},
        scopes=GMAIL_SCOPES,
        redirect_uri=GOOGLE_REDIRECT_URI,
    )
    auth_url, _ = flow.authorization_url(
        access_type="offline",
        prompt="consent",
        include_granted_scopes="true",
    )
    return {"url": auth_url, "redirect_uri": GOOGLE_REDIRECT_URI}


@api_router.get("/oauth/gmail/callback")
async def gmail_oauth_callback(code: str = None, error: str = None):
    """Exchange auth code for tokens and store refresh_token."""
    if error or not code:
        return RedirectResponse(url=f"{APP_PUBLIC_URL}/ajustes?gmail_error={error or 'cancelled'}")
    try:
        cid, csec = await _get_google_credentials_config()
        flow = Flow.from_client_config(
            {"web": {
                "client_id": cid,
                "client_secret": csec,
                "auth_uri": "https://accounts.google.com/o/oauth2/auth",
                "token_uri": "https://oauth2.googleapis.com/token",
                "redirect_uris": [GOOGLE_REDIRECT_URI],
            }},
            scopes=GMAIL_SCOPES,
            redirect_uri=GOOGLE_REDIRECT_URI,
        )
        flow.fetch_token(code=code)
        creds = flow.credentials
        # Get user email
        service = build("oauth2", "v2", credentials=creds)
        user_info = await asyncio.to_thread(service.userinfo().get().execute)
        user_email = user_info.get("email", "")
        # Store tokens
        await db.oauth_tokens.update_one(
            {"provider": "gmail"},
            {"$set": {
                "provider": "gmail",
                "email": user_email,
                "access_token": creds.token,
                "refresh_token": creds.refresh_token,
                "connected_at": datetime.now(timezone.utc).isoformat(),
            }},
            upsert=True,
        )
        logger.info(f"Gmail connected: {user_email}")
        return RedirectResponse(url=f"{APP_PUBLIC_URL}/ajustes?gmail_ok=1")
    except Exception as e:
        logger.error(f"Gmail OAuth callback error: {e}")
        return RedirectResponse(url=f"{APP_PUBLIC_URL}/ajustes?gmail_error={str(e)[:60]}")


@api_router.get("/oauth/gmail/status")
async def gmail_status():
    cid, csec = await _get_google_credentials_config()
    doc = await db.oauth_tokens.find_one({"provider": "gmail"}, {"_id": 0})
    creds_configured = bool(cid and csec)
    if doc and doc.get("refresh_token"):
        return {
            "connected": True,
            "email": doc.get("email", ""),
            "connected_at": doc.get("connected_at"),
            "credentials_configured": creds_configured,
            "redirect_uri": GOOGLE_REDIRECT_URI,
        }
    return {
        "connected": False,
        "email": "",
        "connected_at": None,
        "credentials_configured": creds_configured,
        "redirect_uri": GOOGLE_REDIRECT_URI,
    }


@api_router.delete("/oauth/gmail/disconnect")
async def gmail_disconnect():
    await db.oauth_tokens.delete_one({"provider": "gmail"})
    return {"ok": True}


@api_router.post("/oauth/gmail/test")
async def gmail_test():
    """Send a test email to verify Gmail connection."""
    doc = await db.oauth_tokens.find_one({"provider": "gmail"}, {"_id": 0})
    if not doc or not doc.get("refresh_token"):
        raise HTTPException(status_code=400, detail="Gmail no conectado")
    try:
        html = """
        <div style='font-family:sans-serif;padding:20px'>
          <h2 style='color:#6366f1'>Cinema Productions — Prueba de correo ✓</h2>
          <p>Este es un correo de prueba enviado automáticamente desde tu app de reservas.</p>
          <p>Si lo recibes, los recordatorios automáticos funcionarán correctamente.</p>
        </div>"""
        await _send_gmail(to=doc["email"], subject="✓ Prueba — Cinema Productions", html=html)
        return {"ok": True, "message": f"Correo enviado a {doc['email']}"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ─── WhatsApp Cloud API Endpoints ─────────────────────────────
class WhatsAppTestPayload(BaseModel):
    access_token: Optional[str] = None
    phone_number_id: Optional[str] = None
    recipient: Optional[str] = None
    template_name: Optional[str] = None


@api_router.post("/whatsapp/verify")
async def whatsapp_verify(payload: WhatsAppTestPayload):
    """Ping the Graph API to verify token + phone_number_id are valid."""
    doc = await db.app_settings.find_one({}, {"_id": 0}) or {}
    token = payload.access_token or doc.get("whatsapp_access_token")
    pnid = payload.phone_number_id or doc.get("whatsapp_phone_number_id")
    if not token or not pnid:
        return {"ok": False, "error": "Faltan credenciales de WhatsApp (Access Token y Phone Number ID)"}
    try:
        url = f"https://graph.facebook.com/v20.0/{pnid}"
        async with httpx.AsyncClient(timeout=10) as c:
            r = await c.get(url, headers={"Authorization": f"Bearer {token}"})
        if r.status_code == 200:
            data = r.json()
            return {
                "ok": True,
                "verified_name": data.get("verified_name") or data.get("display_phone_number", ""),
                "display_phone_number": data.get("display_phone_number", ""),
            }
        return {"ok": False, "error": f"Meta API {r.status_code}: {r.text[:200]}"}
    except Exception as e:
        return {"ok": False, "error": f"Error: {e}"}


@api_router.post("/whatsapp/test")
async def whatsapp_test(payload: WhatsAppTestPayload):
    """Send a plain-text test WhatsApp message. Falls back to template if 24h window is closed."""
    doc = await db.app_settings.find_one({}, {"_id": 0}) or {}
    token = payload.access_token or doc.get("whatsapp_access_token")
    pnid = payload.phone_number_id or doc.get("whatsapp_phone_number_id")
    recipient = payload.recipient or doc.get("whatsapp_recipient")
    template = payload.template_name or doc.get("whatsapp_template_name")
    if not token or not pnid or not recipient:
        return {"ok": False, "error": "Faltan credenciales o número destinatario"}
    try:
        await _send_whatsapp_text(
            token, pnid, recipient,
            "*Cinema Productions* — Prueba ✓\n\nWhatsApp conectado correctamente. Recibirás recordatorios de eventos aquí.",
        )
        return {"ok": True, "message": f"Mensaje enviado a +{_normalize_e164(recipient)}"}
    except Exception as e:
        err = str(e)
        # If failure is 131047 (outside 24h window) and template configured -> fallback
        if template and ("131047" in err or "outside" in err.lower() or "re-engagement" in err.lower()):
            try:
                await _send_whatsapp_template(token, pnid, recipient, template)
                return {"ok": True, "message": f"Plantilla '{template}' enviada (fuera de la ventana de 24 h)"}
            except Exception as e2:
                return {"ok": False, "error": f"Texto y plantilla fallaron: {e2}"}
        return {"ok": False, "error": err}


# ─── Web Push Endpoints ───────────────────────────────────────
@api_router.get("/push/vapid-key")
async def get_vapid_key():
    return {"publicKey": VAPID_PUBLIC_KEY}


class PushSubscriptionModel(BaseModel):
    endpoint: str
    expirationTime: Optional[float] = None
    keys: dict


@api_router.post("/push/subscribe")
async def push_subscribe(sub: PushSubscriptionModel):
    await db.push_subscriptions.update_one(
        {"endpoint": sub.endpoint},
        {"$set": {
            "endpoint": sub.endpoint,
            "keys": sub.keys,
            "subscribed_at": datetime.now(timezone.utc).isoformat(),
        }},
        upsert=True,
    )
    return {"ok": True}


@api_router.delete("/push/unsubscribe")
async def push_unsubscribe(endpoint: str):
    await db.push_subscriptions.delete_one({"endpoint": endpoint})
    return {"ok": True}


@api_router.post("/push/test")
async def push_test():
    """Send a test push notification to all subscribers."""
    try:
        await _send_push_to_all(
            title="Cinema Productions — Prueba ✓",
            body="Las notificaciones de escritorio están funcionando.",
            url="/dashboard",
        )
        count = await db.push_subscriptions.count_documents({})
        return {"ok": True, "sent_to": count}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@api_router.post("/telegram/test")
async def telegram_test():
    """Send a test Telegram message. Returns 200 always (K8s ingress strips 4xx bodies)."""
    doc = await db.app_settings.find_one({}, {"_id": 0})
    if not doc or not doc.get("telegram_bot_token") or not doc.get("telegram_chat_id"):
        return {"ok": False, "error": "Configura el token y chat_id de Telegram primero"}
    try:
        await _send_telegram(
            doc["telegram_bot_token"],
            doc["telegram_chat_id"],
            "<b>Cinema Productions</b> — Prueba ✓\n\nTelegram conectado correctamente. Recibirás los recordatorios de eventos aquí.",
        )
        return {"ok": True, "message": "Mensaje enviado a Telegram"}
    except Exception as e:
        return {"ok": False, "error": f"Error de Telegram: {e}"}


@api_router.post("/ntfy/test")
async def ntfy_test():
    """Send a test ntfy.sh notification. Returns 200 always (K8s ingress strips 4xx bodies)."""
    doc = await db.app_settings.find_one({}, {"_id": 0})
    if not doc or not doc.get("ntfy_topic"):
        return {"ok": False, "error": "Configura el tema de ntfy primero"}
    try:
        await _send_ntfy(
            doc["ntfy_topic"],
            "Cinema Productions — Prueba ✓",
            "ntfy conectado correctamente. Recibirás los recordatorios de eventos aquí.",
            priority="high",
        )
        return {"ok": True, "message": f"Notificación enviada al tema: {doc['ntfy_topic']}"}
    except Exception as e:
        return {"ok": False, "error": f"Error de ntfy: {e}"}



_build_state = {"status": "idle", "message": "Listo para actualizar", "started_at": None, "finished_at": None, "progress": 0, "step": 0, "step_label": ""}

# Estado del push a GitHub — alimenta la barra de progreso de la UI
# step / total_steps → permiten mostrar "Paso X de N" en la interfaz.
# detail            → sub-mensaje secundario (ej. archivo/sección en curso).
# started_ts        → timestamp interno (epoch) para calcular segundos transcurridos.
_push_state = {
    "status": "idle",
    "progress": 0,
    "message": "",
    "detail": "",
    "step": 0,
    "total_steps": 8,
    "started_at": None,
    "finished_at": None,
    "started_ts": None,
}

# Etiquetas legibles de cada fase, para el desglose de pasos en la UI.
PUSH_STEP_LABELS = [
    "Conectando con GitHub",
    "Clonando repositorio",
    "Compilando interfaz",
    "Empaquetando archivos",
    "Preparando versión para PC",
    "Registrando cambios (git add)",
    "Creando commit",
    "Subiendo a GitHub",
]

def _set_push_state(progress=None, message=None, status=None, step=None, detail=None):
    """Actualiza el estado del push. Cualquier campo omitido conserva su valor previo."""
    global _push_state
    if status is not None:
        _push_state["status"] = status
    if progress is not None:
        _push_state["progress"] = progress
    if message is not None:
        _push_state["message"] = message
    if detail is not None:
        _push_state["detail"] = detail
    if step is not None:
        _push_state["step"] = step


async def _push_progress_ticker(stop_event, start_ts, lo, hi, base_message, hints=None):
    """Mantiene la barra de progreso en movimiento durante fases largas y bloqueantes
    (por ejemplo `yarn build`). Incrementa el progreso suavemente entre `lo` y `hi`
    y muestra los segundos transcurridos + una pista rotativa de lo que ocurre.

    Esto evita que la barra parezca "congelada" cuando una tarea tarda 1–2 min.
    """
    import time as _time
    hints = hints or ["procesando…"]
    i = 0
    while not stop_event.is_set():
        elapsed = int(_time.time() - start_ts)
        cur = _push_state.get("progress", lo)
        nxt = min(hi, cur + 1) if cur < hi else hi
        hint = hints[i % len(hints)]
        i += 1
        _set_push_state(
            progress=nxt,
            message=f"{base_message} · {elapsed}s transcurridos",
            detail=hint,
        )
        try:
            await asyncio.wait_for(stop_event.wait(), timeout=1.8)
        except asyncio.TimeoutError:
            pass



async def _run_frontend_build():
    global _build_state
    try:
        frontend_dir = str(ROOT_DIR.parent / "frontend")
        _build_state = {**_build_state, "message": "Preparando entorno de compilación…", "progress": 15, "step": 1, "step_label": "Preparando entorno"}
        await asyncio.sleep(0.3)
        _build_state = {**_build_state, "message": "Compilando frontend con yarn build…", "progress": 35, "step": 2, "step_label": "Compilando frontend"}
        process = await asyncio.create_subprocess_exec(
            "yarn", "build",
            cwd=frontend_dir,
            stdout=asyncio.subprocess.PIPE,
            stderr=asyncio.subprocess.PIPE,
            env={
                **os.environ,
                "CI": "false",
                "GENERATE_SOURCEMAP": "false",
                "DISABLE_ESLINT_PLUGIN": "true",
                "REACT_APP_BACKEND_URL": "http://localhost:8001",
                "PUBLIC_URL": "/",
                "NODE_OPTIONS": "--max-old-space-size=4096",
            },
        )
        try:
            stdout, stderr = await asyncio.wait_for(process.communicate(), timeout=600)
        except asyncio.TimeoutError:
            try:
                process.kill()
            except Exception:
                pass
            _build_state = {**_build_state, "status": "error", "message": "Tiempo de espera agotado (10 min). Inténtalo de nuevo.", "progress": 0}
            return

        if process.returncode == 0:
            # Verificar que el build realmente exista
            build_dir = ROOT_DIR.parent / "frontend" / "build"
            if not (build_dir / "index.html").exists():
                _build_state = {**_build_state, "status": "error", "message": "El build terminó pero no se generó index.html. Revisa los logs.", "progress": 0, "step": 0, "step_label": ""}
                logger.error("Build succeeded but no index.html generated")
                return
            _build_state = {**_build_state, "message": "Empaquetando archivos y generando .zip…", "progress": 80, "step": 3, "step_label": "Empaquetando .zip"}
            await asyncio.sleep(0.3)
            _build_state = {
                "status": "ready",
                "message": "✓ App compilada correctamente. Ya puedes descargarla.",
                "started_at": _build_state["started_at"],
                "finished_at": datetime.now(timezone.utc).isoformat(),
                "progress": 100,
                "step": 4,
                "step_label": "Listo para descargar",
            }
            logger.info("Frontend build completed successfully")
        else:
            err_full = stderr.decode("utf-8", errors="replace")
            # Extraer las últimas líneas útiles del error
            err_lines = [l for l in err_full.strip().splitlines() if l.strip() and not l.strip().startswith("$")]
            err_summary = "\n".join(err_lines[-8:]) if err_lines else err_full[:400]
            _build_state = {
                "status": "error",
                "message": f"Error en la compilación:\n{err_summary}",
                "started_at": _build_state["started_at"],
                "finished_at": datetime.now(timezone.utc).isoformat(),
                "progress": 0,
                "step": 0,
                "step_label": "",
            }
            logger.error(f"Frontend build failed: {err_full[:1000]}")
    except FileNotFoundError:
        _build_state = {**_build_state, "status": "error", "message": "yarn no encontrado en el sistema. Instala Node.js + yarn.", "progress": 0, "step": 0, "step_label": ""}
    except Exception as e:
        _build_state = {**_build_state, "status": "error", "message": f"Error inesperado: {str(e)[:200]}", "progress": 0, "step": 0, "step_label": ""}


@api_router.post("/download/package/rebuild")
async def rebuild_package():
    global _build_state
    if _build_state["status"] == "building":
        return {"status": "building", "message": "Ya hay una compilación en progreso. Por favor espera.", "progress": _build_state.get("progress", 10)}
    _build_state = {
        "status": "building",
        "message": "Iniciando compilación… esto tarda 1–3 minutos.",
        "started_at": datetime.now(timezone.utc).isoformat(),
        "finished_at": None,
        "progress": 10,
        "step": 1,
        "step_label": "Iniciando compilación",
    }
    asyncio.create_task(_run_frontend_build())
    return _build_state


@api_router.get("/download/package/build-status")
async def get_build_status():
    return _build_state


# ── Descarga .EXE (Windows, sin Python) desde GitHub Releases ─────────────────
# El .exe se compila automáticamente por GitHub Actions (.github/workflows/build-exe.yml)
# en cada push de tag v* o manualmente (workflow_dispatch). Aquí sólo consultamos
# la Release y devolvemos la URL del asset .exe. Ventajas frente a compilar en
# el servidor con PyInstaller bajo demanda:
#   · descarga instantánea (no hay 3–5 min de espera)
#   · sin consumo de CPU/RAM del servidor
#   · el binario está firmado por el runner de GitHub (SHA verificable)
#   · funciona igual en preview y en la app de escritorio empaquetada
_DEFAULT_EXE_REPO = "AlejandroPiedrasanta/RESERVA-DE-EVENTOS"


def _github_exe_repo() -> str:
    """owner/repo desde env GITHUB_EXE_REPO o fallback constante."""
    v = (os.environ.get("GITHUB_EXE_REPO") or "").strip()
    return v or _DEFAULT_EXE_REPO


async def _find_latest_exe_asset(kind: str = "portable") -> dict | None:
    """Busca el asset .exe más reciente en Releases del repo.
    kind='portable' → excluye archivos con 'setup' en el nombre.
    kind='installer' → sólo archivos que contengan 'setup' (case-insensitive).
    Devuelve dict con keys: name, size, url, tag, published_at,
    browser_download_url, sha256 — o None.
    El sha256 se resuelve leyendo el asset .sha256 hermano (si existe) o
    extrayéndolo del body del release (patrón hexadecimal de 64 chars)."""
    repo = _github_exe_repo()
    headers = {"Accept": "application/vnd.github+json", "User-Agent": "cinema-productions"}
    tok = (os.environ.get("GITHUB_TOKEN") or "").strip()
    if tok:
        headers["Authorization"] = f"Bearer {tok}"

    def _matches(name: str) -> bool:
        n = (name or "").lower()
        if not n.endswith(".exe"):
            return False
        is_setup = "setup" in n or "installer" in n
        return is_setup if kind == "installer" else not is_setup

    async def _extract(rel: dict, client: httpx.AsyncClient) -> dict | None:
        exe_asset = None
        sha_asset = None
        for a in rel.get("assets") or []:
            n = (a.get("name") or "")
            if _matches(n):
                exe_asset = a
        if not exe_asset:
            return None
        # Buscar el .sha256 hermano cuyo nombre empiece con el del .exe
        exe_name_lower = exe_asset["name"].lower()
        for a in rel.get("assets") or []:
            an = (a.get("name") or "").lower()
            if an.endswith(".sha256") and an.startswith(exe_name_lower):
                sha_asset = a
                break
        sha256 = ""
        # 1) Intentar leer del asset .sha256 hermano
        if sha_asset:
            try:
                r = await client.get(sha_asset["browser_download_url"], timeout=8.0)
                if r.status_code == 200:
                    # formato típico: "<hex>  filename"
                    m = re.search(r"\b([a-fA-F0-9]{64})\b", r.text or "")
                    if m:
                        sha256 = m.group(1).lower()
            except Exception as e:
                logger.warning(f"No se pudo leer sha256 asset: {e}")
        # 2) Fallback: extraer del body del release (busca hash junto al nombre del asset)
        if not sha256:
            body = rel.get("body") or ""
            # Buscar hash específico para este exe (línea que mencione el nombre)
            for line in body.splitlines():
                if exe_asset["name"] in line:
                    m = re.search(r"\b([a-fA-F0-9]{64})\b", line)
                    if m:
                        sha256 = m.group(1).lower()
                        break
            # Último fallback: cualquier hash del body (sólo si sólo hay 1 tipo)
            if not sha256:
                m = re.search(r"\b([a-fA-F0-9]{64})\b", body)
                if m and kind == "portable":  # evitar mezclar hashes entre kinds
                    sha256 = m.group(1).lower()
        # Timestamp real del binario subido: preferimos updated_at/created_at
        # del asset (cambia cada vez que el CI re-sube el .exe a un tag rodante
        # como 'latest-exe'), con fallback al published_at del release.
        ts = (exe_asset.get("updated_at") or exe_asset.get("created_at")
              or rel.get("published_at") or rel.get("created_at") or "")
        return {
            "name": exe_asset["name"], "size": exe_asset.get("size") or 0,
            "url": exe_asset["browser_download_url"],
            "browser_download_url": exe_asset["browser_download_url"],
            "tag": rel.get("tag_name") or "", "published_at": rel.get("published_at") or "",
            "sha256": sha256,
            "_ts": ts,
        }

    # ── Escanear TODOS los releases (incluye prereleases como 'latest-exe') y
    #    elegir el asset .exe con el timestamp de subida MÁS RECIENTE. No se
    #    confía en /releases/latest porque GitHub ignora los prereleases, y el
    #    workflow_dispatch publica el binario nuevo como prerelease 'latest-exe':
    #    /releases/latest devolvería un release estable MÁS VIEJO → el usuario
    #    descargaba un .exe desactualizado. Comparar por _ts corrige el bug.
    async with httpx.AsyncClient(timeout=10.0, follow_redirects=True) as client:
        candidates: list[dict] = []
        try:
            r = await client.get(f"https://api.github.com/repos/{repo}/releases?per_page=30", headers=headers)
            if r.status_code == 200:
                for rel in r.json() or []:
                    if rel.get("draft"):
                        continue
                    res = await _extract(rel, client)
                    if res:
                        candidates.append(res)
        except Exception as e:
            logger.warning(f"GitHub releases list falló: {e}")

        if candidates:
            candidates.sort(key=lambda c: c.get("_ts") or "", reverse=True)
            best = candidates[0]
            best.pop("_ts", None)
            return best

        # Fallback: si la lista falló por completo, probar /releases/latest.
        try:
            r = await client.get(f"https://api.github.com/repos/{repo}/releases/latest", headers=headers)
            if r.status_code == 200:
                res = await _extract(r.json(), client)
                if res:
                    res.pop("_ts", None)
                    return res
        except Exception as e:
            logger.warning(f"GitHub releases/latest falló: {e}")
    return None


@api_router.get("/download/desktop-exe/info")
async def download_desktop_exe_info():
    """Metadata del último .exe disponible (tamaño, versión, URL). Si no hay
    release aún, devuelve status='not_available' con instrucciones."""
    asset = await _find_latest_exe_asset("portable")
    repo = _github_exe_repo()
    if not asset:
        return {
            "status": "not_available",
            "repo": repo,
            "workflow_url": f"https://github.com/{repo}/actions/workflows/build-exe.yml",
            "releases_url": f"https://github.com/{repo}/releases",
            "message": (
                "Aún no hay un .exe publicado. Ejecuta el workflow 'Build Windows .exe' "
                "en GitHub Actions o publica un tag v* para generar y publicar el "
                "instalador automáticamente."
            ),
        }
    return {
        "status": "ready",
        "repo": repo,
        "name": asset["name"],
        "size": asset["size"],
        "size_mb": round((asset["size"] or 0) / (1024 * 1024), 1),
        "tag": asset["tag"],
        "published_at": asset["published_at"],
        "url": asset["url"],
        "sha256": asset.get("sha256") or "",
        "sha256_url": f"{asset['url']}.sha256",
    }


@api_router.get("/download/desktop-exe")
async def download_desktop_exe():
    """Redirige al .exe más reciente publicado en GitHub Releases."""
    asset = await _find_latest_exe_asset("portable")
    if not asset:
        repo = _github_exe_repo()
        raise HTTPException(
            status_code=404,
            detail=(
                f"Aún no hay un .exe publicado. Ejecuta el workflow "
                f"'Build Windows .exe' en https://github.com/{repo}/actions "
                f"o publica un tag v* para generar el instalador automáticamente."
            ),
        )
    # 302 → GitHub CDN (descarga directa, sin proxy por nuestro servidor)
    return RedirectResponse(url=asset["url"], status_code=302)


@api_router.get("/download/desktop-installer/info")
async def download_desktop_installer_info():
    """Metadata del último instalador .exe (Inno Setup) publicado en GitHub
    Releases. Filtra assets cuyo nombre contenga 'Setup' o 'installer'."""
    asset = await _find_latest_exe_asset("installer")
    repo = _github_exe_repo()
    if not asset:
        return {
            "status": "not_available",
            "repo": repo,
            "workflow_url": f"https://github.com/{repo}/actions/workflows/build-exe.yml",
            "releases_url": f"https://github.com/{repo}/releases",
            "message": (
                "Aún no hay un instalador publicado. Ejecuta el workflow "
                "'Build Windows .exe' en GitHub Actions — genera CinemaProductions-Setup.exe "
                "automáticamente además del portable."
            ),
        }
    return {
        "status": "ready",
        "repo": repo,
        "name": asset["name"],
        "size": asset["size"],
        "size_mb": round((asset["size"] or 0) / (1024 * 1024), 1),
        "tag": asset["tag"],
        "published_at": asset["published_at"],
        "url": asset["url"],
        "sha256": asset.get("sha256") or "",
        "sha256_url": f"{asset['url']}.sha256",
    }


@api_router.get("/download/desktop-installer")
async def download_desktop_installer():
    """Redirige al instalador (.exe con Inno Setup) más reciente."""
    asset = await _find_latest_exe_asset("installer")
    if not asset:
        repo = _github_exe_repo()
        raise HTTPException(
            status_code=404,
            detail=(
                f"Aún no hay un instalador publicado. Ejecuta el workflow "
                f"'Build Windows .exe' en https://github.com/{repo}/actions."
            ),
        )
    return RedirectResponse(url=asset["url"], status_code=302)


@api_router.get("/download/package")
async def download_package(request: Request):
    import zipfile
    import pyzipper

    build_dir = ROOT_DIR.parent / "frontend" / "build"
    frontend_dir = ROOT_DIR.parent / "frontend"
    src_dir = frontend_dir / "src"

    # ── AUTO-REBUILD si el build está desactualizado ─────────────────────
    # Si algún archivo del src es más nuevo que el build/index.html, recompila
    # automáticamente antes de empaquetar. Esto garantiza que el ejecutable
    # siempre contenga la última versión del código.
    needs_rebuild = False
    if not build_dir.exists() or not (build_dir / "index.html").exists():
        needs_rebuild = True
    else:
        try:
            build_mtime = (build_dir / "index.html").stat().st_mtime
            for src_file in src_dir.rglob("*"):
                if src_file.is_file() and src_file.stat().st_mtime > build_mtime:
                    needs_rebuild = True
                    logger.info(f"Auto-rebuild: {src_file.name} más nuevo que build/")
                    break
        except Exception as e:
            logger.warning(f"No se pudo comprobar mtimes del build: {e}")

    if needs_rebuild:
        logger.warning("Frontend build desactualizado — recompilando (esto tarda 1-2 min)...")
        rebuild = await asyncio.to_thread(
            subprocess.run,
            ["yarn", "build"],
            capture_output=True, text=True, timeout=600, cwd=str(frontend_dir),
        )
        if rebuild.returncode != 0:
            raise HTTPException(
                status_code=500,
                detail=f"Auto-rebuild del frontend falló: {(rebuild.stderr or rebuild.stdout)[-500:]}"
            )
        logger.info("✓ Frontend recompilado exitosamente")

    if not build_dir.exists() or not (build_dir / "index.html").exists():
        raise HTTPException(
            status_code=503,
            detail="El paquete aun no esta listo. Ve a Ajustes → App de Escritorio y pulsa 'Compilar app' primero. Toma 1-3 minutos."
        )

    cloud_url = str(request.base_url).rstrip("/")
    standalone_py = (ROOT_DIR / 'standalone_app.py').read_text()

    # Version UNIFICADA: mismo esquema semver X.Y.Z que version.txt/cloud/EXE/commits.
    # Fuente única de verdad → _compute_next_unified_version().
    auto_version = await _compute_next_unified_version()
    # Sincronizar contador legacy (informativo, ya no autoritativo).
    _cdoc = await db.app_settings.find_one({}, {"desktop_build": 1, "security_config": 1}) or {}
    await db.app_settings.update_one(
        {}, {"$set": {"desktop_build_semver": auto_version}}, upsert=True
    )

    # ── ¿Cifrar ZIP con contraseña? ─────────────────────────────────
    sec_cfg = _cdoc.get("security_config") or {}
    zip_pwd = sec_cfg.get("zip_password") or DEFAULT_ZIP_PASSWORD
    zip_pwd_enabled = bool(sec_cfg.get("zip_password_enabled", True))

    buf = io.BytesIO()

    def _win_lines(s: str) -> str:
        return s.replace('\r\n', '\n').replace('\n', '\r\n')

    # ── Generar icono .ico desde el logo.png del frontend ────────────────
    # Se usa Pillow (ya en requirements) para producir un ICO multi-tamaño
    # que Windows muestre en el .lnk del escritorio. Como el logo.png es
    # una banner ancha (587x77), lo pegamos centrado en un canvas cuadrado
    # con degradado indigo para que se vea como un icono de app real.
    icon_bytes = b""
    try:
        from PIL import Image, ImageDraw
        logo_png = frontend_dir / "public" / "logo.png"
        canvas_size = 256
        canvas = Image.new("RGBA", (canvas_size, canvas_size), (0, 0, 0, 0))
        # Fondo redondeado con degradado indigo → violeta
        draw = ImageDraw.Draw(canvas)
        for y in range(canvas_size):
            t = y / canvas_size
            r = int(79 + (139 - 79) * t)   # 4f → 8b
            g = int(70 + (92 - 70) * t)    # 46 → 5c
            b = int(229 + (246 - 229) * t) # e5 → f6
            draw.line([(0, y), (canvas_size, y)], fill=(r, g, b, 255))
        # Máscara redondeada
        mask = Image.new("L", (canvas_size, canvas_size), 0)
        ImageDraw.Draw(mask).rounded_rectangle(
            (0, 0, canvas_size, canvas_size), radius=48, fill=255
        )
        canvas.putalpha(mask)
        # Pegar el logo escalado y centrado
        if logo_png.exists():
            with Image.open(logo_png) as logo:
                logo = logo.convert("RGBA")
                # Escalar el logo para que quepa dentro del 78% del canvas
                target_w = int(canvas_size * 0.78)
                ratio = target_w / logo.width
                target_h = int(logo.height * ratio)
                logo_resized = logo.resize((target_w, target_h), Image.LANCZOS)
                px = (canvas_size - target_w) // 2
                py = (canvas_size - target_h) // 2
                canvas.paste(logo_resized, (px, py), logo_resized)
        ico_buf = io.BytesIO()
        canvas.save(ico_buf, format="ICO",
                    sizes=[(16, 16), (32, 32), (48, 48), (64, 64), (128, 128), (256, 256)])
        icon_bytes = ico_buf.getvalue()
    except Exception as _ico_err:
        logger.warning(f"No se pudo generar icono.ico: {_ico_err}")

    # ── Estructura del ZIP ───────────────────────────────────────────────
    #   cinema-productions/
    #     ├── START.BAT      <- abrir la app (usuario)
    #     ├── DETENER.BAT    <- cerrar la app (usuario)
    #     └── SISTEMA/       <- TODO el motor (app.py, launcher, libs, build,
    #                            .env, icono, LEEME, backups, etc.)
    #   En la raiz SOLO existen START.BAT y DETENER.BAT.
    ROOT = 'cinema-productions/'
    SYS = 'cinema-productions/SISTEMA/'

    files_to_add = [
        # -- Nivel raíz: SOLO los dos .bat que el usuario usa --
        (ROOT + 'START.BAT',    _win_lines(_START_BAT).encode('utf-8')),
        (ROOT + 'DETENER.BAT',  _win_lines(_STOP_BAT).encode('utf-8')),

        # -- SISTEMA: motor técnico completo --
        (SYS + 'app.py',            standalone_py.encode('utf-8')),
        (SYS + '.env',              _win_lines(_ENV_TEMPLATE).encode('utf-8')),
        (SYS + 'requirements.txt',  _REQUIREMENTS.encode('utf-8')),
        (SYS + 'start.sh',          _START_SH.encode('utf-8')),
        (SYS + 'launcher.pyw',      _LAUNCHER_PYW.encode('utf-8')),
        (SYS + 'config.py',         _CONFIG_PY.encode('utf-8')),
        (SYS + 'config.bat',        _win_lines(_CONFIG_BAT).encode('utf-8')),
        (SYS + 'LEEME.txt',         _win_lines(_LEEME_TXT).encode('utf-8')),
        (SYS + 'README-tecnico.txt', _win_lines(_README).encode('utf-8')),
        (SYS + 'version.txt',       auto_version.encode('utf-8')),
    ]

    # Icono binario (si Pillow lo generó) -> dentro de SISTEMA
    if icon_bytes:
        files_to_add.append((SYS + 'icono.ico', icon_bytes))

    # Adjuntar el mirror actual de saved_themes.json (contiene "Minimalista"
    # como is_default) para que la app de escritorio arranque con el mismo
    # tema por defecto que la web incluso sin internet.
    try:
        themes_payload = await _themes_snapshot_payload()
        themes_json_bytes = json.dumps(themes_payload, indent=2, ensure_ascii=False).encode('utf-8')
        files_to_add.append(
            (SYS + 'themes/saved_themes.json', themes_json_bytes)
        )
    except Exception as _themes_err:
        logger.warning(f"No se pudo adjuntar themes/saved_themes.json al ZIP: {_themes_err}")

    wheels_dir = await asyncio.to_thread(_ensure_desktop_wheels)
    wheel_files = sorted(wheels_dir.glob('*.whl'))
    build_files = [f for f in sorted(build_dir.rglob('*')) if f.is_file()]

    if zip_pwd_enabled and zip_pwd:
        # ZIP cifrado AES-256 (compatible con WinRAR, 7-Zip; NO con el zip de Windows nativo)
        with pyzipper.AESZipFile(buf, 'w', compression=pyzipper.ZIP_DEFLATED,
                                 encryption=pyzipper.WZ_AES) as zf:
            zf.setpassword(zip_pwd.encode('utf-8'))
            for arc_name, content in files_to_add:
                zf.writestr(arc_name, content)
            for whl in wheel_files:
                zf.write(str(whl), SYS + 'libs/' + whl.name)
            for file_path in build_files:
                arc_name = SYS + 'build/' + str(file_path.relative_to(build_dir))
                zf.write(str(file_path), arc_name)
            # Nota informativa dentro del ZIP (sin cifrar sería útil, pero por consistencia
            # dejamos todo cifrado y ponemos la ayuda en README.txt).
    else:
        # ZIP normal sin cifrar
        with zipfile.ZipFile(buf, 'w', zipfile.ZIP_DEFLATED) as zf:
            for arc_name, content in files_to_add:
                zf.writestr(arc_name, content)
            for whl in wheel_files:
                zf.write(str(whl), SYS + 'libs/' + whl.name)
            for file_path in build_files:
                arc_name = SYS + 'build/' + str(file_path.relative_to(build_dir))
                zf.write(str(file_path), arc_name)

    zip_bytes = buf.getvalue()
    filename = f"cinema-productions-{auto_version}.zip"
    safe_name = f"auto_{auto_version.replace('.', '_')}.zip"
    file_path = UPDATES_DIR / safe_name
    file_path.write_bytes(zip_bytes)

    # Auto-register this version in the shared MongoDB database
    await db.app_updates.update_many({}, {"$set": {"is_latest": False}})
    update_doc = {
        "version": auto_version,
        "filename": filename,
        "stored_name": safe_name,
        "notes": "Generada automáticamente al descargar desde Ajustes"
                 + (" (protegido con contraseña)" if zip_pwd_enabled else ""),
        "channel": "stable",
        "file_size": len(zip_bytes),
        "created_at": datetime.now(timezone.utc).isoformat(),
        "is_latest": True,
        "download_url": f"{cloud_url}/api/updates/download",
        "encrypted": zip_pwd_enabled,
    }
    await db.app_updates.insert_one(update_doc)

    headers = {'Content-Disposition': f'attachment; filename={filename}'}
    if zip_pwd_enabled:
        headers['X-Zip-Encrypted'] = 'AES-256'
    return Response(
        content=zip_bytes,
        media_type='application/zip',
        headers=headers,
    )




# ── APP UPDATES ──────────────────────────────────────────────────────────────

@api_router.post("/updates/upload")
async def upload_app_update(
    file: UploadFile = File(...),
    version: str = Form(...),
    notes: str = Form(""),
    channel: str = Form("stable"),
):
    content = await file.read()
    file_id = str(uuid.uuid4())
    safe_name = f"{file_id}_{file.filename}"
    file_path = UPDATES_DIR / safe_name
    file_path.write_bytes(content)

    # Mark all previous as not latest
    await db.app_updates.update_many({}, {"$set": {"is_latest": False}})

    doc = {
        "version": version,
        "filename": file.filename,
        "stored_name": safe_name,
        "notes": notes,
        "channel": channel,
        "file_size": len(content),
        "created_at": datetime.now(timezone.utc).isoformat(),
        "is_latest": True,
    }
    result = await db.app_updates.insert_one(doc)
    return {**{k: v for k, v in doc.items() if k != "_id"}, "id": str(result.inserted_id)}


@api_router.get("/updates/latest")
async def get_latest_update():
    doc = await db.app_updates.find_one({"is_latest": True}, sort=[("created_at", -1)])
    if not doc:
        raise HTTPException(status_code=404, detail="No hay actualizaciones disponibles")
    return {
        "id": str(doc["_id"]),
        "version": doc["version"],
        "filename": doc["filename"],
        "notes": doc.get("notes", ""),
        "channel": doc.get("channel", "stable"),
        "file_size": doc["file_size"],
        "created_at": doc["created_at"],
    }


# Caché en memoria del historial de tags de GitHub (evita llamadas lentas repetidas)
_gh_history_cache = {"data": None, "ts": 0.0, "loading": False}
_GH_HISTORY_TTL = 300  # segundos (5 min)


def _gh_cache_is_fresh() -> bool:
    import time as _time
    return _gh_history_cache["data"] is not None and (_time.time() - _gh_history_cache["ts"]) < _GH_HISTORY_TTL


async def _fetch_github_tag_records(owner: str, repo: str, token: str) -> list:
    """Trae los tags de GitHub + fecha/nota de cada commit EN PARALELO.
    Cacheado 5 min. Se ejecuta en segundo plano para que el historial sea inmediato."""
    import time as _time
    import urllib.request, json as _json

    if _gh_cache_is_fresh():
        return _gh_history_cache["data"]
    _gh_history_cache["loading"] = True

    headers = {"User-Agent": "cinema-productions", "Accept": "application/vnd.github+json"}
    if token:
        headers["Authorization"] = f"Bearer {token}"

    def _get_json(url, timeout):
        req = urllib.request.Request(url, headers=headers)
        with urllib.request.urlopen(req, timeout=timeout) as resp:
            return _json.loads(resp.read().decode())

    try:
        tags_data = await asyncio.to_thread(
            _get_json, f"https://api.github.com/repos/{owner}/{repo}/tags?per_page=30", 6
        )
    except Exception as e:
        logger.warning(f"No se pudo leer tags de GitHub: {e}")
        _gh_history_cache["loading"] = False
        return _gh_history_cache["data"] or []

    tags_data = tags_data[:20]  # límite razonable para mantenerlo rápido

    async def _one(tag):
        tag_name = tag.get("name", "")
        version = tag_name.lstrip("v")
        commit_sha = tag.get("commit", {}).get("sha", "")
        created_at, notes = "", ""
        try:
            cdata = await asyncio.to_thread(
                _get_json, f"https://api.github.com/repos/{owner}/{repo}/commits/{commit_sha}", 5
            )
            created_at = cdata.get("commit", {}).get("author", {}).get("date", "")
            notes = cdata.get("commit", {}).get("message", "")[:200]
        except Exception:
            pass
        return {
            "id": f"gh:{tag_name}",
            "version": version,
            "filename": "",
            "notes": notes,
            "channel": "github",
            "file_size": 0,
            "created_at": created_at,
            "is_latest": False,
            "source": "github_tag",
            "commit_short": commit_sha[:7],
            "author": "",
            "branch": "main",
        }

    try:
        gh_records = list(await asyncio.gather(*[_one(t) for t in tags_data]))
        _gh_history_cache["data"] = gh_records
        _gh_history_cache["ts"] = _time.time()
        return gh_records
    finally:
        _gh_history_cache["loading"] = False


@api_router.get("/updates/history")
async def get_update_history():
    """Historial que combina registros locales + tags de GitHub (cacheados)."""
    cursor = db.app_updates.find({}, sort=[("created_at", -1)])
    docs = await cursor.to_list(200)
    records = [
        {
            "id": str(d["_id"]),
            "version": d.get("version", "?"),
            "filename": d.get("filename", ""),
            "notes": d.get("notes", ""),
            "channel": d.get("channel", "stable"),
            "file_size": d.get("file_size", 0),
            "created_at": d.get("created_at", ""),
            "is_latest": d.get("is_latest", False),
            "source": d.get("source", "package"),
            "commit_short": d.get("commit_short", ""),
            "author": d.get("author", ""),
            "branch": d.get("branch", ""),
        }
        for d in docs
    ]

    # Añadir tags de GitHub (cacheado 5 min + fetch en paralelo para no bloquear)
    cfg = await _get_github_config()
    repo_url = cfg.get("repo_url", "")
    token = cfg.get("token", "")
    owner, repo = _parse_github_url(repo_url)
    if owner and repo:
        seen_versions = {r["version"] for r in records}
        # Nunca bloquear: si la caché está fresca la usamos; si no, refrescamos en
        # segundo plano y devolvemos ya los registros locales (respuesta inmediata).
        if _gh_cache_is_fresh():
            gh_records = _gh_history_cache["data"] or []
        else:
            if not _gh_history_cache["loading"]:
                asyncio.create_task(_fetch_github_tag_records(owner, repo, token))
            gh_records = _gh_history_cache["data"] or []
        for gr in gh_records:
            if gr["version"] not in seen_versions:
                records.append(gr)

    # Ordenar y marcar latest (defensivo ante created_at None/datetime/str mezclados)
    def _sort_key(r):
        v = r.get("created_at", "")
        if v is None:
            return ""
        if hasattr(v, "isoformat"):
            try:
                return v.isoformat()
            except Exception:
                return ""
        return str(v)
    records.sort(key=_sort_key, reverse=True)
    if records:
        for r in records:
            r["is_latest"] = False
        records[0]["is_latest"] = True

    # Serializar cualquier datetime restante a ISO string
    for r in records:
        v = r.get("created_at")
        if hasattr(v, "isoformat"):
            try:
                r["created_at"] = v.isoformat()
            except Exception:
                r["created_at"] = str(v)

    return records


@api_router.get("/updates/download")
async def download_latest_update():
    """Download the currently active (is_latest) update file."""
    doc = await db.app_updates.find_one({"is_latest": True}, sort=[("created_at", -1)])
    if not doc:
        raise HTTPException(status_code=404, detail="No hay actualización activa")
    file_path = UPDATES_DIR / doc["stored_name"]
    if not file_path.exists():
        raise HTTPException(status_code=404, detail="Archivo no encontrado en el servidor")

    def iter_file():
        with open(file_path, "rb") as f:
            while chunk := f.read(1024 * 1024):
                yield chunk

    return StreamingResponse(
        iter_file(),
        media_type="application/octet-stream",
        headers={"Content-Disposition": f'attachment; filename="{doc["filename"]}"'},
    )


@api_router.get("/updates/download/{update_id}")
async def download_update(update_id: str):
    try:
        oid = ObjectId(update_id)
    except Exception:
        raise HTTPException(status_code=400, detail="ID inválido")
    doc = await db.app_updates.find_one({"_id": oid})
    if not doc:
        raise HTTPException(status_code=404, detail="Actualización no encontrada")
    file_path = UPDATES_DIR / doc["stored_name"]
    if not file_path.exists():
        raise HTTPException(status_code=404, detail="Archivo no encontrado en el servidor")

    def iter_file():
        with open(file_path, "rb") as f:
            while chunk := f.read(1024 * 1024):
                yield chunk

    return StreamingResponse(
        iter_file(),
        media_type="application/octet-stream",
        headers={"Content-Disposition": f'attachment; filename="{doc["filename"]}"'},
    )


@api_router.put("/updates/{update_id}/set-latest")
async def set_latest_update(update_id: str):
    try:
        oid = ObjectId(update_id)
    except Exception:
        raise HTTPException(status_code=400, detail="ID inválido")
    await db.app_updates.update_many({}, {"$set": {"is_latest": False}})
    result = await db.app_updates.update_one({"_id": oid}, {"$set": {"is_latest": True}})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Actualización no encontrada")
    return {"message": "Versión marcada como activa"}


@api_router.delete("/updates/{update_id}")
async def delete_update(update_id: str):
    try:
        oid = ObjectId(update_id)
    except Exception:
        raise HTTPException(status_code=400, detail="ID inválido")
    doc = await db.app_updates.find_one({"_id": oid})
    if not doc:
        raise HTTPException(status_code=404, detail="Actualización no encontrada")
    file_path = UPDATES_DIR / doc["stored_name"]
    if file_path.exists():
        file_path.unlink()
    await db.app_updates.delete_one({"_id": oid})
    # If deleted was the latest, promote the next most recent
    if doc.get("is_latest"):
        newer = await db.app_updates.find_one({}, sort=[("created_at", -1)])
        if newer:
            await db.app_updates.update_one({"_id": newer["_id"]}, {"$set": {"is_latest": True}})
    return {"message": "Actualización eliminada"}


async def _read_local_version() -> str:
    try:
        for candidate in (ROOT_DIR / "version.txt", ROOT_DIR.parent / "version.txt"):
            if candidate.exists():
                return candidate.read_text(encoding="utf-8").strip().splitlines()[0].strip()
    except Exception:
        pass
    return ""


# ─────────────────────────────────────────────────────────────────────
# UNIFIED VERSION COMPUTATION
# Fuente única de verdad para TODOS los flujos: cloud release, ZIP,
# EXE, commits, y tags. Formato semver X.Y.Z.
#
# Reglas:
#   1. Base = max(version.txt local, version.txt remoto, max tag semver remoto)
#   2. Next = base con patch + 1
#   3. Si "vNext" ya existe como tag remoto, sigue subiendo patch hasta libre.
#   4. Nunca retrocede. Nunca crea 2-part (siempre X.Y.Z).
# ─────────────────────────────────────────────────────────────────────
_SEMVER_RE = re.compile(r"^v?(\d+)\.(\d+)\.(\d+)$")
_SEMVER_RE_2 = re.compile(r"^v?(\d+)\.(\d+)$")


def _parse_semver(s: str):
    if not s:
        return None
    s = s.strip()
    m = _SEMVER_RE.match(s)
    if m:
        return (int(m.group(1)), int(m.group(2)), int(m.group(3)))
    # Aceptar también 2-partes (X.Y → X.Y.0) para tags como v1.20
    m2 = _SEMVER_RE_2.match(s)
    if m2:
        return (int(m2.group(1)), int(m2.group(2)), 0)
    return None


def _fmt_semver(t) -> str:
    return f"{t[0]}.{t[1]}.{t[2]}"


async def _list_remote_semver_tags() -> list:
    """Lista tags remotos vX.Y.Z usando la GitHub API (sin clonar)."""
    tags_semver = []
    try:
        cfg = await _get_github_config()
        repo_url = cfg.get("repo_url") or ""
        token = (cfg.get("token") or "").strip()
        if not repo_url:
            return []
        m = re.match(r"https?://github\.com/([^/]+)/([^/.]+)", repo_url)
        if not m:
            return []
        owner, repo = m.group(1), m.group(2)
        headers = {"User-Agent": "cinema-productions",
                   "Accept": "application/vnd.github+json"}
        if token:
            headers["Authorization"] = f"Bearer {token}"
        import urllib.request as _u, json as _j

        def _get_tags():
            req = _u.Request(
                f"https://api.github.com/repos/{owner}/{repo}/tags?per_page=100",
                headers=headers,
            )
            with _u.urlopen(req, timeout=8) as resp:
                return _j.loads(resp.read().decode("utf-8", errors="ignore"))

        tags = await asyncio.to_thread(_get_tags)
        if isinstance(tags, list):
            for t in tags:
                sv = _parse_semver(t.get("name") or "")
                # Solo aceptar la serie oficial major=1 (los tags v2001.X,
                # v1.999, etc. son basura de pruebas y rompen la numeración).
                if sv and sv[0] == 1:
                    tags_semver.append(sv)
    except Exception as e:
        logger.warning(f"_list_remote_semver_tags: {e}")
    return tags_semver


async def _compute_next_unified_version() -> str:
    """Calcula la ÚNICA próxima versión semver X.Y.Z canónica.

    Se usa en: ZIP generator, push a GitHub, endpoint /next-version.
    Garantiza que ZIP, EXE, cloud release, commit y tag tienen el MISMO número.
    """
    candidates = []

    local = _parse_semver(await _read_local_version())
    if local:
        candidates.append(local)

    try:
        gh = await _fetch_github_version_txt()
        remote_ver = _parse_semver(gh.get("version") or "")
        if remote_ver:
            candidates.append(remote_ver)
    except Exception:
        pass

    remote_tags = await _list_remote_semver_tags()
    candidates.extend(remote_tags)

    if not candidates:
        base = (1, 0, 0)
    else:
        base = max(candidates)

    # Bump patch + 1, y saltar colisiones con tags existentes.
    tag_set = {_fmt_semver(t) for t in remote_tags}
    nxt = (base[0], base[1], base[2] + 1)
    while _fmt_semver(nxt) in tag_set:
        nxt = (nxt[0], nxt[1], nxt[2] + 1)
    return _fmt_semver(nxt)


def _version_tuple(v: str):
    v = (v or "").strip().lstrip("v")
    parts = []
    for p in v.split("."):
        try:
            parts.append(int(p))
        except ValueError:
            parts.append(-1)
    return tuple(parts) if parts else (0,)


def _is_newer_version(remote: str, local: str) -> bool:
    if not remote or remote == local:
        return False
    try:
        return _version_tuple(remote) > _version_tuple(local)
    except Exception:
        return remote != local


_gh_version_cache_srv: dict = {"ts": 0.0, "version": "", "source_url": ""}
_gh_manifest_cache_srv: dict = {"ts": 0.0, "manifest": None, "source_url": ""}


async def _fetch_github_version_json() -> dict:
    """SEMÁFORO AUTORITATIVO: lee `version.json` desde la Release en GitHub.

    El manifiesto es la ÚNICA fuente de verdad para el auto-updater. Un
    cliente SÓLO debe considerar una versión "disponible" cuando existe un
    version.json publicado que la anuncie — porque publish-manifest.yml
    genera este archivo DESPUÉS de que todos los .exe/binarios están 100%
    subidos a la Release. Esto elimina la carrera:
      version.txt (o tag) anuncia vX.Y.Z antes de que el .exe termine de subir.

    Fuentes en orden de prioridad:
      1. Asset `version.json` de la Release del tag más nuevo (release "latest").
      2. `version.json` en la raíz del repo (rama configurada) — fallback.

    Cachea 60 s. Devuelve {manifest, source_url} o {} si no hay manifiesto.
    """
    import time as _t
    if _t.time() - _gh_manifest_cache_srv["ts"] < 60 and _gh_manifest_cache_srv["manifest"]:
        return {
            "manifest": _gh_manifest_cache_srv["manifest"],
            "source_url": _gh_manifest_cache_srv["source_url"],
        }

    result: dict = {"manifest": None, "source_url": ""}
    try:
        cfg = await _get_github_config()
        owner, repo = _parse_github_url(cfg.get("repo_url", ""))
        if not owner or not repo:
            # Fallback al repo del EXE por defecto
            gh_repo = _github_exe_repo()
            if "/" in gh_repo:
                owner, repo = gh_repo.split("/", 1)
            if not owner or not repo:
                return result

        branch = cfg.get("branch") or "main"
        token = cfg.get("token", "") or (os.environ.get("GITHUB_TOKEN") or "").strip()
        api_headers = {"Accept": "application/vnd.github+json", "User-Agent": "cinema-productions"}
        raw_headers = {"User-Agent": "cinema-productions"}
        if token:
            api_headers["Authorization"] = f"Bearer {token}"
            raw_headers["Authorization"] = f"Bearer {token}"

        async with httpx.AsyncClient(timeout=10) as http:
            # ── 1) Buscar version.json en el asset de la Release más nueva ──
            manifest = None
            source_url = ""
            try:
                r = await http.get(
                    f"https://api.github.com/repos/{owner}/{repo}/releases?per_page=15",
                    headers=api_headers,
                )
                if r.status_code == 200:
                    for rel in (r.json() or []):
                        if rel.get("draft"):
                            continue
                        for asset in (rel.get("assets") or []):
                            if (asset.get("name") or "").lower() != "version.json":
                                continue
                            url = asset.get("browser_download_url") or ""
                            if not url:
                                continue
                            try:
                                rv = await http.get(url, headers=raw_headers, follow_redirects=True)
                                if rv.status_code == 200:
                                    data = rv.json()
                                    # Validación mínima del schema
                                    if isinstance(data, dict) and data.get("version") and isinstance(data.get("assets"), dict):
                                        manifest = data
                                        source_url = url
                                        break
                            except Exception:
                                continue
                        if manifest:
                            break
            except Exception as e:
                logger.warning(f"GitHub releases manifest scan falló: {e}")

            # ── 2) Fallback: version.json en la raíz del repo ─────────────
            if not manifest:
                for b in [branch, "main", "master"]:
                    if not b:
                        continue
                    raw_url = f"https://raw.githubusercontent.com/{owner}/{repo}/{b}/version.json"
                    try:
                        rv = await http.get(raw_url, headers=raw_headers)
                        if rv.status_code == 200:
                            try:
                                data = rv.json()
                            except Exception:
                                continue
                            if isinstance(data, dict) and data.get("version"):
                                manifest = data
                                source_url = raw_url
                                break
                    except Exception:
                        continue

        if manifest:
            result["manifest"] = manifest
            result["source_url"] = source_url
            _gh_manifest_cache_srv["ts"] = _t.time()
            _gh_manifest_cache_srv["manifest"] = manifest
            _gh_manifest_cache_srv["source_url"] = source_url
    except Exception as e:
        logger.warning(f"GitHub version.json fetch failed: {e}")
    return result


async def _fetch_github_version_txt() -> dict:
    """Lee la versión más reciente publicada en el repo GitHub configurado.

    Combina dos fuentes:
      1. `version.txt` en la raíz del repositorio (rama configurada).
      2. Tags publicados en el repo (`vX.Y[.Z]`) — la versión más alta gana.

    De esta forma, si el usuario publica una nueva versión creando un tag
    (por ejemplo `v1.13`) pero `version.txt` no se actualiza inmediatamente,
    el chequeo de actualizaciones sigue reflejando la última versión real.

    Cachea 60 s. Devuelve {version, source_url}. Si falla, version == ''.
    """
    import time as _t
    if _t.time() - _gh_version_cache_srv["ts"] < 60 and _gh_version_cache_srv["version"]:
        return {
            "version": _gh_version_cache_srv["version"],
            "source_url": _gh_version_cache_srv["source_url"],
        }

    result = {"version": "", "source_url": ""}
    try:
        cfg = await _get_github_config()
        repo_url = cfg.get("repo_url", "")
        owner, repo = _parse_github_url(repo_url)
        if not owner or not repo:
            return result
        branch = cfg.get("branch") or "main"
        token = cfg.get("token", "")
        headers = {"Authorization": f"Bearer {token}"} if token else {}

        candidate_branches = []
        for b in (branch, "main", "master"):
            if b and b not in candidate_branches:
                candidate_branches.append(b)

        # ── 1) Leer version.txt del repo ─────────────────────────────────
        txt_version = ""
        txt_source_url = ""
        async with httpx.AsyncClient(timeout=8) as http:
            for b in candidate_branches:
                raw_url = f"https://raw.githubusercontent.com/{owner}/{repo}/{b}/version.txt"
                try:
                    rv = await http.get(raw_url, headers=headers)
                    if rv.status_code == 200:
                        v = (rv.text or "").strip().splitlines()[0].strip().lstrip("v")
                        if v:
                            txt_version = v
                            txt_source_url = raw_url
                            break
                except Exception:
                    continue

            # ── 2) Buscar el tag semver más alto en el repo ───────────────
            tag_version = ""
            tag_source_url = ""
            try:
                api_headers = dict(headers)
                api_headers.setdefault("Accept", "application/vnd.github+json")
                api_headers.setdefault("User-Agent", "cinema-productions")
                tag_re = re.compile(r"^v(1(?:\.\d+){1,2})$")
                best_tuple = None
                # Paginar hasta ~3 páginas por seguridad
                for page in range(1, 4):
                    tags_url = (
                        f"https://api.github.com/repos/{owner}/{repo}/tags"
                        f"?per_page=100&page={page}"
                    )
                    rt = await http.get(tags_url, headers=api_headers)
                    if rt.status_code != 200:
                        break
                    tags = rt.json() or []
                    if not isinstance(tags, list) or not tags:
                        break
                    for t in tags:
                        name = (t.get("name") or "").strip()
                        m = tag_re.match(name)
                        if not m:
                            continue
                        v = m.group(1)
                        tup = _version_tuple(v)
                        if best_tuple is None or tup > best_tuple:
                            best_tuple = tup
                            tag_version = v
                            tag_source_url = (
                                f"https://github.com/{owner}/{repo}/releases/tag/v{v}"
                            )
                    if len(tags) < 100:
                        break
            except Exception as e:
                logger.warning(f"GitHub tags scan failed: {e}")

        # ── 3) Escoger la más alta entre version.txt y tags ──────────────
        candidates = []
        if txt_version:
            candidates.append((txt_version, txt_source_url))
        if tag_version:
            candidates.append((tag_version, tag_source_url))
        if candidates:
            best = max(candidates, key=lambda x: _version_tuple(x[0]))
            result["version"] = best[0]
            result["source_url"] = best[1]

        _gh_version_cache_srv["ts"] = _t.time()
        _gh_version_cache_srv["version"] = result["version"]
        _gh_version_cache_srv["source_url"] = result["source_url"]
    except Exception as e:
        logger.warning(f"GitHub version fetch failed: {e}")
    return result


@api_router.get("/updates/check")
async def check_updates_cloud(refresh: bool = False):
    """Combina versión GitHub + paquete MongoDB.

    NUEVO FLUJO (semáforo version.json):
      1. Prioridad #1 → version.json (manifiesto autoritativo con hashes).
         Este archivo SÓLO existe cuando todos los binarios están 100% en
         la Release, por lo que anunciar esta versión es siempre seguro.
      2. Fallback → version.txt / tags (retrocompatibilidad, para releases
         antiguas sin manifiesto).
    La app de escritorio compara su version.txt local con la versión anunciada.
    """
    if refresh:
        _gh_version_cache_srv["ts"] = 0.0
        _gh_manifest_cache_srv["ts"] = 0.0

    local_version = await _read_local_version()

    # ── 1) Semáforo autoritativo: version.json ─────────────────────────
    manifest_info = await _fetch_github_version_json()
    manifest = manifest_info.get("manifest") or None
    manifest_version = ""
    manifest_source_url = ""
    if isinstance(manifest, dict):
        manifest_version = (manifest.get("version") or "").strip().lstrip("v")
        manifest_source_url = manifest_info.get("source_url", "")

    # ── 2) Fallback: version.txt + tags ────────────────────────────────
    gh = await _fetch_github_version_txt()
    github_version = gh.get("version", "")

    doc = await db.app_updates.find_one({"is_latest": True}, sort=[("created_at", -1)])
    mongo_version = doc.get("version", "") if doc else ""

    # Prioridad: manifest > mongo > github(txt/tags)
    if manifest_version:
        remote_version = manifest_version
        remote_source = "manifest"
    else:
        candidates = [v for v in (github_version, mongo_version) if v]
        remote_version = max(candidates, key=_version_tuple) if candidates else ""
        remote_source = "github_txt" if remote_version == github_version and github_version else (
            "mongo" if remote_version == mongo_version and mongo_version else "none"
        )

    has_update = _is_newer_version(remote_version, local_version) if local_version else False

    payload = {
        "checked": True,
        "has_update": has_update,
        "is_cloud": True,
        "local_version": local_version,
        "github_version": github_version,
        "github_source_url": gh.get("source_url", ""),
        "remote_version": remote_version or None,
        "remote_source": remote_source,
        # Manifest info (semáforo)
        "manifest_available": bool(manifest),
        "manifest_version": manifest_version or None,
        "manifest_source_url": manifest_source_url or None,
        "manifest_assets": (manifest or {}).get("assets") if manifest else None,
    }
    if doc:
        payload.update({
            "id": str(doc["_id"]),
            "filename": doc.get("filename", ""),
            "notes": doc.get("notes", ""),
            "created_at": doc.get("created_at", ""),
            "mongo_version": mongo_version,
        })
    return payload


@api_router.get("/updates/manifest")
async def get_update_manifest(refresh: bool = False):
    """Devuelve el manifest version.json completo (semáforo autoritativo).

    El cliente auto-updater debe:
      1. Llamar a este endpoint (o descargar version.json del release).
      2. Comparar manifest.version con la versión local.
      3. Si es más nueva, descargar el asset correspondiente a su plataforma
         (Windows: assets.windows_installer o windows_portable).
      4. Validar el SHA256 del archivo descargado contra
         manifest.assets.<plataforma>.sha256 ANTES de instalar.
      5. Si el hash NO coincide, abortar y eliminar la descarga.
      6. Sólo entonces ejecutar el swap seguro (ver backend/updater.py).
    """
    if refresh:
        _gh_manifest_cache_srv["ts"] = 0.0
    info = await _fetch_github_version_json()
    manifest = info.get("manifest")
    local_version = await _read_local_version()

    if not manifest:
        return {
            "status": "not_available",
            "local_version": local_version,
            "message": (
                "No hay version.json publicado aún. El auto-updater "
                "no debería notificar actualizaciones hasta que exista "
                "un manifiesto con hashes válidos."
            ),
        }

    remote = (manifest.get("version") or "").strip().lstrip("v")
    return {
        "status": "ready",
        "local_version": local_version,
        "remote_version": remote,
        "has_update": _is_newer_version(remote, local_version) if local_version else False,
        "source_url": info.get("source_url", ""),
        "manifest": manifest,
    }


@api_router.get("/updates/github-version")
async def get_github_version(refresh: bool = False):
    """Devuelve la versión leída directamente desde version.txt del repositorio
    GitHub configurado. Usado por la app de escritorio para saber qué versión
    subiste al repo."""
    if refresh:
        _gh_version_cache_srv["ts"] = 0.0
    local_version = await _read_local_version()
    gh = await _fetch_github_version_txt()
    github_version = gh.get("version", "")
    return {
        "local_version": local_version,
        "github_version": github_version,
        "has_update": _is_newer_version(github_version, local_version) if local_version else False,
        "source_url": gh.get("source_url", ""),
    }


@api_router.post("/updates/dismiss")
async def dismiss_update_cloud():
    return {"message": "OK"}


def _normalize_semver(v: str) -> str:
    """Normaliza cualquier versión al formato X.Y.Z. Ej: '1.20' → '1.20.0'."""
    v = (v or "").strip().lstrip("v")
    if not v:
        return ""
    parts = v.split(".")
    nums = []
    for p in parts:
        try:
            nums.append(str(int(p)))
        except ValueError:
            nums.append("0")
    while len(nums) < 3:
        nums.append("0")
    return ".".join(nums[:3])


# ── PRÓXIMA VERSIÓN AUTOMÁTICA (para el modal "Publicar en GitHub") ──────────
@api_router.get("/github/next-version")
async def get_next_auto_version():
    """Devuelve la versión actual y cuál sería el próximo número si el usuario
    deja el input de versión vacío en el modal 'Publicar en GitHub'.

    UNIFICADA: usa el MISMO semver X.Y.Z que auto-release.yml, el EXE, el ZIP
    y los commits. Fuente: _compute_next_unified_version().

    Response:
    {
      "current_local": "1.0.23",     # version.txt local
      "current_remote": "1.0.23",    # version.txt del repo GitHub
      "current_desktop": "1.0.23",   # último semver publicado (Mongo)
      "next_auto_version": "1.0.24"  # próximo canónico (mismo para ZIP/EXE/tag)
    }
    """
    current_local = await _read_local_version()
    gh = await _fetch_github_version_txt()
    current_remote = gh.get("version", "")

    cdoc = await db.app_settings.find_one(
        {}, {"desktop_build": 1, "desktop_build_semver": 1}
    ) or {}
    current_desktop = cdoc.get("desktop_build_semver") or (
        f"1.{cdoc.get('desktop_build', 0)}" if cdoc.get("desktop_build") else ""
    )

    next_version = await _compute_next_unified_version()

    # Normalizar TODAS las versiones al mismo formato X.Y.Z para evitar
    # confusión visual como "v1.20" vs "v1.0.31" (el usuario ve formatos
    # inconsistentes cuando algunos tags son 2-partes y otros 3-partes).
    return {
        "current_local":     _normalize_semver(current_local),
        "current_remote":    _normalize_semver(current_remote),
        "current_desktop":   _normalize_semver(current_desktop),
        "next_auto_version": _normalize_semver(next_version),
        # Raw sin normalizar (por si algún cliente antiguo lo espera)
        "current_local_raw":  current_local,
        "current_remote_raw": current_remote,
    }


# ── APPEARANCE CLOUD SYNC ────────────────────────────────────────────────────

@api_router.get("/settings/appearance")
async def get_appearance_snapshot():
    doc = await db.app_settings.find_one({}, {"_id": 0, "appearance_snapshot": 1, "appearance_updated_at": 1})
    doc = doc or {}
    return {"snapshot": doc.get("appearance_snapshot"), "updated_at": doc.get("appearance_updated_at")}


@api_router.put("/settings/appearance")
async def save_appearance_snapshot(payload: dict = Body(...)):
    snapshot = payload.get("snapshot") or {}
    updated_at = datetime.now(timezone.utc).isoformat()
    await db.app_settings.update_one(
        {},
        {"$set": {"appearance_snapshot": snapshot, "appearance_updated_at": updated_at}},
        upsert=True,
    )
    return {"updated_at": updated_at}


# ── SAVED THEMES ─────────────────────────────────────────────────────────────
#
# Sincronización de 3 vías (MongoDB → JSON local → GitHub):
#   · MongoDB (colección `saved_themes`) es la fuente principal de la verdad.
#   · Cada cambio (create/update/delete) exporta la lista completa al archivo
#     local `/app/themes/saved_themes.json` (mirror inmediato en disco).
#   · Si hay `GITHUB_TOKEN` configurado (en .env o en app_settings.github_config),
#     ese JSON se sube al repo AlejandroPiedrasanta/RESERVA-DE-EVENTOS en
#     `themes/saved_themes.json` usando la GitHub Contents API (upsert por SHA).
#   · Si la BD arranca vacía, se siembra automáticamente el tema "Minimalista"
#     desde el JSON local (o un snapshot por defecto) y se aplica como
#     `appearance_snapshot` inicial.
# ────────────────────────────────────────────────────────────────────────────

async def _themes_snapshot_payload() -> dict:
    """Construye el payload JSON con todos los temas guardados."""
    docs = await db.saved_themes.find({}, sort=[("created_at", -1)]).to_list(500)
    settings_doc = await db.app_settings.find_one({}, {"default_theme_id": 1}) or {}
    default_id = str(settings_doc.get("default_theme_id") or "")
    return {
        "app": "Cinema Productions — Reserva de Eventos",
        "kind": "saved_themes_backup",
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "count": len(docs),
        "themes": [
            {
                "id": str(d["_id"]),
                "name": d.get("name", ""),
                "snapshot": d.get("snapshot", {}),
                "created_at": d.get("created_at", ""),
                "updated_at": d.get("updated_at", ""),
                "is_default": (str(d["_id"]) == default_id) or bool(d.get("is_default", False)),
            }
            for d in docs
        ],
    }


async def _write_themes_local_json(payload: dict | None = None) -> dict:
    """Escribe el mirror local /app/themes/saved_themes.json."""
    if payload is None:
        payload = await _themes_snapshot_payload()
    try:
        THEMES_JSON_PATH.parent.mkdir(parents=True, exist_ok=True)
        THEMES_JSON_PATH.write_text(
            json.dumps(payload, indent=2, ensure_ascii=False),
            encoding="utf-8",
        )
    except Exception as e:
        logger.warning(f"[themes] no se pudo escribir JSON local: {e}")
    return payload


async def _resolve_github_creds() -> tuple[str, str, str]:
    """Retorna (token, repo_url, branch). Prioriza .env, luego app_settings."""
    token = (os.environ.get("GITHUB_TOKEN") or "").strip()
    repo_url = (os.environ.get("GITHUB_REPO_URL") or "").strip()
    branch = (os.environ.get("GITHUB_BRANCH") or "").strip()
    if not token or not repo_url:
        try:
            cfg = await _get_github_config()
            if not token:
                token = (cfg.get("token") or "").strip()
            if not repo_url:
                repo_url = (cfg.get("repo_url") or "").strip()
            if not branch:
                branch = (cfg.get("branch") or "").strip()
        except Exception:
            pass
    if not repo_url:
        repo_url = SUGGESTED_GITHUB_REPO
    if not branch:
        branch = DEFAULT_GITHUB_BRANCH
    return token, repo_url, branch


async def _push_themes_to_github(payload: dict) -> dict:
    """Upsert `themes/saved_themes.json` en GitHub via Contents API."""
    token, repo_url, branch = await _resolve_github_creds()
    if not token:
        return {"skipped": True, "reason": "no_token"}
    owner, repo = _parse_github_url(repo_url)
    if not owner or not repo:
        return {"skipped": True, "reason": "bad_repo_url"}

    path_in_repo = "themes/saved_themes.json"
    api_url = f"https://api.github.com/repos/{owner}/{repo}/contents/{path_in_repo}"
    content_b64 = base64.b64encode(
        json.dumps(payload, indent=2, ensure_ascii=False).encode("utf-8")
    ).decode("ascii")
    headers = {
        "Authorization": f"Bearer {token}",
        "Accept": "application/vnd.github+json",
        "X-GitHub-Api-Version": "2022-11-28",
        "User-Agent": "cinema-productions-app",
    }
    try:
        async with httpx.AsyncClient(timeout=20) as h:
            sha = None
            try:
                r_get = await h.get(api_url, headers=headers, params={"ref": branch})
                if r_get.status_code == 200:
                    sha = r_get.json().get("sha")
            except Exception:
                pass
            body = {
                "message": f"chore(themes): sync {payload.get('count', 0)} saved theme(s)",
                "content": content_b64,
                "branch": branch,
            }
            if sha:
                body["sha"] = sha
            r = await h.put(api_url, headers=headers, json=body)
            if r.status_code in (200, 201):
                data = r.json()
                sync_info = {
                    "themes_sync.last_github_sha": (data.get("content") or {}).get("sha", ""),
                    "themes_sync.last_github_commit": (data.get("commit") or {}).get("sha", ""),
                    "themes_sync.last_github_at": datetime.now(timezone.utc).isoformat(),
                    "themes_sync.last_status": "ok",
                    "themes_sync.last_error": "",
                }
                await db.app_settings.update_one({}, {"$set": sync_info}, upsert=True)
                return {"ok": True, "commit_sha": (data.get("commit") or {}).get("sha", "")}
            logger.warning(f"[themes/github] {r.status_code}: {r.text[:400]}")
            await db.app_settings.update_one(
                {},
                {"$set": {
                    "themes_sync.last_status": f"error_{r.status_code}",
                    "themes_sync.last_error": r.text[:400],
                    "themes_sync.last_github_at": datetime.now(timezone.utc).isoformat(),
                }},
                upsert=True,
            )
            return {"ok": False, "status": r.status_code, "error": r.text[:400]}
    except Exception as e:
        logger.warning(f"[themes/github] excepción: {e}")
        try:
            await db.app_settings.update_one(
                {},
                {"$set": {
                    "themes_sync.last_status": "error_exception",
                    "themes_sync.last_error": str(e)[:400],
                    "themes_sync.last_github_at": datetime.now(timezone.utc).isoformat(),
                }},
                upsert=True,
            )
        except Exception:
            pass
        return {"ok": False, "error": str(e)}


async def _sync_themes_all_channels(push_github: bool = True) -> dict:
    """Ejecuta la sincronización de 3 vías: Mongo → JSON local → (GitHub)."""
    payload = await _write_themes_local_json()
    gh = {"skipped": True, "reason": "disabled"}
    if push_github:
        gh = await _push_themes_to_github(payload)
    return {
        "local_ok": True,
        "local_path": str(THEMES_JSON_PATH),
        "count": payload.get("count", 0),
        "github": gh,
    }


@api_router.get("/themes")
async def list_saved_themes():
    docs = await db.saved_themes.find({}, sort=[("created_at", -1)]).to_list(500)
    settings_doc = await db.app_settings.find_one({}, {"default_theme_id": 1}) or {}
    default_id = str(settings_doc.get("default_theme_id") or "")
    return [
        {
            "id": str(d["_id"]),
            "name": d.get("name", ""),
            "snapshot": d.get("snapshot", {}),
            "created_at": d.get("created_at", ""),
            "updated_at": d.get("updated_at", ""),
            "is_default": (str(d["_id"]) == default_id) or bool(d.get("is_default", False)),
        }
        for d in docs
    ]


@api_router.post("/themes")
async def create_saved_theme(payload: dict = Body(...)):
    name = (payload.get("name") or "").strip()
    if not name:
        raise HTTPException(status_code=400, detail="El nombre del tema es requerido")
    now = datetime.now(timezone.utc).isoformat()
    doc = {
        "name": name,
        "snapshot": payload.get("snapshot") or {},
        "created_at": now,
        "updated_at": now,
    }
    result = await db.saved_themes.insert_one(doc)
    # sync en background (no bloquea la respuesta)
    asyncio.create_task(_sync_themes_all_channels())
    return {
        "id": str(result.inserted_id),
        "name": name,
        "snapshot": doc["snapshot"],
        "created_at": now,
        "updated_at": now,
    }


@api_router.put("/themes/{theme_id}")
async def update_saved_theme(theme_id: str, payload: dict = Body(...)):
    """Sobreescribe un tema existente con la apariencia actual (Reguardar)."""
    try:
        oid = ObjectId(theme_id)
    except Exception:
        raise HTTPException(status_code=400, detail="ID inválido")
    update: dict = {"updated_at": datetime.now(timezone.utc).isoformat()}
    if "name" in payload:
        n = (payload.get("name") or "").strip()
        if not n:
            raise HTTPException(status_code=400, detail="El nombre no puede quedar vacío")
        update["name"] = n
    if "snapshot" in payload:
        update["snapshot"] = payload.get("snapshot") or {}
    result = await db.saved_themes.update_one({"_id": oid}, {"$set": update})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Tema no encontrado")
    asyncio.create_task(_sync_themes_all_channels())
    doc = await db.saved_themes.find_one({"_id": oid}) or {}
    return {
        "id": str(oid),
        "name": doc.get("name", ""),
        "snapshot": doc.get("snapshot", {}),
        "created_at": doc.get("created_at", ""),
        "updated_at": doc.get("updated_at", ""),
    }


@api_router.delete("/themes/{theme_id}")
async def delete_saved_theme(theme_id: str):
    try:
        oid = ObjectId(theme_id)
    except Exception:
        raise HTTPException(status_code=400, detail="ID inválido")
    # Si era el default, limpiar la referencia
    settings_doc = await db.app_settings.find_one({}, {"default_theme_id": 1}) or {}
    was_default = str(settings_doc.get("default_theme_id") or "") == theme_id
    result = await db.saved_themes.delete_one({"_id": oid})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Tema no encontrado")
    if was_default:
        # Fallback: buscar "Minimalista" o el primer tema restante para no
        # dejar la app sin tema por defecto.
        fallback = await db.saved_themes.find_one({"name": {"$regex": "^minimal", "$options": "i"}})
        if not fallback:
            fallback = await db.saved_themes.find_one({}, sort=[("created_at", 1)])
        if fallback:
            await db.app_settings.update_one(
                {},
                {"$set": {
                    "default_theme_id": str(fallback["_id"]),
                    "default_theme_name": fallback.get("name", ""),
                }},
                upsert=True,
            )
        else:
            await db.app_settings.update_one({}, {"$unset": {"default_theme_id": "", "default_theme_name": ""}}, upsert=True)
    asyncio.create_task(_sync_themes_all_channels())
    return {"message": "Tema eliminado"}


@api_router.post("/themes/{theme_id}/set-default")
async def set_default_theme(theme_id: str):
    try:
        oid = ObjectId(theme_id)
    except Exception:
        raise HTTPException(status_code=400, detail="ID inválido")
    doc = await db.saved_themes.find_one({"_id": oid})
    if not doc:
        raise HTTPException(status_code=404, detail="Tema no encontrado")
    await db.app_settings.update_one(
        {},
        {"$set": {"default_theme_id": str(oid), "default_theme_name": doc.get("name", "")}},
        upsert=True,
    )
    asyncio.create_task(_sync_themes_all_channels())
    return {"success": True, "default_theme_id": str(oid), "name": doc.get("name", "")}


@api_router.post("/themes/sync")
async def themes_sync_now():
    """Fuerza una sincronización manual (Mongo → JSON local → GitHub)."""
    result = await _sync_themes_all_channels()
    return {"success": True, **result}


@api_router.get("/themes/sync/status")
async def themes_sync_status():
    doc = await db.app_settings.find_one({}, {"themes_sync": 1, "default_theme_id": 1, "default_theme_name": 1}) or {}
    ts = doc.get("themes_sync") or {}
    local_mtime = None
    if THEMES_JSON_PATH.exists():
        try:
            local_mtime = datetime.fromtimestamp(
                THEMES_JSON_PATH.stat().st_mtime, tz=timezone.utc
            ).isoformat()
        except Exception:
            pass
    token_env = bool(os.environ.get("GITHUB_TOKEN"))
    return {
        "local_path": str(THEMES_JSON_PATH),
        "local_exists": THEMES_JSON_PATH.exists(),
        "local_mtime": local_mtime,
        "last_github_at": ts.get("last_github_at"),
        "last_github_sha": ts.get("last_github_sha"),
        "last_github_commit": ts.get("last_github_commit"),
        "last_status": ts.get("last_status"),
        "last_error": ts.get("last_error"),
        "github_configured": token_env or bool((await _resolve_github_creds())[0]),
        "default_theme_id": str(doc.get("default_theme_id") or ""),
        "default_theme_name": doc.get("default_theme_name", ""),
    }


# ── APP SECURITY (password lock + page protection) ──────────────────────────

PBKDF2_ITERATIONS = 260000


def _hash_app_password(password: str) -> str:
    salt = secrets.token_hex(16)
    dk = hashlib.pbkdf2_hmac("sha256", password.encode(), bytes.fromhex(salt), PBKDF2_ITERATIONS)
    return f"pbkdf2_sha256${PBKDF2_ITERATIONS}${salt}${dk.hex()}"


def _verify_app_password(password: str, stored: str) -> bool:
    try:
        _, iterations, salt, hash_hex = stored.split("$")
        dk = hashlib.pbkdf2_hmac("sha256", password.encode(), bytes.fromhex(salt), int(iterations))
        return secrets.compare_digest(dk.hex(), hash_hex)
    except Exception:
        return False


@api_router.get("/security/status")
async def get_security_status():
    doc = await db.app_settings.find_one({}, {"_id": 0, "app_password_hash": 1, "app_password_hint": 1, "page_protection_enabled": 1, "security_config": 1}) or {}
    cfg = doc.get("security_config") or {}
    return {
        "password_enabled": bool(doc.get("app_password_hash")),
        "hint": doc.get("app_password_hint") or "",
        "protection_enabled": bool(doc.get("page_protection_enabled")),
        # Nuevos ajustes de seguridad avanzada
        "auto_lock_enabled": bool(cfg.get("auto_lock_enabled", False)),
        "auto_lock_minutes": int(cfg.get("auto_lock_minutes", 5)),
        "max_attempts": int(cfg.get("max_attempts", 5)),
        "lockout_seconds": int(cfg.get("lockout_seconds", 60)),
        "protected_sections": cfg.get("protected_sections", []),
        "failed_attempts": int(cfg.get("failed_attempts", 0)),
        "locked_until": cfg.get("locked_until", ""),
        # ── Nuevos toggles one-click ──
        "block_devtools": bool(cfg.get("block_devtools", False)),
        "blur_on_unfocus": bool(cfg.get("blur_on_unfocus", False)),
        "block_printscreen": bool(cfg.get("block_printscreen", False)),
        "block_drag_drop": bool(cfg.get("block_drag_drop", False)),
        "zip_password_enabled": bool(cfg.get("zip_password_enabled", True)),
    }


@api_router.put("/security/advanced-config")
async def set_advanced_security_config(payload: dict = Body(...)):
    """Configuración avanzada: auto-lock por inactividad, límite de intentos, secciones protegidas."""
    update = {}
    if "auto_lock_enabled" in payload:
        update["security_config.auto_lock_enabled"] = bool(payload["auto_lock_enabled"])
    if "auto_lock_minutes" in payload:
        m = int(payload["auto_lock_minutes"])
        if m < 1 or m > 120:
            raise HTTPException(status_code=400, detail="auto_lock_minutes debe estar entre 1 y 120")
        update["security_config.auto_lock_minutes"] = m
    if "max_attempts" in payload:
        n = int(payload["max_attempts"])
        if n < 3 or n > 20:
            raise HTTPException(status_code=400, detail="max_attempts debe estar entre 3 y 20")
        update["security_config.max_attempts"] = n
    if "lockout_seconds" in payload:
        s = int(payload["lockout_seconds"])
        if s < 10 or s > 3600:
            raise HTTPException(status_code=400, detail="lockout_seconds debe estar entre 10 y 3600")
        update["security_config.lockout_seconds"] = s
    if "protected_sections" in payload:
        sections = payload["protected_sections"] or []
        if not isinstance(sections, list):
            raise HTTPException(status_code=400, detail="protected_sections debe ser una lista")
        valid = ["/base-de-datos", "/ajustes", "/socios", "/reservaciones", "/apariencia", "/actualizaciones", "/calendario"]
        sections = [s for s in sections if s in valid]
        update["security_config.protected_sections"] = sections
    # ── Nuevos toggles one-click ──
    for k in ("block_devtools", "blur_on_unfocus", "block_printscreen", "block_drag_drop", "zip_password_enabled"):
        if k in payload:
            update[f"security_config.{k}"] = bool(payload[k])
    if update:
        await db.app_settings.update_one({}, {"$set": update}, upsert=True)
    return {"success": True, "updated_keys": list(update.keys())}


# ── ZIP password (para la app compilada) ─────────────────────────────
# Valor por defecto configurable vía variable de entorno; el dueño de la app
# puede sobreescribirlo en runtime desde security_config.zip_password.
DEFAULT_ZIP_PASSWORD = os.environ.get("ZIP_DEFAULT_PASSWORD", "2868")

async def _get_zip_password() -> str:
    doc = await db.app_settings.find_one({}, {"security_config": 1}) or {}
    cfg = doc.get("security_config") or {}
    return cfg.get("zip_password") or DEFAULT_ZIP_PASSWORD


@api_router.get("/security/zip-password")
async def get_zip_password():
    """Devuelve la contraseña actual del ZIP compilado (visible para el dueño de la app)."""
    doc = await db.app_settings.find_one({}, {"security_config": 1}) or {}
    cfg = doc.get("security_config") or {}
    pwd = cfg.get("zip_password") or DEFAULT_ZIP_PASSWORD
    return {
        "password": pwd,
        "is_default": pwd == DEFAULT_ZIP_PASSWORD,
        "enabled": bool(cfg.get("zip_password_enabled", True)),
    }


@api_router.post("/security/zip-password")
async def set_zip_password(payload: dict = Body(...)):
    """Cambia la contraseña que se usará al comprimir la app de escritorio."""
    new_pwd = (payload.get("new_password") or "").strip()
    if len(new_pwd) < 3:
        raise HTTPException(status_code=400, detail="La contraseña debe tener al menos 3 caracteres")
    if len(new_pwd) > 64:
        raise HTTPException(status_code=400, detail="La contraseña no puede exceder 64 caracteres")
    await db.app_settings.update_one(
        {},
        {"$set": {"security_config.zip_password": new_pwd}},
        upsert=True,
    )
    return {"success": True, "password": new_pwd}


@api_router.post("/security/zip-password/reset")
async def reset_zip_password():
    """Restaura la contraseña ZIP al valor de fábrica (2868)."""
    await db.app_settings.update_one(
        {},
        {"$set": {"security_config.zip_password": DEFAULT_ZIP_PASSWORD}},
        upsert=True,
    )
    return {"success": True, "password": DEFAULT_ZIP_PASSWORD}


@api_router.post("/security/set-password")
async def set_app_password_endpoint(payload: dict = Body(...)):
    password = (payload.get("password") or "").strip()
    hint = (payload.get("hint") or "").strip()
    current = payload.get("current_password") or ""
    if len(password) < 4:
        raise HTTPException(status_code=400, detail="La contraseña debe tener al menos 4 caracteres")
    doc = await db.app_settings.find_one({}, {"app_password_hash": 1}) or {}
    if doc.get("app_password_hash") and not _verify_app_password(current, doc["app_password_hash"]):
        raise HTTPException(status_code=401, detail="La contraseña actual es incorrecta")
    await db.app_settings.update_one(
        {},
        {"$set": {"app_password_hash": _hash_app_password(password), "app_password_hint": hint}},
        upsert=True,
    )
    return {"message": "Contraseña guardada"}


@api_router.post("/security/verify")
async def verify_app_password_endpoint(payload: dict = Body(...)):
    doc = await db.app_settings.find_one({}, {"app_password_hash": 1, "security_config": 1}) or {}
    if not doc.get("app_password_hash"):
        return {"valid": True}

    cfg = doc.get("security_config") or {}
    max_attempts = int(cfg.get("max_attempts", 5))
    lockout_seconds = int(cfg.get("lockout_seconds", 60))
    failed = int(cfg.get("failed_attempts", 0))
    locked_until_str = cfg.get("locked_until", "")

    # Verificar si está en bloqueo temporal
    if locked_until_str:
        try:
            locked_until = datetime.fromisoformat(locked_until_str.replace("Z", "+00:00"))
            now = datetime.now(timezone.utc)
            if now < locked_until:
                remaining = int((locked_until - now).total_seconds())
                raise HTTPException(
                    status_code=429,
                    detail=f"Demasiados intentos fallidos. Espera {remaining} segundos.",
                    headers={"Retry-After": str(remaining)},
                )
        except (ValueError, AttributeError):
            pass

    if not _verify_app_password(payload.get("password") or "", doc["app_password_hash"]):
        # Incrementar contador de fallos
        new_failed = failed + 1
        update = {"security_config.failed_attempts": new_failed}
        if new_failed >= max_attempts:
            # Bloquear temporalmente
            locked_until = (datetime.now(timezone.utc) + timedelta(seconds=lockout_seconds)).isoformat()
            update["security_config.locked_until"] = locked_until
            update["security_config.failed_attempts"] = 0
            await db.app_settings.update_one({}, {"$set": update}, upsert=True)
            raise HTTPException(
                status_code=429,
                detail=f"Demasiados intentos. Bloqueado por {lockout_seconds} segundos.",
            )
        else:
            await db.app_settings.update_one({}, {"$set": update}, upsert=True)
            remaining = max_attempts - new_failed
            raise HTTPException(
                status_code=401,
                detail=f"Contraseña incorrecta. Te quedan {remaining} intento{'s' if remaining != 1 else ''}.",
            )

    # Éxito → resetear contador
    await db.app_settings.update_one(
        {},
        {"$set": {"security_config.failed_attempts": 0, "security_config.locked_until": ""}},
        upsert=True,
    )
    return {"valid": True}


@api_router.post("/security/remove-password")
async def remove_app_password_endpoint(payload: dict = Body(...)):
    doc = await db.app_settings.find_one({}, {"app_password_hash": 1}) or {}
    if not doc.get("app_password_hash"):
        raise HTTPException(status_code=400, detail="No hay contraseña configurada")
    if not _verify_app_password(payload.get("current_password") or "", doc["app_password_hash"]):
        raise HTTPException(status_code=401, detail="La contraseña actual es incorrecta")
    await db.app_settings.update_one({}, {"$unset": {"app_password_hash": "", "app_password_hint": ""}})
    return {"message": "Contraseña eliminada"}


@api_router.put("/security/protection")
async def set_page_protection_endpoint(payload: dict = Body(...)):
    enabled = bool(payload.get("enabled"))
    await db.app_settings.update_one({}, {"$set": {"page_protection_enabled": enabled}}, upsert=True)
    return {"enabled": enabled}


# ═══════════════════════════════════════════════════════════════════
# GitHub Integration & AI Context (Continuity for next AI)
# ═══════════════════════════════════════════════════════════════════
import subprocess
import re as _re

REPO_ROOT = Path(__file__).parent.parent  # /app

from ai_context_default import DEFAULT_AI_CONTEXT


def _parse_github_url(url: str):
    """Extrae owner/repo de una URL de GitHub."""
    if not url:
        return None, None
    m = _re.match(r"^https?://github\.com/([^/]+)/([^/.]+?)(?:\.git)?/?$", url.strip())
    if not m:
        return None, None
    return m.group(1), m.group(2)


# Repo GitHub que viene DE FÁBRICA en el proyecto. Se auto-persiste al arrancar
# la app si no hay configuración previa en la base de datos.
SUGGESTED_GITHUB_REPO = "https://github.com/AlejandroPiedrasanta/RESERVA-DE-EVENTOS"
DEFAULT_GITHUB_BRANCH = "main"

# Conexiones MongoDB que vienen DE FÁBRICA en el proyecto. Se exponen en
# /api/settings/database/factory-presets y el frontend las carga automáticamente
# junto con las guardadas localmente (localStorage). El usuario puede modificar
# o eliminarlas desde la UI (solo se persisten en su navegador).
FACTORY_SAVED_CONNECTIONS = [
    {
        "name": "MongoDB Atlas (por defecto)",
        "url": "mongodb+srv://reu1:cinemaproductions@cluster0.ozg25wu.mongodb.net/?appName=Cluster0",
        "color": "emerald",
        "factory": True,
    },
]


async def _get_github_config():
    """Devuelve la config de GitHub tal cual está en la DB. Sin defaults forzados."""
    doc = await db.app_settings.find_one({}, {"github_config": 1}) or {}
    return doc.get("github_config") or {}


@api_router.get("/github/config")
async def get_github_config():
    cfg = await _get_github_config()
    repo_url = cfg.get("repo_url", "")
    return {
        "repo_url": repo_url,
        "has_token": bool(cfg.get("token")),
        "last_commit_sha": cfg.get("last_commit_sha", ""),
        "last_check_at": cfg.get("last_check_at", ""),
        "branch": cfg.get("branch", DEFAULT_GITHUB_BRANCH),
        "is_configured": bool(repo_url),
        "suggested_repo": SUGGESTED_GITHUB_REPO,
        "username": cfg.get("username", ""),
        "avatar_url": cfg.get("avatar_url", ""),
        "connected_at": cfg.get("connected_at", ""),
        "last_push_at": cfg.get("last_push_at", ""),
        "last_push_message": cfg.get("last_push_message", ""),
    }


@api_router.post("/github/config")
async def save_github_config(payload: dict = Body(...)):
    repo_url = (payload.get("repo_url") or "").strip()
    token = (payload.get("token") or "").strip()
    branch = (payload.get("branch") or "main").strip()

    owner, repo = _parse_github_url(repo_url)
    if repo_url and not owner:
        raise HTTPException(status_code=400, detail="URL de GitHub inválida. Formato esperado: https://github.com/usuario/repo")

    update = {"repo_url": repo_url, "branch": branch}
    if token:
        update["token"] = token
    elif payload.get("clear_token"):
        update["token"] = ""

    await db.app_settings.update_one(
        {},
        {"$set": {f"github_config.{k}": v for k, v in update.items()}},
        upsert=True
    )

    # Intentar sincronizar el remoto del repo local
    if repo_url:
        try:
            subprocess.run(
                ["git", "-C", str(REPO_ROOT), "remote", "set-url", "origin", repo_url],
                capture_output=True, timeout=10, check=False
            )
        except Exception as e:
            logger.warning(f"No se pudo actualizar remote: {e}")

    return {"success": True, "repo_url": repo_url, "branch": branch}


@api_router.post("/github/connect")
async def github_connect(payload: dict = Body(...)):
    """Conecta con GitHub usando un Personal Access Token.
    Valida el token contra la API de GitHub, obtiene el usuario y guarda todo
    en app_settings.github_config junto con el repo por defecto de fábrica."""
    import urllib.request, json as _json, ssl as _ssl
    # En el .exe (PyInstaller) urllib no encuentra el store de CAs del sistema y
    # falla la verificacion SSL contra api.github.com. Usamos el bundle de certifi
    # (que SI se empaqueta) para tener certificados validos en cualquier entorno.
    try:
        import certifi as _certifi
        _ssl_ctx = _ssl.create_default_context(cafile=_certifi.where())
    except Exception:
        _ssl_ctx = _ssl.create_default_context()

    token = (payload.get("token") or "").strip()
    repo_url = (payload.get("repo_url") or SUGGESTED_GITHUB_REPO).strip()
    branch = (payload.get("branch") or DEFAULT_GITHUB_BRANCH).strip()

    if not token:
        raise HTTPException(status_code=400, detail="Token requerido. Crea uno en https://github.com/settings/tokens con scope 'repo'.")

    # Validar contra GitHub API
    try:
        req = urllib.request.Request(
            "https://api.github.com/user",
            headers={
                "Authorization": f"Bearer {token}",
                "Accept": "application/vnd.github+json",
                "User-Agent": "cinema-productions",
            },
        )
        with urllib.request.urlopen(req, timeout=10, context=_ssl_ctx) as resp:
            user_data = _json.loads(resp.read().decode())
    except urllib.error.HTTPError as e:
        detail = "Token inválido o expirado" if e.code == 401 else f"GitHub API error: {e.code}"
        raise HTTPException(status_code=400, detail=detail)
    except Exception as e:
        raise HTTPException(status_code=502, detail=f"No se pudo contactar GitHub: {e}")

    username = user_data.get("login", "")
    avatar = user_data.get("avatar_url", "")

    # Guardar en la BD
    await db.app_settings.update_one(
        {},
        {"$set": {
            "github_config.token": token,
            "github_config.repo_url": repo_url,
            "github_config.branch": branch,
            "github_config.username": username,
            "github_config.avatar_url": avatar,
            "github_config.connected_at": datetime.now(timezone.utc).isoformat(),
        }},
        upsert=True,
    )

    # Configurar git remoto con auth embebida
    try:
        auth_url = repo_url.replace("https://", f"https://{username}:{token}@") if username else repo_url
        subprocess.run(
            ["git", "-C", str(REPO_ROOT), "remote", "set-url", "origin", auth_url],
            capture_output=True, timeout=10, check=False,
        )
    except Exception as e:
        logger.warning(f"No se pudo configurar remote autenticado: {e}")

    return {
        "success": True,
        "username": username,
        "avatar_url": avatar,
        "repo_url": repo_url,
        "branch": branch,
    }


@api_router.post("/github/disconnect")
async def github_disconnect():
    """Desconecta la cuenta de GitHub (borra token, username, avatar). Deja el repo_url intacto."""
    await db.app_settings.update_one(
        {},
        {"$unset": {
            "github_config.token": "",
            "github_config.username": "",
            "github_config.avatar_url": "",
            "github_config.connected_at": "",
        }},
        upsert=True,
    )
    return {"success": True}


def _human_bytes(n: int) -> str:
    """Formatea bytes a KB/MB/GB legible."""
    try:
        n = float(n or 0)
    except Exception:
        n = 0.0
    if n < 1024:
        return f"{int(n)} B"
    if n < 1024 * 1024:
        return f"{n / 1024:.1f} KB"
    if n < 1024 * 1024 * 1024:
        return f"{n / (1024 * 1024):.1f} MB"
    return f"{n / (1024 * 1024 * 1024):.2f} GB"


def _gh_api_headers(token: str) -> dict:
    h = {
        "Accept": "application/vnd.github+json",
        "X-GitHub-Api-Version": "2022-11-28",
        "User-Agent": "cinema-productions-app",
    }
    if token:
        h["Authorization"] = f"Bearer {token}"
    return h


async def _trigger_exe_build_workflow(owner_repo=None, token: str = "", repo_url: str = "",
                                      branch: str = "main", version: str = "") -> dict:
    """Dispara el workflow 'build-exe.yml' vía workflow_dispatch API.
    Se usa porque el commit del push lleva '[skip ci]', lo que impide que el
    push del tag dispare el build automáticamente. workflow_dispatch NO se ve
    afectado por '[skip ci]', así que garantiza que el .exe SIEMPRE se compile.
    Devuelve {ok: bool, status?, error?}."""
    if not token:
        return {"ok": False, "error": "no_token"}
    owner, repo = (None, None)
    if owner_repo and "/" in owner_repo:
        owner, repo = owner_repo.split("/", 1)
    if not owner or not repo:
        owner, repo = _parse_github_url(repo_url)
    if not owner or not repo:
        parts = _github_exe_repo().split("/")
        if len(parts) == 2:
            owner, repo = parts[0], parts[1]
    if not owner or not repo:
        return {"ok": False, "error": "no_repo"}

    ver = (version or "").strip().lstrip("vV").strip()
    release_tag = f"v{ver}" if ver else "latest-exe"
    inputs = {"release_tag": release_tag, "prerelease": "false"}
    if ver:
        inputs["app_version"] = ver

    url = f"https://api.github.com/repos/{owner}/{repo}/actions/workflows/build-exe.yml/dispatches"
    headers = _gh_api_headers(token)
    body = {"ref": branch or "main", "inputs": inputs}
    try:
        async with httpx.AsyncClient(timeout=20.0) as client:
            r = await client.post(url, headers=headers, json=body)
            if r.status_code in (204, 201, 200):
                return {"ok": True, "status": r.status_code, "release_tag": release_tag}
            # Fallback: si el ref de la rama fallara, reintenta con el tag como ref
            if r.status_code == 422 and ver:
                body_retry = {"ref": release_tag, "inputs": inputs}
                r2 = await client.post(url, headers=headers, json=body_retry)
                if r2.status_code in (204, 201, 200):
                    return {"ok": True, "status": r2.status_code, "release_tag": release_tag}
                return {"ok": False, "status": r2.status_code, "error": (r2.text or "")[:300]}
            return {"ok": False, "status": r.status_code, "error": (r.text or "")[:300]}
    except Exception as e:
        return {"ok": False, "error": str(e)[:200]}


async def _trigger_refresh_deps_workflow(token: str = "", repo_url: str = "",
                                         branch: str = "main") -> dict:
    """Dispara el workflow 'refresh-deps.yml' vía workflow_dispatch API.
    Se llama después de un push que modificó frontend/package.json,
    frontend/yarn.lock o backend/requirements.txt. Necesario porque el commit
    del push lleva '[skip ci]' y GitHub NO dispara workflows con esa marca.
    Este workflow regenera el tarball de node_modules + sha16 y los publica al
    release 'deps-latest' para que bootstrap.sh entre por fast-path (~30s)."""
    if not token:
        return {"ok": False, "error": "no_token"}
    owner, repo = _parse_github_url(repo_url)
    if not owner or not repo:
        return {"ok": False, "error": "no_repo"}
    url = f"https://api.github.com/repos/{owner}/{repo}/actions/workflows/refresh-deps.yml/dispatches"
    headers = _gh_api_headers(token)
    body = {"ref": branch or "main"}
    try:
        async with httpx.AsyncClient(timeout=20.0) as client:
            r = await client.post(url, headers=headers, json=body)
            if r.status_code in (204, 201, 200):
                return {"ok": True, "status": r.status_code}
            return {"ok": False, "status": r.status_code, "error": (r.text or "")[:300]}
    except Exception as e:
        return {"ok": False, "error": str(e)[:200]}


@api_router.post("/github/build-exe")
async def github_build_exe(payload: dict = Body(default={})):
    """Dispara manualmente la compilación del .exe en GitHub Actions.
    Útil como botón 'Reintentar compilación' si un build no arrancó.
    Body opcional: { version: 'X.Y.Z' }."""
    token, repo_url, branch = await _resolve_github_creds()
    if not token:
        raise HTTPException(status_code=400, detail="Conecta tu cuenta de GitHub primero.")
    version = (payload or {}).get("version") or ""
    if not version:
        try:
            version = await _compute_next_unified_version()
        except Exception:
            version = ""
    res = await _trigger_exe_build_workflow(token=token, repo_url=repo_url, branch=branch, version=version)
    if not res.get("ok"):
        raise HTTPException(status_code=502, detail=f"No se pudo iniciar la compilación: {res.get('error') or res.get('status')}")
    owner, repo = _parse_github_url(repo_url)
    return {
        "success": True,
        "release_tag": res.get("release_tag"),
        "actions_url": f"https://github.com/{owner}/{repo}/actions/workflows/build-exe.yml" if owner else "",
        "message": "Compilación del .exe iniciada en GitHub Actions.",
    }



@api_router.get("/github/storage")
async def github_storage():
    """Devuelve el uso de espacio del repositorio, el plan de GitHub de la
    cuenta conectada y la lista de builds .exe publicados en Releases.
    Sirve para el panel 'Almacenamiento del repositorio' en Soporte avanzado."""
    token, repo_url, branch = await _resolve_github_creds()
    owner, repo = _parse_github_url(repo_url)
    if not owner or not repo:
        # Fallback al repo de fábrica de los .exe
        parts = _github_exe_repo().split("/")
        if len(parts) == 2:
            owner, repo = parts[0], parts[1]
    if not owner or not repo:
        raise HTTPException(status_code=400, detail="No hay repositorio configurado.")

    headers = _gh_api_headers(token)
    result = {
        "connected": bool(token),
        "repo_full_name": f"{owner}/{repo}",
        "repo": None,
        "plan": None,
        "builds": [],
        "builds_count": 0,
        "builds_total_bytes": 0,
        "builds_total_human": "0 B",
        "errors": [],
    }

    async with httpx.AsyncClient(timeout=15.0, follow_redirects=True) as client:
        # 1) Info del repositorio (tamaño en KB)
        try:
            r = await client.get(f"https://api.github.com/repos/{owner}/{repo}", headers=headers)
            if r.status_code == 200:
                d = r.json()
                size_bytes = int(d.get("size") or 0) * 1024  # size viene en KB
                result["repo"] = {
                    "full_name": d.get("full_name") or f"{owner}/{repo}",
                    "private": bool(d.get("private")),
                    "size_kb": int(d.get("size") or 0),
                    "size_bytes": size_bytes,
                    "size_human": _human_bytes(size_bytes),
                    "default_branch": d.get("default_branch") or branch,
                    "html_url": d.get("html_url") or f"https://github.com/{owner}/{repo}",
                }
            else:
                result["errors"].append(f"repo:{r.status_code}")
        except Exception as e:
            result["errors"].append(f"repo_exc:{str(e)[:80]}")

        # 2) Plan de la cuenta (requiere token)
        if token:
            try:
                r = await client.get("https://api.github.com/user", headers=headers)
                if r.status_code == 200:
                    u = r.json()
                    plan = u.get("plan") or {}
                    space_kb = int(plan.get("space") or 0)
                    result["plan"] = {
                        "login": u.get("login") or "",
                        "name": (plan.get("name") or "free").capitalize(),
                        "space_kb": space_kb,
                        "space_human": _human_bytes(space_kb * 1024) if space_kb else "—",
                        "private_repos": plan.get("private_repos"),
                        "collaborators": plan.get("collaborators"),
                        "owned_private_repos": u.get("owned_private_repos"),
                        "public_repos": u.get("public_repos"),
                        "total_private_repos": u.get("total_private_repos"),
                    }
                else:
                    result["errors"].append(f"user:{r.status_code}")
            except Exception as e:
                result["errors"].append(f"user_exc:{str(e)[:80]}")

        # 3) Releases → builds .exe
        try:
            r = await client.get(
                f"https://api.github.com/repos/{owner}/{repo}/releases?per_page=100",
                headers=headers,
            )
            if r.status_code == 200:
                builds = []
                total = 0
                for rel in r.json() or []:
                    for a in rel.get("assets") or []:
                        name = (a.get("name") or "")
                        low = name.lower()
                        # Contamos ejecutables y sus checksums asociados
                        if low.endswith(".exe") or low.endswith(".exe.sha256"):
                            size = int(a.get("size") or 0)
                            total += size
                            is_setup = "setup" in low or "installer" in low
                            builds.append({
                                "asset_id": a.get("id"),
                                "name": name,
                                "size": size,
                                "size_human": _human_bytes(size),
                                "kind": ".sha256" if low.endswith(".sha256") else ("installer" if is_setup else "portable"),
                                "release_id": rel.get("id"),
                                "release_name": rel.get("name") or rel.get("tag_name") or "",
                                "tag": rel.get("tag_name") or "",
                                "published_at": a.get("updated_at") or rel.get("published_at") or "",
                                "download_url": a.get("browser_download_url") or "",
                            })
                # Ordenar: más recientes primero, .exe antes que .sha256
                builds.sort(key=lambda b: (b.get("published_at") or ""), reverse=True)
                result["builds"] = builds
                result["builds_count"] = len([b for b in builds if b["kind"] != ".sha256"])
                result["builds_total_bytes"] = total
                result["builds_total_human"] = _human_bytes(total)
            else:
                result["errors"].append(f"releases:{r.status_code}")
        except Exception as e:
            result["errors"].append(f"releases_exc:{str(e)[:80]}")

    return result


@api_router.delete("/github/builds")
async def github_delete_builds(payload: dict = Body(default={})):
    """Borra builds .exe (y sus .sha256) de los Releases de GitHub para liberar
    espacio Y DEJAR EL REPO LIMPIO. Body opcional:
      - asset_ids: [int]      → borra solo esos assets.
      - release_id: int       → borra TODOS los .exe/.sha256 de ese release.
      - (vacío)               → borra TODOS los builds .exe/.sha256 del repo.
      - delete_releases: bool → (default True) si un release se queda SIN assets,
                                borra también el release y su tag git, para que
                                la sección de Releases del repo quede limpia.
    Requiere cuenta de GitHub conectada (token con scope 'repo')."""
    token, repo_url, branch = await _resolve_github_creds()
    if not token:
        raise HTTPException(status_code=400, detail="Conecta tu cuenta de GitHub primero para poder borrar builds.")

    owner, repo = _parse_github_url(repo_url)
    if not owner or not repo:
        parts = _github_exe_repo().split("/")
        if len(parts) == 2:
            owner, repo = parts[0], parts[1]
    if not owner or not repo:
        raise HTTPException(status_code=400, detail="No hay repositorio configurado.")

    payload = payload or {}
    target_asset_ids = set(payload.get("asset_ids") or [])
    target_release_id = payload.get("release_id")
    delete_releases = payload.get("delete_releases", True)

    headers = _gh_api_headers(token)
    deleted = []
    freed_bytes = 0
    errors = []
    deleted_releases = []
    deleted_tags = []

    async with httpx.AsyncClient(timeout=25.0, follow_redirects=True) as client:
        # Reunir releases + assets .exe/.sha256 candidatos
        try:
            r = await client.get(
                f"https://api.github.com/repos/{owner}/{repo}/releases?per_page=100",
                headers=headers,
            )
            if r.status_code != 200:
                raise HTTPException(status_code=502, detail=f"No se pudo leer Releases (GitHub {r.status_code}).")
            releases = r.json() or []
            candidates = []
            # Mapa release_id → info para la limpieza posterior
            rel_map = {}
            for rel in releases:
                rel_map[rel.get("id")] = {
                    "tag": rel.get("tag_name") or "",
                    "all_assets": [a.get("id") for a in (rel.get("assets") or [])],
                    "name": rel.get("name") or rel.get("tag_name") or "",
                }
                for a in rel.get("assets") or []:
                    low = (a.get("name") or "").lower()
                    if not (low.endswith(".exe") or low.endswith(".exe.sha256")):
                        continue
                    if target_asset_ids and a.get("id") not in target_asset_ids:
                        continue
                    if target_release_id is not None and rel.get("id") != target_release_id:
                        continue
                    a["_release_id"] = rel.get("id")
                    candidates.append(a)
        except HTTPException:
            raise
        except Exception as e:
            raise HTTPException(status_code=502, detail=f"Error al listar builds: {str(e)[:120]}")

        if not candidates:
            return {"success": True, "deleted_count": 0, "freed_bytes": 0,
                    "freed_human": "0 B", "deleted_releases": [], "deleted_tags": [],
                    "message": "No había builds .exe para borrar.", "errors": []}

        deleted_asset_ids = set()
        touched_release_ids = set()
        for a in candidates:
            aid = a.get("id")
            try:
                dr = await client.delete(
                    f"https://api.github.com/repos/{owner}/{repo}/releases/assets/{aid}",
                    headers=headers,
                )
                if dr.status_code in (204, 200):
                    freed_bytes += int(a.get("size") or 0)
                    deleted.append({"asset_id": aid, "name": a.get("name"), "size": int(a.get("size") or 0)})
                    deleted_asset_ids.add(aid)
                    touched_release_ids.add(a.get("_release_id"))
                else:
                    errors.append(f"{a.get('name')}:{dr.status_code}")
            except Exception as e:
                errors.append(f"{a.get('name')}:{str(e)[:60]}")

        # ── Limpieza: borrar releases que quedaron SIN assets + su tag git ──
        if delete_releases:
            for rid in touched_release_ids:
                info = rel_map.get(rid) or {}
                remaining = [aid for aid in info.get("all_assets", []) if aid not in deleted_asset_ids]
                if remaining:
                    continue  # aún tiene otros assets → no borrar el release
                # 1) Borrar el release
                try:
                    rr = await client.delete(
                        f"https://api.github.com/repos/{owner}/{repo}/releases/{rid}",
                        headers=headers,
                    )
                    if rr.status_code in (204, 200):
                        deleted_releases.append(info.get("name") or str(rid))
                    else:
                        errors.append(f"release_{rid}:{rr.status_code}")
                        continue
                except Exception as e:
                    errors.append(f"release_{rid}:{str(e)[:60]}")
                    continue
                # 2) Borrar el tag git asociado (para limpiar la lista de tags)
                tag = info.get("tag") or ""
                if tag:
                    try:
                        tr = await client.delete(
                            f"https://api.github.com/repos/{owner}/{repo}/git/refs/tags/{tag}",
                            headers=headers,
                        )
                        if tr.status_code in (204, 200):
                            deleted_tags.append(tag)
                        elif tr.status_code != 422:  # 422 = el tag ya no existe
                            errors.append(f"tag_{tag}:{tr.status_code}")
                    except Exception as e:
                        errors.append(f"tag_{tag}:{str(e)[:60]}")

    parts_msg = [f"{len(deleted)} archivo(s)", f"{_human_bytes(freed_bytes)} liberados"]
    if deleted_releases:
        parts_msg.append(f"{len(deleted_releases)} release(s) borrados")
    if deleted_tags:
        parts_msg.append(f"{len(deleted_tags)} tag(s) borrados")
    return {
        "success": True,
        "deleted_count": len(deleted),
        "deleted": deleted,
        "freed_bytes": freed_bytes,
        "freed_human": _human_bytes(freed_bytes),
        "deleted_releases": deleted_releases,
        "deleted_tags": deleted_tags,
        "message": "Repositorio limpio · " + " · ".join(parts_msg) + ".",
        "errors": errors,
    }



@api_router.get("/github/push-preview")
async def github_push_preview():
    """Devuelve la lista de categorías/archivos que serían subidos al repositorio.

    Sirve para el modal "Elegir qué subir": el frontend muestra un checkbox por
    categoría y llama a POST /github/push-all con `include={...}` para filtrar.

    Cada categoría incluye:
      - id, label, description
      - files: número aproximado de archivos
      - size_bytes: peso aproximado en disco
      - default: si viene marcado por defecto en el modal
      - slow: True para las tareas lentas (yarn build) → aviso visual
    """
    root_dir = ROOT_DIR.parent

    def _dir_stats(path: Path, ignore_names=None) -> tuple:
        ignore_names = ignore_names or {
            "__pycache__", "node_modules", ".cache", ".pytest_cache",
            "build", ".git",
        }
        if not path.exists():
            return 0, 0
        files, size = 0, 0
        try:
            for p in path.rglob("*"):
                # Excluir cualquier ruta que contenga un directorio ignorado
                if any(part in ignore_names for part in p.parts):
                    continue
                if p.is_file():
                    files += 1
                    try:
                        size += p.stat().st_size
                    except Exception:
                        pass
        except Exception:
            pass
        return files, size

    def _root_files_stats() -> tuple:
        excluded_dirs = {"backend", "frontend", ".git", "node_modules",
                         "backups", "uploads", "memory", "test_reports",
                         "tests", "repo", "__pycache__", ".cache",
                         "desktop_wheels"}
        included_dirs = {"scripts", "docs", ".github", "public"}
        files, size = 0, 0
        try:
            for item in root_dir.iterdir():
                if item.name.startswith(".env"):
                    continue
                if item.is_file():
                    files += 1
                    try:
                        size += item.stat().st_size
                    except Exception:
                        pass
                elif item.is_dir() and item.name in included_dirs and item.name not in excluded_dirs:
                    f, s = _dir_stats(item)
                    files += f
                    size += s
        except Exception:
            pass
        return files, size

    backend_files, backend_size = _dir_stats(root_dir / "backend")
    frontend_files, frontend_size = _dir_stats(root_dir / "frontend")
    root_files, root_size = _root_files_stats()
    standalone_exists = (ROOT_DIR / "standalone_app.py").exists()
    version_exists = (ROOT_DIR.parent / "version.txt").exists() or (ROOT_DIR / "version.txt").exists()

    categories = [
        {
            "id": "backend",
            "label": "Backend (Python)",
            "description": "Código del servidor FastAPI (backend/)",
            "files": backend_files,
            "size_bytes": backend_size,
            "default": True,
            "slow": False,
        },
        {
            "id": "frontend_src",
            "label": "Frontend (código fuente)",
            "description": "React source, package.json, tailwind, public/ (frontend/)",
            "files": frontend_files,
            "size_bytes": frontend_size,
            "default": True,
            "slow": False,
        },
        {
            "id": "root_files",
            "label": "Archivos raíz",
            "description": "README.md, bootstrap.sh, .github/workflows, scripts/, etc.",
            "files": root_files,
            "size_bytes": root_size,
            "default": True,
            "slow": False,
        },
        {
            "id": "standalone_app",
            "label": "App de escritorio (app.py)",
            "description": "backend/standalone_app.py → app.py (raíz del repo)",
            "files": 1 if standalone_exists else 0,
            "size_bytes": 0,
            "default": True,
            "slow": False,
        },
        {
            "id": "version_txt",
            "label": "version.txt",
            "description": "Etiqueta la versión (usada por la app de escritorio para saber si hay update)",
            "files": 1,
            "size_bytes": 0,
            "default": True,
            "slow": False,
        },
        {
            "id": "build_frontend",
            "label": "Compilar frontend (yarn build)",
            "description": "Ejecuta yarn build y sube build/ a la raíz. Suele tardar 1–2 min. GitHub Actions ya lo hace al crear el .exe, así que puedes desmarcarlo para ir más rápido.",
            "files": 0,
            "size_bytes": 0,
            "default": False,
            "slow": True,
        },
    ]
    total_files = sum(c["files"] for c in categories if c.get("default"))
    total_size = sum(c["size_bytes"] for c in categories if c.get("default"))
    return {
        "categories": categories,
        "totals_defaults": {"files": total_files, "size_bytes": total_size},
    }


# Cache simple en memoria para no re-clonar en cada apertura del modal.
_push_diff_cache = {"data": None, "ts": 0.0}


def _categorize_repo_path(rel_path: str) -> str:
    """Mapea una ruta relativa del repo a la categoría del modal 'Elegir qué subir'."""
    p = rel_path.replace("\\", "/").lstrip("/")
    if p.startswith("backend/"):
        return "backend"
    if p.startswith("frontend/"):
        return "frontend_src"
    if p == "app.py":
        return "standalone_app"
    if p == "version.txt":
        return "version_txt"
    if p.startswith("build/"):
        return "build_frontend"
    return "root_files"


@api_router.get("/github/push-diff")
async def github_push_diff(refresh: bool = False):
    """Compara el código LOCAL con la versión ACTUAL del repositorio en GitHub y
    devuelve la lista REAL de archivos que cambiarían (nuevos, modificados y
    eliminados), agrupados por categoría.

    Esto alimenta la vista 'Ver cambios detallados' del modal para que el usuario
    vea EXACTAMENTE qué se va a subir antes de publicar.

    Response:
    {
      "ok": true,
      "changed": 12,
      "summary": {"added": 3, "modified": 8, "deleted": 1},
      "by_category": {"backend": {"added":1,"modified":2,"deleted":0,"total":3}, ...},
      "files": [{"path": "backend/server.py", "status": "M", "category": "backend", "size_bytes": 1234}, ...],
      "truncated": false,
      "remote_branch": "main",
      "cached": false
    }
    """
    import shutil, tempfile, time as _t

    # Cache de 90s para respuestas instantáneas al reabrir el modal.
    if not refresh and _push_diff_cache["data"] and (_t.time() - _push_diff_cache["ts"] < 90):
        cached = dict(_push_diff_cache["data"])
        cached["cached"] = True
        return cached

    cfg = await _get_github_config()
    token = cfg.get("token", "")
    repo_url = cfg.get("repo_url", "")
    branch = cfg.get("branch", DEFAULT_GITHUB_BRANCH)
    username = cfg.get("username", "")
    if not token or not repo_url or not username:
        raise HTTPException(status_code=400, detail="Conecta tu cuenta de GitHub primero para comparar los cambios.")

    auth_url = repo_url.replace("https://", f"https://{username}:{token}@")
    work_dir = Path(tempfile.mkdtemp(prefix="cp_diff_"))

    def _run(args, cwd=None, timeout=120):
        r = subprocess.run(args, cwd=str(cwd) if cwd else None,
                            capture_output=True, text=True, timeout=timeout)
        return r.returncode, (r.stdout or "") + (r.stderr or "")

    async def _arun(args, cwd=None, timeout=120):
        return await asyncio.to_thread(_run, args, cwd=cwd, timeout=timeout)

    try:
        # 1. Clonar remoto (shallow + sin blobs = rápido)
        rc, out = await _arun(
            ["git", "clone", "--depth", "1", "--filter=blob:none", "--branch", branch, auth_url, str(work_dir)],
            timeout=180,
        )
        if rc != 0:
            rc2, out2 = await _arun(["git", "clone", "--depth", "1", "--filter=blob:none", auth_url, str(work_dir)], timeout=180)
            if rc2 != 0:
                raise HTTPException(status_code=502, detail="No se pudo clonar el repositorio para comparar.")

        # 2. Espejar el contenido local (mismos filtros que el push real)
        ignore_patterns = shutil.ignore_patterns(
            "__pycache__", "*.pyc", ".pytest_cache",
            "node_modules", "build", ".cache",
            ".env", ".env.local", ".env.production", ".env.development",
            ".db_override", "backups", "uploads",
            "cinema_data.json", "cinema_data.json.bak",
            "*.log", ".DS_Store", "desktop_wheels",
        )
        root_dir = ROOT_DIR.parent
        for d in ("backend", "frontend"):
            src = root_dir / d
            dst = work_dir / d
            if src.exists():
                if dst.exists():
                    await asyncio.to_thread(shutil.rmtree, str(dst))
                await asyncio.to_thread(shutil.copytree, str(src), str(dst), ignore=ignore_patterns)

        excluded_root_dirs = {"backend", "frontend", ".git", "node_modules",
                              "backups", "uploads", "memory", "test_reports",
                              "tests", "repo", "__pycache__", ".cache", "desktop_wheels"}
        for item in root_dir.iterdir():
            if item.name in excluded_root_dirs or item.name.startswith(".env"):
                continue
            if item.name == "version.txt":
                continue
            if item.is_file():
                try:
                    shutil.copy2(str(item), str(work_dir / item.name))
                except Exception:
                    pass
            elif item.is_dir() and item.name in {"scripts", "docs", ".github", "public"}:
                dst_sub = work_dir / item.name
                if dst_sub.exists():
                    await asyncio.to_thread(shutil.rmtree, str(dst_sub))
                await asyncio.to_thread(shutil.copytree, str(item), str(dst_sub), ignore=ignore_patterns)

        # App de escritorio (app.py) y version.txt
        standalone_src = ROOT_DIR / "standalone_app.py"
        if standalone_src.exists():
            try:
                shutil.copy2(str(standalone_src), str(work_dir / "app.py"))
            except Exception:
                pass
        try:
            local_ver = await _read_local_version()
            next_ver = await _compute_next_unified_version()
            (work_dir / "version.txt").write_text(next_ver or local_ver or "", encoding="utf-8")
        except Exception:
            pass

        # 3. git status --porcelain para el diff REAL
        await _arun(["git", "add", "-A"], cwd=work_dir)
        for forced in ("app.py", "version.txt"):
            if (work_dir / forced).exists():
                await _arun(["git", "add", "-f", forced], cwd=work_dir)
        rc_st, status_out = await _arun(["git", "status", "--porcelain"], cwd=work_dir)

        files = []
        summary = {"added": 0, "modified": 0, "deleted": 0}
        by_category = {}
        MAX_FILES = 800
        for line in status_out.split("\n"):
            if not line.strip():
                continue
            code = line[:2]
            rest = line[3:].strip()
            # Renombrados: "old -> new"
            if " -> " in rest:
                rest = rest.split(" -> ", 1)[1]
            rest = rest.strip().strip('"')
            xy = code.replace(" ", "")
            if "D" in xy:
                status = "D"
                summary["deleted"] += 1
            elif "A" in xy or "?" in xy:
                status = "A"
                summary["added"] += 1
            else:
                status = "M"
                summary["modified"] += 1

            cat = _categorize_repo_path(rest)
            bc = by_category.setdefault(cat, {"added": 0, "modified": 0, "deleted": 0, "total": 0})
            bc["total"] += 1
            bc["added" if status == "A" else "deleted" if status == "D" else "modified"] += 1

            if len(files) < MAX_FILES:
                size_bytes = 0
                if status != "D":
                    try:
                        size_bytes = (work_dir / rest).stat().st_size
                    except Exception:
                        size_bytes = 0
                files.append({"path": rest, "status": status, "category": cat, "size_bytes": size_bytes})

        # Orden: primero por categoría, luego eliminados/modificados/nuevos, luego path
        _st_order = {"D": 0, "M": 1, "A": 2}
        files.sort(key=lambda f: (f["category"], _st_order.get(f["status"], 3), f["path"]))

        changed = summary["added"] + summary["modified"] + summary["deleted"]
        result = {
            "ok": True,
            "changed": changed,
            "summary": summary,
            "by_category": by_category,
            "files": files,
            "truncated": changed > MAX_FILES,
            "remote_branch": branch,
            "cached": False,
            "generated_at": datetime.now(timezone.utc).isoformat(),
        }
        _push_diff_cache["data"] = result
        _push_diff_cache["ts"] = _t.time()
        return result
    finally:
        await asyncio.to_thread(shutil.rmtree, str(work_dir), ignore_errors=True)



@api_router.post("/github/push-all")
async def github_push_all(payload: dict = Body(default={})):
    """Lanza el push-all en background para evitar timeouts de Cloudflare (>100s).
    El frontend hace polling a /github/push-status para conocer el resultado final."""
    # Verificaciones rápidas antes de arrancar el worker
    cfg = await _get_github_config()
    if not cfg.get("token"):
        raise HTTPException(status_code=400, detail="Sin cuenta conectada. Usa 'Conectar con GitHub' primero.")
    if not cfg.get("repo_url") or not cfg.get("username"):
        raise HTTPException(status_code=400, detail="Repositorio o usuario no configurado.")

    if _push_state.get("status") == "running":
        return {
            "status": "already_running",
            "message": "Ya hay un push en curso. Espera a que termine.",
            "progress": _push_state.get("progress", 0),
        }

    # Limpia el resultado anterior
    _push_state["result"] = None
    _push_state["error"] = None
    import time as _time_init
    _set_push_state(progress=1, message="En cola…", status="running", step=0, detail="Preparando el entorno")
    _push_state["total_steps"] = len(PUSH_STEP_LABELS)
    _push_state["started_at"] = datetime.now(timezone.utc).isoformat()
    _push_state["started_ts"] = _time_init.time()
    _push_state["finished_at"] = None

    asyncio.create_task(_do_github_push_all(payload))
    return {"status": "started", "message": "Push iniciado en background. Sigue el progreso en /github/push-status."}


async def _do_github_push_all(payload: dict):
    """Sube TODO el código actual al repositorio de GitHub, funcionando como
    el botón "Save to GitHub" de Emergent.

    Flujo:
    1. Clona el repo remoto en una carpeta temporal (--depth 1)
    2. Copia el backend/, frontend/ y archivos raíz actuales dentro del clone,
       preservando el .git del clone. NUNCA copia .env, node_modules, __pycache__,
       backups/, uploads/. SÍ incluye frontend/build/ (compilado) para que la
       app de escritorio en otras PCs reciba la interfaz actualizada.
    3. Configura git identity con la cuenta conectada
    4. git add -A && git commit && git push origin <branch>
    5. Limpia la carpeta temporal.

    Requiere:
    - github_config.token (via POST /github/connect)
    - github_config.repo_url y username
    """
    import shutil, tempfile

    cfg = await _get_github_config()
    token = cfg.get("token", "")
    repo_url = cfg.get("repo_url", "")
    branch = cfg.get("branch", DEFAULT_GITHUB_BRANCH)
    username = cfg.get("username", "")

    if not token:
        raise HTTPException(status_code=400, detail="Sin cuenta conectada. Usa 'Conectar con GitHub' primero.")
    if not repo_url or not username:
        raise HTTPException(status_code=400, detail="Repositorio o usuario no configurado.")

    message = (payload.get("message") or f"Auto-save from Cinema Productions — {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}").strip()

    # ── Filtro de inclusión (viene del modal "Elegir qué subir") ───────────
    # Si no viene ningún filtro, mantener comportamiento antiguo (todo TRUE
    # excepto compilar frontend, que ahora es opt-in porque GitHub Actions ya
    # lo hace al crear el .exe y así el push es mucho más rápido).
    include_raw = payload.get("include")
    if isinstance(include_raw, dict):
        include = {
            "backend":        bool(include_raw.get("backend", True)),
            "frontend_src":   bool(include_raw.get("frontend_src", True)),
            "root_files":     bool(include_raw.get("root_files", True)),
            "standalone_app": bool(include_raw.get("standalone_app", True)),
            "version_txt":    bool(include_raw.get("version_txt", True)),
            "build_frontend": bool(include_raw.get("build_frontend", False)),
        }
    else:
        # Compatibilidad hacia atrás con el flag `build_frontend` que ya existía.
        include = {
            "backend": True,
            "frontend_src": True,
            "root_files": True,
            "standalone_app": True,
            "version_txt": True,
            "build_frontend": bool(payload.get("build_frontend", False)),
        }

    # URL con auth para el clone/push
    auth_url = repo_url.replace("https://", f"https://{username}:{token}@")

    # Directorio temporal (limpio) para clonar
    work_dir = Path(tempfile.mkdtemp(prefix="cp_push_"))
    _set_push_state(progress=5, message="Conectando con GitHub…", status="running", step=1,
                    detail="Verificando credenciales del repositorio")
    _push_state["started_at"] = datetime.now(timezone.utc).isoformat()
    _push_state["finished_at"] = None

    # Helper para ejecutar comandos git
    def _run(args, cwd=None, timeout=120, sensitive=False):
        result = subprocess.run(
            args, cwd=str(cwd) if cwd else None,
            capture_output=True, text=True, timeout=timeout,
        )
        out = (result.stdout or "") + (result.stderr or "")
        # Ocultar el token de los logs
        if sensitive and token:
            out = out.replace(token, "***")
        return result.returncode, out

    # Versión NO bloqueante: ejecuta el comando en un hilo aparte para que el
    # event loop siga libre y `/github/push-status` responda EN VIVO durante
    # operaciones largas (clone, yarn build, push). Antes bloqueaba el loop y la
    # barra parecía congelada 1–2 min.
    async def _arun(args, cwd=None, timeout=120, sensitive=False):
        return await asyncio.to_thread(_run, args, cwd=cwd, timeout=timeout, sensitive=sensitive)

    try:
        # ── 1. Clonar el repo remoto ─────────────────────────────────
        _set_push_state(progress=8, message="Clonando repositorio desde GitHub…", step=2,
                        detail="Descargando la última versión de la rama")
        rc, out = await _arun(
            ["git", "clone", "--depth", "1", "--branch", branch, auth_url, str(work_dir)],
            timeout=180, sensitive=True,
        )
        if rc != 0:
            # Puede ser porque la rama no existe aún → clonar sin branch
            rc2, out2 = await _arun(["git", "clone", "--depth", "1", auth_url, str(work_dir)], timeout=180, sensitive=True)
            if rc2 != 0:
                raise HTTPException(status_code=502, detail=f"git clone falló: {out2[-400:]}")
            # Crear la rama solicitada
            await _arun(["git", "checkout", "-b", branch], cwd=work_dir)

        # ── 2. Configurar identidad ──────────────────────────────────
        _set_push_state(progress=20, message="Repositorio clonado. Preparando compilación…", step=3,
                        detail="Configurando identidad de Git")
        await _arun(["git", "config", "user.email", f"{username}@users.noreply.github.com"], cwd=work_dir)
        await _arun(["git", "config", "user.name", username], cwd=work_dir)

        # ── 3. Copiar contenido actual sobre el clone (sin tocar .git) ──
        # Directorios completos a espejar (según include)
        mirror_dirs = []
        if include.get("backend"):
            mirror_dirs.append("backend")
        if include.get("frontend_src"):
            mirror_dirs.append("frontend")

        # Patrones a ignorar SIEMPRE (nunca subir a GitHub)
        # NOTA: NO subimos frontend/build (fuente compilado nested); en su lugar
        # publicamos la "versión para PC" PLANA en la raíz (app.py + build/), que
        # es lo que el actualizador de escritorio sabe aplicar directamente.
        ignore_patterns = shutil.ignore_patterns(
            "__pycache__", "*.pyc", ".pytest_cache",
            "node_modules", "build", ".cache",
            ".env", ".env.local", ".env.production", ".env.development",
            ".db_override", "backups", "uploads",
            "cinema_data.json", "cinema_data.json.bak",
            "*.log", ".DS_Store", "desktop_wheels",
        )

        # ── 2b. Compilar el frontend (opt-in: solo si el usuario lo marcó) ──
        if include.get("build_frontend"):
            frontend_src = ROOT_DIR.parent / "frontend"
            if frontend_src.exists():
                import time as _time_b
                _set_push_state(progress=30, message="Compilando interfaz (yarn build)…", step=4,
                                detail="Esto suele tardar 1–2 min · no cierres esta ventana")
                # Ticker: mantiene la barra en movimiento mientras compila y muestra
                # los segundos transcurridos + una pista rotativa de la fase.
                _build_stop = asyncio.Event()
                _build_hints = [
                    "Optimizando componentes de React…",
                    "Minificando CSS y JavaScript…",
                    "Generando bundles de producción…",
                    "Aplicando Tailwind y estilos…",
                    "Creando archivos estáticos…",
                ]
                _ticker_task = asyncio.create_task(
                    _push_progress_ticker(_build_stop, _time_b.time(), lo=30, hi=58,
                                          base_message="Compilando interfaz (yarn build)",
                                          hints=_build_hints)
                )
                try:
                    rc_b, out_b = await _arun(["yarn", "build"], cwd=frontend_src, timeout=600)
                finally:
                    _build_stop.set()
                    try:
                        await _ticker_task
                    except Exception:
                        pass
                if rc_b != 0:
                    shutil.rmtree(str(work_dir), ignore_errors=True)
                    _set_push_state(progress=0, message="Error al compilar el frontend", status="error",
                                    detail="Revisa los logs del build")
                    raise HTTPException(
                        status_code=500,
                        detail=f"yarn build falló, no se subió nada: {out_b[-400:]}",
                    )
                _set_push_state(progress=60, message="Interfaz compilada. Empaquetando archivos…", step=4,
                                detail="Copiando código fuente al repositorio")


        for idx, d in enumerate(mirror_dirs, start=1):
            src = ROOT_DIR.parent / d
            dst = work_dir / d
            if src.exists():
                _set_push_state(progress=62 + idx * 2, message="Empaquetando archivos…", step=4,
                                detail=f"Copiando carpeta «{d}» ({idx}/{len(mirror_dirs)})")
                if dst.exists():
                    shutil.rmtree(str(dst))
                await asyncio.to_thread(shutil.copytree, str(src), str(dst), ignore=ignore_patterns)

        # Archivos raíz del repo — copiar TODOS los archivos individuales
        # (no directorios; excepto los ya cubiertos arriba)
        excluded_root_dirs = {"backend", "frontend", ".git", "node_modules",
                              "backups", "uploads", "memory", "test_reports",
                              "tests", "repo", "__pycache__", ".cache",
                              "desktop_wheels"}
        root_dir = ROOT_DIR.parent
        if include.get("root_files"):
            for item in root_dir.iterdir():
                if item.name in excluded_root_dirs:
                    continue
                if item.name.startswith(".env"):
                    continue  # nunca subir .env
                # version.txt se maneja aparte según su propio flag
                if item.name == "version.txt":
                    continue
                if item.is_file():
                    # Copiar cualquier archivo raíz (README.md, bootstrap.sh, yarn.lock, etc.)
                    try:
                        shutil.copy2(str(item), str(work_dir / item.name))
                    except Exception as e:
                        logger.warning(f"No se pudo copiar {item.name}: {e}")
                elif item.is_dir():
                    # Copiar carpetas pequeñas de config si existen (scripts, docs, .github, etc.)
                    if item.name in {"scripts", "docs", ".github", "public"}:
                        dst_sub = work_dir / item.name
                        if dst_sub.exists():
                            shutil.rmtree(str(dst_sub))
                        shutil.copytree(str(item), str(dst_sub), ignore=ignore_patterns)

        # ── 3b. Publicar la "VERSIÓN PARA PC" (layout PLANO) en la RAÍZ del repo ──
        # La app de escritorio instalada tiene layout plano: install_dir/app.py y
        # install_dir/build/. Su actualizador copia los archivos de la RAÍZ del repo
        # directo a la carpeta de instalación. Por eso publicamos aquí:
        #   raíz/app.py    (= backend/standalone_app.py)
        #   raíz/build/    (= frontend/build compilado)
        #   raíz/version.txt
        # Así CUALQUIER actualizador (el ya instalado o el nuevo) aplica los cambios
        # correctamente y de forma automática y remota, sin reinstalar.
        # ── Resolver la VERSIÓN de esta publicación ──
        # Fuente única de verdad: _compute_next_unified_version() → semver X.Y.Z.
        # Garantiza que ZIP, EXE, cloud release, commit y tag comparten el MISMO
        # número. El usuario puede seguir sobrescribiendo con custom_version
        # desde el modal ("2.0", "Navidad", etc.) — se respeta tal cual.
        custom_version = (payload.get("version") or "").strip().lstrip("vV").strip()
        version_name = (payload.get("version_name") or "").strip()
        try:
            if custom_version:
                version_str = custom_version
            else:
                version_str = await _compute_next_unified_version()
                await db.app_settings.update_one(
                    {}, {"$set": {"desktop_build_semver": version_str}}, upsert=True
                )
        except Exception as ver_err:
            logger.warning(f"No se pudo resolver versión: {ver_err}")
            version_str = custom_version or datetime.now(timezone.utc).strftime("%Y.%m.%d.%H%M")

        try:
            standalone_src = ROOT_DIR / "standalone_app.py"
            _set_push_state(progress=75, message=f"Preparando versión v{version_str} para PC…", step=5,
                            detail="Publicando app.py, build/ y version.txt para la app de escritorio")
            if include.get("standalone_app") and standalone_src.exists():
                shutil.copy2(str(standalone_src), str(work_dir / "app.py"))
            if include.get("build_frontend"):
                fe_build = ROOT_DIR.parent / "frontend" / "build"
                if fe_build.exists() and (fe_build / "index.html").exists():
                    root_build = work_dir / "build"
                    if root_build.exists():
                        shutil.rmtree(str(root_build))
                    await asyncio.to_thread(shutil.copytree, str(fe_build), str(root_build))
            if include.get("version_txt"):
                (work_dir / "version.txt").write_text(version_str, encoding="utf-8")
        except Exception as e:
            logger.warning(f"No se pudo publicar la versión PC plana: {e}")

        # ── 4. Add + Commit + Push ───────────────────────────────────
        _set_push_state(progress=85, message="Registrando cambios (git add)…", step=6,
                        detail="Comparando archivos con el repositorio")
        await _arun(["git", "add", "-A"], cwd=work_dir)
        # Forzar el add de la versión PC (build/ está en .gitignore como /build)
        forced_items = []
        if include.get("build_frontend"):
            forced_items.append("build")
        if include.get("standalone_app"):
            forced_items.append("app.py")
        if include.get("version_txt"):
            forced_items.append("version.txt")
        for forced in forced_items:
            if (work_dir / forced).exists():
                await _arun(["git", "add", "-f", forced], cwd=work_dir)

        # ── 3c. FORZAR inclusión de archivos críticos de infraestructura ───
        # Estos archivos SIEMPRE deben ir al repo si existen en el checkout
        # local, INDEPENDIENTEMENTE de las categorías marcadas en el modal:
        #   · bootstrap.sh         (arranque + reconciliación de deps)
        #   · .github/workflows/*  (CI/CD atómico: build → manifest → tag)
        #   · backend/updater.py   (auto-updater cliente con validación SHA256)
        #   · version.json         (semáforo de manifest, si existe local)
        # Sin esto, un push con `root_files` desmarcado dejaba el repo sin
        # los fixes de infraestructura y los bugs (deps faltantes, race con
        # el .exe, etc.) volvían a aparecer en el siguiente clone.
        _force_paths = [
            "bootstrap.sh",
            ".github/workflows/auto-release.yml",
            ".github/workflows/build-exe.yml",
            ".github/workflows/refresh-deps.yml",
            "backend/updater.py",
            "version.json",
        ]
        _forced_count = 0
        for rel in _force_paths:
            src_p = ROOT_DIR.parent / rel
            dst_p = work_dir / rel
            if not src_p.exists():
                continue
            try:
                dst_p.parent.mkdir(parents=True, exist_ok=True)
                shutil.copy2(str(src_p), str(dst_p))
                rc_af, _ = await _arun(["git", "add", "-f", rel], cwd=work_dir)
                if rc_af == 0:
                    _forced_count += 1
            except Exception as e:
                logger.warning(f"[push-all] no se pudo forzar {rel}: {e}")
        if _forced_count:
            logger.info(f"[push-all] {_forced_count} archivo(s) crítico(s) forzado(s) al push")

        rc_st, status_out = await _arun(["git", "status", "--porcelain"], cwd=work_dir)
        if not status_out.strip():
            _set_push_state(progress=100, message="Sin cambios que subir — ya está sincronizado", status="done",
                            step=len(PUSH_STEP_LABELS), detail="El repositorio ya estaba al día")
            _push_state["finished_at"] = datetime.now(timezone.utc).isoformat()
            _push_state["result"] = {
                "success": True,
                "nothing_to_commit": True,
                "message": "Sin cambios que subir — el repositorio ya está sincronizado",
            }
            return

        # Contar archivos cambiados para el resumen
        changed = len([l for l in status_out.strip().split("\n") if l.strip()])

        _set_push_state(progress=90, message="Creando commit…", step=7,
                        detail=f"{changed} archivo(s) con cambios · versión v{version_str}")
        # Añadimos [skip ci] para que auto-release.yml NO haga un segundo bump.
        # El tag vX.Y.Z que creamos abajo dispara build-exe.yml directamente,
        # que compila EXE/installer/linux/macos con EXACTAMENTE la MISMA versión.
        commit_message = f"{message}\n\nv{version_str} [skip ci]"
        rc_c, out_c = await _arun(["git", "commit", "-m", commit_message], cwd=work_dir)
        if rc_c != 0:
            _set_push_state(progress=0, message="Error al crear el commit", status="error")
            raise HTTPException(status_code=500, detail=f"git commit falló: {out_c[-400:]}")

        _set_push_state(progress=95, message="Subiendo a GitHub…", step=8,
                        detail=f"Enviando {changed} archivo(s) a la rama «{branch}»")
        rc_p, out_p = await _arun(["git", "push", "origin", branch], cwd=work_dir, timeout=300, sensitive=True)
        if rc_p != 0:
            _set_push_state(progress=0, message="Error al subir a GitHub", status="error")
            raise HTTPException(status_code=502, detail=f"git push falló: {out_p[-500:]}")

        # ── 5. Obtener SHA del commit ────────────────────────────────
        rc_sha, sha = await _arun(["git", "rev-parse", "HEAD"], cwd=work_dir)
        new_sha = sha.strip() if rc_sha == 0 else ""

        # ── 6. Crear registro en app_updates (para que aparezca en el historial
        #    y sea detectado por la app de escritorio como nueva versión) ──
        try:
            # Reutiliza la versión ya resuelta arriba (custom o auto 1.NN).
            # Marcar todos los anteriores como no-latest
            await db.app_updates.update_many({}, {"$set": {"is_latest": False}})
            await db.app_updates.insert_one({
                "version": version_str,
                "version_name": version_name or None,
                "commit_sha": new_sha,
                "commit_short": new_sha[:7],
                "source": "github_push",
                "branch": branch,
                "notes": message,
                "channel": "github",
                "files_count": changed,
                "created_at": datetime.now(timezone.utc).isoformat(),
                "is_latest": True,
                "author": username,
                "repo_url": repo_url,
            })
            registered_version = version_str

            # ── 6b. Crear TAG en GitHub para versionar el historial ──
            # Los tags v1.NN quedan como "releases" persistentes en GitHub que
            # las apps de escritorio pueden consultar directamente sin necesitar
            # una base de datos compartida.
            tag_name = f"v{version_str}"
            try:
                # Crea tag anotado y lo pushea
                rc_tag, out_tag = await _arun(
                    ["git", "tag", "-a", tag_name, "-m", message], cwd=work_dir
                )
                if rc_tag == 0:
                    rc_push_tag, out_push_tag = await _arun(
                        ["git", "push", "origin", tag_name],
                        cwd=work_dir, timeout=120, sensitive=True,
                    )
                    if rc_push_tag != 0:
                        logger.warning(f"No se pudo pushear tag {tag_name}: {out_push_tag[-200:]}")
                else:
                    logger.warning(f"No se pudo crear tag {tag_name}: {out_tag[-200:]}")
            except Exception as tag_err:
                logger.warning(f"Error creando tag: {tag_err}")

            # ── 6c. Disparar la compilación del .exe EXPLÍCITAMENTE ──────────
            # IMPORTANTE: el commit lleva "[skip ci]" para que auto-release.yml
            # no haga un segundo bump. Pero GitHub evalúa "[skip ci]" sobre todo
            # el push, así que el push del tag NO dispara build-exe.yml (el .exe
            # nunca se construía). Por eso lo lanzamos vía workflow_dispatch API,
            # que NO se ve afectado por "[skip ci]".
            try:
                _set_push_state(progress=98, message="Iniciando compilación del .exe en GitHub Actions…", step=8,
                                detail="Disparando el workflow «Build Windows .exe»")
                dispatch = await _trigger_exe_build_workflow(
                    owner_repo=None, token=token, repo_url=repo_url,
                    branch=branch, version=version_str,
                )
                if dispatch.get("ok"):
                    logger.info(f"[push-all] build-exe.yml disparado para v{version_str}")
                else:
                    logger.warning(f"[push-all] no se pudo disparar build-exe.yml: {dispatch.get('error')}")
            except Exception as disp_err:
                logger.warning(f"[push-all] excepción al disparar build-exe.yml: {disp_err}")

            # ── 6d. Disparar refresh-deps.yml SIEMPRE ────────────────────────
            # El commit lleva "[skip ci]" así que refresh-deps.yml no corre
            # automáticamente. Lo forzamos vía workflow_dispatch en CADA push
            # para garantizar que el release 'deps-latest' quede siempre fresco
            # como backup del tarball committeado. El workflow es idempotente y
            # rápido (~5-8 min) — no ejecuta build si ya está actualizado.
            try:
                _set_push_state(progress=99, message="Refrescando cache de dependencias en GitHub…", step=8,
                                detail="Disparando el workflow «refresh-deps»")
                rd = await _trigger_refresh_deps_workflow(token=token, repo_url=repo_url, branch=branch)
                if rd.get("ok"):
                    logger.info("[push-all] refresh-deps.yml disparado")
                else:
                    logger.warning(f"[push-all] no se pudo disparar refresh-deps.yml: {rd.get('error')}")
            except Exception as rd_err:
                logger.warning(f"[push-all] excepción al disparar refresh-deps.yml: {rd_err}")
        except Exception as reg_err:
            logger.warning(f"No se pudo registrar el push como update: {reg_err}")
            registered_version = None

        # Guardar en la BD
        await db.app_settings.update_one(
            {},
            {"$set": {
                "github_config.last_commit_sha": new_sha,
                "github_config.last_push_at": datetime.now(timezone.utc).isoformat(),
                "github_config.last_push_message": message,
                "github_config.last_push_files": changed,
            }},
            upsert=True,
        )

        _set_push_state(progress=100, message="¡Subido a GitHub! La versión PC ya está publicada.", status="done",
                        step=len(PUSH_STEP_LABELS), detail=f"Commit {new_sha[:7] if new_sha else ''} · v{registered_version or version_str}")
        _push_state["finished_at"] = datetime.now(timezone.utc).isoformat()
        _push_state["result"] = {
            "success": True,
            "nothing_to_commit": False,
            "commit_sha": new_sha,
            "commit_short": new_sha[:7] if new_sha else "",
            "branch": branch,
            "message": message,
            "files_changed": changed,
            "repo_url": repo_url,
            "version": registered_version,
        }
        return
    except subprocess.TimeoutExpired as e:
        _set_push_state(progress=0, message=f"Timeout ({e.timeout}s). Verifica tu conexión.", status="error")
        _push_state["error"] = f"Timeout ({e.timeout}s)"
        _push_state["finished_at"] = datetime.now(timezone.utc).isoformat()
        return
    except HTTPException as he:
        _set_push_state(progress=0, message=str(he.detail)[:200], status="error")
        _push_state["error"] = str(he.detail)[:400]
        _push_state["finished_at"] = datetime.now(timezone.utc).isoformat()
        return
    except Exception as e:
        _set_push_state(progress=0, message=f"Error inesperado: {type(e).__name__}", status="error")
        _push_state["error"] = f"{type(e).__name__}: {e}"[:400]
        _push_state["finished_at"] = datetime.now(timezone.utc).isoformat()
        return
    finally:
        try:
            shutil.rmtree(str(work_dir), ignore_errors=True)
        except Exception:
            pass


@api_router.get("/github/push-status")
async def github_push_status():
    """Estado actual del push a GitHub (para la barra de progreso en la UI).
    Incluye `result` cuando termina OK y `error` cuando falla.
    Añade step/total_steps/detail/elapsed_seconds para un progreso más detallado."""
    import time as _time_st
    started_ts = _push_state.get("started_ts")
    elapsed = int(_time_st.time() - started_ts) if started_ts and _push_state.get("status") == "running" else None
    step = _push_state.get("step", 0)
    total = _push_state.get("total_steps", len(PUSH_STEP_LABELS))
    step_label = PUSH_STEP_LABELS[step - 1] if 1 <= step <= len(PUSH_STEP_LABELS) else ""
    return {
        "status": _push_state.get("status", "idle"),
        "progress": _push_state.get("progress", 0),
        "message": _push_state.get("message", ""),
        "detail": _push_state.get("detail", ""),
        "step": step,
        "total_steps": total,
        "step_label": step_label,
        "steps": PUSH_STEP_LABELS,
        "elapsed_seconds": elapsed,
        "started_at": _push_state.get("started_at"),
        "finished_at": _push_state.get("finished_at"),
        "result": _push_state.get("result"),
        "error": _push_state.get("error"),
    }


# ── DIAGNÓSTICO DE LA APP ─────────────────────────────────────────
@api_router.get("/diagnostic")
async def run_diagnostic():
    """Ejecuta chequeos de salud sobre la app: dependencias, DB, GitHub, seguridad."""
    import importlib.util, sys as _sys

    checks = []

    def add(id_, label, ok, detail="", severity="error", fixable=False):
        checks.append({
            "id": id_,
            "label": label,
            "ok": bool(ok),
            "detail": detail,
            "severity": severity if not ok else "ok",
            "fixable": bool(fixable),
        })

    # 1) Dependencias Python críticas
    critical_py = ["fastapi", "motor", "pymongo", "apscheduler", "bcrypt", "pyzipper", "requests"]
    missing_py = [p for p in critical_py if importlib.util.find_spec(p) is None]
    add("python_deps",
        f"Dependencias Python ({len(critical_py)} críticas)",
        not missing_py,
        f"Faltantes: {', '.join(missing_py)}" if missing_py else f"Todas instaladas: {', '.join(critical_py)}",
        fixable=True,
    )

    # 2) Dependencias Node
    node_modules = ROOT_DIR.parent / "frontend" / "node_modules"
    yarn_lock = ROOT_DIR.parent / "frontend" / "yarn.lock"
    add("node_deps",
        "Dependencias Node (yarn install)",
        node_modules.exists() and yarn_lock.exists(),
        "node_modules OK y yarn.lock presente" if node_modules.exists() else "Falta node_modules — ejecuta yarn install",
        fixable=True,
    )

    # 3) Servidor MongoDB
    try:
        await db.command("ping")
        server_info = await db.client.server_info()
        mongo_version = server_info.get("version", "?")
        add("mongo_conn", f"MongoDB conectado (v{mongo_version})", True, f"URL: {_mongo_url[:35]}...")
    except Exception as e:
        add("mongo_conn", "MongoDB conectado", False, str(e)[:200], fixable=True)

    # 4) Solo la BD por defecto (sin bases externas rondando)
    try:
        db_list = await db.client.list_database_names()
        user_dbs = [d for d in db_list if d not in ("admin", "config", "local")]
        expected = DB_NAME
        extras = [d for d in user_dbs if d != expected]
        add("mongo_isolation",
            "Aislamiento de base de datos",
            len(extras) == 0,
            f"Solo '{expected}' (correcto)" if not extras else f"Bases extra detectadas: {', '.join(extras)}",
            severity="warning",
        )
    except Exception as e:
        add("mongo_isolation", "Aislamiento de base de datos", False, str(e)[:120])

    # 5) Sin usuarios/autenticación en la instancia local
    try:
        users_info = await db.client.admin.command({"usersInfo": 1})
        users = users_info.get("users", [])
        add("mongo_no_auth",
            "MongoDB sin cuentas externas",
            len(users) == 0,
            "Sin usuarios asociados (instancia limpia)" if not users else f"{len(users)} usuario(s) detectado(s)",
            severity="warning",
        )
    except Exception:
        # Muchas instalaciones locales no exponen usersInfo — no es error
        add("mongo_no_auth", "MongoDB sin cuentas externas", True, "N/A (instancia local sin auth)")

    # 6) Repo GitHub configurado con el default
    cfg = await _get_github_config()
    repo_url = cfg.get("repo_url", "")
    is_default_repo = repo_url == SUGGESTED_GITHUB_REPO
    add("github_default_repo",
        "Repositorio GitHub por defecto",
        is_default_repo,
        f"OK: {repo_url}" if is_default_repo else (f"Otro repo: {repo_url}" if repo_url else "Sin repo configurado"),
        severity="warning",
        fixable=True,
    )

    # 7) Token de GitHub
    add("github_token",
        "Cuenta de GitHub vinculada",
        bool(cfg.get("token")),
        f"Conectado como @{cfg.get('username', '?')}" if cfg.get("token") else "Sin token — usa 'Conectar con GitHub'",
        severity="warning",
    )

    # 8) Variables de entorno críticas
    env_ok = bool(os.environ.get("MONGO_URL")) and bool(os.environ.get("DB_NAME"))
    add("env_vars", "Variables de entorno (backend/.env)", env_ok,
        "MONGO_URL y DB_NAME configuradas" if env_ok else "Faltan variables críticas",
        fixable=True)

    # 9) Seguridad — contraseña ZIP no está vacía
    sec_cfg = (await db.app_settings.find_one({}, {"security_config": 1}) or {}).get("security_config") or {}
    zip_pwd = sec_cfg.get("zip_password") or DEFAULT_ZIP_PASSWORD
    add("zip_password",
        "Contraseña ZIP configurada",
        len(zip_pwd) >= 3,
        f"Longitud: {len(zip_pwd)} chars ({'DEFAULT' if zip_pwd == DEFAULT_ZIP_PASSWORD else 'personalizada'})",
        severity="warning",
        fixable=True,
    )

    # 10) Directorio de backups escribible
    try:
        BACKUP_DIR.mkdir(parents=True, exist_ok=True)
        test_file = BACKUP_DIR / ".diagnostic_test"
        test_file.write_text("ok")
        test_file.unlink()
        add("backup_writable", "Directorio de backups escribible", True, str(BACKUP_DIR))
    except Exception as e:
        add("backup_writable", "Directorio de backups escribible", False, str(e)[:120], fixable=True)

    # 11) Frontend build (para el paquete desktop)
    build_dir = ROOT_DIR.parent / "frontend" / "build"
    has_build = build_dir.exists() and (build_dir / "index.html").exists()
    add("frontend_build",
        "Build de frontend (para app de escritorio)",
        has_build,
        "Build presente" if has_build else "Sin build — se puede compilar automáticamente",
        severity="info",
        fixable=True,
    )

    # Score global
    total = len(checks)
    ok_count = sum(1 for c in checks if c["ok"])
    errors = sum(1 for c in checks if not c["ok"] and c["severity"] == "error")
    warnings = sum(1 for c in checks if not c["ok"] and c["severity"] == "warning")

    return {
        "checks": checks,
        "summary": {
            "total": total,
            "ok": ok_count,
            "errors": errors,
            "warnings": warnings,
            "score": round((ok_count / total) * 100) if total else 0,
        },
        "generated_at": datetime.now(timezone.utc).isoformat(),
    }


@api_router.post("/diagnostic/fix")
async def diagnostic_fix(payload: dict = Body(...)):
    """Aplica una corrección concreta al item indicado por id."""
    global client, db
    check_id = (payload.get("id") or "").strip()
    if not check_id:
        raise HTTPException(status_code=400, detail="Falta 'id' del chequeo a corregir")

    fixed = False
    detail = ""

    if check_id == "python_deps":
        # Reinstala los que faltan
        result = subprocess.run(
            ["pip", "install", "--quiet", "-r", str(ROOT_DIR / "requirements.txt")],
            capture_output=True, text=True, timeout=180,
        )
        fixed = result.returncode == 0
        detail = (result.stderr or result.stdout or "")[-400:]
    elif check_id == "node_deps":
        result = subprocess.run(
            ["yarn", "install"], cwd=str(ROOT_DIR.parent / "frontend"),
            capture_output=True, text=True, timeout=300,
        )
        fixed = result.returncode == 0
        detail = (result.stderr or result.stdout or "")[-400:]
    elif check_id == "github_default_repo":
        await db.app_settings.update_one(
            {},
            {"$set": {
                "github_config.repo_url": SUGGESTED_GITHUB_REPO,
                "github_config.branch": DEFAULT_GITHUB_BRANCH,
            }},
            upsert=True,
        )
        fixed = True
        detail = f"Repositorio restaurado a {SUGGESTED_GITHUB_REPO}"
    elif check_id == "backup_writable":
        try:
            BACKUP_DIR.mkdir(parents=True, exist_ok=True)
            fixed = True
            detail = f"Directorio creado: {BACKUP_DIR}"
        except Exception as e:
            detail = str(e)
    elif check_id == "mongo_conn":
        # Reintenta la conexión con la URL original
        try:
            original_url = os.environ.get("MONGO_URL", "")
            if original_url:
                new_client = AsyncIOMotorClient(original_url)
                await new_client[DB_NAME].command("ping")
                client = new_client
                db = new_client[DB_NAME]
                fixed = True
                detail = "Reconectado a MongoDB usando la URL original"
            else:
                detail = "No hay MONGO_URL en el entorno"
        except Exception as e:
            detail = f"Reconexión falló: {e}"
    elif check_id == "env_vars":
        # Recrea el .env con valores por defecto seguros (backup del anterior si existe)
        try:
            env_path = ROOT_DIR / ".env"
            if env_path.exists():
                bak = ROOT_DIR / ".env.bak"
                bak.write_text(env_path.read_text())
            env_path.write_text(
                "MONGO_URL=mongodb://localhost:27017\n"
                "DB_NAME=reserva_eventos\n"
                "CORS_ORIGINS=*\n"
            )
            fixed = True
            detail = f".env restaurado a valores por defecto (backup: {env_path}.bak)"
        except Exception as e:
            detail = str(e)
    elif check_id == "zip_password":
        await db.app_settings.update_one(
            {},
            {"$set": {"security_config.zip_password": DEFAULT_ZIP_PASSWORD}},
            upsert=True,
        )
        fixed = True
        detail = f"Contraseña ZIP restaurada a {DEFAULT_ZIP_PASSWORD}"
    elif check_id == "frontend_build":
        # Compila el frontend
        result = subprocess.run(
            ["yarn", "build"], cwd=str(ROOT_DIR.parent / "frontend"),
            capture_output=True, text=True, timeout=600,
        )
        fixed = result.returncode == 0
        detail = ("Build compilado correctamente" if fixed else (result.stderr or result.stdout))[-400:]
    else:
        raise HTTPException(status_code=400, detail=f"El chequeo '{check_id}' no es corregible automáticamente")

    return {"success": fixed, "id": check_id, "detail": detail}


@api_router.post("/diagnostic/fix-all")
async def diagnostic_fix_all():
    """Ejecuta el diagnóstico y aplica correcciones automáticas a todo lo que sea fixable.

    Devuelve un resumen con lo que se corrigió y lo que no.
    """
    diag = await run_diagnostic()
    results = []
    fixed_count = 0
    failed_count = 0

    for check in diag["checks"]:
        if check["ok"] or not check.get("fixable"):
            continue
        try:
            res = await diagnostic_fix({"id": check["id"]})
            results.append({
                "id": check["id"],
                "label": check["label"],
                "success": res.get("success", False),
                "detail": res.get("detail", "")[:200],
            })
            if res.get("success"):
                fixed_count += 1
            else:
                failed_count += 1
        except HTTPException as e:
            results.append({
                "id": check["id"],
                "label": check["label"],
                "success": False,
                "detail": str(e.detail)[:200],
            })
            failed_count += 1
        except Exception as e:
            results.append({
                "id": check["id"],
                "label": check["label"],
                "success": False,
                "detail": str(e)[:200],
            })
            failed_count += 1

    # Ejecutar diagnóstico final
    final_diag = await run_diagnostic()

    return {
        "success": failed_count == 0,
        "fixed": fixed_count,
        "failed": failed_count,
        "results": results,
        "final_score": final_diag["summary"]["score"],
        "final_errors": final_diag["summary"]["errors"],
        "final_warnings": final_diag["summary"]["warnings"],
    }


def _get_local_commit_sha():
    try:
        result = subprocess.run(
            ["git", "-C", str(REPO_ROOT), "rev-parse", "HEAD"],
            capture_output=True, text=True, timeout=10
        )
        if result.returncode == 0:
            return result.stdout.strip()
    except Exception as e:
        logger.warning(f"Error obteniendo SHA local: {e}")
    return ""


def _is_frozen_bundle() -> bool:
    """True cuando la app corre desde un ejecutable empaquetado (PyInstaller
    onefile/onedir) — en este caso las actualizaciones vía `git` no aplican
    porque no existe un repositorio funcional junto al binario congelado."""
    import sys as _sys
    return bool(getattr(_sys, "frozen", False) or hasattr(_sys, "_MEIPASS"))


def _bundle_kind() -> str:
    """Distingue 'portable' vs 'installer' cuando corremos como .exe.
    Portable => el ejecutable está dentro de una carpeta temporal _MEIPASS
    junto a un onefile; el instalador de Inno Setup deja el .exe en
    'Program Files' o similar y define un valor típico en la ruta."""
    import sys as _sys
    exe = Path(getattr(_sys, "executable", "") or "").resolve()
    exe_str = str(exe).lower()
    if "program files" in exe_str or "\\programdata\\" in exe_str:
        return "installer"
    return "portable"


@api_router.get("/github/check-updates")
async def check_github_updates():
    cfg = await _get_github_config()
    repo_url = cfg.get("repo_url", "") or f"https://github.com/{_github_exe_repo()}"

    owner, repo = _parse_github_url(repo_url)
    if not owner:
        raise HTTPException(status_code=400, detail="URL de GitHub inválida")

    branch = cfg.get("branch") or "main"
    token = cfg.get("token", "")
    headers = {"Accept": "application/vnd.github+json"}
    if token:
        headers["Authorization"] = f"Bearer {token}"

    local_sha = _get_local_commit_sha()

    # Obtener commits del remoto
    api_url = f"https://api.github.com/repos/{owner}/{repo}/commits"
    async with httpx.AsyncClient(timeout=15) as http:
        try:
            r = await http.get(api_url, headers=headers, params={"sha": branch, "per_page": 20})
        except Exception as e:
            raise HTTPException(status_code=502, detail=f"Error conectando a GitHub: {e}")

    if r.status_code == 404:
        raise HTTPException(status_code=404, detail="Repositorio no encontrado (¿es privado y falta token?)")
    if r.status_code == 401:
        raise HTTPException(status_code=401, detail="Token de GitHub inválido o insuficiente")
    if r.status_code != 200:
        raise HTTPException(status_code=502, detail=f"GitHub API respondió {r.status_code}: {r.text[:200]}")

    commits_data = r.json()
    if not commits_data:
        return {"has_updates": False, "commits": [], "local_sha": local_sha, "remote_sha": "", "branch": branch}

    remote_sha = commits_data[0]["sha"]

    # Determinar si hay updates: local_sha debe encontrarse en la lista de commits del remoto
    # Si local está en la lista, contamos cuántos hay ANTES (más nuevos) que él
    local_found_at = -1
    for i, c in enumerate(commits_data):
        if c["sha"] == local_sha:
            local_found_at = i
            break

    if local_found_at == 0:
        # Local == remote latest → nada nuevo
        has_updates = False
        new_commits = []
    elif local_found_at > 0:
        # Hay `local_found_at` commits nuevos en el remoto
        has_updates = True
        new_commits_data = commits_data[:local_found_at]
    else:
        # Local NO se encuentra en top-20 commits del remoto:
        # Casos posibles:
        #   a) Local está muy desactualizado (remoto tiene >20 commits nuevos) → has_updates=True
        #   b) Primer arranque, local_sha vacío → has_updates=True
        #   c) Local es un SHA de otra rama/divergente → no podemos determinarlo con la API,
        #      pero es más útil marcarlo como "hay updates" para que el usuario resincronice.
        # En todos los casos, si remote_sha existe y difiere de local_sha, marcamos update.
        has_updates = bool(remote_sha) and (remote_sha != local_sha)
        new_commits_data = commits_data if has_updates else []

    new_commits = []
    if has_updates:
        for c in (new_commits_data if 'new_commits_data' in locals() else commits_data[:5]):
            commit_info = c.get("commit", {})
            author = commit_info.get("author", {})
            new_commits.append({
                "sha": c["sha"][:7],
                "full_sha": c["sha"],
                "message": (commit_info.get("message") or "").split("\n")[0][:200],
                "author": author.get("name", "desconocido"),
                "date": author.get("date", ""),
                "url": c.get("html_url", "")
            })

    # Guardar último chequeo
    await db.app_settings.update_one(
        {},
        {"$set": {
            "github_config.last_check_at": datetime.now(timezone.utc).isoformat(),
            "github_config.last_remote_sha": remote_sha,
        }},
        upsert=True
    )

    # Leer versión local (archivo version.txt) y versión remota (via GitHub raw)
    local_version = ""
    try:
        for candidate in (ROOT_DIR / "version.txt", ROOT_DIR.parent / "version.txt"):
            if candidate.exists():
                local_version = candidate.read_text(encoding="utf-8").strip()
                break
    except Exception:
        local_version = ""

    remote_version = ""
    try:
        raw_url = f"https://raw.githubusercontent.com/{owner}/{repo}/{branch}/version.txt"
        async with httpx.AsyncClient(timeout=8) as http:
            rv = await http.get(raw_url, headers={"Authorization": f"Bearer {token}"} if token else {})
            if rv.status_code == 200:
                remote_version = (rv.text or "").strip()
    except Exception:
        remote_version = ""

    # ── FUENTE DE VERDAD: la VERSIÓN (semver X.Y.Z), no el SHA ──────────────
    # El SHA local puede divergir del remoto (rama distinta, historial reescrito,
    # o local_sha fuera de los últimos 20 commits) y provocar un falso positivo
    # "Nueva versión disponible: vX" cuando en realidad es la MISMA versión.
    # Si conocemos ambas versiones y son idénticas (normalizadas), NO hay update.
    ln = _normalize_semver(local_version)
    rn = _normalize_semver(remote_version)
    if ln and rn:
        # Fuente de verdad: la versión semántica. Solo hay update si el remoto
        # es ESTRICTAMENTE más nuevo — evita falsos positivos por SHA divergente
        # y evita ofrecer un "update" que en realidad sería un downgrade.
        has_updates = _is_newer_version(rn, ln)
        if not has_updates:
            new_commits = []

    return {
        "has_updates": has_updates,
        "local_sha": local_sha,
        "local_sha_short": local_sha[:7] if local_sha else "",
        "remote_sha": remote_sha,
        "remote_sha_short": remote_sha[:7],
        "local_version": local_version,
        "remote_version": remote_version,
        "branch": branch,
        "commits_ahead": len(new_commits),
        "commits": new_commits,
        "repo_url": repo_url,
    }


# ── Estado de progreso de descarga/instalación desktop (para barra en el modal)
import threading as _threading_mod
_update_progress: dict = {
    "active": False, "stage": "idle", "downloaded": 0, "total": 0,
    "percent": 0, "name": "", "error": None,
}
_update_progress_lock = _threading_mod.Lock()


def _reset_update_progress(name: str = "", total: int = 0):
    with _update_progress_lock:
        _update_progress.update({
            "active": True, "stage": "starting", "downloaded": 0,
            "total": int(total or 0), "percent": 0, "name": name, "error": None,
        })


def _set_update_progress(downloaded=None, total=None, stage=None, error=None):
    with _update_progress_lock:
        if downloaded is not None:
            _update_progress["downloaded"] = int(downloaded)
        if total is not None and total:
            _update_progress["total"] = int(total)
        if stage is not None:
            _update_progress["stage"] = stage
        if error is not None:
            _update_progress["error"] = error
            _update_progress["active"] = False
        if stage in ("done", "error"):
            _update_progress["active"] = stage != "error"
        t = _update_progress["total"]
        d = _update_progress["downloaded"]
        _update_progress["percent"] = int(d * 100 / t) if t > 0 else (100 if stage == "done" else 0)
        if stage == "done":
            _update_progress["percent"] = 100
            _update_progress["active"] = False


@api_router.get("/github/update-progress")
async def github_update_progress():
    """Estado actual de la descarga/instalación desktop para la barra del modal."""
    with _update_progress_lock:
        return dict(_update_progress)


@api_router.post("/github/apply-update")
async def apply_github_update(payload: dict = Body(default={})):
    cfg = await _get_github_config()
    repo_url = cfg.get("repo_url", "") or f"https://github.com/{_github_exe_repo()}"

    # ── Caso .exe (portable o instalador) ────────────────────────────
    # En Windows empaquetado con PyInstaller no existe .git ni supervisor,
    # así que `git fetch/reset` fallaría. En su lugar devolvemos las URLs
    # de los binarios más recientes publicados en GitHub Releases para que
    # la UI muestre el botón "Descargar nueva versión".
    if _is_frozen_bundle():
        import platform
        import time
        kind = _bundle_kind()
        portable = await _find_latest_exe_asset("portable")
        installer = await _find_latest_exe_asset("installer")
        asset = installer if kind == "installer" else portable
        if not asset:
            return {
                "success": False,
                "is_desktop": True,
                "status": "desktop_update",
                "mode": "desktop_bundle",
                "bundle_kind": kind,
                "message": (
                    "No hay una versión nueva publicada aún. Ejecuta el workflow "
                    "'Build Windows .exe' en GitHub Actions o publica un tag v* "
                    "para generar el instalador y el portable."
                ),
                "portable": portable,
                "installer": installer,
            }

        # ── Guard anti-bucle/downgrade: el binario publicado más reciente debe
        #    ser ESTRICTAMENTE más nuevo que el instalado. `version.txt` en main
        #    puede ir por delante del último exe publicado (aún compilándose en
        #    GitHub Actions); sin este guard, se reinstalaría la MISMA versión en
        #    bucle. Comparamos el tag del asset (vX.Y.Z) contra version.txt local.
        _local_v = _normalize_semver(await _read_local_version())
        _asset_v = _normalize_semver(asset.get("tag") or "")
        if _local_v and _asset_v and not _is_newer_version(_asset_v, _local_v):
            return {
                "success": False,
                "is_desktop": True,
                "status": "up_to_date",
                "mode": "desktop_bundle",
                "bundle_kind": kind,
                "message": (
                    f"Ya tienes instalado el binario más reciente (v{_local_v}). "
                    "La versión anunciada aún se está compilando/publicando en "
                    "GitHub Releases; vuelve a intentarlo en unos minutos."
                ),
                "installed_version": _local_v,
                "available_exe_version": _asset_v,
                "portable": portable,
                "installer": installer,
            }

        # ── Auto-actualización en Windows: descarga verificada (SHA256) EN
        #    SEGUNDO PLANO con progreso consultable + instalación automática.
        #    · installer → lanza Setup.exe /SILENT (Inno Setup: CloseApplications
        #      force + RestartApplications yes → cierra y relanza la app).
        #    · portable  → apply_update_windows(): batch DETACHED que espera a que
        #      el .exe se libere, lo reemplaza y relanza la app.
        #    La request retorna de inmediato con status "installing" para que el
        #    modal/página hagan polling a /github/update-progress (barra).
        if platform.system().lower() == "windows" and asset.get("sha256"):
            import threading

            _reset_update_progress(name=asset["name"], total=asset.get("size") or 0)

            def _download_and_install(_kind=kind, _asset=asset):
                try:
                    import sys as _sys
                    from updater import download_and_verify, apply_update_windows

                    def _cb(done, total):
                        _set_update_progress(downloaded=done, total=total, stage="downloading")

                    pkg_path = download_and_verify(
                        _asset["url"], _asset["sha256"], filename=_asset["name"],
                        timeout=600, progress_cb=_cb,
                    )
                    _set_update_progress(stage="installing")
                    time.sleep(0.5)
                    if _kind == "installer":
                        # /SILENT muestra progreso del instalador; /NORESTART evita
                        # reinicio del SO (el relanzado de la app lo hace Inno Setup).
                        subprocess.Popen([pkg_path, "/SILENT", "/NORESTART"], close_fds=True)
                    else:
                        # Portable: reemplazo en caliente vía batch detached.
                        app_name = os.path.basename(getattr(_sys, "executable", "") or "") or "CinemaProductions.exe"
                        apply_update_windows(pkg_path, app_name=app_name)
                    _set_update_progress(stage="done")
                    time.sleep(1.0)
                    os._exit(0)
                except Exception as e:
                    logger.warning(f"Auto-instalación ({_kind}) falló: {e}")
                    _set_update_progress(stage="error", error=str(e))

            threading.Thread(target=_download_and_install, daemon=True).start()
            return {
                "success": True,
                "is_desktop": True,
                "restarted": True,
                "status": "installing",
                "mode": "desktop_bundle",
                "bundle_kind": kind,
                "message": "Descarga verificada. Instalando la nueva versión y reiniciando…",
                "download_name": asset["name"],
                "tag": asset.get("tag"),
            }

        friendly = (
            "Estás usando la versión portable: no puede reemplazarse a sí misma "
            "en caliente. La descarga de la nueva versión comenzará ahora; cierra "
            "la app y sustituye el .exe (o instala el instalador para futuras "
            "actualizaciones automáticas)."
            if kind == "portable"
            else "Descarga e inicia el instalador. Se actualizará sobre la instalación actual y reiniciará la app."
        )
        return {
            "success": False,
            "is_desktop": True,
            "download_now": True,
            "status": "desktop_update",
            "mode": "desktop_bundle",
            "bundle_kind": kind,
            "message": friendly,
            "download_url": asset["url"],
            "download_name": asset["name"],
            "download_size_mb": round((asset.get("size") or 0) / (1024 * 1024), 1),
            "tag": asset.get("tag"),
            "portable": portable,
            "installer": installer,
        }

    branch = cfg.get("branch") or "main"
    force = bool(payload.get("force", True))  # por defecto reset --hard

    logs = []
    try:
        # 1. Fetch
        fetch = subprocess.run(
            ["git", "-C", str(REPO_ROOT), "fetch", "origin", branch],
            capture_output=True, text=True, timeout=60
        )
        logs.append(f"$ git fetch origin {branch}\n{fetch.stdout}{fetch.stderr}")
        if fetch.returncode != 0:
            raise HTTPException(status_code=500, detail=f"Error en git fetch: {fetch.stderr}")

        # 2. Reset o pull
        if force:
            reset = subprocess.run(
                ["git", "-C", str(REPO_ROOT), "reset", "--hard", f"origin/{branch}"],
                capture_output=True, text=True, timeout=60
            )
            logs.append(f"$ git reset --hard origin/{branch}\n{reset.stdout}{reset.stderr}")
            if reset.returncode != 0:
                raise HTTPException(status_code=500, detail=f"Error en git reset: {reset.stderr}")
        else:
            pull = subprocess.run(
                ["git", "-C", str(REPO_ROOT), "pull", "origin", branch],
                capture_output=True, text=True, timeout=60
            )
            logs.append(f"$ git pull origin {branch}\n{pull.stdout}{pull.stderr}")
            if pull.returncode != 0:
                raise HTTPException(status_code=500, detail=f"Error en git pull: {pull.stderr}")

        new_sha = _get_local_commit_sha()
        await db.app_settings.update_one(
            {},
            {"$set": {
                "github_config.last_commit_sha": new_sha,
                "github_config.last_update_at": datetime.now(timezone.utc).isoformat(),
            }},
            upsert=True
        )

        # 3. Reiniciar servicios en background (para reflejar cambios)
        async def _restart_later():
            await asyncio.sleep(2)
            try:
                subprocess.run(["sudo", "supervisorctl", "restart", "frontend"], timeout=30, capture_output=True)
                subprocess.run(["sudo", "supervisorctl", "restart", "backend"], timeout=30, capture_output=True)
            except Exception as e:
                logger.warning(f"Error reiniciando servicios: {e}")

        asyncio.create_task(_restart_later())

        return {
            "success": True,
            "new_sha": new_sha,
            "new_sha_short": new_sha[:7] if new_sha else "",
            "logs": "\n".join(logs),
            "message": "Actualización aplicada. Los servicios se reiniciarán en 2 segundos.",
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error aplicando actualización: {e}")


# ─── AI Context (contexto para la próxima IA) ─────────────────
@api_router.get("/ai-context")
async def get_ai_context():
    doc = await db.app_settings.find_one({}, {"ai_context": 1}) or {}
    ctx = doc.get("ai_context") or {}
    if not ctx.get("content"):
        # Inicializar con el contexto por defecto
        await db.app_settings.update_one(
            {},
            {"$set": {"ai_context": {
                "content": DEFAULT_AI_CONTEXT,
                "updated_at": datetime.now(timezone.utc).isoformat(),
            }}},
            upsert=True
        )
        return {"content": DEFAULT_AI_CONTEXT, "updated_at": datetime.now(timezone.utc).isoformat(), "is_default": True}
    return {
        "content": ctx.get("content", ""),
        "updated_at": ctx.get("updated_at", ""),
        "is_default": False,
    }


@api_router.post("/ai-context")
async def save_ai_context(payload: dict = Body(...)):
    content = payload.get("content", "")
    if not isinstance(content, str):
        raise HTTPException(status_code=400, detail="content debe ser string")
    await db.app_settings.update_one(
        {},
        {"$set": {"ai_context": {
            "content": content,
            "updated_at": datetime.now(timezone.utc).isoformat(),
        }}},
        upsert=True
    )
    return {"success": True, "updated_at": datetime.now(timezone.utc).isoformat()}


@api_router.post("/ai-context/reset")
async def reset_ai_context():
    await db.app_settings.update_one(
        {},
        {"$set": {"ai_context": {
            "content": DEFAULT_AI_CONTEXT,
            "updated_at": datetime.now(timezone.utc).isoformat(),
        }}},
        upsert=True
    )
    return {"success": True, "content": DEFAULT_AI_CONTEXT}

# ═══════════════════════════════════════════════════════════════════════════
# 15.  Error Reporting  ·  logs visibles + reporte automático a GitHub Issues
# ───────────────────────────────────────────────────────────────────────────
#  Objetivo: que NADA falle en silencio. Todo error (backend, frontend, exe de
#  escritorio o actualización) queda:
#    1. Registrado en el log del servidor (logger.error).
#    2. Guardado en la colección `error_reports` (dedup por fingerprint).
#    3. Reportado como GitHub Issue en el repo del usuario (si hay token y el
#       auto-reporte está activo) para poder repararlo.
#    4. Visible para el usuario vía /api/errors (panel "Incidencias").
# ═══════════════════════════════════════════════════════════════════════════
ERROR_LOG_FILE = ROOT_DIR / "error_reports.log"


def _error_fingerprint(source: str, message: str, stack: str) -> str:
    """Huella estable para deduplicar el mismo error (ignora rutas/números)."""
    first_line = ""
    if stack:
        for ln in str(stack).splitlines():
            ln = ln.strip()
            if ln:
                first_line = ln
                break
    base = f"{source}|{(message or '')[:200]}|{first_line[:200]}"
    return hashlib.sha256(base.encode("utf-8", "replace")).hexdigest()[:16]


async def _get_error_settings() -> dict:
    doc = await db.app_settings.find_one({}, {"error_reporting": 1}) or {}
    cfg = doc.get("error_reporting") or {}
    return {
        "auto_report_github": cfg.get("auto_report_github", True),
        "notify_user": cfg.get("notify_user", True),
    }


async def _create_github_issue(title: str, body: str, labels: list[str]) -> Optional[str]:
    """Crea un GitHub Issue en el repo configurado. Devuelve la URL o None."""
    token, repo_url, _branch = await _resolve_github_creds()
    if not token:
        return None
    owner, repo = _parse_github_url(repo_url)
    if not owner or not repo:
        return None
    url = f"https://api.github.com/repos/{owner}/{repo}/issues"
    headers = {
        "Authorization": f"Bearer {token}",
        "Accept": "application/vnd.github+json",
        "X-GitHub-Api-Version": "2022-11-28",
        "User-Agent": "cinema-productions-app",
    }
    try:
        async with httpx.AsyncClient(timeout=20) as c:
            r = await c.post(url, headers=headers, json={"title": title[:250], "body": body[:60000], "labels": labels})
        if r.status_code in (200, 201):
            return r.json().get("html_url")
        logger.warning(f"[errors] GitHub issue no creado: HTTP {r.status_code} {r.text[:200]}")
    except Exception as e:
        logger.warning(f"[errors] GitHub issue error: {e}")
    return None


async def _persist_error_report(payload: dict, request_meta: dict | None = None) -> dict:
    """Guarda/deduplica un error y (opcional) crea GitHub Issue. Core reutilizable."""
    now = datetime.now(timezone.utc)
    source = (payload.get("source") or "unknown").strip()[:60]
    message = (payload.get("message") or "Error desconocido").strip()[:2000]
    stack = (payload.get("stack") or "")[:8000]
    level = (payload.get("level") or "error").strip()[:20]
    context = payload.get("context") or {}
    version = str(payload.get("version") or "")[:40]
    platform_info = str(payload.get("platform") or "")[:120]
    fingerprint = _error_fingerprint(source, message, stack)

    # Log en archivo + consola (siempre, aunque falle GitHub / Mongo)
    logger.error(f"[error_report] source={source} fp={fingerprint} :: {message}")
    try:
        with open(ERROR_LOG_FILE, "a", encoding="utf-8") as f:
            f.write(f"{now.isoformat()} [{level}] {source} ({fingerprint}) {message}\n")
            if stack:
                f.write(stack[:4000] + "\n")
    except Exception:
        pass

    settings = await _get_error_settings()
    existing = await db.error_reports.find_one({"fingerprint": fingerprint, "resolved": {"$ne": True}})

    if existing:
        await db.error_reports.update_one(
            {"_id": existing["_id"]},
            {"$set": {"last_seen": now.isoformat(), "version": version or existing.get("version", "")},
             "$inc": {"count": 1}},
        )
        doc = await db.error_reports.find_one({"_id": existing["_id"]})
        return {"id": str(doc["_id"]), "fingerprint": fingerprint, "count": doc.get("count", 1),
                "github_issue_url": doc.get("github_issue_url"), "deduped": True,
                "notify_user": settings["notify_user"]}

    record = {
        "source": source, "message": message, "stack": stack, "level": level,
        "context": context, "version": version, "platform": platform_info,
        "fingerprint": fingerprint, "count": 1, "resolved": False,
        "first_seen": now.isoformat(), "last_seen": now.isoformat(),
        "created_at": now.isoformat(), "github_issue_url": None,
        "request": request_meta or {},
    }
    res = await db.error_reports.insert_one(record)
    rid = str(res.inserted_id)

    issue_url = None
    if settings["auto_report_github"]:
        title = f"[{source}] {message[:120]}"
        body_lines = [
            f"**Reporte automático de error** · `{fingerprint}`",
            "",
            f"- **Origen:** {source}",
            f"- **Nivel:** {level}",
            f"- **Versión app:** {version or 'n/a'}",
            f"- **Plataforma:** {platform_info or 'n/a'}",
            f"- **Fecha (UTC):** {now.isoformat()}",
            "",
            "### Mensaje", "```", message, "```",
        ]
        if stack:
            body_lines += ["### Stack trace", "```", stack[:6000], "```"]
        if context:
            try:
                body_lines += ["### Contexto", "```json", json.dumps(context, indent=2, ensure_ascii=False, default=_json_safe)[:3000], "```"]
            except Exception:
                pass
        body_lines += ["", "_Generado automáticamente por CinemaProductions para poder reparar el error._"]
        issue_url = await _create_github_issue(title, "\n".join(body_lines), ["bug", "auto-report", f"src:{source}"])
        if issue_url:
            await db.error_reports.update_one({"_id": res.inserted_id}, {"$set": {"github_issue_url": issue_url}})

    return {"id": rid, "fingerprint": fingerprint, "count": 1, "github_issue_url": issue_url,
            "deduped": False, "notify_user": settings["notify_user"]}


@api_router.post("/errors/report")
async def report_error(payload: dict = Body(...), request: Request = None):
    """Recibe un error del frontend / exe / updater y lo procesa."""
    meta = {}
    try:
        if request is not None:
            meta = {"ua": request.headers.get("user-agent", "")[:200],
                    "ip": (request.client.host if request.client else "")}
    except Exception:
        pass
    result = await _persist_error_report(payload, meta)
    return {"success": True, **result}


@api_router.get("/updates/last-result")
async def updates_last_result(clear: bool = False):
    """En el servidor de preview no hay swap de .exe; siempre 'sin fallo'.
    (El exe de escritorio implementa la versión real que lee el flag de fallo.)"""
    return {"failed": False}


@api_router.get("/errors")
async def list_errors(limit: int = 50, include_resolved: bool = False):
    """Lista de incidencias para que el usuario VEA qué falló (no fallos silenciosos)."""
    q = {} if include_resolved else {"resolved": {"$ne": True}}
    cur = db.error_reports.find(q).sort("last_seen", -1).limit(min(max(limit, 1), 200))
    items = [doc_to_dict(d) async for d in cur]
    open_count = await db.error_reports.count_documents({"resolved": {"$ne": True}})
    return {"items": items, "open_count": open_count}


@api_router.get("/errors/settings")
async def get_error_settings():
    return await _get_error_settings()


@api_router.put("/errors/settings")
async def update_error_settings(payload: dict = Body(...)):
    update = {}
    if "auto_report_github" in payload:
        update["error_reporting.auto_report_github"] = bool(payload["auto_report_github"])
    if "notify_user" in payload:
        update["error_reporting.notify_user"] = bool(payload["notify_user"])
    if update:
        await db.app_settings.update_one({}, {"$set": update}, upsert=True)
    return await _get_error_settings()


@api_router.post("/errors/{error_id}/resolve")
async def resolve_error(error_id: str):
    try:
        oid = ObjectId(error_id)
    except Exception:
        raise HTTPException(status_code=400, detail="ID inválido")
    r = await db.error_reports.update_one({"_id": oid}, {"$set": {"resolved": True, "resolved_at": datetime.now(timezone.utc).isoformat()}})
    if r.matched_count == 0:
        raise HTTPException(status_code=404, detail="Incidencia no encontrada")
    return {"success": True}


@api_router.delete("/errors/{error_id}")
async def delete_error(error_id: str):
    try:
        oid = ObjectId(error_id)
    except Exception:
        raise HTTPException(status_code=400, detail="ID inválido")
    await db.error_reports.delete_one({"_id": oid})
    return {"success": True}


@api_router.post("/errors/clear-resolved")
async def clear_resolved_errors():
    r = await db.error_reports.delete_many({"resolved": True})
    return {"success": True, "deleted": r.deleted_count}


# ── Global exception handler: ningún 500 pasa desapercibido ──────────────
from fastapi.exceptions import RequestValidationError
from starlette.exceptions import HTTPException as StarletteHTTPException


@app.exception_handler(Exception)
async def _global_exception_handler(request: Request, exc: Exception):
    """Captura CUALQUIER excepción no controlada: la registra, la reporta y
    devuelve un mensaje CLARO al usuario (nunca un fallo silencioso)."""
    import traceback as _tb
    stack = "".join(_tb.format_exception(type(exc), exc, exc.__traceback__))
    try:
        result = await _persist_error_report({
            "source": "backend",
            "message": f"{type(exc).__name__}: {exc}",
            "stack": stack,
            "level": "error",
            "context": {"path": str(request.url.path), "method": request.method},
        }, {"ua": request.headers.get("user-agent", "")[:200]})
        ref = result.get("fingerprint")
        issue = result.get("github_issue_url")
    except Exception:
        ref, issue = None, None
    return JSONResponse(
        status_code=500,
        content={
            "error": True,
            "detail": f"Ocurrió un error en el servidor: {type(exc).__name__}",
            "message": str(exc)[:500],
            "reference": ref,
            "github_issue_url": issue,
            "reported": bool(ref),
        },
    )




app.include_router(api_router)

# ── Subscription + Emergent Google Auth + PayPal ──────────────────
# Subscription / auth / users module removed — app funciona sin cuentas.

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)
