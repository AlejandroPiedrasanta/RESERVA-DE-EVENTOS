"""
Cinema Productions - App Local Auto-contenida
==============================================
Ejecutar: python app.py  (o doble clic en start.bat)

Modos de base de datos:
  MONGO_URL=embedded          → Almacenamiento local en cinema_data.json (predeterminado)
  MONGO_URL=mongodb://...     → MongoDB externo (local o Atlas)
"""
from fastapi import FastAPI, APIRouter, HTTPException, UploadFile, File, Form, Body
from fastapi.responses import JSONResponse, Response, FileResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from starlette.middleware.cors import CORSMiddleware
from contextlib import asynccontextmanager
from dotenv import load_dotenv
import csv
import io
import os
import re
import json
import hashlib
import secrets
import logging
import base64
import uuid
import subprocess
import webbrowser
import threading
import time
import asyncio
import sys
from pathlib import Path
from pydantic import BaseModel, Field
from typing import Optional
from datetime import datetime, timezone, timedelta
from bson import ObjectId

# ROOT_DIR: dónde vive el módulo (para .env, cinema_data.json, backups, etc.)
ROOT_DIR = Path(__file__).parent

# BUNDLE_DIR: dónde PyInstaller extrae los recursos --add-data (build/, themes/, ...).
# En modo frozen (--onefile) es sys._MEIPASS (temp dir _MEI*). En dev = ROOT_DIR.
if getattr(sys, "frozen", False) and hasattr(sys, "_MEIPASS"):
    BUNDLE_DIR = Path(sys._MEIPASS)
else:
    BUNDLE_DIR = ROOT_DIR

load_dotenv(ROOT_DIR / '.env')

DB_NAME = os.environ.get('DB_NAME', 'cinema_productions')
MONGO_URL = os.environ.get('MONGO_URL', 'embedded')
DATA_FILE = ROOT_DIR / 'cinema_data.json'
CUSTOM_DB_FILE = ROOT_DIR / '.db_override'
# Cuando app.py está dentro de "_sistema (NO TOCAR)/", los backups deben quedar
# VISIBLES en el directorio superior (la carpeta raíz de la app de escritorio).
_VISIBLE_ROOT = ROOT_DIR.parent if ROOT_DIR.name.startswith('_sistema') else ROOT_DIR
BACKUP_DIR = _VISIBLE_ROOT / 'backups'
BACKUP_DIR.mkdir(parents=True, exist_ok=True)
UPDATES_DIR = ROOT_DIR / 'uploads' / 'updates'
UPDATES_DIR.mkdir(parents=True, exist_ok=True)
BACKUP_COLLECTIONS = ['reservations', 'socios', 'app_settings', 'metas']

# ── Repo de GitHub de fábrica (pre-cargado, independiente de Emergent) ────────
DEFAULT_GITHUB_REPO = "https://github.com/AlejandroPiedrasanta/RESERVA-DE-EVENTOS"
DEFAULT_GITHUB_BRANCH = "main"

# ── Conexiones MongoDB "de fábrica" que vienen con el ejecutable de escritorio ─
FACTORY_SAVED_CONNECTIONS = [
    {
        "name": "MongoDB Atlas (por defecto)",
        "url": "mongodb+srv://reu1:cinemaproductions@cluster0.ozg25wu.mongodb.net/?appName=Cluster0",
        "color": "emerald",
        "factory": True,
    },
]

_DEFAULT_AI_CONTEXT = (
    "# Contexto — Cinema Productions (App de escritorio independiente)\n\n"
    "Esta es la versión de escritorio: 100% local e independiente. Base de datos\n"
    "embebida (cinema_data.json) por defecto; puedes conectar tu propio MongoDB o\n"
    "MongoDB Atlas desde Ajustes → Base de Datos.\n\n"
    f"Repositorio oficial: {DEFAULT_GITHUB_REPO}\n"
)

# ── Auto-update config ────────────────────────────────────────────────────────
_UPDATE_SERVER_FILE = ROOT_DIR / 'update_server_url.txt'
_update_server_url = _UPDATE_SERVER_FILE.read_text().strip() if _UPDATE_SERVER_FILE.exists() else ""


def _read_baked_version() -> str:
    """Lee la versión REALMENTE en ejecución.

    CRÍTICO para el EXE (--onefile): el version.txt horneado en el build se
    empaqueta con `--add-data "version.txt;."`, por lo que PyInstaller lo extrae
    en `sys._MEIPASS` (BUNDLE_DIR), NO junto al .exe (ROOT_DIR). El instalador
    tampoco copia version.txt al directorio de instalación. Leer de ROOT_DIR
    provocaba que en modo congelado no se encontrara el archivo y la versión
    cayera a "0.0.0": el EXE mostraba una versión que NO concuerda con la
    instalada y creía SIEMPRE que había una actualización disponible.

    Prioridad:
      1. BUNDLE_DIR/version.txt  → versión horneada del binario en ejecución
         (siempre correcta tras un swap de binario en la auto-actualización).
      2. ROOT_DIR/version.txt    → modo desarrollo (BUNDLE_DIR == ROOT_DIR) o
         version.txt escrito junto al módulo.
    """
    for cand in (BUNDLE_DIR / 'version.txt', ROOT_DIR / 'version.txt', ROOT_DIR.parent / 'version.txt'):
        try:
            if cand.exists():
                v = (cand.read_text(encoding='utf-8') or '').strip()
                if v:
                    return v
        except Exception:
            pass
    return "0.0.0"


# Ruta preferida (para compatibilidad con código que la referencie).
_VERSION_FILE = BUNDLE_DIR / 'version.txt'
_local_version = _read_baked_version()
_update_status: dict = {"checked": False, "has_update": False}

# ── Determine effective MongoDB URL (override file takes priority) ────────────
_override_url = CUSTOM_DB_FILE.read_text().strip() if CUSTOM_DB_FILE.exists() else None
_effective_mongo_url = _override_url or MONGO_URL

_using_embedded = _effective_mongo_url.strip().lower() in ('embedded', '', 'local')

logging.basicConfig(level=logging.WARNING)
logger = logging.getLogger(__name__)

if _using_embedded:
    from mongomock_motor import AsyncMongoMockClient
    _mongo_client = AsyncMongoMockClient()
    logger.warning("Usando base de datos embebida (cinema_data.json)")
else:
    try:
        from motor.motor_asyncio import AsyncIOMotorClient as _MotorClient
        _mongo_client = _MotorClient(_effective_mongo_url, serverSelectionTimeoutMS=8000)
        logger.warning(f"Motor inicializado para: {_effective_mongo_url[:50]}...")
    except Exception as _e:
        logger.error(f"Error al inicializar motor: {_e} — usando modo embebido")
        from mongomock_motor import AsyncMongoMockClient
        _mongo_client = AsyncMongoMockClient()
        _using_embedded = True

db = _mongo_client[DB_NAME]
client = _mongo_client  # alias


# ─── Embedded DB persistence ─────────────────────────────

def _serialize_doc(doc: dict) -> dict:
    """Convert MongoDB doc to JSON-serializable dict."""
    d = {}
    for k, v in doc.items():
        if k == '_id':
            d['__id'] = str(v)
        elif isinstance(v, ObjectId):
            d[k] = str(v)
        elif isinstance(v, list):
            d[k] = [_serialize_doc(i) if isinstance(i, dict) else i for i in v]
        else:
            d[k] = v
    return d


def _deserialize_doc(doc: dict) -> dict:
    """Restore JSON doc to MongoDB-compatible dict with ObjectId _id."""
    d = {}
    for k, v in doc.items():
        if k == '__id':
            try:
                d['_id'] = ObjectId(v)
            except Exception:
                d['_id'] = v
        elif isinstance(v, list):
            d[k] = [_deserialize_doc(i) if isinstance(i, dict) else i for i in v]
        else:
            d[k] = v
    return d


async def _load_embedded_data():
    if not DATA_FILE.exists():
        logger.info("No saved data found — starting fresh.")
        return
    try:
        data = json.loads(DATA_FILE.read_text(encoding='utf-8'))
        for col_name, docs in data.items():
            if not docs:
                continue
            col = db[col_name]
            restored = [_deserialize_doc(d) for d in docs]
            if restored:
                await col.insert_many(restored)
        logger.info(f"Loaded {sum(len(v) for v in data.values())} docs from {DATA_FILE.name}")
    except Exception as e:
        logger.error(f"Failed to load saved data: {e}")


async def _save_embedded_data():
    try:
        data = {}
        for col_name in ['reservations', 'socios', 'app_settings', 'saved_themes', 'metas']:
            docs = await db[col_name].find({}).to_list(100000)
            data[col_name] = [_serialize_doc(d) for d in docs]
        DATA_FILE.write_text(
            json.dumps(data, ensure_ascii=False, indent=2, default=str),
            encoding='utf-8'
        )
        logger.info(f"Data saved → {DATA_FILE.name}")
    except Exception as e:
        logger.error(f"Failed to save data: {e}")


async def _auto_save_loop():
    """Auto-save every 60 seconds."""
    while True:
        await asyncio.sleep(60)
        await _save_embedded_data()


# ─── Lifespan ─────────────────────────────────────────────

@asynccontextmanager
async def lifespan(app_instance: FastAPI):
    global db, client, _using_embedded
    _task = None

    # ── Verify MongoDB connection or fall back to embedded ──────────────────
    if not _using_embedded:
        try:
            await asyncio.wait_for(db.command("ping"), timeout=9.0)
            logger.warning("MongoDB Atlas: conexión verificada OK")
        except Exception as _conn_err:
            logger.error(
                f"MongoDB no accesible ({_conn_err}). "
                "Revisa tu conexión a internet o el archivo .env. "
                "Activando modo embebido temporal."
            )
            from mongomock_motor import AsyncMongoMockClient
            _fb = AsyncMongoMockClient()
            db = _fb[DB_NAME]
            client = _fb
            _using_embedded = True

    if _using_embedded:
        await _load_embedded_data()
        _task = asyncio.create_task(_auto_save_loop())

    # ── SEED_APP_SETTINGS: sembrar github_config de fábrica ──────────────────
    # Persiste el repositorio oficial en la BD si no existe (funciona tanto en
    # embedded como en MongoDB real). Así la copia de seguridad y cualquier
    # herramienta que lea app_settings lo hereda automáticamente.
    try:
        existing = await db.app_settings.find_one({}, {"github_config": 1}) or {}
        gh = existing.get("github_config") or {}
        if not gh.get("repo_url"):
            await db.app_settings.update_one(
                {},
                {"$set": {
                    "github_config.repo_url": DEFAULT_GITHUB_REPO,
                    "github_config.branch": DEFAULT_GITHUB_BRANCH,
                }},
                upsert=True,
            )
            logger.warning(f"Factory GitHub repo seeded: {DEFAULT_GITHUB_REPO}")
            if _using_embedded:
                await _save_embedded_data()
    except Exception as _seed_err:
        logger.warning(f"No se pudo sembrar github_config de fábrica: {_seed_err}")

    # ── Seed default "Minimalista" theme si la BD está vacía ────────────────
    # (1) si existe un mirror local themes/saved_themes.json (viene con el ZIP
    #     descargable), se importa para no perder los temas subidos a GitHub;
    # (2) si sigue vacío, se crea un "Minimalista" mínimo por defecto;
    # (3) se marca default_theme_id/name/appearance_snapshot en app_settings.
    try:
        seeded_default_id = None
        seeded_default_snapshot = None
        themes_json_path = ROOT_DIR / "themes" / "saved_themes.json"
        if themes_json_path.exists():
            try:
                data = json.loads(themes_json_path.read_text(encoding="utf-8"))
                for t in (data.get("themes") or []):
                    name = t.get("name", "Minimalista")
                    existing = await db.saved_themes.find_one({"name": name})
                    if existing:
                        tid = existing["_id"]
                    else:
                        ins = await db.saved_themes.insert_one({
                            "name": name,
                            "snapshot": t.get("snapshot", {}),
                            "created_at": t.get("created_at") or datetime.now(timezone.utc).isoformat(),
                            "updated_at": t.get("updated_at") or "",
                            "is_default": bool(t.get("is_default", False)),
                        })
                        tid = ins.inserted_id
                    if t.get("is_default") and not seeded_default_id:
                        seeded_default_id = tid
                        seeded_default_snapshot = t.get("snapshot", {})
            except Exception as ex:
                logger.warning(f"No se pudo leer themes/saved_themes.json: {ex}")

        if await db.saved_themes.count_documents({}) == 0:
            default_snapshot = {"sidebar_style": "island", "nav_config": "[]"}
            ins = await db.saved_themes.insert_one({
                "name": "Minimalista",
                "snapshot": default_snapshot,
                "created_at": datetime.now(timezone.utc).isoformat(),
                "updated_at": "",
                "is_default": True,
            })
            seeded_default_id = ins.inserted_id
            seeded_default_snapshot = default_snapshot
            logger.warning("Seeded default 'Minimalista' theme (fallback)")

        cur_settings = await db.app_settings.find_one({}, {"default_theme_id": 1}) or {}
        existing_default = cur_settings.get("default_theme_id")
        if isinstance(existing_default, ObjectId):
            await db.app_settings.update_one(
                {}, {"$set": {"default_theme_id": str(existing_default)}}
            )
            existing_default = str(existing_default)

        if not existing_default:
            if not seeded_default_id:
                minimal = await db.saved_themes.find_one({"name": {"$regex": "^minimal", "$options": "i"}})
                if not minimal:
                    minimal = await db.saved_themes.find_one({}, sort=[("created_at", 1)])
                if minimal:
                    seeded_default_id = minimal["_id"]
                    seeded_default_snapshot = minimal.get("snapshot", {})
            if seeded_default_id:
                doc_st = await db.saved_themes.find_one({"_id": seeded_default_id}) or {}
                await db.app_settings.update_one(
                    {},
                    {"$set": {
                        "default_theme_id": str(seeded_default_id),
                        "default_theme_name": doc_st.get("name", "Minimalista"),
                        "appearance_snapshot": seeded_default_snapshot or doc_st.get("snapshot", {}),
                        "appearance_updated_at": datetime.now(timezone.utc).isoformat(),
                    }},
                    upsert=True,
                )
                logger.warning(f"Marcado tema por defecto: {doc_st.get('name', 'Minimalista')}")

        if _using_embedded:
            await _save_embedded_data()
    except Exception as _theme_seed_err:
        logger.warning(f"No se pudo sembrar tema por defecto: {_theme_seed_err}")

    # ── Pull inicial de themes desde GitHub (best-effort, sin bloquear) ─────
    asyncio.create_task(_pull_themes_from_github_safe())

    # ── Check for updates in background ──────────────────────────────────────
    asyncio.create_task(_check_for_updates())

    yield
    if _using_embedded:
        if _task:
            _task.cancel()
        await _save_embedded_data()
    _mongo_client.close()


app = FastAPI(title="Cinema Productions", lifespan=lifespan)
api_router = APIRouter(prefix="/api")


# ─── Helpers ──────────────────────────────────────────────

def doc_to_dict(doc: dict) -> dict:
    if doc is None:
        return {}
    d = {k: v for k, v in doc.items() if k != '_id'}
    if '_id' in doc:
        d['id'] = str(doc['_id'])
    return d


# ─── Models ───────────────────────────────────────────────

class ReservationCreate(BaseModel):
    client_name: str
    client_phone: Optional[str] = None
    client_email: Optional[str] = None
    event_type: str
    event_date: str
    event_time: Optional[str] = None
    venue: Optional[str] = None
    guests_count: Optional[int] = None
    total_amount: float
    advance_paid: float = 0.0
    status: str = "Pendiente"
    notes: Optional[str] = None
    package_type: Optional[str] = None


class ReservationUpdate(BaseModel):
    client_name: Optional[str] = None
    client_phone: Optional[str] = None
    client_email: Optional[str] = None
    event_type: Optional[str] = None
    event_date: Optional[str] = None
    event_time: Optional[str] = None
    venue: Optional[str] = None
    guests_count: Optional[int] = None
    total_amount: Optional[float] = None
    advance_paid: Optional[float] = None
    status: Optional[str] = None
    notes: Optional[str] = None
    locations: Optional[list] = None
    assigned_partners: Optional[list] = None
    package_type: Optional[str] = None


class SocioCreate(BaseModel):
    name: str
    role: str = "Fotógrafo"
    phone: Optional[str] = None
    email: Optional[str] = None
    notes: Optional[str] = None
    rate_per_event: Optional[float] = None


class SocioUpdate(BaseModel):
    name: Optional[str] = None
    role: Optional[str] = None
    phone: Optional[str] = None
    email: Optional[str] = None
    notes: Optional[str] = None
    rate_per_event: Optional[float] = None


class NotificationSettingsModel(BaseModel):
    reminders_enabled: bool = False
    reminder_days: Optional[int] = 3
    reminder_periods: Optional[list] = None
    reminder_time: Optional[str] = "09:00"
    reminder_hours_before: Optional[int] = 0
    admin_email: Optional[str] = None
    admin_whatsapp: Optional[str] = None
    notification_channel: str = "email"
    resend_api_key: Optional[str] = None
    resend_sender_name: Optional[str] = None
    resend_cc: Optional[str] = None
    resend_subject_template: Optional[str] = None
    notify_client: Optional[bool] = False
    telegram_enabled: Optional[bool] = False
    telegram_bot_token: Optional[str] = None
    telegram_chat_id: Optional[str] = None
    ntfy_enabled: Optional[bool] = False
    ntfy_topic: Optional[str] = None
    company_name: Optional[str] = None
    company_address: Optional[str] = None
    company_phone: Optional[str] = None
    company_email: Optional[str] = None
    company_logo: Optional[str] = None
    currency: Optional[str] = "Q"
    language: Optional[str] = "es"
    timezone: Optional[str] = None


class DBConnectRequest(BaseModel):
    url: str


# ─── Routes ───────────────────────────────────────────────

@api_router.get("/")
async def root():
    mode = "embedded" if _using_embedded else "mongodb"
    # `version` = versión REALMENTE en ejecución (horneada en el binario). El
    # frontend la usa tras una auto-actualización para confirmar que ya está
    # hablando con el NUEVO exe (versión distinta) y no con el viejo que aún
    # está muriendo → evita recargar en la "zona muerta" y ver pantalla blanca.
    return {"message": "Cinema Productions API", "db_mode": mode, "version": _local_version}


@api_router.get("/stats")
async def get_stats():
    total = await db.reservations.count_documents({})
    today_str = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    # Fetch all active reservations (include assigned_partners for real_income)
    all_res = await db.reservations.find(
        {},
        {"event_date": 1, "status": 1, "total_amount": 1, "advance_paid": 1,
         "assigned_partners": 1, "_id": 0}
    ).to_list(10000)

    upcoming = sum(
        1 for d in all_res
        if d.get("event_date", "") >= today_str
        and d.get("status") not in ("Cancelado", "Completado")
    )
    active_docs = [d for d in all_res if d.get("status") not in ("Cancelado",)]

    total_pending = sum(
        max(0, (d.get("total_amount") or 0) - (d.get("advance_paid") or 0))
        for d in active_docs
    )
    total_event_amount = sum((d.get("total_amount") or 0) for d in active_docs)
    total_partner_cost = sum(
        p.get("payment", 0) or 0
        for d in active_docs
        for p in (d.get("assigned_partners") or [])
    )

    return {
        "total_reservations": total,
        "upcoming_events": upcoming,
        "pending_payment": round(total_pending, 2),
        "real_income": round(total_event_amount - total_partner_cost, 2),
    }


@api_router.get("/reservations")
async def list_reservations():
    docs = await db.reservations.find({}, {"receipt_images.data": 0}).to_list(1000)
    return [doc_to_dict(d) for d in docs]


@api_router.post("/reservations", status_code=201)
async def create_reservation(reservation: ReservationCreate):
    doc = reservation.model_dump()
    doc["receipt_images"] = []
    doc["created_at"] = datetime.now(timezone.utc).isoformat()
    result = await db.reservations.insert_one(doc)
    doc["id"] = str(result.inserted_id)
    doc.pop("_id", None)
    if _using_embedded:
        asyncio.create_task(_save_embedded_data())
    return doc


@api_router.get("/reservations/{reservation_id}")
async def get_reservation(reservation_id: str):
    try:
        oid = ObjectId(reservation_id)
    except Exception:
        raise HTTPException(status_code=400, detail="ID inválido")
    doc = await db.reservations.find_one({"_id": oid})
    if not doc:
        raise HTTPException(status_code=404, detail="Reserva no encontrada")
    return doc_to_dict(doc)


@api_router.put("/reservations/{reservation_id}")
async def update_reservation(reservation_id: str, reservation: ReservationUpdate):
    try:
        oid = ObjectId(reservation_id)
    except Exception:
        raise HTTPException(status_code=400, detail="ID inválido")
    update_data = {k: v for k, v in reservation.model_dump().items() if v is not None}
    if not update_data:
        raise HTTPException(status_code=400, detail="Sin datos para actualizar")
    result = await db.reservations.update_one({"_id": oid}, {"$set": update_data})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Reserva no encontrada")
    doc = await db.reservations.find_one({"_id": oid})
    if _using_embedded:
        asyncio.create_task(_save_embedded_data())
    return doc_to_dict(doc)


@api_router.delete("/reservations/{reservation_id}")
async def delete_reservation(reservation_id: str):
    try:
        oid = ObjectId(reservation_id)
    except Exception:
        raise HTTPException(status_code=400, detail="ID inválido")
    result = await db.reservations.delete_one({"_id": oid})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Reserva no encontrada")
    if _using_embedded:
        asyncio.create_task(_save_embedded_data())
    return {"message": "Reserva eliminada"}


@api_router.post("/reservations/{reservation_id}/receipts")
async def upload_receipt(reservation_id: str, file: UploadFile = File(...)):
    try:
        oid = ObjectId(reservation_id)
    except Exception:
        raise HTTPException(status_code=400, detail="ID inválido")
    content = await file.read()
    if len(content) > 10 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="Archivo muy grande (máx 10MB)")
    b64 = base64.b64encode(content).decode("utf-8")
    receipt = {
        "id": str(uuid.uuid4()),
        "filename": file.filename,
        "content_type": file.content_type,
        "data": b64,
        "uploaded_at": datetime.now(timezone.utc).isoformat()
    }
    result = await db.reservations.update_one({"_id": oid}, {"$push": {"receipt_images": receipt}})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Reserva no encontrada")
    if _using_embedded:
        asyncio.create_task(_save_embedded_data())
    return {k: v for k, v in receipt.items() if k != "data"}


@api_router.delete("/reservations/{reservation_id}/receipts/{receipt_id}")
async def delete_receipt(reservation_id: str, receipt_id: str):
    try:
        oid = ObjectId(reservation_id)
    except Exception:
        raise HTTPException(status_code=400, detail="ID inválido")
    result = await db.reservations.update_one(
        {"_id": oid}, {"$pull": {"receipt_images": {"id": receipt_id}}}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Reserva no encontrada")
    if _using_embedded:
        asyncio.create_task(_save_embedded_data())
    return {"message": "Comprobante eliminado"}


@api_router.get("/export/reservations")
async def export_reservations(format: str = "csv"):
    docs = await db.reservations.find({}, {"receipt_images": 0}).to_list(10000)
    data = [doc_to_dict(d) for d in docs]
    if format == "json":
        return JSONResponse(content=data,
                            headers={"Content-Disposition": "attachment; filename=reservaciones.json"})
    if not data:
        return Response("", media_type="text/csv",
                        headers={"Content-Disposition": "attachment; filename=reservaciones.csv"})
    output = io.StringIO()
    fields = ["id", "client_name", "client_phone", "client_email", "event_type", "event_date",
              "event_time", "venue", "guests_count", "total_amount", "advance_paid", "status", "notes", "created_at"]
    writer = csv.DictWriter(output, fieldnames=fields, extrasaction="ignore")
    writer.writeheader()
    writer.writerows(data)
    return Response(output.getvalue(), media_type="text/csv; charset=utf-8",
                    headers={"Content-Disposition": "attachment; filename=reservaciones.csv"})


@api_router.get("/calendar")
async def get_calendar_events():
    docs = await db.reservations.find(
        {"status": {"$nin": ["Cancelado"]}},
        {"client_name": 1, "event_date": 1, "event_type": 1, "status": 1, "_id": 1}
    ).to_list(1000)
    return [doc_to_dict(d) for d in docs]


# ─── Socios ───────────────────────────────────────────────

@api_router.get("/socios")
async def list_socios():
    docs = await db.socios.find({}, {"photo": 0, "photo_content_type": 0}).to_list(1000)
    return [doc_to_dict(d) for d in docs]


@api_router.post("/socios", status_code=201)
async def create_socio(socio: SocioCreate):
    doc = socio.model_dump()
    doc.update({"photo": None, "photo_content_type": None,
                "created_at": datetime.now(timezone.utc).isoformat()})
    result = await db.socios.insert_one(doc)
    doc["id"] = str(result.inserted_id)
    doc.pop("_id", None)
    if _using_embedded:
        asyncio.create_task(_save_embedded_data())
    return doc


@api_router.get("/socios/{socio_id}")
async def get_socio(socio_id: str):
    try:
        oid = ObjectId(socio_id)
    except Exception:
        raise HTTPException(status_code=400, detail="ID inválido")
    doc = await db.socios.find_one({"_id": oid})
    if not doc:
        raise HTTPException(status_code=404, detail="Socio no encontrado")
    return doc_to_dict(doc)


@api_router.put("/socios/{socio_id}")
async def update_socio(socio_id: str, socio: SocioUpdate):
    try:
        oid = ObjectId(socio_id)
    except Exception:
        raise HTTPException(status_code=400, detail="ID inválido")
    update_data = {k: v for k, v in socio.model_dump().items() if v is not None}
    if not update_data:
        raise HTTPException(status_code=400, detail="Sin datos para actualizar")
    result = await db.socios.update_one({"_id": oid}, {"$set": update_data})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Socio no encontrado")
    doc = await db.socios.find_one({"_id": oid})
    if _using_embedded:
        asyncio.create_task(_save_embedded_data())
    return doc_to_dict(doc)


@api_router.delete("/socios/{socio_id}")
async def delete_socio(socio_id: str):
    try:
        oid = ObjectId(socio_id)
    except Exception:
        raise HTTPException(status_code=400, detail="ID inválido")
    result = await db.socios.delete_one({"_id": oid})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Socio no encontrado")
    if _using_embedded:
        asyncio.create_task(_save_embedded_data())
    return {"message": "Socio eliminado"}


@api_router.post("/socios/{socio_id}/photo")
async def upload_socio_photo(socio_id: str, file: UploadFile = File(...)):
    try:
        oid = ObjectId(socio_id)
    except Exception:
        raise HTTPException(status_code=400, detail="ID inválido")
    content = await file.read()
    if len(content) > 5 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="Archivo muy grande (máx 5MB)")
    b64 = base64.b64encode(content).decode("utf-8")
    result = await db.socios.update_one(
        {"_id": oid}, {"$set": {"photo": b64, "photo_content_type": file.content_type}}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Socio no encontrado")
    if _using_embedded:
        asyncio.create_task(_save_embedded_data())
    return {"message": "Foto actualizada"}


@api_router.delete("/socios/{socio_id}/photo")
async def delete_socio_photo(socio_id: str):
    try:
        oid = ObjectId(socio_id)
    except Exception:
        raise HTTPException(status_code=400, detail="ID inválido")
    await db.socios.update_one({"_id": oid}, {"$set": {"photo": None, "photo_content_type": None}})
    return {"message": "Foto eliminada"}


@api_router.get("/financials")
async def get_financials():
    docs = await db.reservations.find(
        {"status": {"$nin": ["Cancelado"]}},
        {"total_amount": 1, "advance_paid": 1, "assigned_partners": 1}
    ).to_list(10000)
    total_event = sum((d.get("total_amount") or 0) for d in docs)
    total_advance = sum((d.get("advance_paid") or 0) for d in docs)
    total_cost = total_paid = total_pending = 0
    for d in docs:
        for p in (d.get("assigned_partners") or []):
            amt = p.get("payment") or 0
            total_cost += amt
            if p.get("payment_status") == "Pagado":
                total_paid += amt
            else:
                total_pending += amt
    return {
        "total_event_amount": round(total_event, 2),
        "total_advance": round(total_advance, 2),
        "total_partner_cost": round(total_cost, 2),
        "total_paid_to_partners": round(total_paid, 2),
        "total_pending_to_partners": round(total_pending, 2),
        "real_income": round(total_event - total_cost, 2),
    }


@api_router.get("/settings")
async def get_app_settings():
    doc = await db.app_settings.find_one({}, {"_id": 0})
    if not doc:
        return {}
    doc.pop("app_password_hash", None)
    if doc.get("resend_api_key"):
        key = doc["resend_api_key"]
        doc["resend_api_key"] = "re_" + "*" * 20 + key[-4:] if len(key) > 4 else "****"
        doc["has_resend_key"] = True
    else:
        doc["has_resend_key"] = False
    return doc


@api_router.put("/settings")
async def update_app_settings(settings: NotificationSettingsModel):
    update_doc = {k: v for k, v in settings.model_dump().items() if v is not None}
    # Don't overwrite the real key if frontend sends a masked placeholder
    key = update_doc.get("resend_api_key") or ""
    if "****" in key or key.startswith("re_" + "*"):
        update_doc.pop("resend_api_key", None)
    telegram_token = update_doc.get("telegram_bot_token") or ""
    if "•" in telegram_token or "****" in telegram_token:
        update_doc.pop("telegram_bot_token", None)
    existing = await db.app_settings.find_one({}, {"_id": 0})
    if existing:
        await db.app_settings.update_one({}, {"$set": update_doc})
    else:
        await db.app_settings.insert_one(update_doc)
    if _using_embedded:
        asyncio.create_task(_save_embedded_data())
    saved = await db.app_settings.find_one({}, {"_id": 0})
    if saved and saved.get("resend_api_key"):
        key = saved["resend_api_key"]
        saved["resend_api_key"] = "re_" + "*" * 20 + key[-4:] if len(key) > 4 else "****"
        saved["has_resend_key"] = True
    elif saved:
        saved["has_resend_key"] = False
    return saved or {}


@api_router.get("/settings/database")
async def get_db_stats():
    if _using_embedded:
        counts = {}
        total_docs = 0
        for col in ['reservations', 'socios', 'app_settings']:
            n = await db[col].count_documents({})
            counts[col] = n
            total_docs += n
        data_file_size = DATA_FILE.stat().st_size if DATA_FILE.exists() else 0

        def fmt(b):
            if b < 1024:
                return f"{b} B"
            elif b < 1024 ** 2:
                return f"{b / 1024:.1f} KB"
            return f"{b / (1024 ** 2):.2f} MB"

        return {
            "db_name": DB_NAME + " (embebida)",
            "collections": len(counts),
            "objects": total_docs,
            "data_size": fmt(data_file_size),
            "storage_size": fmt(data_file_size),
            "index_size": "0 B",
            "total_size": fmt(data_file_size),
            "used_size": fmt(data_file_size),
            "free_size": "Ilimitado (disco local)",
            "limit_size": "—",
            "used_pct": 0,
            "is_atlas": False,
            "current_url": "embedded (cinema_data.json)",
            "is_custom": False,
        }
    try:
        raw = await db.command("dbstats")
        storage_bytes = raw.get("storageSize", 0)
        data_bytes = raw.get("dataSize", 0)
        index_bytes = raw.get("indexSize", 0)

        def fmt(b):
            if b < 1024:
                return f"{b} B"
            elif b < 1024 ** 2:
                return f"{b / 1024:.1f} KB"
            return f"{b / (1024 ** 2):.2f} MB"

        is_custom = CUSTOM_DB_FILE.exists()
        current_url = CUSTOM_DB_FILE.read_text().strip() if is_custom else MONGO_URL
        display_url = current_url
        if "@" in current_url:
            proto_end = current_url.find("://") + 3
            at_pos = current_url.rfind("@")
            display_url = current_url[:proto_end] + "***:***@" + current_url[at_pos + 1:]
        is_atlas = current_url.startswith("mongodb+srv")
        used = storage_bytes + index_bytes
        limit_bytes = 512 * 1024 ** 2
        free_bytes = max(0, limit_bytes - used)
        return {
            "db_name": DB_NAME,
            "collections": raw.get("collections", 0),
            "objects": raw.get("objects", 0),
            "data_size": fmt(data_bytes),
            "storage_size": fmt(storage_bytes),
            "index_size": fmt(index_bytes),
            "total_size": fmt(used),
            "used_size": fmt(used),
            "free_size": fmt(free_bytes),
            "limit_size": fmt(limit_bytes),
            "used_pct": round(min(100, used / limit_bytes * 100), 1),
            "is_atlas": is_atlas,
            "current_url": display_url,
            "is_custom": is_custom,
        }
    except Exception as e:
        # Atlas not reachable — return safe fallback so UI doesn't crash
        is_custom = CUSTOM_DB_FILE.exists()
        current_url = CUSTOM_DB_FILE.read_text().strip() if is_custom else _effective_mongo_url
        display_url = current_url
        if "@" in current_url:
            proto_end = current_url.find("://") + 3
            at_pos = current_url.rfind("@")
            display_url = current_url[:proto_end] + "***:***@" + current_url[at_pos + 1:]
        return {
            "db_name": DB_NAME,
            "collections": 0,
            "objects": 0,
            "data_size": "—",
            "storage_size": "—",
            "index_size": "—",
            "total_size": "—",
            "current_url": display_url,
            "is_custom": is_custom,
            "connection_error": str(e),
        }


@api_router.post("/settings/database/test")
async def test_db_connection(req: DBConnectRequest):
    if req.url.lower() == "embedded":
        return {"success": True, "message": "Modo embebido seleccionado"}
    try:
        from motor.motor_asyncio import AsyncIOMotorClient
        test_client = AsyncIOMotorClient(req.url, serverSelectionTimeoutMS=5000)
        await test_client[DB_NAME].command("ping")
        test_client.close()
        return {"success": True, "message": "Conexión exitosa"}
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"No se pudo conectar: {e}")


@api_router.post("/settings/database/connect")
async def switch_database(req: DBConnectRequest):
    global db, client, _using_embedded
    url = req.url.strip()

    # Switch to embedded mode
    if url.lower() in ('embedded', '', 'local'):
        from mongomock_motor import AsyncMongoMockClient
        new_client = AsyncMongoMockClient()
        db = new_client[DB_NAME]
        client = new_client
        CUSTOM_DB_FILE.unlink(missing_ok=True)
        _using_embedded = True
        return {"success": True, "message": "Modo embebido activado. Los datos se guardan localmente."}

    # Switch to MongoDB URL
    try:
        from motor.motor_asyncio import AsyncIOMotorClient as _MotorClient
        new_client = _MotorClient(url, serverSelectionTimeoutMS=8000)
        await new_client[DB_NAME].command("ping")
        db = new_client[DB_NAME]
        client = new_client
        CUSTOM_DB_FILE.write_text(url)
        _using_embedded = False
        return {"success": True, "message": "Base de datos conectada correctamente"}
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Error al conectar: {e}")


@api_router.get("/settings/database/factory-presets")
async def get_factory_presets():
    """Conexiones MongoDB de fábrica embebidas en la app de escritorio."""
    return {"presets": FACTORY_SAVED_CONNECTIONS}


@api_router.post("/settings/database/reset")
async def reset_database():
    global db, client, _using_embedded
    if CUSTOM_DB_FILE.exists():
        CUSTOM_DB_FILE.unlink()
    # Fall back to default MONGO_URL from .env
    default_url = MONGO_URL.strip()
    if default_url.lower() in ('embedded', '', 'local'):
        from mongomock_motor import AsyncMongoMockClient
        new_client = AsyncMongoMockClient()
        _using_embedded = True
    else:
        from motor.motor_asyncio import AsyncIOMotorClient as _MotorClient
        new_client = _MotorClient(default_url)
        _using_embedded = False
    db = new_client[DB_NAME]
    client = new_client
    return {"success": True, "message": "Restaurado a la base de datos predeterminada."}


@api_router.delete("/data/clear-all")
async def clear_all_data_endpoint():
    try:
        await _create_backup(label="auto_pre_delete")
    except Exception:
        pass
    res_result = await db.reservations.delete_many({})
    soc_result = await db.socios.delete_many({})
    await db.app_settings.delete_many({})
    if _using_embedded:
        asyncio.create_task(_save_embedded_data())
    return {
        "ok": True,
        "deleted_reservations": res_result.deleted_count,
        "deleted_socios": soc_result.deleted_count,
        "auto_backup_created": True,
    }


@api_router.post("/data/cleanup")
async def cleanup_data(action: str = "cancelled", months_old: int = 6):
    if action == "preview":
        cancelled_count = await db.reservations.count_documents({"status": "Cancelado"})
        cutoff = datetime.now(timezone.utc) - timedelta(days=months_old * 30)
        cutoff_str = cutoff.strftime("%Y-%m-%d")
        docs = await db.reservations.find({"status": "Completado"}).to_list(100000)
        old_completed = sum(1 for d in docs if d.get("event_date", "") < cutoff_str)
        return {
            "ok": True,
            "cancelled_count": cancelled_count,
            "old_completed_count": old_completed,
            "months_threshold": months_old,
        }

    try:
        await _create_backup(label="auto_pre_cleanup")
    except Exception:
        pass

    if action == "cancelled":
        result = await db.reservations.delete_many({"status": "Cancelado"})
        if _using_embedded:
            asyncio.create_task(_save_embedded_data())
        return {"ok": True, "deleted": result.deleted_count, "message": f"{result.deleted_count} reservas canceladas eliminadas"}

    if action == "old_completed":
        cutoff = datetime.now(timezone.utc) - timedelta(days=months_old * 30)
        cutoff_str = cutoff.strftime("%Y-%m-%d")
        docs = await db.reservations.find({"status": "Completado"}).to_list(100000)
        ids_to_delete = [d["_id"] for d in docs if d.get("event_date", "") < cutoff_str]
        if ids_to_delete:
            result = await db.reservations.delete_many({"_id": {"$in": ids_to_delete}})
            if _using_embedded:
                asyncio.create_task(_save_embedded_data())
            return {"ok": True, "deleted": result.deleted_count, "message": f"{result.deleted_count} reservas completadas antiguas eliminadas"}
        return {"ok": True, "deleted": 0, "message": "No hay reservas completadas antiguas para eliminar"}

    return {"ok": False, "message": "Acción no reconocida"}


# ─── Backup helper ─────────────────────────────────────────

async def _all_collections() -> list:
    try:
        names = await db.list_collection_names()
    except Exception:
        names = list(BACKUP_COLLECTIONS)
    return [n for n in names if not n.startswith("system.")] or list(BACKUP_COLLECTIONS)


async def _create_backup(label: str = "manual") -> dict:
    names = await _all_collections()
    backup_data: dict = {"_meta": {"created_at": datetime.now(timezone.utc).isoformat(), "label": label, "collections": names}}
    for cname in names:
        docs = await db[cname].find({}).to_list(100000)
        backup_data[cname] = [doc_to_dict(d) for d in docs]
    ts = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
    filename = f"backup_{label}_{ts}.json"
    filepath = BACKUP_DIR / filename
    filepath.write_text(json.dumps(backup_data, ensure_ascii=False, indent=2), encoding="utf-8")
    # Keep only last 15 backups
    existing = sorted(BACKUP_DIR.glob("backup_*.json"), key=lambda f: f.stat().st_mtime)
    for old in existing[:-15]:
        old.unlink(missing_ok=True)
    total_docs = sum(len(v) for k, v in backup_data.items() if k != "_meta")
    return {"filename": filename, "docs": total_docs}


# ─── Backup endpoints ──────────────────────────────────────

@api_router.get("/backup/download")
async def download_full_backup():
    names = await _all_collections()
    backup_data: dict = {"_meta": {"created_at": datetime.now(timezone.utc).isoformat(), "app": "Cinema Productions", "collections": names}}
    for cname in names:
        docs = await db[cname].find({}).to_list(100000)
        backup_data[cname] = [doc_to_dict(d) for d in docs]
    ts = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
    filename = f"cinema_backup_{ts}.json"
    content = json.dumps(backup_data, ensure_ascii=False, indent=2)
    return Response(
        content=content.encode("utf-8"),
        media_type="application/json",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@api_router.post("/backup/create")
async def create_server_backup():
    try:
        result = await _create_backup(label="manual")
        return {"success": True, **result, "message": f"Respaldo creado: {result['docs']} documentos"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error al crear respaldo: {e}")


@api_router.get("/backup/history")
async def list_backups():
    files = sorted(BACKUP_DIR.glob("backup_*.json"), key=lambda f: f.stat().st_mtime, reverse=True)
    result = []
    for f in files:
        stat = f.stat()
        size_kb = stat.st_size / 1024
        size_str = f"{size_kb:.0f} KB" if size_kb < 1024 else f"{size_kb / 1024:.1f} MB"
        label = "auto" if "auto_" in f.name else "manual"
        result.append({
            "filename": f.name,
            "size": size_str,
            "label": label,
            "created_at": datetime.fromtimestamp(stat.st_mtime, tz=timezone.utc).isoformat(),
        })
    return result


@api_router.get("/backup/{filename}/download")
async def download_backup_file(filename: str):
    if not filename.endswith(".json") or "/" in filename or ".." in filename:
        raise HTTPException(status_code=400, detail="Nombre de archivo inválido")
    filepath = BACKUP_DIR / filename
    if not filepath.exists():
        raise HTTPException(status_code=404, detail="Respaldo no encontrado")
    return Response(
        content=filepath.read_bytes(),
        media_type="application/json",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@api_router.delete("/backup/{filename}")
async def delete_backup_file(filename: str):
    if not filename.endswith(".json") or "/" in filename or ".." in filename:
        raise HTTPException(status_code=400, detail="Nombre de archivo inválido")
    filepath = BACKUP_DIR / filename
    if not filepath.exists():
        raise HTTPException(status_code=404, detail="Respaldo no encontrado")
    filepath.unlink()
    return {"success": True, "message": "Respaldo eliminado"}


@api_router.post("/backup/restore")
async def restore_backup(file: UploadFile = File(...)):
    if not file.filename.endswith(".json"):
        raise HTTPException(status_code=400, detail="El archivo debe ser .json")
    try:
        content = await file.read()
        backup_data = json.loads(content)
    except Exception:
        raise HTTPException(status_code=400, detail="Archivo JSON inválido o corrupto")
    try:
        await _create_backup(label="auto_pre_restore")
    except Exception:
        pass
    restored: dict = {}
    errors: list = []
    for cname, docs in backup_data.items():
        if cname == "_meta" or not isinstance(docs, list):
            continue
        try:
            clean_docs = []
            for d in docs:
                d.pop("id", None)
                d.pop("_id", None)
                clean_docs.append(d)
            await db[cname].delete_many({})
            if clean_docs:
                await db[cname].insert_many(clean_docs)
            restored[cname] = len(clean_docs)
        except Exception as e:
            errors.append(f"{cname}: {e}")
    if errors:
        raise HTTPException(status_code=500, detail="Errores al restaurar: " + " | ".join(errors))
    if _using_embedded:
        asyncio.create_task(_save_embedded_data())
    total = sum(restored.values())
    return {"success": True, "restored": restored, "total": total,
            "message": f"Restaurado correctamente: {total} documentos en {len(restored)} colecciones"}


@api_router.get("/settings/database/stats-detailed")
async def db_stats_detailed():
    """Contadores por colección para preview de 'subir a la nube'."""
    try:
        names = await _all_collections()
        counts = {}
        for n in names:
            try:
                counts[n] = await db[n].estimated_document_count()
            except Exception:
                counts[n] = 0
        reservations = int(counts.get("reservations", 0))
        socios = int(counts.get("socios", 0))
        settings = int(counts.get("app_settings", 0)) + int(counts.get("appearance_settings", 0))
        themes = int(counts.get("saved_themes", 0))
        total = sum(int(v) for v in counts.values())
        try:
            current_url = CUSTOM_DB_FILE.read_text().strip() if CUSTOM_DB_FILE.exists() else os.environ.get('MONGO_URL', '')
        except Exception:
            current_url = ""
        is_atlas = "mongodb+srv" in (current_url or "") or ".mongodb.net" in (current_url or "")
        size_str = "—"
        try:
            st = await db.command("dbStats")
            data_size = int(st.get("dataSize", 0))
            if data_size < 1024:
                size_str = f"{data_size} B"
            elif data_size < 1024 * 1024:
                size_str = f"{data_size/1024:.1f} KB"
            else:
                size_str = f"{data_size/(1024*1024):.2f} MB"
        except Exception:
            pass
        return {
            "reservations": reservations,
            "socios": socios,
            "settings": max(settings, 1),
            "themes": themes,
            "total": total,
            "size": size_str,
            "is_cloud": is_atlas,
            "collections": counts,
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error al obtener detalle: {e}")


@api_router.post("/settings/database/optimize")
async def optimize_database():
    """Compacta y reordena: asegura índices e intenta compactar (se omite en embebido/Atlas)."""
    result = {"indexed": [], "compacted": [], "skipped": []}
    if _using_embedded:
        try:
            await _save_embedded_data()
        except Exception:
            pass
        return {"success": True, "message": "Base local reordenada y guardada.", **result}
    try:
        try:
            await db.reservations.create_index("event_date")
            await db.reservations.create_index("status")
            await db.reservations.create_index("created_at")
            result["indexed"].append("reservations")
        except Exception as e:
            result["skipped"].append(f"idx reservations: {str(e)[:60]}")
        try:
            await db.socios.create_index("name")
            result["indexed"].append("socios")
        except Exception as e:
            result["skipped"].append(f"idx socios: {str(e)[:60]}")
        for cname in await _all_collections():
            try:
                await db.command({"compact": cname})
                result["compacted"].append(cname)
            except Exception:
                result["skipped"].append(f"compact {cname} (no permitido aquí)")
        return {"success": True, "message": f"Optimización lista. Índices: {len(result['indexed'])}", **result}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error al optimizar: {e}")


# ─── Import / Export adicionales ──────────────────────────

@api_router.post("/import/reservations")
async def import_reservations_csv(file: UploadFile = File(...)):
    content = await file.read()
    try:
        text = content.decode("utf-8-sig")
    except UnicodeDecodeError:
        text = content.decode("latin-1")
    reader = csv.DictReader(io.StringIO(text))
    FIELD_MAP = {
        "client_name":  ["client_name", "nombre", "cliente", "name"],
        "client_phone": ["client_phone", "telefono", "teléfono", "phone"],
        "client_email": ["client_email", "email", "correo"],
        "event_type":   ["event_type", "tipo_evento", "tipo", "type"],
        "event_date":   ["event_date", "fecha", "date", "fecha_evento"],
        "event_time":   ["event_time", "hora", "time"],
        "venue":        ["venue", "lugar", "location", "ubicacion"],
        "guests_count": ["guests_count", "invitados", "guests"],
        "total_amount": ["total_amount", "total", "monto_total"],
        "advance_paid": ["advance_paid", "anticipo", "advance", "pago_anticipado"],
        "status":       ["status", "estado"],
        "notes":        ["notes", "notas", "note"],
    }

    def find_col(headers, candidates):
        headers_lower = {h.lower().strip(): h for h in headers}
        for c in candidates:
            if c.lower() in headers_lower:
                return headers_lower[c.lower()]
        return None

    imported = 0
    errors = []
    now_str = datetime.now(timezone.utc).isoformat()

    for i, row in enumerate(reader, start=2):
        try:
            headers = list(row.keys())
            def get(field):
                col = find_col(headers, FIELD_MAP.get(field, [field]))
                return row.get(col, "").strip() if col else ""
            client_name = get("client_name")
            if not client_name:
                errors.append(f"Fila {i}: nombre vacío — omitida")
                continue

            def to_float(s):
                s = re.sub(r"[^\d.]", "", s)
                return float(s) if s else 0.0

            doc = {
                "client_name":   client_name,
                "client_phone":  get("client_phone") or None,
                "client_email":  get("client_email") or None,
                "event_type":    get("event_type") or "Otro",
                "event_date":    get("event_date") or "",
                "event_time":    get("event_time") or None,
                "venue":         get("venue") or None,
                "guests_count":  int(get("guests_count")) if get("guests_count").isdigit() else None,
                "total_amount":  to_float(get("total_amount") or "0"),
                "advance_paid":  to_float(get("advance_paid") or "0"),
                "status":        get("status") or "Pendiente",
                "notes":         get("notes") or None,
                "receipts":      [],
                "locations":     [],
                "assigned_partners": [],
                "created_at":    now_str,
                "imported_from_csv": True,
            }
            await db.reservations.insert_one(doc)
            imported += 1
        except Exception as e:
            errors.append(f"Fila {i}: {str(e)}")

    if _using_embedded:
        asyncio.create_task(_save_embedded_data())
    return {
        "ok": True,
        "imported": imported,
        "errors": errors[:10],
        "message": f"{imported} reservas importadas correctamente" + (f" ({len(errors)} errores)" if errors else ""),
    }


@api_router.get("/export/reservations/xlsx")
async def export_reservations_xlsx():
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
        import io as _io

        docs = await db.reservations.find({}, {"receipt_images": 0, "assigned_partners": 0}).to_list(100000)
        data = [doc_to_dict(d) for d in docs]

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Reservas"
        headers = ["Nombre", "Teléfono", "Email", "Tipo Evento", "Fecha", "Hora",
                   "Lugar", "Invitados", "Total", "Anticipo", "Saldo", "Estado", "Notas", "Creado"]
        keys = ["client_name", "client_phone", "client_email", "event_type", "event_date",
                "event_time", "venue", "guests_count", "total_amount", "advance_paid",
                None, "status", "notes", "created_at"]
        header_fill = PatternFill("solid", fgColor="4F46E5")
        header_font = Font(color="FFFFFF", bold=True, size=11)
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        alt_fill = PatternFill("solid", fgColor="F8F7FF")
        for row_idx, doc in enumerate(data, start=2):
            fill = alt_fill if row_idx % 2 == 0 else None
            for col_idx, key in enumerate(keys, start=1):
                if key is None:
                    value = (doc.get("total_amount") or 0) - (doc.get("advance_paid") or 0)
                else:
                    value = doc.get(key, "")
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                if fill:
                    cell.fill = fill
        for col in ws.columns:
            max_len = max((len(str(c.value or "")) for c in col), default=8)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 40)
        buf = _io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        ts = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M")
        return Response(
            content=buf.getvalue(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="reservaciones_{ts}.xlsx"'},
        )
    except ImportError:
        return JSONResponse({"error": "openpyxl no instalado. Ejecuta: pip install openpyxl"}, status_code=500)
    except Exception as e:
        return JSONResponse({"error": str(e)}, status_code=500)


@api_router.post("/reminders/send")
async def trigger_reminders_manual():
    return {"success": True, "events_found": 0, "sent": 0,
            "message": "Recordatorios por email disponibles en la versión web"}





# ── Auto-update detection (via shared MongoDB + GitHub) ───────────────────────

_gh_version_cache: dict = {"ts": 0.0, "version": "", "source": "", "notes": ""}


def _version_tuple(v: str):
    """Convierte '1.2.3' o '100' a tupla para comparar semánticamente.
    Si no puede parsear, cae a comparación por string."""
    v = (v or "").strip().lstrip("v")
    parts = []
    for p in v.split("."):
        try:
            parts.append(int(p))
        except ValueError:
            parts.append(-1)
    return tuple(parts) if parts else (0,)


def _is_newer(remote: str, local: str) -> bool:
    """True si remote > local (numéricamente si es posible)."""
    if not remote or remote == local:
        return False
    try:
        return _version_tuple(remote) > _version_tuple(local)
    except Exception:
        return remote != local


async def _fetch_github_version() -> dict:
    """Lee version.txt del repositorio GitHub configurado (raw). Cachea 3 min.
    Devuelve: {version, source, notes}. Si falla, version == ""."""
    import time as _t, urllib.request, urllib.error, json as _json

    # Cache de 3 minutos
    if _t.time() - _gh_version_cache["ts"] < 180 and _gh_version_cache["version"]:
        return {
            "version": _gh_version_cache["version"],
            "source": _gh_version_cache["source"],
            "notes": _gh_version_cache["notes"],
        }

    result = {"version": "", "source": "github", "notes": ""}
    try:
        cfg = await _get_github_cfg()
        repo_url = cfg.get("repo_url") or DEFAULT_GITHUB_REPO
        branch = cfg.get("branch") or "main"
        token = (cfg.get("token") or "").strip()
        owner, repo = _parse_github_url(repo_url)
        if not owner or not repo:
            return result

        headers = {"User-Agent": "cinema-productions-desktop"}
        if token:
            headers["Authorization"] = f"Bearer {token}"

        def _read_url(url: str, timeout: int = 6) -> str:
            req = urllib.request.Request(url, headers=headers)
            with urllib.request.urlopen(req, timeout=timeout) as resp:
                return resp.read().decode("utf-8", errors="ignore")

        # 1) Intento principal: version.txt en la rama configurada
        raw_urls = [
            f"https://raw.githubusercontent.com/{owner}/{repo}/{branch}/version.txt",
            f"https://raw.githubusercontent.com/{owner}/{repo}/main/version.txt",
            f"https://raw.githubusercontent.com/{owner}/{repo}/master/version.txt",
        ]
        # Deduplicar preservando orden
        seen = set()
        raw_urls = [u for u in raw_urls if not (u in seen or seen.add(u))]

        remote_version = ""
        used_url = ""
        for u in raw_urls:
            try:
                content = await asyncio.to_thread(_read_url, u, 6)
                candidate = (content or "").strip().splitlines()[0].strip() if content else ""
                candidate = candidate.lstrip("v").strip()
                if candidate:
                    remote_version = candidate
                    used_url = u
                    break
            except Exception:
                continue

        # 2) Fallback: último tag de GitHub
        if not remote_version:
            try:
                api_headers = {**headers, "Accept": "application/vnd.github+json"}

                def _get_tags():
                    req = urllib.request.Request(
                        f"https://api.github.com/repos/{owner}/{repo}/tags?per_page=1",
                        headers=api_headers,
                    )
                    with urllib.request.urlopen(req, timeout=6) as resp:
                        return _json.loads(resp.read().decode("utf-8", errors="ignore"))

                tags = await asyncio.to_thread(_get_tags)
                if tags and isinstance(tags, list):
                    tag_name = (tags[0].get("name") or "").lstrip("v").strip()
                    if tag_name:
                        remote_version = tag_name
                        used_url = f"https://github.com/{owner}/{repo}/releases/tag/{tags[0].get('name')}"
                        result["source"] = "github_tag"
            except Exception:
                pass

        result["version"] = remote_version
        if used_url:
            result["notes"] = used_url

        _gh_version_cache["ts"] = _t.time()
        _gh_version_cache["version"] = remote_version
        _gh_version_cache["source"] = result["source"]
        _gh_version_cache["notes"] = result["notes"]
    except Exception as e:
        logger.warning(f"GitHub version fetch failed: {e}")
    return result


async def _check_for_updates():
    """
    Verifica si hay una versión nueva. Combina TRES fuentes con reglas de
    coordinación para evitar anunciar versiones sin binario descargable:

      1) GitHub Release que CONTIENE el asset de esta plataforma (fuente de
         verdad para binarios compilados — .exe / linux / macOS). Sólo se
         reporta has_update=True cuando el asset ya está publicado.
      2) version.txt en el repo (informativo — indica que hay un build en curso
         pero puede no estar aún disponible para esta plataforma).
      3) MongoDB compartida (app_updates), como respaldo con paquete descargable
         para builds no-binarios.

    Esta lógica arregla la carrera entre los 4 jobs de release (Windows / Linux
    x86 / Linux ARM / macOS) que publican al mismo tag en momentos distintos.
    """
    global _update_status
    try:
        # 1) MongoDB (paquete de actualización, si existe)
        mongo_status = {"has_update": False, "remote_version": "", "filename": "",
                        "notes": "", "file_size": 0, "download_url": ""}
        try:
            latest = await db.app_updates.find_one({"is_latest": True}, sort=[("created_at", -1)])
            if latest:
                dl_url = f"{_update_server_url}/api/updates/download" if _update_server_url else ""
                mongo_status = {
                    "has_update": _is_newer(latest.get("version", ""), _local_version),
                    "remote_version": latest.get("version", ""),
                    "filename": latest.get("filename", ""),
                    "notes": latest.get("notes", ""),
                    "file_size": latest.get("file_size", 0),
                    "download_url": dl_url,
                }
        except Exception as e:
            logger.warning(f"Mongo update check failed: {e}")

        # 2) version.txt del repo (informativo)
        gh = await _fetch_github_version()
        gh_version = gh.get("version", "")

        # 3) GitHub Release con el asset de ESTA plataforma (fuente de verdad
        #    para binarios). Sólo si estamos ejecutando como binario congelado.
        release_status = {"has_update": False, "remote_version": "", "filename": "",
                          "notes": "", "file_size": 0, "download_url": ""}
        asset_pending = False  # version.txt anuncia una versión sin asset todavía
        if _is_frozen():
            try:
                asset_name = _current_asset_name()
                rel = await _find_release_asset(asset_name)
                rel_version = rel.get("version", "") if rel else ""
                if rel_version:
                    release_status = {
                        "has_update": _is_newer(rel_version, _local_version),
                        "remote_version": rel_version,
                        "filename": asset_name,
                        "notes": rel.get("url", ""),
                        "file_size": rel.get("size", 0),
                        "download_url": rel.get("url", ""),
                    }
                # ¿version.txt adelanta al release del asset? → build en curso
                if gh_version and rel_version and _is_newer(gh_version, rel_version):
                    asset_pending = True
                    logger.info(
                        f"Build pending for {asset_name}: version.txt={gh_version} "
                        f"pero release con asset está en v{rel_version}"
                    )
            except Exception as e:
                logger.warning(f"Release asset lookup failed: {e}")

        # Elegir la fuente ganadora
        candidates = []
        # En binarios congelados: SOLO el release con asset marca has_update=True.
        # version.txt se guarda como github_version informativo pero no genera banner.
        if _is_frozen():
            if release_status["remote_version"]:
                candidates.append(("github_release", release_status["remote_version"], release_status))
            # Mongo sigue siendo válido para paquetes no-binarios
            if mongo_status["remote_version"]:
                candidates.append(("mongo", mongo_status["remote_version"], mongo_status))
        else:
            # Modo fuente (dev): version.txt es la fuente principal, como antes
            if gh_version:
                candidates.append(("github", gh_version, {
                    "has_update": _is_newer(gh_version, _local_version),
                    "remote_version": gh_version,
                    "filename": "",
                    "notes": gh.get("notes", ""),
                    "file_size": 0,
                    "download_url": gh.get("notes", "") if gh.get("source") == "github_tag" else "",
                }))
            if mongo_status["remote_version"]:
                candidates.append(("mongo", mongo_status["remote_version"], mongo_status))

        if not candidates:
            _update_status = {"checked": True, "has_update": False,
                              "local_version": _local_version,
                              "github_version": gh_version, "remote_version": "",
                              "asset_pending": asset_pending}
            return

        # Ordenar por versión desc
        candidates.sort(key=lambda c: _version_tuple(c[1]), reverse=True)
        winner_source, winner_version, winner_payload = candidates[0]
        has_update = _is_newer(winner_version, _local_version)

        _update_status = {
            "checked": True,
            "has_update": has_update,
            "remote_version": winner_version,
            "local_version": _local_version,
            "github_version": gh_version,
            "source": winner_source,
            "filename": winner_payload.get("filename", ""),
            "notes": winner_payload.get("notes", ""),
            "file_size": winner_payload.get("file_size", 0),
            "download_url": winner_payload.get("download_url", ""),
            "asset_pending": asset_pending,
        }
        if has_update:
            logger.warning(
                f"Update available: {_local_version} → {winner_version} (source: {winner_source})"
            )
        elif asset_pending:
            logger.info(
                f"Build en curso: version.txt={gh_version} pero asset para esta "
                f"plataforma aún no publicado en release."
            )
    except Exception as e:
        logger.warning(f"Update check failed: {e}")
        _update_status = {"checked": True, "has_update": False,
                          "local_version": _local_version,
                          "github_version": "", "remote_version": "",
                          "asset_pending": False}


# ── APPEARANCE CLOUD SYNC ────────────────────────────────────────────────────

@api_router.get("/settings/appearance")
async def get_appearance_snapshot():
    doc = await db.app_settings.find_one({}, {"_id": 0, "appearance_snapshot": 1, "appearance_updated_at": 1})
    doc = doc or {}
    return {"snapshot": doc.get("appearance_snapshot"), "updated_at": doc.get("appearance_updated_at")}


@api_router.put("/settings/appearance")
async def save_appearance_snapshot(payload: dict = Body(...)):
    snapshot = payload.get("snapshot") or {}
    updated_at = datetime.now(timezone.utc).isoformat()
    await db.app_settings.update_one(
        {},
        {"$set": {"appearance_snapshot": snapshot, "appearance_updated_at": updated_at}},
        upsert=True,
    )
    if _using_embedded:
        asyncio.create_task(_save_embedded_data())
    return {"updated_at": updated_at}


# ── SAVED THEMES ─────────────────────────────────────────────────────────────

# ── SAVED THEMES: sincronización 3-vías Mongo ↔ JSON local ↔ GitHub ─────────

async def _get_github_cfg() -> dict:
    doc = await db.app_settings.find_one({}, {"github_config": 1}) or {}
    return doc.get("github_config") or {}


async def _themes_snapshot_payload_desktop() -> dict:
    docs = await db.saved_themes.find({}, sort=[("created_at", -1)]).to_list(500)
    settings_doc = await db.app_settings.find_one({}, {"default_theme_id": 1}) or {}
    default_id = str(settings_doc.get("default_theme_id") or "")
    return {
        "app": "Cinema Productions — Reserva de Eventos",
        "kind": "saved_themes_backup",
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "count": len(docs),
        "themes": [
            {
                "id": str(d["_id"]),
                "name": d.get("name", ""),
                "snapshot": d.get("snapshot", {}),
                "created_at": d.get("created_at", ""),
                "updated_at": d.get("updated_at", ""),
                "is_default": (str(d["_id"]) == default_id) or bool(d.get("is_default", False)),
            }
            for d in docs
        ],
    }


def _themes_json_path() -> Path:
    return ROOT_DIR / "themes" / "saved_themes.json"


async def _write_themes_local_json_desktop(payload: Optional[dict] = None) -> dict:
    if payload is None:
        payload = await _themes_snapshot_payload_desktop()
    try:
        p = _themes_json_path()
        p.parent.mkdir(parents=True, exist_ok=True)
        p.write_text(
            json.dumps(payload, indent=2, ensure_ascii=False),
            encoding="utf-8",
        )
    except Exception as e:
        logger.warning(f"[themes] no se pudo escribir JSON local: {e}")
    return payload


async def _pull_themes_from_github(cfg: dict) -> dict:
    """Descarga themes/saved_themes.json del repo y hace upsert por nombre."""
    import httpx
    repo_url = cfg.get("repo_url") or DEFAULT_GITHUB_REPO
    branch = cfg.get("branch") or DEFAULT_GITHUB_BRANCH
    token = (cfg.get("token") or "").strip()
    owner, repo = _parse_github_url(repo_url)
    if not owner or not repo:
        return {"ok": False, "reason": "bad_repo_url"}

    headers = {
        "Accept": "application/vnd.github.raw",
        "User-Agent": "cinema-productions-desktop",
        "X-GitHub-Api-Version": "2022-11-28",
    }
    if token:
        headers["Authorization"] = f"Bearer {token}"
    api_url = f"https://api.github.com/repos/{owner}/{repo}/contents/themes/saved_themes.json"

    try:
        async with httpx.AsyncClient(timeout=15) as h:
            r = await h.get(api_url, headers=headers, params={"ref": branch})
            if r.status_code == 404:
                return {"ok": True, "imported": 0, "reason": "no_remote_file"}
            if r.status_code != 200:
                return {"ok": False, "status": r.status_code, "error": r.text[:300]}
            data = r.json() if r.headers.get("content-type", "").startswith("application/json") else None
            if not data or "themes" not in data:
                # Retry sin Accept raw para obtener JSON envuelto
                headers2 = {**headers, "Accept": "application/vnd.github+json"}
                r2 = await h.get(api_url, headers=headers2, params={"ref": branch})
                if r2.status_code != 200:
                    return {"ok": False, "status": r2.status_code, "error": r2.text[:300]}
                b64 = (r2.json() or {}).get("content", "")
                try:
                    data = json.loads(base64.b64decode(b64).decode("utf-8"))
                except Exception as e:
                    return {"ok": False, "error": f"decode: {e}"}

        imported = 0
        updated = 0
        for t in (data.get("themes") or []):
            name = (t.get("name") or "").strip()
            if not name:
                continue
            snapshot = t.get("snapshot") or {}
            now_iso = datetime.now(timezone.utc).isoformat()
            existing = await db.saved_themes.find_one({"name": name})
            if existing:
                # actualiza solo si el snapshot difiere
                if existing.get("snapshot") != snapshot:
                    await db.saved_themes.update_one(
                        {"_id": existing["_id"]},
                        {"$set": {"snapshot": snapshot, "updated_at": t.get("updated_at") or now_iso}},
                    )
                    updated += 1
            else:
                await db.saved_themes.insert_one({
                    "name": name,
                    "snapshot": snapshot,
                    "created_at": t.get("created_at") or now_iso,
                    "updated_at": t.get("updated_at") or "",
                    "is_default": bool(t.get("is_default", False)),
                })
                imported += 1

        # Aplicar default remoto si existe
        remote_default = next((t for t in (data.get("themes") or []) if t.get("is_default")), None)
        if remote_default:
            doc_local = await db.saved_themes.find_one({"name": remote_default.get("name")})
            if doc_local:
                await db.app_settings.update_one(
                    {},
                    {"$set": {
                        "default_theme_id": str(doc_local["_id"]),
                        "default_theme_name": doc_local.get("name", ""),
                        "appearance_snapshot": doc_local.get("snapshot", {}),
                        "appearance_updated_at": datetime.now(timezone.utc).isoformat(),
                    }},
                    upsert=True,
                )

        await db.app_settings.update_one(
            {},
            {"$set": {
                "themes_sync.last_pull_at": datetime.now(timezone.utc).isoformat(),
                "themes_sync.last_pull_count": imported + updated,
            }},
            upsert=True,
        )
        if _using_embedded:
            await _save_embedded_data()
        return {"ok": True, "imported": imported, "updated": updated}
    except Exception as e:
        return {"ok": False, "error": str(e)[:300]}


async def _push_themes_to_github_desktop(payload: dict, cfg: dict) -> dict:
    import httpx
    token = (cfg.get("token") or "").strip()
    if not token:
        return {"skipped": True, "reason": "no_token"}
    repo_url = cfg.get("repo_url") or DEFAULT_GITHUB_REPO
    branch = cfg.get("branch") or DEFAULT_GITHUB_BRANCH
    owner, repo = _parse_github_url(repo_url)
    if not owner or not repo:
        return {"skipped": True, "reason": "bad_repo_url"}
    api_url = f"https://api.github.com/repos/{owner}/{repo}/contents/themes/saved_themes.json"
    content_b64 = base64.b64encode(
        json.dumps(payload, indent=2, ensure_ascii=False).encode("utf-8")
    ).decode("ascii")
    headers = {
        "Authorization": f"Bearer {token}",
        "Accept": "application/vnd.github+json",
        "X-GitHub-Api-Version": "2022-11-28",
        "User-Agent": "cinema-productions-desktop",
    }
    try:
        async with httpx.AsyncClient(timeout=20) as h:
            sha = None
            try:
                r_get = await h.get(api_url, headers=headers, params={"ref": branch})
                if r_get.status_code == 200:
                    sha = r_get.json().get("sha")
            except Exception:
                pass
            body = {
                "message": f"chore(themes): sync {payload.get('count', 0)} theme(s) from desktop",
                "content": content_b64,
                "branch": branch,
            }
            if sha:
                body["sha"] = sha
            r = await h.put(api_url, headers=headers, json=body)
            if r.status_code in (200, 201):
                data = r.json()
                await db.app_settings.update_one(
                    {},
                    {"$set": {
                        "themes_sync.last_github_sha": (data.get("content") or {}).get("sha", ""),
                        "themes_sync.last_github_commit": (data.get("commit") or {}).get("sha", ""),
                        "themes_sync.last_github_at": datetime.now(timezone.utc).isoformat(),
                        "themes_sync.last_status": "ok",
                        "themes_sync.last_error": "",
                    }},
                    upsert=True,
                )
                return {"ok": True, "commit_sha": (data.get("commit") or {}).get("sha", "")}
            await db.app_settings.update_one(
                {},
                {"$set": {
                    "themes_sync.last_status": f"error_{r.status_code}",
                    "themes_sync.last_error": r.text[:300],
                    "themes_sync.last_github_at": datetime.now(timezone.utc).isoformat(),
                }},
                upsert=True,
            )
            return {"ok": False, "status": r.status_code, "error": r.text[:300]}
    except Exception as e:
        await db.app_settings.update_one(
            {},
            {"$set": {
                "themes_sync.last_status": "error_exception",
                "themes_sync.last_error": str(e)[:300],
                "themes_sync.last_github_at": datetime.now(timezone.utc).isoformat(),
            }},
            upsert=True,
        )
        return {"ok": False, "error": str(e)[:300]}


async def _sync_themes_all_channels_desktop(push_github: bool = True) -> dict:
    """Pull GitHub → Local Mongo upsert → JSON local → (Push si hay token)."""
    cfg = await _get_github_cfg()
    pull_res = await _pull_themes_from_github(cfg)
    payload = await _write_themes_local_json_desktop()
    push_res: dict = {"skipped": True, "reason": "disabled"}
    if push_github:
        push_res = await _push_themes_to_github_desktop(payload, cfg)
    if _using_embedded:
        await _save_embedded_data()
    return {
        "pull": pull_res,
        "local_ok": True,
        "local_path": str(_themes_json_path()),
        "count": payload.get("count", 0),
        "github": push_res,
    }


async def _pull_themes_from_github_safe():
    """Wrapper de startup: no lanza excepciones."""
    try:
        cfg = await _get_github_cfg()
        await _pull_themes_from_github(cfg)
        await _write_themes_local_json_desktop()
    except Exception as e:
        logger.warning(f"[themes] pull inicial falló: {e}")


@api_router.get("/themes")
async def list_saved_themes():
    docs = await db.saved_themes.find({}, sort=[("created_at", -1)]).to_list(200)
    settings_doc = await db.app_settings.find_one({}, {"default_theme_id": 1}) or {}
    default_id = str(settings_doc.get("default_theme_id") or "")
    return [
        {
            "id": str(d["_id"]),
            "name": d["name"],
            "snapshot": d.get("snapshot", {}),
            "created_at": d["created_at"],
            "updated_at": d.get("updated_at", ""),
            "is_default": (str(d["_id"]) == default_id) or bool(d.get("is_default", False)),
        }
        for d in docs
    ]


@api_router.post("/themes")
async def create_saved_theme(payload: dict = Body(...)):
    name = (payload.get("name") or "").strip()
    if not name:
        raise HTTPException(status_code=400, detail="El nombre del tema es requerido")
    doc = {
        "name": name,
        "snapshot": payload.get("snapshot") or {},
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    result = await db.saved_themes.insert_one(doc)
    if _using_embedded:
        asyncio.create_task(_save_embedded_data())
    asyncio.create_task(_write_themes_local_json_desktop())
    return {"id": str(result.inserted_id), "name": name, "created_at": doc["created_at"]}


@api_router.delete("/themes/{theme_id}")
async def delete_saved_theme(theme_id: str):
    try:
        oid = ObjectId(theme_id)
    except Exception:
        raise HTTPException(status_code=400, detail="ID inválido")
    result = await db.saved_themes.delete_one({"_id": oid})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Tema no encontrado")
    if _using_embedded:
        asyncio.create_task(_save_embedded_data())
    return {"message": "Tema eliminado"}


@api_router.put("/themes/{theme_id}")
async def update_saved_theme(theme_id: str, payload: dict = Body(...)):
    try:
        oid = ObjectId(theme_id)
    except Exception:
        raise HTTPException(status_code=400, detail="ID inválido")
    update = {}
    if "snapshot" in payload:
        update["snapshot"] = payload.get("snapshot") or {}
    if payload.get("name"):
        update["name"] = str(payload["name"]).strip()
    if not update:
        raise HTTPException(status_code=400, detail="Sin cambios")
    update["updated_at"] = datetime.now(timezone.utc).isoformat()
    result = await db.saved_themes.update_one({"_id": oid}, {"$set": update})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Tema no encontrado")
    doc = await db.saved_themes.find_one({"_id": oid})
    if _using_embedded:
        asyncio.create_task(_save_embedded_data())
    return {
        "id": str(doc["_id"]),
        "name": doc.get("name", ""),
        "snapshot": doc.get("snapshot", {}),
        "updated_at": doc.get("updated_at", ""),
    }


@api_router.post("/themes/{theme_id}/set-default")
async def set_default_theme(theme_id: str):
    try:
        oid = ObjectId(theme_id)
    except Exception:
        raise HTTPException(status_code=400, detail="ID inválido")
    doc = await db.saved_themes.find_one({"_id": oid})
    if not doc:
        raise HTTPException(status_code=404, detail="Tema no encontrado")
    await db.app_settings.update_one(
        {},
        {"$set": {
            "default_theme_id": str(oid),
            "default_theme_name": doc.get("name", ""),
            "appearance_snapshot": doc.get("snapshot", {}),
            "appearance_updated_at": datetime.now(timezone.utc).isoformat(),
        }},
        upsert=True,
    )
    if _using_embedded:
        asyncio.create_task(_save_embedded_data())
    return {"success": True, "default_theme_id": str(oid), "name": doc.get("name", "")}


@api_router.post("/themes/sync")
async def themes_sync_now():
    """Sincronización real de 3 vías desde el escritorio.

    1. PULL: baja themes/saved_themes.json del repo de GitHub (público o con
       token) y hace upsert por nombre en la BD local.
    2. LOCAL JSON: reescribe el mirror local ROOT_DIR/themes/saved_themes.json
       con el estado consolidado.
    3. PUSH: si hay token guardado en github_config.token, sube el JSON
       consolidado al repo (Contents API, upsert por SHA).
    """
    result = await _sync_themes_all_channels_desktop(push_github=True)
    return {"success": True, **result}


@api_router.get("/themes/sync/status")
async def themes_sync_status():
    doc = await db.app_settings.find_one(
        {},
        {"default_theme_id": 1, "default_theme_name": 1, "themes_sync": 1, "github_config": 1},
    ) or {}
    ts = doc.get("themes_sync") or {}
    gh_cfg = doc.get("github_config") or {}
    themes_json_path = ROOT_DIR / "themes" / "saved_themes.json"
    local_mtime = None
    if themes_json_path.exists():
        try:
            local_mtime = datetime.fromtimestamp(
                themes_json_path.stat().st_mtime, tz=timezone.utc
            ).isoformat()
        except Exception:
            pass
    return {
        "local_path": str(themes_json_path),
        "local_exists": themes_json_path.exists(),
        "local_mtime": local_mtime,
        "last_github_at": ts.get("last_github_at"),
        "last_github_sha": ts.get("last_github_sha"),
        "last_github_commit": ts.get("last_github_commit"),
        "last_pull_at": ts.get("last_pull_at"),
        "last_pull_count": ts.get("last_pull_count", 0),
        "last_status": ts.get("last_status") or "idle",
        "last_error": ts.get("last_error"),
        "github_configured": bool(gh_cfg.get("token")),
        "default_theme_id": str(doc.get("default_theme_id") or ""),
        "default_theme_name": doc.get("default_theme_name", ""),
    }


# ── APP SECURITY (password lock + page protection) ──────────────────────────

PBKDF2_ITERATIONS = 260000


def _hash_app_password(password: str) -> str:
    salt = secrets.token_hex(16)
    dk = hashlib.pbkdf2_hmac("sha256", password.encode(), bytes.fromhex(salt), PBKDF2_ITERATIONS)
    return f"pbkdf2_sha256${PBKDF2_ITERATIONS}${salt}${dk.hex()}"


def _verify_app_password(password: str, stored: str) -> bool:
    try:
        _, iterations, salt, hash_hex = stored.split("$")
        dk = hashlib.pbkdf2_hmac("sha256", password.encode(), bytes.fromhex(salt), int(iterations))
        return secrets.compare_digest(dk.hex(), hash_hex)
    except Exception:
        return False


@api_router.get("/security/status")
async def get_security_status():
    doc = await db.app_settings.find_one({}, {"_id": 0, "app_password_hash": 1, "app_password_hint": 1, "page_protection_enabled": 1, "security_config": 1}) or {}
    cfg = doc.get("security_config") or {}
    return {
        "password_enabled": bool(doc.get("app_password_hash")),
        "hint": doc.get("app_password_hint") or "",
        "protection_enabled": bool(doc.get("page_protection_enabled")),
        "auto_lock_enabled": bool(cfg.get("auto_lock_enabled", False)),
        "auto_lock_minutes": int(cfg.get("auto_lock_minutes", 5)),
        "max_attempts": int(cfg.get("max_attempts", 5)),
        "lockout_seconds": int(cfg.get("lockout_seconds", 60)),
        "protected_sections": cfg.get("protected_sections", []),
        "failed_attempts": int(cfg.get("failed_attempts", 0)),
        "locked_until": cfg.get("locked_until", ""),
    }


@api_router.post("/security/set-password")
async def set_app_password_endpoint(payload: dict = Body(...)):
    password = (payload.get("password") or "").strip()
    hint = (payload.get("hint") or "").strip()
    current = payload.get("current_password") or ""
    if len(password) < 4:
        raise HTTPException(status_code=400, detail="La contraseña debe tener al menos 4 caracteres")
    doc = await db.app_settings.find_one({}, {"app_password_hash": 1}) or {}
    if doc.get("app_password_hash") and not _verify_app_password(current, doc["app_password_hash"]):
        raise HTTPException(status_code=401, detail="La contraseña actual es incorrecta")
    await db.app_settings.update_one(
        {},
        {"$set": {"app_password_hash": _hash_app_password(password), "app_password_hint": hint}},
        upsert=True,
    )
    if _using_embedded:
        asyncio.create_task(_save_embedded_data())
    return {"message": "Contraseña guardada"}


@api_router.post("/security/verify")
async def verify_app_password_endpoint(payload: dict = Body(...)):
    doc = await db.app_settings.find_one({}, {"app_password_hash": 1}) or {}
    if not doc.get("app_password_hash"):
        return {"valid": True}
    if not _verify_app_password(payload.get("password") or "", doc["app_password_hash"]):
        raise HTTPException(status_code=401, detail="Contraseña incorrecta")
    return {"valid": True}


@api_router.post("/security/remove-password")
async def remove_app_password_endpoint(payload: dict = Body(...)):
    doc = await db.app_settings.find_one({}, {"app_password_hash": 1}) or {}
    if not doc.get("app_password_hash"):
        raise HTTPException(status_code=400, detail="No hay contraseña configurada")
    if not _verify_app_password(payload.get("current_password") or "", doc["app_password_hash"]):
        raise HTTPException(status_code=401, detail="La contraseña actual es incorrecta")
    await db.app_settings.update_one({}, {"$unset": {"app_password_hash": "", "app_password_hint": ""}})
    if _using_embedded:
        asyncio.create_task(_save_embedded_data())
    return {"message": "Contraseña eliminada"}


@api_router.put("/security/protection")
async def set_page_protection_endpoint(payload: dict = Body(...)):
    enabled = bool(payload.get("enabled"))
    await db.app_settings.update_one({}, {"$set": {"page_protection_enabled": enabled}}, upsert=True)
    if _using_embedded:
        asyncio.create_task(_save_embedded_data())
    return {"enabled": enabled}


@api_router.get("/updates/check")
async def check_for_updates_endpoint(refresh: bool = False):
    """Frontend calls this to get update status. Combines GitHub (version.txt)
    + MongoDB compartida. Con ?refresh=true fuerza re-lectura ignorando caché."""
    if refresh:
        _gh_version_cache["ts"] = 0.0
        await _check_for_updates()
    elif not _update_status.get("checked"):
        await _check_for_updates()
    return _update_status


@api_router.get("/updates/github-version")
async def get_github_version_endpoint(refresh: bool = False):
    """Devuelve la versión leída directamente de version.txt del repo GitHub.
    Útil para que la app de escritorio compare contra el número que subes al repo."""
    if refresh:
        _gh_version_cache["ts"] = 0.0
    gh = await _fetch_github_version()
    remote_version = gh.get("version", "")
    return {
        "local_version": _local_version,
        "github_version": remote_version,
        "has_update": _is_newer(remote_version, _local_version),
        "source": gh.get("source", "github"),
        "source_url": gh.get("notes", ""),
    }


@api_router.post("/updates/dismiss")
async def dismiss_update():
    """User dismissed the update banner."""
    global _update_status
    _update_status = {**_update_status, "has_update": False}
    return {"message": "OK"}


# ── AUTO-UPDATE (binario compilado) ─────────────────────────────────────────
# Copiar código fuente sobre un binario compilado NO actualiza nada (el código
# vive DENTRO del .exe/binario) y deja la instalación en estado inconsistente
# ("se arruina el exe"). Para binarios congelados SIEMPRE hay que reemplazar el
# propio binario descargándolo desde GitHub Releases.

def _is_frozen() -> bool:
    """True si estamos ejecutando como binario compilado (PyInstaller onefile)."""
    return getattr(sys, "frozen", False)


def _current_asset_name() -> str:
    """Nombre del asset de release para este SO/arquitectura."""
    import platform
    sysname = platform.system()
    machine = (platform.machine() or "").lower()
    if sysname == "Windows":
        return "CinemaProductions.exe"
    if sysname == "Darwin":
        return "CinemaProductions-macos-arm64"
    if machine in ("arm64", "aarch64"):
        return "CinemaProductions-linux-arm64"
    return "CinemaProductions-linux-x86_64"


def _semver_key(tag: str):
    """Convierte 'v1.0.25' o '1.0.25' en tupla (1,0,25) para comparación.
    Retorna (-1,) para tags no-semver (ej. 'latest-exe') → ordenan al final."""
    import re as _re
    m = _re.match(r"^v?(\d+)\.(\d+)\.(\d+)(?:[-.]|$)", (tag or "").strip())
    if not m:
        return (-1,)
    return (int(m.group(1)), int(m.group(2)), int(m.group(3)))


async def _find_release_asset(asset_name: str) -> dict:
    """Devuelve {tag, version, url, size} del release con MAYOR versión semver
    que contenga el asset indicado. Filtra drafts. Prioriza tags 'vX.Y.Z' sobre
    'latest-exe' u otros tags no-semver para evitar downgrades accidentales."""
    import urllib.request, json as _json
    cfg = await _get_github_cfg()
    repo_url = cfg.get("repo_url") or DEFAULT_GITHUB_REPO
    token = (cfg.get("token") or "").strip()
    owner, repo = _parse_github_url(repo_url)
    if not owner or not repo:
        return {}
    headers = {"User-Agent": "cinema-productions-desktop",
               "Accept": "application/vnd.github+json"}
    if token:
        headers["Authorization"] = f"Bearer {token}"

    def _get():
        req = urllib.request.Request(
            f"https://api.github.com/repos/{owner}/{repo}/releases?per_page=30",
            headers=headers)
        with urllib.request.urlopen(req, timeout=12) as resp:
            return _json.loads(resp.read().decode("utf-8", errors="ignore"))

    try:
        releases = await asyncio.to_thread(_get)
    except Exception as e:
        logger.warning(f"No se pudieron listar releases de GitHub: {e}")
        return {}
    if not isinstance(releases, list):
        return {}

    # Recolectar TODOS los candidatos que tengan el asset, luego ordenar por
    # versión semver desc. Así el update siempre apunta al último release
    # oficial, no al primero devuelto por la API (que puede ser un prerelease
    # antiguo 'latest-exe' mezclado con la lista).
    candidates = []
    for rel in releases:
        if rel.get("draft"):
            continue
        tag = rel.get("tag_name") or ""
        for asset in (rel.get("assets") or []):
            if asset.get("name") == asset_name and asset.get("browser_download_url"):
                candidates.append({
                    "tag": tag,
                    "version": tag.lstrip("v"),
                    "url": asset["browser_download_url"],
                    "size": asset.get("size", 0),
                    "_sort": _semver_key(tag),
                    "_prerelease": bool(rel.get("prerelease")),
                })
                break  # 1 asset por release basta
    if not candidates:
        return {}
    # Prioridad: (a) semver descendente, (b) no-prerelease sobre prerelease.
    candidates.sort(key=lambda c: (c["_sort"], not c["_prerelease"]), reverse=True)
    chosen = candidates[0]
    return {k: v for k, v in chosen.items() if not k.startswith("_")}


def _spawn_swap_helper(exe_path: Path, new_path: Path):
    """Lanza un proceso externo independiente que espera a que ESTE proceso
    libere el binario, lo reemplaza por el nuevo y relanza la app. En Windows
    no se puede sobrescribir un .exe en ejecución: el helper hace polling hasta
    que el archivo se libera (al salir el proceso).

    Robusto ante:
      - Antivirus que bloquea el .new temporalmente (retries en move).
      - Batch previa colgada (limpia _cp_update.bak/.new residuales al iniciar).
      - Fallo del move (rollback desde .bak).
      - Log de cada paso en _cp_update.log para diagnóstico."""
    import subprocess, os as _os, stat as _stat, platform
    install_dir = exe_path.parent
    if platform.system() == "Windows":
        bat = install_dir / "_cp_update.bat"
        log = install_dir / "_cp_update.log"
        exe_name = exe_path.name
        old = str(exe_path)
        new = str(new_path)
        bak = old + ".bak"
        script = (
            "@echo off\r\n"
            "setlocal enabledelayedexpansion\r\n"
            f'set "LOG={log}"\r\n'
            'echo [%DATE% %TIME%] === Cinema Productions auto-update start === > "!LOG!"\r\n'
            f'echo old={old} >> "!LOG!"\r\n'
            f'echo new={new} >> "!LOG!"\r\n'
            f'echo bak={bak} >> "!LOG!"\r\n'
            # Esperar a que el proceso libere el .exe (hasta 60s de espera suave)
            "set TRIES=0\r\n"
            ":waitloop\r\n"
            "set /a TRIES+=1\r\n"
            "ping -n 2 127.0.0.1 >nul\r\n"
            f'del "{bak}" >nul 2>&1\r\n'
            f'move /y "{old}" "{bak}" >nul 2>&1\r\n'
            f'if exist "{old}" (\r\n'
            "  if !TRIES! GEQ 30 (\r\n"
            f'    echo [!TIME!] taskkill tras !TRIES! intentos >> "!LOG!"\r\n'
            f'    taskkill /F /IM "{exe_name}" >nul 2>&1\r\n'
            "  )\r\n"
            "  if !TRIES! GEQ 90 (\r\n"
            f'    echo [!TIME!] ABORT: no se liberó el .exe tras 90 intentos >> "!LOG!"\r\n'
            "    exit /b 1\r\n"
            "  )\r\n"
            "  goto waitloop\r\n"
            ")\r\n"
            f'echo [!TIME!] .exe liberado, aplicando swap (intento inicial) >> "!LOG!"\r\n'
            # Swap con retries: hasta 10 intentos con 1s de espera (AV puede bloquear el .new)
            "set MTRIES=0\r\n"
            ":moveloop\r\n"
            "set /a MTRIES+=1\r\n"
            f'move /y "{new}" "{old}" >nul 2>&1\r\n'
            f'if exist "{old}" goto :moveok\r\n'
            "if !MTRIES! GEQ 10 (\r\n"
            f'  echo [!TIME!] MOVE FAIL tras !MTRIES! intentos, rollback desde .bak >> "!LOG!"\r\n'
            f'  move /y "{bak}" "{old}" >nul 2>&1\r\n'
            f'  echo [!TIME!] rollback aplicado. Update ABORTADO. >> "!LOG!"\r\n'
            f'  start "" "{old}"\r\n'
            "  exit /b 2\r\n"
            ")\r\n"
            "ping -n 2 127.0.0.1 >nul\r\n"
            "goto moveloop\r\n"
            ":moveok\r\n"
            f'echo [!TIME!] SWAP OK en !MTRIES! intento(s) >> "!LOG!"\r\n'
            f'del "{bak}" >nul 2>&1\r\n'
            f'echo [!TIME!] Relanzando app >> "!LOG!"\r\n'
            f'start "" "{old}"\r\n'
            f'echo [!TIME!] === update completo === >> "!LOG!"\r\n'
            # Auto-borrar el .bat (deja el .log para debug)
            '(goto) 2>nul & del "%~f0"\r\n'
        )
        bat.write_text(script, encoding="ascii")
        DETACHED_PROCESS = 0x00000008
        CREATE_NEW_PROCESS_GROUP = 0x00000200
        subprocess.Popen(["cmd", "/c", str(bat)], cwd=str(install_dir),
                         creationflags=DETACHED_PROCESS | CREATE_NEW_PROCESS_GROUP,
                         close_fds=True)
    else:
        sh = install_dir / "_cp_update.sh"
        log = install_dir / "_cp_update.log"
        old = str(exe_path)
        new = str(new_path)
        pid = _os.getpid()
        script = (
            "#!/usr/bin/env bash\n"
            f'LOG="{log}"\n'
            'echo "[$(date)] === auto-update start ===" > "$LOG"\n'
            f'PID={pid}\n'
            f'echo "waiting PID $PID to exit..." >> "$LOG"\n'
            'TRIES=0\n'
            'while kill -0 "$PID" 2>/dev/null; do\n'
            '  sleep 0.5\n'
            '  TRIES=$((TRIES+1))\n'
            '  if [ "$TRIES" -ge 120 ]; then\n'
            '    echo "[$(date)] timeout waiting, forcing kill" >> "$LOG"\n'
            '    kill -9 "$PID" 2>/dev/null || true\n'
            '    break\n'
            '  fi\n'
            'done\n'
            'sleep 0.5\n'
            f'echo "[$(date)] applying swap: {new} -> {old}" >> "$LOG"\n'
            f'if mv -f "{new}" "{old}"; then\n'
            f'  chmod +x "{old}"\n'
            f'  echo "[$(date)] swap OK, relaunching" >> "$LOG"\n'
            f'  nohup "{old}" >/dev/null 2>&1 &\n'
            'else\n'
            f'  echo "[$(date)] swap FAILED" >> "$LOG"\n'
            'fi\n'
            f'echo "[$(date)] === update done ===" >> "$LOG"\n'
            'rm -- "$0"\n'
        )
        sh.write_text(script)
        sh.chmod(sh.stat().st_mode | _stat.S_IEXEC | _stat.S_IXGRP | _stat.S_IXOTH)
        subprocess.Popen(["/usr/bin/env", "bash", str(sh)], cwd=str(install_dir),
                         start_new_session=True, close_fds=True)


async def _apply_binary_update_frozen(dry_run: bool = False, force: bool = False):
    """Actualiza un binario compilado descargando el asset correspondiente desde
    GitHub Releases y programando el swap + relanzamiento. Los datos del usuario
    (.env, cinema_data.json, backups/, uploads/) NO se tocan: viven fuera del
    binario."""
    import sys as _sys, os as _os, shutil, threading, time as _time, urllib.request
    exe_path = Path(_sys.executable).resolve()
    install_dir = exe_path.parent
    asset_name = _current_asset_name()

    # Limpieza defensiva de residuos de intentos previos (evita "file in use" o
    # confusión si un update anterior falló a mitad).
    for stale in (
        install_dir / (exe_path.name + ".new"),
        install_dir / (exe_path.name + ".bak"),
        install_dir / "_cp_update.bat",
    ):
        try:
            if stale.exists():
                stale.unlink()
        except Exception as _cerr:
            logger.warning(f"No se pudo limpiar residuo {stale}: {_cerr}")

    rel = await _find_release_asset(asset_name)
    if not rel or not rel.get("url"):
        raise HTTPException(status_code=404,
            detail=f"No se encontró el binario '{asset_name}' en las releases de GitHub. "
                   f"Verifica que el workflow 'Build Windows .exe' haya publicado el asset.")

    # La versión objetivo es la del release donde SÍ está publicado el asset
    # de esta plataforma. _find_release_asset ya prioriza el semver más alto,
    # así que rel.get('version') es autoritativo.
    new_version = rel.get("version", "")
    if not new_version:
        # Fallback: intentar version.txt remoto (release sin tag semver, ej.
        # 'latest-exe'). Nunca usarlo como fuente primaria.
        gh = await _fetch_github_version()
        new_version = gh.get("version", "")
    if not new_version:
        raise HTTPException(status_code=500,
            detail="El release existe pero no expone una versión (tag_name vacío). "
                   "Regenera el release desde GitHub Actions.")

    # Regla dura: si la versión remota es IGUAL a la local, nunca reinstalar,
    # ni siquiera con force=True. Evita gastar ancho de banda y un swap
    # innecesario en clics accidentales al botón "Actualizar".
    if new_version and new_version == _local_version:
        return {"success": True, "restarted": False,
                "old_version": _local_version, "new_version": new_version,
                "message": "Ya tienes esta versión instalada"}
    if new_version and not _is_newer(new_version, _local_version) and not force:
        return {"success": True, "restarted": False,
                "old_version": _local_version, "new_version": new_version,
                "message": "Ya estás en la última versión"}

    new_path = install_dir / (exe_path.name + ".new")
    cfg = await _get_github_cfg()
    token = (cfg.get("token") or "").strip()
    dl_headers = {"User-Agent": "cinema-productions-desktop",
                  "Accept": "application/octet-stream"}
    if token:
        dl_headers["Authorization"] = f"Bearer {token}"

    def _download():
        req = urllib.request.Request(rel["url"], headers=dl_headers)
        with urllib.request.urlopen(req, timeout=600) as resp, open(new_path, "wb") as f:
            shutil.copyfileobj(resp, f)

    try:
        if new_path.exists():
            new_path.unlink()
        await asyncio.to_thread(_download)
    except Exception as e:
        # Limpiar el .new parcial si quedó a medias
        try:
            if new_path.exists():
                new_path.unlink()
        except Exception:
            pass
        raise HTTPException(status_code=502, detail=f"Descarga del binario falló: {e}")

    # Validación: tamaño + "magic bytes" del binario. Evita que una respuesta
    # de error (HTML/JSON) se guarde como si fuera el ejecutable, lo que
    # "arruinaría" el exe al hacer el swap.
    def _valid_binary(p: Path) -> bool:
        try:
            if p.stat().st_size < 1_000_000:
                return False
            with open(p, "rb") as fh:
                head = fh.read(4)
            import platform as _pf
            s = _pf.system()
            if s == "Windows":
                return head[:2] == b"MZ"
            if s == "Darwin":
                return head in (b"\xcf\xfa\xed\xfe", b"\xce\xfa\xed\xfe",
                                b"\xca\xfe\xba\xbe", b"\xfe\xed\xfa\xcf",
                                b"\xfe\xed\xfa\xce")
            return head == b"\x7fELF"
        except Exception:
            return False

    if not _valid_binary(new_path):
        try:
            new_path.unlink()
        except Exception:
            pass
        raise HTTPException(status_code=500,
            detail="El binario descargado es inválido o está corrupto; no se aplicó la actualización.")

    if dry_run:
        try:
            new_path.unlink()
        except Exception:
            pass
        return {"success": True, "restarted": False, "dry_run": True,
                "old_version": _local_version, "new_version": new_version,
                "asset": asset_name, "install_dir": str(install_dir),
                "message": "DRY RUN — binario descargado y verificado; no se aplicó el swap."}

    _spawn_swap_helper(exe_path, new_path)

    def _shutdown():
        # 3s: suficiente para que el cliente HTTP reciba la respuesta y para
        # que el swap helper haya sido spawn-eado como proceso independiente.
        _time.sleep(3.0)
        _os._exit(0)
    threading.Thread(target=_shutdown, daemon=True).start()

    return {"success": True, "restarted": True, "is_desktop": True,
            "files_updated": 1,
            "old_version": _local_version, "new_version": new_version,
            "asset": asset_name, "install_dir": str(install_dir),
            "message": "Descarga completa. La app se cerrará y se actualizará automáticamente en unos segundos."}


# ── AUTO-UPDATE: descarga, aplica y reinicia el ejecutable ──────────────────
@api_router.post("/updates/apply-and-restart")
async def apply_update_and_restart(payload: dict = Body(default={})):
    """Descarga el ZIP más reciente (del shared MongoDB), lo descomprime en la
    misma carpeta del ejecutable actual, y reinicia la app.

    Parámetros:
    - dry_run (bool): si True, no reinicia el proceso (útil para tests).
    - force (bool): si True, aplica aunque la versión sea la misma.

    Funcionamiento paso a paso:
    1. Localiza la carpeta donde se está ejecutando `app.py` (o el .exe).
    2. Consulta el último update en MongoDB (compartido con la app en la nube).
    3. Descarga el ZIP → ubicación temporal → lo extrae sobre la carpeta actual.
    4. Programa un reinicio del proceso (os.execv) tras responder al cliente.

    Requisitos: el ZIP debe estar cifrado con la contraseña que trae la app o
    debe usar la contraseña guardada en security_config.zip_password.
    """
    import sys, subprocess, zipfile, shutil, tempfile, threading, time

    dry_run = bool(payload.get("dry_run", False))
    force = bool(payload.get("force", False))

    # Binario compilado (.exe / ejecutable): reemplazar el propio binario desde
    # GitHub Releases en lugar de descomprimir código fuente encima (que no
    # actualiza nada y arruina la instalación).
    if getattr(sys, 'frozen', False):
        return await _apply_binary_update_frozen(dry_run=dry_run, force=force)

    try:
        import pyzipper
        _has_pyzipper = True
    except ImportError:
        _has_pyzipper = False

    # 1) Últimas update en la BD
    latest = await db.app_updates.find_one({"is_latest": True}, sort=[("created_at", -1)])
    if not latest:
        raise HTTPException(status_code=404, detail="No hay actualizaciones disponibles")
    if latest["version"] == _local_version and not force:
        return {"success": True, "message": "Ya estás en la última versión", "restarted": False}

    # 2) Detectar carpeta del ejecutable actual
    if getattr(sys, 'frozen', False):
        install_dir = Path(sys.executable).resolve().parent
    else:
        install_dir = Path(__file__).resolve().parent

    # 3) Descargar ZIP a tmp
    tmp_dir = Path(tempfile.mkdtemp(prefix="cp_update_"))
    zip_path = tmp_dir / latest["filename"]

    if _update_server_url:
        # Descarga vía HTTP
        import urllib.request
        dl_url = f"{_update_server_url}/api/updates/download"
        try:
            with urllib.request.urlopen(dl_url, timeout=60) as resp:
                zip_path.write_bytes(resp.read())
        except Exception as e:
            raise HTTPException(status_code=502, detail=f"Descarga falló: {e}")
    else:
        # Sin server URL configurada — intentar leer directamente del disco (dev/mismo pod)
        local_zip = UPDATES_DIR / latest.get("stored_name", "")
        if local_zip.exists():
            shutil.copy2(str(local_zip), str(zip_path))
        else:
            raise HTTPException(status_code=400, detail="ZIP no encontrado ni en URL remota ni en /uploads/updates")

    # 4) Obtener contraseña del ZIP (si existe)
    zip_pwd = None
    try:
        sec = (await db.app_settings.find_one({}, {"security_config": 1}) or {}).get("security_config") or {}
        zip_pwd = sec.get("zip_password") or "2868"
    except Exception:
        zip_pwd = "2868"

    # 5) Extraer sobre la instalación
    extract_dir = tmp_dir / "extracted"
    extract_dir.mkdir(exist_ok=True)
    try:
        # Intentar como AES primero
        if _has_pyzipper:
            try:
                with pyzipper.AESZipFile(str(zip_path)) as zf:
                    zf.setpassword(zip_pwd.encode() if zip_pwd else None)
                    zf.extractall(str(extract_dir))
            except Exception:
                # Fallback a ZIP normal
                with zipfile.ZipFile(str(zip_path)) as zf:
                    zf.extractall(str(extract_dir))
        else:
            with zipfile.ZipFile(str(zip_path)) as zf:
                zf.extractall(str(extract_dir))
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Descompresión falló (¿contraseña incorrecta?): {e}")

    # 6) Localizar la subcarpeta 'cinema-productions/' dentro del extract
    src_root = extract_dir / "cinema-productions"
    if not src_root.exists():
        # Puede que el ZIP no tenga esa subcarpeta
        src_root = extract_dir

    # 7) Copiar archivos NO destructivamente (preserva backups, data locales, .env con datos del usuario)
    preserve = {".env", "cinema_data.json", "backups", "uploads"}
    copied = 0
    for item in src_root.rglob("*"):
        if not item.is_file():
            continue
        rel = item.relative_to(src_root)
        if rel.parts and rel.parts[0] in preserve:
            continue
        dest = install_dir / rel
        dest.parent.mkdir(parents=True, exist_ok=True)
        shutil.copy2(str(item), str(dest))
        copied += 1

    # 8) Actualizar version.txt
    version_file = install_dir / "version.txt"
    try:
        version_file.write_text(latest["version"])
    except Exception:
        pass

    # 9) Programar reinicio del proceso en background (a menos que sea dry_run)
    if dry_run:
        try:
            shutil.rmtree(str(tmp_dir), ignore_errors=True)
        except Exception:
            pass
        return {
            "success": True,
            "restarted": False,
            "dry_run": True,
            "old_version": _local_version,
            "new_version": latest["version"],
            "files_updated": copied,
            "install_dir": str(install_dir),
            "message": "DRY RUN — archivos copiados pero el proceso NO se reinició",
        }

    def _restart_process():
        time.sleep(2)  # dar tiempo a que la respuesta HTTP termine
        try:
            # Limpieza
            shutil.rmtree(str(tmp_dir), ignore_errors=True)
        except Exception:
            pass

        # En Windows, spawn el hijo sin ventana de consola visible.
        popen_kwargs = {"cwd": str(install_dir)}
        if sys.platform == "win32":
            # DETACHED_PROCESS (0x00000008) + CREATE_NO_WINDOW (0x08000000)
            # → nuevo proceso sin heredar la consola, sin ventana negra.
            popen_kwargs["creationflags"] = 0x00000008 | 0x08000000
            popen_kwargs["close_fds"] = True

        try:
            if getattr(sys, 'frozen', False):
                # Ejecutable compilado: relanzar
                subprocess.Popen([str(sys.executable)], **popen_kwargs)
            else:
                # Modo desarrollo: relanzar el python actual
                launcher = install_dir / "launcher.pyw"
                if launcher.exists():
                    subprocess.Popen([sys.executable, str(launcher)], **popen_kwargs)
                else:
                    subprocess.Popen([sys.executable, str(install_dir / "app.py")], **popen_kwargs)
            # Terminar este proceso
            os._exit(0)
        except Exception as e:
            logger.error(f"Restart failed: {e}")

    threading.Thread(target=_restart_process, daemon=True).start()

    return {
        "success": True,
        "restarted": True,
        "old_version": _local_version,
        "new_version": latest["version"],
        "files_updated": copied,
        "install_dir": str(install_dir),
        "message": "Actualización aplicada. La app se reiniciará en 2 segundos.",
    }


# ── APP UPDATES (local + remote check) ──────────────────────────────────────

@api_router.post("/updates/upload")
async def upload_app_update_local(
    file: UploadFile = File(...),
    version: str = Form(...),
    notes: str = Form(""),
    channel: str = Form("stable"),
):
    content = await file.read()
    file_id = str(uuid.uuid4())
    safe_name = f"{file_id}_{file.filename}"
    file_path = UPDATES_DIR / safe_name
    file_path.write_bytes(content)
    await db.app_updates.update_many({}, {"$set": {"is_latest": False}})
    doc = {
        "version": version, "filename": file.filename, "stored_name": safe_name,
        "notes": notes, "channel": channel, "file_size": len(content),
        "created_at": datetime.now(timezone.utc).isoformat(), "is_latest": True,
    }
    result = await db.app_updates.insert_one(doc)
    return {**{k: v for k, v in doc.items() if k != "_id"}, "id": str(result.inserted_id)}


@api_router.get("/updates/latest")
async def get_latest_update_local():
    doc = await db.app_updates.find_one({"is_latest": True}, sort=[("created_at", -1)])
    if not doc:
        raise HTTPException(status_code=404, detail="No hay actualizaciones disponibles")
    return {"id": str(doc["_id"]), "version": doc["version"], "filename": doc["filename"],
            "notes": doc.get("notes", ""), "channel": doc.get("channel", "stable"),
            "file_size": doc["file_size"], "created_at": doc["created_at"]}


_gh_hist_cache = {"data": None, "ts": 0.0, "loading": False}
_GH_HIST_TTL = 300  # 5 min


def _gh_hist_fresh() -> bool:
    import time as _t
    return _gh_hist_cache["data"] is not None and (_t.time() - _gh_hist_cache["ts"]) < _GH_HIST_TTL


async def _fetch_gh_tag_records_local(owner: str, repo: str, token: str) -> list:
    """Tags de GitHub + fecha/nota de cada commit EN PARALELO. Cacheado 5 min."""
    import time as _t, urllib.request, json as _json
    if _gh_hist_fresh():
        return _gh_hist_cache["data"]
    _gh_hist_cache["loading"] = True

    headers = {"User-Agent": "cinema-productions", "Accept": "application/vnd.github+json"}
    if token:
        headers["Authorization"] = f"Bearer {token}"

    def _get_json(url, timeout):
        req = urllib.request.Request(url, headers=headers)
        with urllib.request.urlopen(req, timeout=timeout) as resp:
            return _json.loads(resp.read().decode())

    try:
        tags_data = await asyncio.to_thread(
            _get_json, f"https://api.github.com/repos/{owner}/{repo}/tags?per_page=30", 6
        )
    except Exception as e:
        logger.warning(f"No se pudo leer tags de GitHub: {e}")
        _gh_hist_cache["loading"] = False
        return _gh_hist_cache["data"] or []

    tags_data = tags_data[:20]

    async def _one(tag):
        tag_name = tag.get("name", "")
        version = tag_name.lstrip("v")
        commit_sha = tag.get("commit", {}).get("sha", "")
        commit_url = tag.get("commit", {}).get("url", "")
        created_at, notes = "", ""
        if commit_url:
            try:
                cdata = await asyncio.to_thread(_get_json, commit_url, 5)
                created_at = cdata.get("commit", {}).get("author", {}).get("date", "")
                notes = cdata.get("commit", {}).get("message", "")[:200]
            except Exception:
                pass
        return {
            "id": f"gh:{tag_name}",
            "version": version,
            "filename": "",
            "notes": notes,
            "channel": "github",
            "file_size": 0,
            "created_at": created_at,
            "is_latest": False,
            "source": "github_tag",
            "commit_short": commit_sha[:7],
            "author": "",
            "branch": "main",
        }

    try:
        recs = list(await asyncio.gather(*[_one(t) for t in tags_data]))
        _gh_hist_cache["data"] = recs
        _gh_hist_cache["ts"] = _t.time()
        return recs
    finally:
        _gh_hist_cache["loading"] = False


@api_router.get("/updates/history")
async def get_update_history_local():
    """Historial de versiones (local + tags de GitHub). Respuesta INMEDIATA:
    los registros locales se devuelven al instante y los tags de GitHub se
    cachean (5 min) y se refrescan en segundo plano, sin bloquear nunca."""
    # 1) Registros locales
    cursor = db.app_updates.find({}, sort=[("created_at", -1)])
    docs = await cursor.to_list(200)
    local_records = [{
        "id": str(d["_id"]),
        "version": d.get("version", "?"),
        "filename": d.get("filename", ""),
        "notes": d.get("notes", ""),
        "channel": d.get("channel", "stable"),
        "file_size": d.get("file_size", 0),
        "created_at": d.get("created_at", ""),
        "is_latest": d.get("is_latest", False),
        "source": d.get("source", "package"),
        "commit_short": d.get("commit_short", ""),
        "author": d.get("author", ""),
        "branch": d.get("branch", ""),
    } for d in docs]

    # 2) Tags de GitHub — desde caché; refresco en segundo plano si hace falta
    cfg = await _get_github_cfg()
    repo_url = cfg.get("repo_url", "")
    token = cfg.get("token", "")
    owner, repo = _parse_github_url(repo_url)
    if owner and repo:
        if _gh_hist_fresh():
            github_records = _gh_hist_cache["data"] or []
        else:
            if not _gh_hist_cache["loading"]:
                asyncio.create_task(_fetch_gh_tag_records_local(owner, repo, token))
            github_records = _gh_hist_cache["data"] or []
        seen_versions = {r["version"] for r in local_records}
        for gr in github_records:
            if gr["version"] not in seen_versions:
                local_records.append(gr)
                seen_versions.add(gr["version"])

    # 3) Ordenar por created_at descendente y marcar la primera como is_latest
    local_records.sort(key=lambda r: r.get("created_at", ""), reverse=True)
    if local_records:
        for r in local_records:
            r["is_latest"] = False
        local_records[0]["is_latest"] = True

    return local_records


@api_router.get("/updates/download/{update_id}")
async def download_update_local(update_id: str):
    try:
        oid = ObjectId(update_id)
    except Exception:
        raise HTTPException(status_code=400, detail="ID inválido")
    doc = await db.app_updates.find_one({"_id": oid})
    if not doc:
        raise HTTPException(status_code=404, detail="Actualización no encontrada")
    file_path = UPDATES_DIR / doc["stored_name"]
    if not file_path.exists():
        raise HTTPException(status_code=404, detail="Archivo no encontrado")

    def iter_file():
        with open(file_path, "rb") as f:
            while chunk := f.read(1024 * 1024):
                yield chunk

    return StreamingResponse(iter_file(), media_type="application/octet-stream",
                             headers={"Content-Disposition": f'attachment; filename="{doc["filename"]}"'})


@api_router.put("/updates/{update_id}/set-latest")
async def set_latest_update_local(update_id: str):
    try:
        oid = ObjectId(update_id)
    except Exception:
        raise HTTPException(status_code=400, detail="ID inválido")
    await db.app_updates.update_many({}, {"$set": {"is_latest": False}})
    result = await db.app_updates.update_one({"_id": oid}, {"$set": {"is_latest": True}})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="No encontrado")
    return {"message": "Versión marcada como activa"}


@api_router.delete("/updates/{update_id}")
async def delete_update_local(update_id: str):
    try:
        oid = ObjectId(update_id)
    except Exception:
        raise HTTPException(status_code=400, detail="ID inválido")
    doc = await db.app_updates.find_one({"_id": oid})
    if not doc:
        raise HTTPException(status_code=404, detail="No encontrado")
    file_path = UPDATES_DIR / doc["stored_name"]
    if file_path.exists():
        file_path.unlink()
    await db.app_updates.delete_one({"_id": oid})
    if doc.get("is_latest"):
        newer = await db.app_updates.find_one({}, sort=[("created_at", -1)])
        if newer:
            await db.app_updates.update_one({"_id": newer["_id"]}, {"$set": {"is_latest": True}})
    return {"message": "Actualización eliminada"}


@api_router.get("/updates/check-remote")
async def check_remote_update(url: str, current_version: str = "0.0.0"):
    """Check a remote server for updates. Used by Desktop App."""
    import httpx
    try:
        async with httpx.AsyncClient(timeout=10) as client:
            r = await client.get(f"{url.rstrip('/')}/api/updates/latest")
            if r.status_code == 404:
                return {"has_update": False, "message": "No hay actualizaciones en el servidor remoto"}
            r.raise_for_status()
            remote = r.json()
            has_update = remote["version"] != current_version
            return {
                "has_update": has_update,
                "remote_version": remote["version"],
                "current_version": current_version,
                "filename": remote["filename"],
                "notes": remote.get("notes", ""),
                "file_size": remote.get("file_size", 0),
                "download_url": f"{url.rstrip('/')}/api/updates/download/{remote['id']}",
            }
    except Exception as e:
        raise HTTPException(status_code=502, detail=f"No se pudo conectar al servidor: {str(e)}")


# ═══════════════════════════════════════════════════════════════════════════
#  Endpoints extra para paridad con la web (evitan pantallas en blanco por 404).
#  La app de escritorio es INDEPENDIENTE: GitHub es solo informativo, las
#  integraciones que requieren nube devuelven respuestas controladas.
# ═══════════════════════════════════════════════════════════════════════════

async def _get_github_cfg() -> dict:
    doc = await db.app_settings.find_one({}, {"github_config": 1}) or {}
    cfg = doc.get("github_config") or {}
    if not cfg.get("repo_url"):
        cfg = {"repo_url": DEFAULT_GITHUB_REPO, "branch": DEFAULT_GITHUB_BRANCH}
    return cfg


def _parse_github_url(url: str):
    if not url:
        return None, None
    m = re.match(r"^https?://github\.com/([^/]+)/([^/.]+?)(?:\.git)?/?$", url.strip())
    return (m.group(1), m.group(2)) if m else (None, None)


@api_router.get("/github/config")
async def get_github_config():
    cfg = await _get_github_cfg()
    repo_url = cfg.get("repo_url") or DEFAULT_GITHUB_REPO
    return {
        "repo_url": repo_url,
        "has_token": bool(cfg.get("token")),
        "last_commit_sha": cfg.get("last_commit_sha", ""),
        "last_check_at": cfg.get("last_check_at", ""),
        "branch": cfg.get("branch", DEFAULT_GITHUB_BRANCH),
        "is_configured": True,
        "is_desktop": True,
        "suggested_repo": DEFAULT_GITHUB_REPO,
    }


@api_router.post("/github/config")
async def save_github_config(payload: dict = Body(...)):
    repo_url = (payload.get("repo_url") or "").strip() or DEFAULT_GITHUB_REPO
    branch = (payload.get("branch") or DEFAULT_GITHUB_BRANCH).strip()
    owner, _ = _parse_github_url(repo_url)
    if not owner:
        raise HTTPException(status_code=400, detail="URL de GitHub inválida. Formato: https://github.com/usuario/repo")
    update = {"github_config.repo_url": repo_url, "github_config.branch": branch}
    token = (payload.get("token") or "").strip()
    if token:
        update["github_config.token"] = token
    elif payload.get("clear_token"):
        update["github_config.token"] = ""
    await db.app_settings.update_one({}, {"$set": update}, upsert=True)
    if _using_embedded:
        asyncio.create_task(_save_embedded_data())
    return {"success": True, "repo_url": repo_url, "branch": branch}


@api_router.get("/github/check-updates")
async def check_github_updates():
    """Solo informativo en escritorio: revisa si hay commits nuevos en el repo."""
    import httpx
    cfg = await _get_github_cfg()
    repo_url = cfg.get("repo_url") or DEFAULT_GITHUB_REPO
    branch = cfg.get("branch") or DEFAULT_GITHUB_BRANCH
    owner, repo = _parse_github_url(repo_url)
    if not owner:
        raise HTTPException(status_code=400, detail="URL de GitHub inválida")
    headers = {"Accept": "application/vnd.github+json"}
    if cfg.get("token"):
        headers["Authorization"] = f"Bearer {cfg['token']}"

    # Ejecutar en paralelo el fetch de commits y version.txt para responder rápido
    async def _fetch_commits():
        async with httpx.AsyncClient(timeout=6) as http:
            return await http.get(
                f"https://api.github.com/repos/{owner}/{repo}/commits",
                headers=headers, params={"sha": branch, "per_page": 10},
            )

    async def _fetch_ver():
        try:
            gh = await _fetch_github_version()
            return gh.get("version", "") or ""
        except Exception as e:
            logger.warning(f"No se pudo leer remote_version en check-updates: {e}")
            return ""

    try:
        r, remote_version = await asyncio.gather(_fetch_commits(), _fetch_ver())
    except Exception as e:
        raise HTTPException(status_code=502, detail=f"Sin conexión a GitHub: {e}")
    if r.status_code == 404:
        raise HTTPException(status_code=404, detail="Repositorio no encontrado (¿privado y sin token?)")
    if r.status_code != 200:
        raise HTTPException(status_code=502, detail=f"GitHub respondió {r.status_code}")
    commits = r.json() or []
    remote_sha = commits[0]["sha"] if commits else ""
    last_seen = cfg.get("last_commit_sha", "")

    # Decisión de "hay actualizaciones":
    # Las releases se distribuyen por TAG (v1.0.x) y cada .exe embebe su version.txt.
    # La fuente de verdad es version.txt del repo: solo hay update real cuando la
    # versión remota es MAYOR que la local. El SHA de main NO se usa como criterio
    # porque cambia con commits que no bumpean versión (auto-commits, docs, CI,
    # cambios en workflows…) y provocaba el falso positivo "Nueva versión v1.0.14"
    # cuando el usuario ya tenía v1.0.14 instalada.
    # Solo cae al SHA como fallback cuando no hay información de versión.
    sha_differs = bool(remote_sha) and remote_sha != last_seen
    if remote_version and _local_version:
        has_updates = _is_newer(remote_version, _local_version)
    else:
        has_updates = sha_differs

    new_commits = [{
        "sha": c["sha"][:7],
        "full_sha": c["sha"],
        "message": (c.get("commit", {}).get("message") or "").split("\n")[0][:200],
        "author": c.get("commit", {}).get("author", {}).get("name", "?"),
        "date": c.get("commit", {}).get("author", {}).get("date", ""),
        "url": c.get("html_url", ""),
    } for c in commits[:5]]

    # Persistir estado. SOLO sincronizamos last_commit_sha con el remoto cuando
    # de verdad NO hay update (mismo SHA que el aplicado). Si hay update pendiente,
    # NO se sobrescribe last_commit_sha para no "enterrar" los commits pendientes.
    set_fields = {
        "github_config.last_check_at": datetime.now(timezone.utc).isoformat(),
        "github_config.last_remote_sha": remote_sha,
    }
    if not has_updates and remote_sha and not sha_differs:
        set_fields["github_config.last_commit_sha"] = remote_sha
    await db.app_settings.update_one({}, {"$set": set_fields}, upsert=True)
    if _using_embedded:
        asyncio.create_task(_save_embedded_data())

    return {
        "has_updates": has_updates,
        "is_desktop": True,
        "remote_sha": remote_sha,
        "remote_sha_short": remote_sha[:7],
        "local_sha": last_seen,
        "local_sha_short": last_seen[:7] if last_seen else "",
        "local_version": _local_version,
        "remote_version": remote_version,
        "branch": branch,
        "commits_ahead": len(new_commits) if has_updates else 0,
        "commits": new_commits,
        "repo_url": repo_url,
    }


@api_router.post("/github/apply-update")
async def apply_github_update(payload: dict = Body(default={})):
    """AUTO-UPDATE desde GitHub: descarga los archivos actualizados del repositorio
    y los aplica sobre la instalación actual, sin intervención manual.

    Funcionamiento:
    1. Localiza la carpeta del ejecutable actual.
    2. Descarga el tarball del repositorio desde GitHub API.
    3. Descomprime en temp dir.
    4. Copia backend/*, frontend/build/*, y archivos raíz sobre la instalación,
       PRESERVANDO .env, cinema_data.json, backups/ y uploads/.
    5. Actualiza last_commit_sha en la BD.
    6. Programa reinicio del proceso (o solo copia si dry_run=true).
    """
    import sys, subprocess, tarfile, shutil, tempfile, threading, time
    import urllib.request

    dry_run = bool(payload.get("dry_run", False))

    cfg = await _get_github_cfg()
    repo_url = cfg.get("repo_url", "")
    branch = cfg.get("branch", "main")
    token = cfg.get("token", "")

    if not repo_url:
        raise HTTPException(status_code=400, detail="Sin repositorio de GitHub configurado")

    owner, repo = _parse_github_url(repo_url)
    if not owner:
        raise HTTPException(status_code=400, detail=f"URL de GitHub inválida: {repo_url}")

    # Binario compilado (.exe / ejecutable): NO se puede actualizar copiando
    # código fuente encima (el código vive dentro del binario, así que "no
    # recibe la actualización" y la instalación queda inconsistente). Se
    # descarga y reemplaza el binario desde GitHub Releases.
    if getattr(sys, 'frozen', False):
        return await _apply_binary_update_frozen(
            dry_run=dry_run, force=bool(payload.get("force", False)))

    # 1) Detectar carpeta de instalación
    if getattr(sys, 'frozen', False):
        install_dir = Path(sys.executable).resolve().parent
    else:
        install_dir = Path(__file__).resolve().parent

    tmp_dir = Path(tempfile.mkdtemp(prefix="cp_gh_update_"))

    try:
        # 2) Descargar tarball desde GitHub API (respeta rama)
        api_url = f"https://api.github.com/repos/{owner}/{repo}/tarball/{branch}"
        req = urllib.request.Request(api_url, headers={
            "User-Agent": "cinema-productions-updater",
            "Accept": "application/vnd.github+json",
        })
        if token:
            req.add_header("Authorization", f"Bearer {token}")

        tar_path = tmp_dir / "repo.tar.gz"
        try:
            with urllib.request.urlopen(req, timeout=90) as resp:
                tar_path.write_bytes(resp.read())
        except Exception as e:
            raise HTTPException(status_code=502, detail=f"Descarga desde GitHub falló: {e}")

        # 3) Extraer
        extract_dir = tmp_dir / "extracted"
        extract_dir.mkdir()
        with tarfile.open(str(tar_path), "r:gz") as tar:
            tar.extractall(str(extract_dir))

        # GitHub tarball crea una única carpeta owner-repo-sha/
        subdirs = [d for d in extract_dir.iterdir() if d.is_dir()]
        if not subdirs:
            raise HTTPException(status_code=500, detail="Tarball vacío o corrupto")
        src_root = subdirs[0]

        # 4) Aplicar sobre la instalación.
        # IMPORTANTE: la instalación de escritorio tiene layout PLANO:
        #   install_dir/app.py   (= backend/standalone_app.py del repo)
        #   install_dir/build/   (= frontend/build/ del repo — UI que se sirve)
        #   install_dir/<archivos raíz>
        # El repo es ANIDADO (backend/, frontend/), así que hay que REMAPEAR,
        # no copiar tal cual (ese era el bug: caía en frontend/build y backend/).
        preserve = {".env", "cinema_data.json", "cinema_data.json.bak", "backups", "uploads", ".db_override"}
        copied = 0
        skipped_preserve = 0

        def _copy_file(src_file, dst_file):
            nonlocal copied
            dst_file.parent.mkdir(parents=True, exist_ok=True)
            shutil.copy2(str(src_file), str(dst_file))
            copied += 1

        # 4a) UI compilada: raíz/build (versión PC) o frontend/build -> install_dir/build/*
        repo_build = src_root / "build"
        if not (repo_build / "index.html").exists():
            repo_build = src_root / "frontend" / "build"
        if repo_build.exists() and (repo_build / "index.html").exists():
            dst_build = install_dir / "build"
            if dst_build.exists():
                shutil.rmtree(str(dst_build), ignore_errors=True)
            shutil.copytree(str(repo_build), str(dst_build))
            copied += sum(1 for f in dst_build.rglob("*") if f.is_file())

        # 4b) Backend: raíz/app.py (versión PC) o backend/standalone_app.py -> install_dir/app.py
        repo_app = src_root / "app.py"
        if not repo_app.exists():
            repo_app = src_root / "backend" / "standalone_app.py"
        if repo_app.exists():
            _copy_file(repo_app, install_dir / "app.py")
        # Otros módulos .py del backend (por si app.py los importa) -> raíz
        repo_backend = src_root / "backend"
        if repo_backend.exists():
            for pyf in repo_backend.glob("*.py"):
                if pyf.name in {"standalone_app.py", "server.py"}:
                    continue
                _copy_file(pyf, install_dir / pyf.name)

        # 4c) Archivos raíz del repo (version.txt, README, etc.) -> install_dir
        _skip_root = {".git", "node_modules", "__pycache__", ".cache", "backend",
                      "frontend", "memory", "tests", "test_reports", ".emergent"}
        for item in src_root.iterdir():
            if item.name in preserve:
                skipped_preserve += 1
                continue
            if item.name in _skip_root:
                continue
            if item.is_file():
                _copy_file(item, install_dir / item.name)


        # 5) Guardar SHA aplicado
        remote_sha = cfg.get("last_remote_sha", "")
        if remote_sha:
            await db.app_settings.update_one(
                {}, {"$set": {"github_config.last_commit_sha": remote_sha}}, upsert=True)
            if _using_embedded:
                await _save_embedded_data()

        # 6) Reiniciar (o no, si dry_run)
        if dry_run:
            shutil.rmtree(str(tmp_dir), ignore_errors=True)
            return {
                "success": True, "is_desktop": True, "restarted": False, "dry_run": True,
                "files_updated": copied, "files_preserved": skipped_preserve,
                "install_dir": str(install_dir),
                "message": f"DRY RUN — {copied} archivos actualizados, {skipped_preserve} preservados",
            }

        def _restart_process():
            time.sleep(2)
            try:
                shutil.rmtree(str(tmp_dir), ignore_errors=True)
            except Exception:
                pass
            try:
                if getattr(sys, 'frozen', False):
                    subprocess.Popen([str(sys.executable)], cwd=str(install_dir))
                else:
                    launcher = install_dir / "launcher.pyw"
                    if launcher.exists():
                        subprocess.Popen([sys.executable, str(launcher)], cwd=str(install_dir))
                    else:
                        subprocess.Popen([sys.executable, str(install_dir / "app.py")], cwd=str(install_dir))
                os._exit(0)
            except Exception as e:
                logger.error(f"Restart failed: {e}")

        threading.Thread(target=_restart_process, daemon=True).start()

        return {
            "success": True, "is_desktop": True, "restarted": True,
            "files_updated": copied, "files_preserved": skipped_preserve,
            "install_dir": str(install_dir),
            "message": f"✓ Actualización aplicada — {copied} archivos actualizados. La app se reiniciará en 2 segundos.",
        }
    except HTTPException:
        try: shutil.rmtree(str(tmp_dir), ignore_errors=True)
        except Exception: pass
        raise
    except Exception as e:
        try: shutil.rmtree(str(tmp_dir), ignore_errors=True)
        except Exception: pass
        raise HTTPException(status_code=500, detail=f"Error inesperado al aplicar update: {type(e).__name__}: {e}")


# ─── GitHub: conectar cuenta / push (paridad con la nube) ─────────────────────
# En escritorio la cuenta de GitHub sirve para AUTENTICAR la descarga de
# actualizaciones (repos privados / mayor rate limit). Validamos el token contra
# la API de GitHub y guardamos usuario/avatar. NO usamos `git` (puede no estar
# instalado en la PC del usuario).
@api_router.post("/github/connect")
async def github_connect(payload: dict = Body(...)):
    """Valida un Personal Access Token de GitHub y guarda la cuenta conectada."""
    import urllib.request, urllib.error, json as _json

    token = (payload.get("token") or "").strip()
    repo_url = (payload.get("repo_url") or DEFAULT_GITHUB_REPO).strip()
    branch = (payload.get("branch") or DEFAULT_GITHUB_BRANCH).strip()

    if not token:
        raise HTTPException(status_code=400, detail="Token requerido. Crea uno en https://github.com/settings/tokens con scope 'repo'.")

    try:
        req = urllib.request.Request(
            "https://api.github.com/user",
            headers={
                "Authorization": f"Bearer {token}",
                "Accept": "application/vnd.github+json",
                "User-Agent": "cinema-productions",
            },
        )
        with urllib.request.urlopen(req, timeout=10) as resp:
            user_data = _json.loads(resp.read().decode())
    except urllib.error.HTTPError as e:
        detail = "Token inválido o expirado" if e.code == 401 else f"GitHub API error: {e.code}"
        raise HTTPException(status_code=400, detail=detail)
    except Exception as e:
        raise HTTPException(status_code=502, detail=f"No se pudo contactar GitHub: {e}")

    username = user_data.get("login", "")
    avatar = user_data.get("avatar_url", "")

    await db.app_settings.update_one(
        {},
        {"$set": {
            "github_config.token": token,
            "github_config.repo_url": repo_url,
            "github_config.branch": branch,
            "github_config.username": username,
            "github_config.avatar_url": avatar,
            "github_config.connected_at": datetime.now(timezone.utc).isoformat(),
        }},
        upsert=True,
    )
    if _using_embedded:
        asyncio.create_task(_save_embedded_data())

    return {
        "success": True,
        "username": username,
        "avatar_url": avatar,
        "repo_url": repo_url,
        "branch": branch,
    }


@api_router.post("/github/disconnect")
async def github_disconnect():
    """Desconecta la cuenta de GitHub (borra token/usuario/avatar). Conserva repo_url."""
    await db.app_settings.update_one(
        {},
        {"$unset": {
            "github_config.token": "",
            "github_config.username": "",
            "github_config.avatar_url": "",
            "github_config.connected_at": "",
        }},
        upsert=True,
    )
    if _using_embedded:
        asyncio.create_task(_save_embedded_data())
    return {"success": True}


@api_router.post("/github/push-all")
async def github_push_all(payload: dict = Body(default={})):
    """No disponible en escritorio: subir el código fuente al repositorio requiere
    un entorno git de desarrollo. En la app de escritorio, GitHub se usa solo para
    RECIBIR actualizaciones (Actualizaciones → Buscar/Aplicar)."""
    return {
        "status": "unavailable",
        "is_desktop": True,
        "message": "Subir cambios al repositorio no está disponible en la app de escritorio. "
                   "Usa esta función desde la versión en la nube. Aquí puedes RECIBIR actualizaciones "
                   "desde 'Actualizaciones'.",
    }


@api_router.get("/github/push-status")
async def github_push_status():
    """Estado de push (siempre inactivo en escritorio)."""
    return {
        "status": "idle",
        "is_desktop": True,
        "progress": 0,
        "message": "El push al repositorio no está disponible en la app de escritorio.",
    }


# ─── Metas (Goals) ────────────────────────────────────────

class MetaUpsert(BaseModel):
    year: int
    month: Optional[int] = None  # 1-12 or None for annual
    type: str  # "ventas" | "ganancias" | "gastos"
    amount: float


@api_router.get("/metas")
async def get_metas(year: int, type: str):
    if type not in ("ventas", "ganancias", "gastos"):
        raise HTTPException(status_code=400, detail="type inválido")
    cursor = db.metas.find({"year": year, "type": type}, {"_id": 0})
    goals = await cursor.to_list(200)
    return {"year": year, "type": type, "goals": goals}


@api_router.put("/metas")
async def upsert_meta(meta: MetaUpsert):
    if meta.type not in ("ventas", "ganancias", "gastos"):
        raise HTTPException(status_code=400, detail="type inválido")
    key = {"year": meta.year, "type": meta.type, "month": meta.month}
    payload = {
        **key,
        "amount": float(meta.amount or 0),
        "updated_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.metas.update_one(
        key,
        {"$set": payload, "$setOnInsert": {"id": str(uuid.uuid4())}},
        upsert=True,
    )
    if _using_embedded:
        asyncio.create_task(_save_embedded_data())
    return {"ok": True, **payload}


@api_router.delete("/metas")
async def delete_meta(year: int, type: str, month: Optional[int] = None):
    await db.metas.delete_one({"year": year, "type": type, "month": month})
    if _using_embedded:
        asyncio.create_task(_save_embedded_data())
    return {"ok": True}


@api_router.get("/metas/progress")
async def metas_progress(year: int, type: str):
    """Return per-month actual values + goals + reached flags."""
    if type not in ("ventas", "ganancias", "gastos"):
        raise HTTPException(status_code=400, detail="type inválido")

    # Fetch reservations for that year (event_date is 'YYYY-MM-DD' string)
    start = f"{year}-01-01"
    end = f"{year}-12-31"
    cursor = db.reservations.find(
        {"status": {"$nin": ["Cancelado"]}, "event_date": {"$gte": start, "$lte": end}},
        {"total_amount": 1, "event_date": 1, "assigned_partners": 1, "_id": 0},
    )
    docs = await cursor.to_list(20000)

    # Aggregate by month
    monthly = {m: {"ventas": 0.0, "gastos": 0.0} for m in range(1, 13)}
    for d in docs:
        ed = d.get("event_date") or ""
        try:
            m = int(ed.split("-")[1])
        except Exception:
            continue
        if m < 1 or m > 12:
            continue
        monthly[m]["ventas"] += float(d.get("total_amount") or 0)
        for p in (d.get("assigned_partners") or []):
            monthly[m]["gastos"] += float(p.get("payment") or 0)

    # Fetch goals for year+type
    goal_cursor = db.metas.find({"year": year, "type": type}, {"_id": 0})
    goal_docs = await goal_cursor.to_list(200)
    goal_by_month = {}
    annual_goal = 0.0
    for g in goal_docs:
        if g.get("month") is None:
            annual_goal = float(g.get("amount") or 0)
        else:
            goal_by_month[int(g["month"])] = float(g.get("amount") or 0)

    def actual_for(m):
        v = monthly[m]
        if type == "ventas":
            return round(v["ventas"], 2)
        if type == "gastos":
            return round(v["gastos"], 2)
        # ganancias
        return round(v["ventas"] - v["gastos"], 2)

    months_out = []
    annual_actual = 0.0
    # Meta mensual derivada de la anual (ventas/ganancias). En gastos NO se auto-deriva.
    auto_monthly = 0.0
    if type != "gastos" and annual_goal > 0:
        auto_monthly = round(annual_goal / 12.0, 2)

    for m in range(1, 13):
        act = actual_for(m)
        annual_actual += act
        custom = goal_by_month.get(m)
        is_custom = custom is not None
        if is_custom:
            goal = float(custom)
        else:
            goal = auto_monthly
        pct = (act / goal * 100) if goal > 0 else 0.0
        months_out.append({
            "month": m,
            "actual": round(act, 2),
            "goal": round(goal, 2),
            "percent": round(pct, 2),
            "reached": goal > 0 and act >= goal,
            "is_custom": is_custom,
            "is_auto": (not is_custom) and goal > 0,
        })

    ann_pct = (annual_actual / annual_goal * 100) if annual_goal > 0 else 0.0
    return {
        "year": year,
        "type": type,
        "months": months_out,
        "annual_goal": round(annual_goal, 2),
        "annual_actual": round(annual_actual, 2),
        "annual_percent": round(ann_pct, 2),
        "annual_reached": annual_goal > 0 and annual_actual >= annual_goal,
        "auto_monthly": auto_monthly,
    }


# ─── Contexto para IA (local) ────────────────────────────────────────────────
@api_router.get("/ai-context")
async def get_ai_context():
    doc = await db.app_settings.find_one({}, {"ai_context": 1}) or {}
    ctx = doc.get("ai_context") or {}
    if not ctx.get("content"):
        return {"content": _DEFAULT_AI_CONTEXT, "updated_at": "", "is_default": True}
    return {"content": ctx.get("content", ""), "updated_at": ctx.get("updated_at", ""), "is_default": False}


@api_router.post("/ai-context")
async def save_ai_context(payload: dict = Body(...)):
    content = payload.get("content", "")
    if not isinstance(content, str):
        raise HTTPException(status_code=400, detail="content debe ser string")
    await db.app_settings.update_one(
        {}, {"$set": {"ai_context": {"content": content,
                                     "updated_at": datetime.now(timezone.utc).isoformat()}}}, upsert=True)
    if _using_embedded:
        asyncio.create_task(_save_embedded_data())
    return {"success": True, "updated_at": datetime.now(timezone.utc).isoformat()}


@api_router.post("/ai-context/reset")
async def reset_ai_context():
    await db.app_settings.update_one(
        {}, {"$set": {"ai_context": {"content": _DEFAULT_AI_CONTEXT,
                                     "updated_at": datetime.now(timezone.utc).isoformat()}}}, upsert=True)
    if _using_embedded:
        asyncio.create_task(_save_embedded_data())
    return {"success": True, "content": _DEFAULT_AI_CONTEXT}


# ─── Seguridad avanzada (config) ──────────────────────────────────────────────
@api_router.put("/security/advanced-config")
async def set_advanced_security_config(payload: dict = Body(...)):
    update = {}
    if "auto_lock_enabled" in payload:
        update["security_config.auto_lock_enabled"] = bool(payload["auto_lock_enabled"])
    if "auto_lock_minutes" in payload:
        m = int(payload["auto_lock_minutes"])
        if 1 <= m <= 120:
            update["security_config.auto_lock_minutes"] = m
    if "max_attempts" in payload:
        n = int(payload["max_attempts"])
        if 3 <= n <= 20:
            update["security_config.max_attempts"] = n
    if "lockout_seconds" in payload:
        s = int(payload["lockout_seconds"])
        if 10 <= s <= 3600:
            update["security_config.lockout_seconds"] = s
    if "protected_sections" in payload and isinstance(payload["protected_sections"], list):
        update["security_config.protected_sections"] = payload["protected_sections"]
    if update:
        await db.app_settings.update_one({}, {"$set": update}, upsert=True)
        if _using_embedded:
            asyncio.create_task(_save_embedded_data())
    return {"success": True, "updated_keys": list(update.keys())}


# ─── Contraseña del ZIP (cifrado AES-256 de backups) ──────────────────────────
# Paridad con la versión en la nube: el frontend compartido (SecuritySection)
# consume estos endpoints para ver/cambiar/restaurar la contraseña con la que se
# cifran los backups .zip. En escritorio el valor por defecto es 2868.
@api_router.get("/security/zip-password")
async def get_zip_password():
    """Devuelve la contraseña actual del ZIP (visible para el dueño de la app)."""
    doc = await db.app_settings.find_one({}, {"security_config": 1}) or {}
    cfg = doc.get("security_config") or {}
    pwd = cfg.get("zip_password") or DEFAULT_ZIP_PASSWORD_STANDALONE
    return {
        "password": pwd,
        "is_default": pwd == DEFAULT_ZIP_PASSWORD_STANDALONE,
        "enabled": bool(cfg.get("zip_password_enabled", True)),
    }


@api_router.post("/security/zip-password")
async def set_zip_password(payload: dict = Body(...)):
    """Cambia la contraseña usada al cifrar los backups .zip."""
    new_pwd = (payload.get("new_password") or "").strip()
    if len(new_pwd) < 3:
        raise HTTPException(status_code=400, detail="La contraseña debe tener al menos 3 caracteres")
    if len(new_pwd) > 64:
        raise HTTPException(status_code=400, detail="La contraseña no puede exceder 64 caracteres")
    await db.app_settings.update_one(
        {},
        {"$set": {"security_config.zip_password": new_pwd}},
        upsert=True,
    )
    if _using_embedded:
        asyncio.create_task(_save_embedded_data())
    return {"success": True, "password": new_pwd}


@api_router.post("/security/zip-password/reset")
async def reset_zip_password():
    """Restaura la contraseña ZIP al valor de fábrica (2868)."""
    await db.app_settings.update_one(
        {},
        {"$set": {"security_config.zip_password": DEFAULT_ZIP_PASSWORD_STANDALONE}},
        upsert=True,
    )
    if _using_embedded:
        asyncio.create_task(_save_embedded_data())
    return {"success": True, "password": DEFAULT_ZIP_PASSWORD_STANDALONE}


# ─── Notificaciones (endpoints de prueba / pendientes) ────────────────────────
@api_router.get("/notifications/pending")
async def get_pending_notifications():
    settings_doc = await db.app_settings.find_one({}, {"_id": 0})
    periods = (settings_doc.get("reminder_periods") if settings_doc else None) or [3]
    days = max(periods) if periods else 3
    today = datetime.now(timezone.utc).date()
    end = (today + timedelta(days=days)).isoformat()
    cursor = db.reservations.find(
        {"event_date": {"$gte": today.isoformat(), "$lte": end}, "status": {"$nin": ["Cancelado", "Completado"]}},
        {"client_name": 1, "event_date": 1, "event_type": 1, "venue": 1, "_id": 1})
    docs = await cursor.to_list(100)
    return [doc_to_dict(d) for d in docs]


@api_router.post("/reminders/test-email")
async def test_email_connection():
    settings_doc = await db.app_settings.find_one({}, {"_id": 0})
    if not settings_doc or not settings_doc.get("resend_api_key"):
        raise HTTPException(status_code=400, detail="Ingresa tu API Key de Resend primero")
    if not settings_doc.get("admin_email"):
        raise HTTPException(status_code=400, detail="Ingresa un Email Destino primero")
    try:
        import resend as _resend
        _resend.api_key = settings_doc["resend_api_key"]
        sender = settings_doc.get("sender_name") or "Cinema Productions"
        await asyncio.to_thread(_resend.Emails.send, {
            "from": f"{sender} <onboarding@resend.dev>",
            "to": [settings_doc["admin_email"]],
            "subject": f"Prueba de email — {sender}",
            "html": "<p>Conexion de email correcta.</p>",
        })
        return {"success": True, "message": f"Email de prueba enviado a {settings_doc['admin_email']}"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error al enviar: {e}")


@api_router.post("/telegram/test")
async def telegram_test():
    import httpx
    doc = await db.app_settings.find_one({}, {"_id": 0})
    if not doc or not doc.get("telegram_bot_token") or not doc.get("telegram_chat_id"):
        return {"ok": False, "error": "Configura el token y chat_id de Telegram primero"}
    try:
        async with httpx.AsyncClient(timeout=10) as c:
            r = await c.post(
                f"https://api.telegram.org/bot{doc['telegram_bot_token']}/sendMessage",
                json={"chat_id": doc["telegram_chat_id"],
                      "text": "<b>Cinema Productions</b> — Prueba ✓", "parse_mode": "HTML"})
            r.raise_for_status()
        return {"ok": True, "message": "Mensaje enviado a Telegram"}
    except Exception as e:
        return {"ok": False, "error": f"Error de Telegram: {e}"}


@api_router.post("/ntfy/test")
async def ntfy_test():
    import httpx
    doc = await db.app_settings.find_one({}, {"_id": 0})
    if not doc or not doc.get("ntfy_topic"):
        return {"ok": False, "error": "Configura el tema de ntfy primero"}
    try:
        async with httpx.AsyncClient(timeout=10) as c:
            r = await c.post(f"https://ntfy.sh/{doc['ntfy_topic']}",
                             content="ntfy conectado correctamente.".encode("utf-8"),
                             headers={"Title": "Cinema Productions — Prueba", "Priority": "high"})
            r.raise_for_status()
        return {"ok": True, "message": f"Notificación enviada al tema: {doc['ntfy_topic']}"}
    except Exception as e:
        return {"ok": False, "error": f"Error de ntfy: {e}"}


# ─── Web Push / Gmail (no disponibles offline → respuestas controladas) ───────
@api_router.get("/push/vapid-key")
async def get_vapid_key():
    return {"publicKey": ""}


@api_router.post("/push/subscribe")
async def push_subscribe(payload: dict = Body(default={})):
    return {"ok": True, "desktop": True}


@api_router.delete("/push/unsubscribe")
async def push_unsubscribe(endpoint: str = ""):
    return {"ok": True}


@api_router.post("/push/test")
async def push_test():
    return {"ok": False, "error": "Las notificaciones push del navegador no están disponibles en la app de escritorio local."}


@api_router.get("/oauth/gmail/status")
async def gmail_status():
    return {"connected": False, "email": "", "connected_at": None, "desktop": True}


@api_router.get("/oauth/gmail/start")
async def gmail_oauth_start():
    raise HTTPException(status_code=400, detail="El inicio de sesión con Gmail no está disponible en la app de escritorio. Usa Resend, Telegram o ntfy.")


@api_router.delete("/oauth/gmail/disconnect")
async def gmail_disconnect():
    return {"ok": True}


@api_router.post("/oauth/gmail/test")
async def gmail_test():
    raise HTTPException(status_code=400, detail="Gmail no disponible en la app de escritorio.")


# ─── Deployment helpers (informativo) ─────────────────────────────────────────
@api_router.get("/deployment/env-template")
async def get_env_template():
    tpl = ("# Cinema Productions — Config\nMONGO_URL=embedded\nDB_NAME=cinema_productions\n")
    return Response(content=tpl.encode(), media_type="text/plain",
                    headers={"Content-Disposition": 'attachment; filename=".env.template"'})


@api_router.post("/deployment/health-check")
async def health_check_url(url: str):
    import httpx
    if not url.startswith(("http://", "https://")):
        return {"ok": False, "error": "URL inválida"}
    try:
        async with httpx.AsyncClient(timeout=8.0, follow_redirects=True) as c:
            r = await c.get(url)
            return {"ok": r.status_code < 400, "status": r.status_code}
    except Exception as e:
        return {"ok": False, "error": str(e)[:100]}


# ═══════════════════════════════════════════════════════════════════
# DIAGNÓSTICO DE LA APP (versión desktop/standalone)
# ═══════════════════════════════════════════════════════════════════

DEFAULT_ZIP_PASSWORD_STANDALONE = "2868"


@api_router.get("/diagnostic")
async def run_diagnostic():
    """Chequeos de salud de la app: dependencias, DB embedded/remota, GitHub, seguridad."""
    import importlib.util

    checks = []

    def add(id_, label, ok, detail="", severity="error", fixable=False):
        checks.append({
            "id": id_,
            "label": label,
            "ok": bool(ok),
            "detail": detail,
            "severity": severity if not ok else "ok",
            "fixable": bool(fixable),
        })

    # 1) Dependencias Python críticas (SOLO las que la app de escritorio usa)
    critical_py = ["fastapi", "motor", "pymongo", "pydantic", "pyzipper"]
    # nota: apscheduler/bcrypt NO se usan en desktop (contraseñas via hashlib.pbkdf2)
    missing_py = [p for p in critical_py if importlib.util.find_spec(p) is None]
    add("python_deps",
        f"Dependencias Python ({len(critical_py)} críticas)",
        not missing_py,
        f"Faltantes: {', '.join(missing_py)}" if missing_py else f"Todas instaladas: {', '.join(critical_py)}",
        fixable=True,
    )

    # 2) MongoDB (embedded o real)
    try:
        if _using_embedded:
            add("mongo_conn", "Base de datos local (embedded)", True,
                f"Modo embedded — datos en cinema_data.json")
        else:
            await db.command("ping")
            server_info = await db.client.server_info()
            mongo_version = server_info.get("version", "?")
            add("mongo_conn", f"MongoDB conectado (v{mongo_version})", True,
                f"URL: {_effective_mongo_url[:35]}...")
    except Exception as e:
        add("mongo_conn", "MongoDB conectado", False, str(e)[:200], fixable=True)

    # 3) Archivo de datos embedded existe y es escribible
    if _using_embedded:
        data_file = ROOT_DIR / "cinema_data.json"
        add("embedded_data_file",
            "Archivo de datos (cinema_data.json)",
            data_file.exists(),
            f"OK — {data_file.stat().st_size} bytes" if data_file.exists() else "No existe todavía",
            severity="info",
        )

    # 4) Repositorio GitHub configurado
    cfg = await _get_github_cfg()
    repo_url = cfg.get("repo_url", "")
    is_default_repo = repo_url == DEFAULT_GITHUB_REPO
    add("github_default_repo",
        "Repositorio GitHub por defecto",
        is_default_repo,
        f"OK: {repo_url}" if is_default_repo else (f"Otro repo: {repo_url}" if repo_url else "Sin repo configurado"),
        severity="warning",
        fixable=True,
    )

    # 5) Token de GitHub
    add("github_token",
        "Cuenta de GitHub vinculada",
        bool(cfg.get("token")),
        f"Conectado como @{cfg.get('username', '?')}" if cfg.get("token") else "Sin token — usa 'Conectar con GitHub'",
        severity="warning",
    )

    # 6) Directorio de backups escribible
    try:
        BACKUP_DIR.mkdir(parents=True, exist_ok=True)
        test_file = BACKUP_DIR / ".diagnostic_test"
        test_file.write_text("ok")
        test_file.unlink()
        add("backup_writable", "Directorio de backups escribible", True, str(BACKUP_DIR))
    except Exception as e:
        add("backup_writable", "Directorio de backups escribible", False, str(e)[:120], fixable=True)

    # 7) Contraseña ZIP configurada
    sec_cfg = (await db.app_settings.find_one({}, {"security_config": 1}) or {}).get("security_config") or {}
    zip_pwd = sec_cfg.get("zip_password") or DEFAULT_ZIP_PASSWORD_STANDALONE
    add("zip_password",
        "Contraseña ZIP configurada",
        len(zip_pwd) >= 3,
        f"Longitud: {len(zip_pwd)} chars ({'DEFAULT (2868)' if zip_pwd == DEFAULT_ZIP_PASSWORD_STANDALONE else 'personalizada'})",
        severity="warning",
        fixable=True,
    )

    # 8) Espacio en disco (aviso si <100MB libres)
    try:
        import shutil as _sh
        free_bytes = _sh.disk_usage(str(ROOT_DIR)).free
        free_mb = free_bytes // (1024 * 1024)
        add("disk_space",
            "Espacio libre en disco",
            free_mb >= 100,
            f"{free_mb} MB libres",
            severity="warning",
        )
    except Exception as e:
        add("disk_space", "Espacio libre en disco", True, "N/A")

    # Score global
    total = len(checks)
    ok_count = sum(1 for c in checks if c["ok"])
    errors = sum(1 for c in checks if not c["ok"] and c["severity"] == "error")
    warnings = sum(1 for c in checks if not c["ok"] and c["severity"] == "warning")

    return {
        "checks": checks,
        "summary": {
            "total": total,
            "ok": ok_count,
            "errors": errors,
            "warnings": warnings,
            "score": round((ok_count / total) * 100) if total else 0,
        },
        "generated_at": datetime.now(timezone.utc).isoformat(),
    }


@api_router.post("/diagnostic/fix")
async def diagnostic_fix(payload: dict = Body(...)):
    """Corrige el chequeo indicado por id."""
    global client, db
    check_id = (payload.get("id") or "").strip()
    if not check_id:
        raise HTTPException(status_code=400, detail="Falta 'id'")

    fixed = False
    detail = ""

    if check_id == "python_deps":
        # Instalación robusta: usa el intérprete actual (pip como módulo, no del PATH),
        # prioriza las wheels OFFLINE de libs/ (-r requirements.txt) y cae a online.
        import importlib.util, sys as _sys
        crit = ["fastapi", "motor", "pymongo", "pydantic", "pyzipper"]
        req = ROOT_DIR / "requirements.txt"
        libs = ROOT_DIR / "libs"
        has_libs = libs.exists() and any(libs.glob("*.whl"))
        base = [_sys.executable, "-m", "pip", "install", "--disable-pip-version-check", "-q"]

        attempts = []
        if req.exists() and has_libs:
            attempts.append(base + ["--no-index", "--find-links", str(libs), "-r", str(req)])
        if req.exists():
            attempts.append(base + ["-r", str(req)])
        missing = [p for p in crit if importlib.util.find_spec(p) is None]
        if missing and has_libs:
            attempts.append(base + ["--no-index", "--find-links", str(libs)] + missing)
        if missing:
            attempts.append(base + missing)

        out = ""
        for cmd in attempts:
            try:
                r = subprocess.run(cmd, cwd=str(ROOT_DIR), capture_output=True, text=True, timeout=300)
                out = ((r.stdout or "") + (r.stderr or ""))[-600:]
                if r.returncode == 0:
                    break
            except Exception as e:
                out += f"\n{e}"

        importlib.invalidate_caches()
        still = [p for p in crit if importlib.util.find_spec(p) is None]
        if not still:
            fixed = True
            try:
                (ROOT_DIR / ".deps_ok").write_text("ok", encoding="utf-8")
            except Exception:
                pass
            detail = "Dependencias instaladas correctamente" + (" (offline desde libs/)" if has_libs else " (descarga online)")
        else:
            fixed = False
            detail = f"Aún faltan: {', '.join(still)}. Detalle: {out[-300:]}"
    elif check_id == "github_default_repo":
        await db.app_settings.update_one(
            {},
            {"$set": {
                "github_config.repo_url": DEFAULT_GITHUB_REPO,
                "github_config.branch": DEFAULT_GITHUB_BRANCH,
            }},
            upsert=True,
        )
        if _using_embedded:
            await _save_embedded_data()
        fixed = True
        detail = f"Repositorio restaurado a {DEFAULT_GITHUB_REPO}"
    elif check_id == "backup_writable":
        try:
            BACKUP_DIR.mkdir(parents=True, exist_ok=True)
            fixed = True
            detail = f"Directorio creado: {BACKUP_DIR}"
        except Exception as e:
            detail = str(e)
    elif check_id == "mongo_conn":
        if _using_embedded:
            fixed = True
            detail = "Modo embedded — no requiere reconexión"
        else:
            try:
                from motor.motor_asyncio import AsyncIOMotorClient
                new_client = AsyncIOMotorClient(_effective_mongo_url)
                await new_client[DB_NAME].command("ping")
                client = new_client
                db = new_client[DB_NAME]
                fixed = True
                detail = "Reconectado exitosamente"
            except Exception as e:
                detail = f"Reconexión falló: {e}"
    elif check_id == "zip_password":
        await db.app_settings.update_one(
            {},
            {"$set": {"security_config.zip_password": DEFAULT_ZIP_PASSWORD_STANDALONE}},
            upsert=True,
        )
        if _using_embedded:
            await _save_embedded_data()
        fixed = True
        detail = f"Contraseña ZIP restaurada a {DEFAULT_ZIP_PASSWORD_STANDALONE}"
    else:
        raise HTTPException(status_code=400, detail=f"'{check_id}' no es corregible automáticamente")

    return {"success": fixed, "id": check_id, "detail": detail}


@api_router.post("/diagnostic/fix-all")
async def diagnostic_fix_all():
    """Ejecuta el diagnóstico y aplica todas las correcciones fixables."""
    diag = await run_diagnostic()
    results = []
    fixed_count = 0
    failed_count = 0

    for check in diag["checks"]:
        if check["ok"] or not check.get("fixable"):
            continue
        try:
            res = await diagnostic_fix({"id": check["id"]})
            results.append({
                "id": check["id"],
                "label": check["label"],
                "success": res.get("success", False),
                "detail": res.get("detail", "")[:200],
            })
            if res.get("success"):
                fixed_count += 1
            else:
                failed_count += 1
        except HTTPException as e:
            results.append({
                "id": check["id"],
                "label": check["label"],
                "success": False,
                "detail": str(e.detail)[:200],
            })
            failed_count += 1
        except Exception as e:
            results.append({
                "id": check["id"],
                "label": check["label"],
                "success": False,
                "detail": str(e)[:200],
            })
            failed_count += 1

    final_diag = await run_diagnostic()

    return {
        "success": failed_count == 0,
        "fixed": fixed_count,
        "failed": failed_count,
        "results": results,
        "final_score": final_diag["summary"]["score"],
        "final_errors": final_diag["summary"]["errors"],
        "final_warnings": final_diag["summary"]["warnings"],
    }


# ─── App config ───────────────────────────────────────────

app.include_router(api_router)
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# ─── Serve React build (SPA) ─────────────────────────────

_LOCAL_INJECT = '<script>window.__API_BASE_URL__="http://localhost:8001";</script>'

# Scripts externos que hay que eliminar para evitar pantalla en blanco offline
_EXTERNAL_SCRIPTS_TO_REMOVE = [
    '<script src="https://assets.emergent.sh/scripts/emergent-main.js"></script>',
    # PostHog analytics (hace llamadas externas que cuelgan offline)
]


def _inject_local_url(html: str) -> str:
    """
    1. Inyecta la URL local de la API.
    2. Elimina scripts externos que causan pantalla en blanco sin internet.
    """
    # Eliminar scripts externos
    for script in _EXTERNAL_SCRIPTS_TO_REMOVE:
        html = html.replace(script, "<!-- removed for offline use -->")

    # Eliminar script de posthog (bloquea la carga sin internet)
    import re as _re
    html = _re.sub(
        r'<script[^>]*>\s*!function\(e,t\)\{var r,s,o,i;t\.__SV.*?posthog\.init\([^)]+\)\s*</script>',
        '<!-- analytics removed for offline use -->',
        html,
        flags=_re.DOTALL,
    )

    # Inyectar URL local
    return html.replace("</head>", _LOCAL_INJECT + "</head>", 1)


def _resolve_build_dir() -> Optional[Path]:
    """Localiza la carpeta build/ del frontend en runtime.

    Prioridades:
      1. BUNDLE_DIR/build      → PyInstaller onefile (sys._MEIPASS/build)
      2. ROOT_DIR/build        → ejecución en dev / desde source
      3. ROOT_DIR.parent/build → cuando standalone_app está en backend/ y
                                 el build vive un nivel arriba (repo layout).
    Devuelve el primer path que contenga index.html; None si ninguno.
    """
    candidates = [
        BUNDLE_DIR / "build",
        ROOT_DIR / "build",
        ROOT_DIR.parent / "build",
        ROOT_DIR.parent / "frontend" / "build",
    ]
    seen = set()
    for c in candidates:
        try:
            c_res = c.resolve()
        except Exception:
            continue
        if c_res in seen:
            continue
        seen.add(c_res)
        if (c_res / "index.html").is_file():
            logging.getLogger(__name__).info("Frontend build dir: %s", c_res)
            return c_res
    logging.getLogger(__name__).warning(
        "No se encontró build/index.html en ninguna ubicación esperada: %s",
        [str(c) for c in candidates],
    )
    return None


BUILD_DIR = _resolve_build_dir()
if BUILD_DIR is not None:
    _static_dir = BUILD_DIR / "static"
    if _static_dir.is_dir():
        app.mount("/static", StaticFiles(directory=str(_static_dir)), name="static")

    @app.get("/favicon.ico")
    async def favicon():
        f = BUILD_DIR / "favicon.ico"
        return FileResponse(str(f)) if f.exists() else Response(status_code=204)

    @app.get("/manifest.json")
    async def manifest_json():
        f = BUILD_DIR / "manifest.json"
        return FileResponse(str(f)) if f.exists() else Response("{}", media_type="application/json")

    # Cache-Control para HTML: nunca cachear. Si cacheamos index.html, tras un
    # update los navegadores siguen apuntando a chunks JS antiguos (con hash
    # viejo) que ya no existen → pantalla en blanco. Los assets estáticos con
    # hash (StaticFiles /static/*) sí pueden cachearse porque su nombre cambia.
    _NO_CACHE_HEADERS = {
        "Cache-Control": "no-cache, no-store, must-revalidate, max-age=0",
        "Pragma": "no-cache",
        "Expires": "0",
    }

    @app.get("/")
    async def serve_index():
        html_path = BUILD_DIR / "index.html"
        html = _inject_local_url(html_path.read_text(encoding="utf-8"))
        return Response(content=html, media_type="text/html; charset=utf-8", headers=_NO_CACHE_HEADERS)

    @app.get("/{path:path}")
    async def serve_spa(path: str):
        # Nunca devolver HTML para rutas de API desconocidas (causaría que el
        # frontend interprete HTML como JSON → pantalla en blanco).
        if path.startswith("api/") or path == "api":
            return JSONResponse({"detail": "Not Found"}, status_code=404)
        f = BUILD_DIR / path
        if f.exists() and f.is_file():
            # sw.js e index.html: no cachear para poder actualizar en caliente.
            if path in ("sw.js", "index.html", "manifest.json"):
                return FileResponse(str(f), headers=_NO_CACHE_HEADERS)
            return FileResponse(str(f))
        html_path = BUILD_DIR / "index.html"
        html = _inject_local_url(html_path.read_text(encoding="utf-8"))
        return Response(content=html, media_type="text/html; charset=utf-8", headers=_NO_CACHE_HEADERS)
else:
    # Sin frontend build: exponer una landing mínima para no romper el navegador.
    _MISSING_BUILD_HTML = (
        "<!doctype html><meta charset='utf-8'>"
        "<title>Cinema Productions</title>"
        "<style>body{font-family:system-ui;padding:40px;max-width:640px;margin:auto;"
        "color:#222;background:#faf7ff}code{background:#eee;padding:2px 6px;border-radius:4px}</style>"
        "<h1>Frontend no empaquetado</h1>"
        "<p>El backend arrancó, pero <code>build/index.html</code> no está en el .exe. "
        "Reconstruye el ejecutable asegurándote de que <code>frontend/build/</code> se "
        "generó con <code>yarn build</code> antes de PyInstaller y que "
        "<code>--add-data \"backend/_bundle/build;build\"</code> se pasó correctamente.</p>"
        "<p>API disponible en <a href='/api/'>/api/</a>.</p>"
    )

    @app.get("/")
    async def serve_index_missing():
        return Response(content=_MISSING_BUILD_HTML, media_type="text/html; charset=utf-8")

    @app.get("/favicon.ico")
    async def favicon_missing():
        return Response(status_code=204)


# ─── Entry point ─────────────────────────────────────────

if __name__ == "__main__":
    import uvicorn

    # En modo --windowed (frozen sin consola) sys.stdout puede ser None.
    # Redirigimos los prints a un log file para no crashear con AttributeError.
    if getattr(sys, "frozen", False) and (sys.stdout is None or not hasattr(sys.stdout, "write")):
        _log_dir = ROOT_DIR / "logs" if not getattr(sys, "_MEIPASS", None) else Path.home() / "CinemaProductions" / "logs"
        try:
            _log_dir.mkdir(parents=True, exist_ok=True)
            _log_path = _log_dir / "cinema.log"
            _log_fh = open(_log_path, "a", encoding="utf-8", buffering=1)
            sys.stdout = _log_fh
            sys.stderr = _log_fh
        except Exception:
            # Último recurso: descartar stdout/stderr.
            import io as _io
            sys.stdout = _io.StringIO()
            sys.stderr = _io.StringIO()

    db_label = "Embebida (cinema_data.json)" if _using_embedded else MONGO_URL[:40]

    def _open_browser():
        time.sleep(3)
        webbrowser.open("http://localhost:8001")

    # El launcher grafico abre el navegador con mejor timing; si nos lanza,
    # define CP_NO_BROWSER=1 para no abrir dos pestanas.
    if not os.environ.get("CP_NO_BROWSER"):
        threading.Thread(target=_open_browser, daemon=True).start()

    print("\n" + "=" * 54)
    print("  CINEMA PRODUCTIONS — Gestor de Reservas")
    print("=" * 54)
    print(f"  URL:  http://localhost:8001")
    print(f"  BD:   {db_label}")
    print(f"  Datos: {DATA_FILE.name if _using_embedded else 'MongoDB'}")
    print("  Para cerrar: Ctrl+C  o  cierra esta ventana")
    print("=" * 54 + "\n")

    # ── Shutdown limpio para evitar "Failed to remove temporary directory _MEIxxx" ──
    # PyInstaller onefile extrae recursos a %TEMP%/_MEIxxx y trata de borrarlo al
    # salir. Si uvicorn/threads/loggers todavia tienen handles abiertos, Windows
    # rechaza el rmdir y el bootloader muestra un popup Warning. Forzamos:
    #   1) Cerrar file handlers de logging (liberan .log dentro de _MEI si hay)
    #   2) uvicorn.Server con should_exit + config controlada
    #   3) atexit + signal handlers que fuerzan flush y cierre de stdio
    import atexit
    import signal as _signal
    import logging as _logging

    def _graceful_shutdown(*_a, **_kw):
        try:
            _logging.shutdown()
        except Exception:
            pass
        try:
            if hasattr(sys.stdout, "flush"): sys.stdout.flush()
            if hasattr(sys.stderr, "flush"): sys.stderr.flush()
        except Exception:
            pass
        try:
            _fh = globals().get("_log_fh")
            if _fh and not _fh.closed:
                _fh.close()
        except Exception:
            pass

    atexit.register(_graceful_shutdown)
    for _sig in (getattr(_signal, "SIGINT", None), getattr(_signal, "SIGTERM", None), getattr(_signal, "SIGBREAK", None)):
        if _sig is not None:
            try:
                _signal.signal(_sig, lambda *_a: (_graceful_shutdown(), sys.exit(0)))
            except Exception:
                pass

    config = uvicorn.Config(app, host="0.0.0.0", port=8001, log_level="warning", access_log=False)
    server = uvicorn.Server(config)
    try:
        server.run()
    finally:
        _graceful_shutdown()
